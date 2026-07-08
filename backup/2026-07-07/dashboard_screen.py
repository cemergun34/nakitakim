"""
Dashboard ekranı — PHP admin_panel.php / admin.php'nin PyQt6 karşılığı.
12 KPI kartı + Yıllık Gelir-Gider grafik.
"""
from __future__ import annotations
from datetime import datetime
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QScrollArea,
    QGridLayout, QFrame, QPushButton, QComboBox, QSizePolicy,
    QDateEdit, QSpacerItem, QDialog, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QDate
from PyQt6.QtGui import QFont

from ui.theme import COLORS, FONTS
from ui.components.kpi_card import KPICard
from ui.components.detay_dialog import DetayDialog
from ui.components.gelir_gider_chart import GelirGiderChart
from services.dashboard_service import get_all_dashboard_data
from services import detay_service
from utils.format import fmt_para, fmt_signed


class PaytrKPICard(KPICard):
    """
    PHP admin.php #btnSanalPosHareketleri kartının PyQt6 karşılığı.

    Layout (PHP HTML birebir):
      H3  : 'Sanal Pos Paytr'  (üst sol)
      div : Fark değeri  (balance — yeşil badge)
      row : [İşlem  |  Ödeme]  (iki sütun)
      son : 'Son güncelleme: DD.MM.YYYY'  (alt sağ opak)
    """

    def __init__(self, click_cb=None, parent=None):
        super().__init__(
            title="Sanal Pos Paytr",
            value="Yükleniyor...",
            color="#212121",
            color2="#000000",
            click_cb=click_cb,
            parent=parent,
        )
        # setFixedHeight üst sınıftan geliyor; biraz daha yüksek yap
        self.setFixedHeight(160)

        # Üst sınıfın 'value_lbl' Fark değerini gösterir (yeşil renk)
        self.value_lbl.setStyleSheet(
            "color: rgb(170,255,204); font-size: 20px; font-weight: 700;"
            " background: transparent; letter-spacing: -0.5px;"
        )
        # Title rengi beyaz
        self.title_lbl.setStyleSheet(
            "color: #ffffff; font-size: 12px; font-weight: 600;"
            " background: transparent;"
        )

        # Ekstra satır: [İşlem | Ödeme]
        from PyQt6.QtWidgets import QHBoxLayout, QVBoxLayout, QLabel
        row_w = QHBoxLayout()
        row_w.setSpacing(4)

        def _blok(key: str, lbl: str, clr: str) -> QLabel:
            col = QVBoxLayout()
            col.setSpacing(0)
            hdr = QLabel(lbl)
            hdr.setAlignment(Qt.AlignmentFlag.AlignCenter)
            hdr.setStyleSheet(
                "color:rgba(255,255,255,.70);font-size:11px;background:transparent;"
            )
            val = QLabel("-")
            val.setAlignment(Qt.AlignmentFlag.AlignCenter)
            val.setStyleSheet(
                f"color:{clr};font-size:12px;font-weight:600;background:transparent;"
            )
            col.addWidget(hdr)
            col.addWidget(val)
            setattr(self, key, val)
            row_w.addLayout(col)
            return val

        _blok("_islem_lbl", "İşlem",  "#ffe0a0")
        _blok("_odeme_lbl", "Ödeme",  "#d0ffd0")

        # Son güncelleme etiketi
        self._son_guncelleme_lbl = QLabel("")
        self._son_guncelleme_lbl.setStyleSheet(
            "color:rgba(255,255,255,.80);font-size:10px;background:transparent;"
        )

        # Mevcut layout'a ekle
        lay = self.layout()
        lay.addLayout(row_w)
        lay.addWidget(self._son_guncelleme_lbl)

    def set_paytr(
        self,
        fark_fmt: str,
        islem_fmt: str,
        odeme_fmt: str,
        son_guncelleme: str = "",
    ):
        """
        PHP'deki:
          #paytrToplamBadge   → fark_fmt
          #paytrDashIslem     → islem_fmt
          #paytrDashOdeme     → odeme_fmt
          #paytrSonGuncelleme → son_guncelleme
        """
        self.value_lbl.setText(fark_fmt)
        self._islem_lbl.setText(islem_fmt)
        self._odeme_lbl.setText(odeme_fmt)
        self._son_guncelleme_lbl.setText(son_guncelleme)


class FizikselPosKPICard(KPICard):
    """
    PHP admin.php #btnFizikselPosHareketleri kartının PyQt6 karşılığı.

    Layout:
      H3  : 'Fiziksel Pos Womsis'
      div : Net Tutar  (balance — mint yeşil)
      row : [İşlem Tutarı | Net Tutar]  (sarı / yeşil)
      son : 'Yerel tablodan anlık veriler'

    Renk: #1a3a5c → #0d2137 (PHP: linear-gradient(135deg, #1a3a5c, #0d2137))
    """

    def __init__(self, click_cb=None, parent=None):
        super().__init__(
            title="Fiziksel Pos Womsis",
            value="₺0,00",
            color="#1a3a5c",
            color2="#0d2137",
            click_cb=click_cb,
            parent=parent,
        )
        self.setFixedHeight(160)

        # Net Tutar badge — PHP: #womsisToplamBadge  color: rgb(170,255,204)
        self.value_lbl.setStyleSheet(
            "color: rgb(170,255,204); font-size:20px; font-weight:700;"
            " background:transparent; letter-spacing:-0.5px;"
        )
        self.title_lbl.setStyleSheet(
            "color:#ffffff; font-size:12px; font-weight:600; background:transparent;"
        )

        from PyQt6.QtWidgets import QHBoxLayout as _HBL, QVBoxLayout as _VBL, QLabel as _LBL
        row_w = _HBL()
        row_w.setSpacing(4)

        def _blok(key: str, lbl: str, clr: str):
            col = _VBL()
            col.setSpacing(0)
            hdr = _LBL(lbl)
            hdr.setAlignment(Qt.AlignmentFlag.AlignCenter)
            hdr.setStyleSheet("color:rgba(255,255,255,.70);font-size:11px;background:transparent;")
            val = _LBL("-")
            val.setAlignment(Qt.AlignmentFlag.AlignCenter)
            val.setStyleSheet(f"color:{clr};font-size:12px;font-weight:600;background:transparent;")
            col.addWidget(hdr)
            col.addWidget(val)
            setattr(self, key, val)
            row_w.addLayout(col)

        _blok("_islem_lbl", "İşlem",  "#ffe0a0")   # PHP: #womsisDashIslem
        _blok("_odeme_lbl", "Net",    "#d0ffd0")   # PHP: #womsisDashOdeme

        self._son_lbl = _LBL("Yerel tablodan anlık veriler")
        self._son_lbl.setStyleSheet(
            "color:rgba(255,255,255,.80);font-size:10px;background:transparent;"
        )

        lay = self.layout()
        lay.addLayout(row_w)
        lay.addWidget(self._son_lbl)

    def set_fiziksel_pos(
        self,
        net_fmt: str,
        islem_fmt: str,
        odeme_fmt: str,
        alt_yazi: str = "Yerel tablodan anlık veriler",
    ):
        """
        PHP:
          #womsisToplamBadge  → net_fmt
          #womsisDashIslem    → islem_fmt
          #womsisDashOdeme    → odeme_fmt
          #womsisSonGuncelleme→ alt_yazi
        """
        self.value_lbl.setText(net_fmt)
        self._islem_lbl.setText(islem_fmt)
        self._odeme_lbl.setText(odeme_fmt)
        self._son_lbl.setText(alt_yazi)


class DashboardLoader(QThread):
    """Arka planda veri yükler — UI donmaması için."""
    data_ready = pyqtSignal(dict)
    error = pyqtSignal(str)

    def __init__(self, userid: int, musterino: int, yil: int):
        super().__init__()
        self.userid = userid
        self.musterino = musterino
        self.yil = yil

    def run(self):
        try:
            data = get_all_dashboard_data(self.userid, self.musterino, self.yil)
            self.data_ready.emit(data)
        except Exception as e:
            self.error.emit(str(e))


class DashboardScreen(QWidget):
    """
    Ana dashboard ekranı.
    12 KPI kartını 2 satır × 6 sütun grid içinde gösterir.
    """

    def __init__(self, user: dict, parent=None):
        super().__init__(parent)
        self._user = user
        self._userid = user.get("GercekUserId", user.get("Kayitno", 1))
        self._musterino = user.get("GercekUserId", user.get("Kayitno", 1))
        self._yil = datetime.now().year
        self._loader = None
        self._cards: dict[str, KPICard] = {}
        self._setup_ui()
        self._load_data()

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(20)

        # ── Başlık çubuğu ──────────────────────────────────────────────────────
        header = QHBoxLayout()
        title = QLabel("Dashboard")
        title.setStyleSheet(f"""
            color: {COLORS['text_primary']};
            font-size: 24px;
            font-weight: 700;
            letter-spacing: 1px;
        """)
        header.addWidget(title)
        header.addStretch()
        # Şirket / kullanıcı
        sirket = self._user.get("Kurum_Durumu", "IQ Finans")
        kullanici = self._user.get("Adi", "")
        profil = QLabel(f"🏢 {sirket}  •  👤 {kullanici}")
        profil.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 12px; margin-right: 15px;")
        header.addWidget(profil)

        root.addLayout(header)

        # ── Bilgi banner ────────────────────────────────────────────────────────
        self.banner = QLabel(f"FİNANS DURUM BİLGİSİ ( {self._yil} )")
        self.banner.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.banner.setFixedHeight(38)
        self.banner.setStyleSheet(f"""
            background: {COLORS['text_primary']};
            color: white;
            font-size: 13px;
            font-weight: 700;
            letter-spacing: 3px;
            border-radius: 10px;
            padding: 0 20px;
        """)
        root.addWidget(self.banner)

        # Banner altı araç çubuğu (Yıl Seçimi sağda)
        banner_sub = QHBoxLayout()
        banner_sub.addStretch()

        self.yil_combo = QComboBox()
        current_year = datetime.now().year
        for y in range(current_year + 1, current_year - 6, -1):
            self.yil_combo.addItem(str(y), y)
        # ⚠ Sinyali bağlamadan ÖNCE setCurrentText yapılmalı —
        # aksi hâlde _on_yil_changed erken tetiklenir ve currentData() None döner
        self.yil_combo.setCurrentText(str(self._yil))
        # Başlangıç değerini combo'dan doğrula (setCurrentText bulamazsa ilk item seçilir)
        combo_val = self.yil_combo.currentData()
        if combo_val is not None:
            self._yil = combo_val
        self.yil_combo.setFixedHeight(32)
        self.yil_combo.setFixedWidth(120)
        self.yil_combo.setStyleSheet(f"""
            QComboBox {{
                background: white;
                border: 1.5px solid {COLORS['border']};
                border-radius: 8px;
                padding: 0 10px;
                font-size: 12px;
                font-weight: 600;
                color: {COLORS['text_primary']};
            }}
            QComboBox:focus {{
                border-color: {COLORS['btn_primary']};
            }}
            QComboBox::drop-down {{ border: none; }}
        """)
        self.yil_combo.currentIndexChanged.connect(self._on_yil_changed)
        banner_sub.addWidget(self.yil_combo)
        root.addLayout(banner_sub)

        # ── Scroll alan ─────────────────────────────────────────────────────────
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setStyleSheet("QScrollArea { background: transparent; }")

        container = QWidget()
        container.setStyleSheet("background: transparent;")
        vbox = QVBoxLayout(container)
        vbox.setContentsMargins(0, 0, 0, 0)
        vbox.setSpacing(20)

        # ── KPI Grid (2 × 6) ───────────────────────────────────────────────────
        grid = QGridLayout()
        grid.setSpacing(16)

        KPI_DEFS = [
            # (key,              başlık,              renk1,                renk2,        satır, sütun)
            ("nakit_kasa_gelir", "Nakit Kasa Gelir",  COLORS["green"],      "#059669",    0, 0),
            ("nakit_kasa_odeme", "Nakit Kasa Ödeme",  COLORS["teal"],       "#0284C7",    0, 1),
            ("kesilen_fatura",   "Kesilen Fatura",     COLORS["purple"],     "#6D28D9",    0, 2),
            ("gelen_fatura",     "Gelen Fatura",       COLORS["pink"],       "#DB2777",    0, 3),
            ("gider_pusulasi",   "Gider Pusulası",     "#16A34A",            "#15803D",    0, 4),
            ("kurum_odemeleri",  "Kurum Ödemeleri",    COLORS["dark_blue"],  "#162C47",    0, 5),
            ("maas_kira_smm",    "Maaş Kira SMM",      "#1a3a5c",            "#0d2137",    1, 0),
            ("bankalar_bakiye",  "Bankalar Bakiye",    COLORS["grey"],       "#4B5563",    1, 1),
            ("sanal_pos",        "Sanal Pos",          "#111827",            "#030712",    1, 2),
            ("fiziksel_pos",     "Fiziksel Pos",       "#1F2937",            "#111827",    1, 3),
            ("kredi_karti",      "Kredi Kartları",     "#D97706",            "#B45309",    1, 4),
            ("genel_hesap",      "Genel Hesap Tablosu","#EA580C",            "#C2410C",    1, 5),
        ]

        for key, title, c1, c2, row, col in KPI_DEFS:
            if key == "sanal_pos":
                card = PaytrKPICard(
                    click_cb=lambda k=key: self._on_card_click(k),
                )
            elif key == "fiziksel_pos":
                # PHP: #btnFizikselPosHareketleri (koyu mavi gradient)
                card = FizikselPosKPICard(
                    click_cb=lambda k=key: self._on_card_click(k),
                )
            else:
                card = KPICard(
                    title=title,
                    value="Yükleniyor...",
                    color=c1,
                    color2=c2,
                    click_cb=lambda k=key: self._on_card_click(k),
                )
            card.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            self._cards[key] = card
            grid.addWidget(card, row, col)

        vbox.addLayout(grid)

        # ── Tarih filtresi + Excel ──────────────────────────────────────────────
        filter_bar = QHBoxLayout()
        filter_bar.setSpacing(10)

        ilk_lbl = QLabel("İlk Tarih:")
        ilk_lbl.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 12px;")
        filter_bar.addWidget(ilk_lbl)

        self.ilk_tarih = QDateEdit()
        self.ilk_tarih.setCalendarPopup(True)
        self.ilk_tarih.setDate(QDate(self._yil, 1, 1))
        self.ilk_tarih.setFixedHeight(34)
        self.ilk_tarih.setStyleSheet(self._input_style())
        filter_bar.addWidget(self.ilk_tarih)

        son_lbl = QLabel("Son Tarih:")
        son_lbl.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 12px;")
        filter_bar.addWidget(son_lbl)

        self.son_tarih = QDateEdit()
        self.son_tarih.setCalendarPopup(True)
        self.son_tarih.setDate(QDate.currentDate())
        self.son_tarih.setFixedHeight(34)
        self.son_tarih.setStyleSheet(self._input_style())
        filter_bar.addWidget(self.son_tarih)

        filter_bar.addStretch()

        self.excel_btn = QPushButton("📥 EXCEL İNDİR")
        self.excel_btn.setFixedHeight(34)
        self.excel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.excel_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLORS['btn_excel']};
                color: white;
                font-size: 12px;
                font-weight: 700;
                border: none;
                border-radius: 8px;
                padding: 0 18px;
                letter-spacing: 1px;
            }}
            QPushButton:hover {{ background: #047857; }}
        """)
        filter_bar.addWidget(self.excel_btn)

        vbox.addLayout(filter_bar)

        # ── Grafik ─────────────────────────────────────────────────────────────
        self.chart = GelirGiderChart()
        vbox.addWidget(self.chart)

        scroll.setWidget(container)
        root.addWidget(scroll)

    def _input_style(self) -> str:
        return f"""
            QDateEdit {{
                background: white;
                border: 1.5px solid {COLORS['border']};
                border-radius: 8px;
                padding: 0 10px;
                font-size: 12px;
                color: {COLORS['text_primary']};
            }}
        """

    def _load_data(self):
        if self._loader and self._loader.isRunning():
            # Eski loader'ın sinyallerini kes — yanıt gelirse UI'yı bozmasın
            try:
                self._loader.data_ready.disconnect()
                self._loader.error.disconnect()
            except Exception:
                pass
            self._loader.quit()

        self._loader = DashboardLoader(self._userid, self._musterino, self._yil)
        self._loader.data_ready.connect(self._on_data_ready)
        self._loader.error.connect(self._on_error)
        self._loader.start()

    def refresh(self):
        """Sol menüden sayfaya geçince veriler yeniden yüklenir."""
        self.banner.setText(f"FİNANS DURUM BİLGİSİ ( {self._yil} )")
        self._load_data()

    def _on_data_ready(self, data: dict):
        c = self._cards
        
        # Grafik verilerini güncelle
        self.chart.load_data(data.get("monthly_chart", []))

        # Nakit Kasa Gelir
        kasa = data.get("nakit_kasa", {})
        c["nakit_kasa_gelir"].set_value(
            fmt_para(kasa.get("gelir", 0)),
            f"Gelir: {fmt_para(kasa.get('gelir', 0))}  Gider: {fmt_para(kasa.get('gider', 0))}"
        )

        # Nakit Kasa Ödeme
        gider = data.get("gider", {})
        c["nakit_kasa_odeme"].set_value(
            fmt_para(gider.get("gider", 0)),
            "genel_hesap_hareketleri · gider"
        )

        # Kesilen / Gelen Fatura
        fat = data.get("faturalar", {})
        c["kesilen_fatura"].set_value(
            fmt_para(fat.get("kesilen", 0)),
            "Detaylar için tıklayın..."
        )
        c["gelen_fatura"].set_value(
            fmt_para(fat.get("gelen", 0)),
            "Detaylar için tıklayın..."
        )

        # Gider Pusulası
        gp = data.get("gider_pusulasi", {})
        c["gider_pusulasi"].set_value(
            fmt_para(gp.get("gider", 0)),
            "Parça Alımı (Cihaz)"
        )

        # Kurum Ödemeleri
        kurum = data.get("kurum_odemeleri", {})
        kurum_moy = kurum.get("moy", 0)
        kurum_alt = "Vergi ödeme detayları"
        if kurum_moy > 0:
            kurum_alt += f"  •  Moy: {fmt_para(kurum_moy)}"
        c["kurum_odemeleri"].set_value(
            fmt_para(kurum.get("toplam", 0)),
            kurum_alt
        )

        # Maaş Kira SMM — PHP: #dashMaasKiraSmmmToplam
        # Formül: SUM(ABS(gaytutar - vergkestutar))  her satır için
        from services.vergi_muhtasar_service import get_dashboard_toplam as _mks_toplam
        mks_r = _mks_toplam(self._userid, musterino=str(self._musterino), yil=self._yil)
        if mks_r.get("success") and mks_r["fark_toplam"] > 0:
            mks_val_str = mks_r["fark_toplam_fmt"]   # '10.853.717,00 ₺'
        else:
            mks = data.get("maas_kira_smm", {})
            mks_val_str = fmt_para(mks.get("toplam", 0))
        c["maas_kira_smm"].set_value(
            mks_val_str,
            "Personel, Kira ve Müşavirlik Giderleri"
        )

        # Bankalar Bakiye
        c["bankalar_bakiye"].set_value("₺0,00", "Detaylar için tıklayın...")

        # Sanal Pos (PayTR) — PHP: #paytrToplamBadge / #paytrDashIslem / #paytrDashOdeme / #paytrSonGuncelleme
        pos = data.get("sanal_pos", {})
        pos_card = self._cards.get("sanal_pos")
        if isinstance(pos_card, PaytrKPICard):
            # Fark = odeme - islem
            islem_val = pos.get("islem", 0)
            odeme_val = pos.get("odeme", 0)
            try:
                fark_val = float(odeme_val) - float(islem_val)
            except (TypeError, ValueError):
                fark_val = 0.0
            fark_sign = "+" if fark_val >= 0 else "-"
            fark_abs = abs(fark_val)
            fark_fmt = (
                f"{fark_sign}₺"
                + f"{fark_abs:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            )
            islem_fmt = fmt_para(islem_val)
            odeme_fmt = fmt_para(odeme_val)
            son = pos.get("son_guncelleme", "")
            pos_card.set_paytr(fark_fmt, islem_fmt, odeme_fmt, son)
        else:
            pos_card.set_value(
                fmt_para(pos.get("islem", 0)),
                f"İşlem: {fmt_para(pos.get('islem', 0))}  Ödeme: {fmt_para(pos.get('odeme', 0))}"
            )

        # Fiziksel Pos (Womsis) — PHP: #womsisToplamBadge / #womsisDashIslem / #womsisDashOdeme
        fp_card = self._cards.get("fiziksel_pos")
        if isinstance(fp_card, FizikselPosKPICard):
            from services.fiziksel_pos_service import get_dashboard_ozet
            fp = get_dashboard_ozet(self._userid)
            fp_card.set_fiziksel_pos(
                net_fmt=fp.get("toplam_net_fmt",   "₺0,00"),
                islem_fmt=fp.get("toplam_islem_fmt", "₺0,00"),
                odeme_fmt=fp.get("toplam_net_fmt",   "₺0,00"),
                alt_yazi="Yerel tablodan anlık veriler",
            )
        else:
            if fp_card:
                fp_card.set_value("₺0,00", "Yerel tablodan anlık veriler")


        # Kredi Kartları
        kk = data.get("kredi_karti", {})
        borc  = kk.get("borc", 0)
        odeme = kk.get("odeme", 0)
        net   = kk.get("net",   0)
        c["kredi_karti"].set_value(
            fmt_para(borc),
            f"Harcama: {fmt_para(borc)}  Ödeme: {fmt_para(abs(odeme))}  Net: {fmt_para(net)}"
        )

        # Genel Hesap
        gh = data.get("genel_hesap", {})
        c["genel_hesap"].set_value(
            fmt_para(gh.get("net", 0)),
            "genel_hesap_hareketleri · genelHesap"
        )

    def _on_error(self, msg: str):
        self.banner.setText(f"⚠  Hata: {msg}")
        self.banner.setStyleSheet(f"""
            background: {COLORS['red']};
            color: white;
            font-size: 12px;
            font-weight: 600;
            border-radius: 10px;
            padding: 0 20px;
        """)
        for card in self._cards.values():
            card.set_value("Hata")

    # ─── KART TIKLAMA ────────────────────────────────────────────────────────

    def _on_card_click(self, key: str):
        """KPI kartına tıklandığında ilgili detay diyaloğunu açar."""
        uid = self._userid
        mno = self._musterino
        yil = self._yil

        if key in ("nakit_kasa_gelir", "nakit_kasa_odeme"):
            def detay_fn(_sube_adi):
                # genel_hesap_hareketleri'nden kasa kayıtları — dashboard ile aynı kaynak
                return detay_service.get_nakit_kasa_detay(uid, mno, yil)
            dlg = DetayDialog(
                baslik="Nakit Kasa Hareketleri",
                ozet_rows=[],
                detay_fn=detay_fn,
                tablo_tipi="genel_hesap",
                direct_detay=True,
                yil=yil,
                parent=self,
            )

        elif key in ("genel_hesap",):
            ozet = detay_service.get_genel_hesap_sube_ozet(uid, mno, yil)
            def detay_fn(sube_adi):
                return detay_service.get_genel_hesap_detay(uid, mno, yil, sube_adi=sube_adi)
            dlg = DetayDialog(
                baslik="Genel Hesap Tablosu — Şube Özeti",
                ozet_rows=ozet,
                detay_fn=detay_fn,
                tablo_tipi="genel_hesap",
                userid=uid,
                yil=yil,
                parent=self,
            )

        elif key == "kesilen_fatura":
            ozet = detay_service.get_fatura_sube_ozet(uid, mno, yil, "gelir")
            def _kes_detay(sube_adi, _u=uid, _m=mno, _y=yil):
                return detay_service.get_fatura_detay_by_sube(_u, _m, _y, "gelir", sube_adi)
            dlg = DetayDialog(
                baslik="Kesilen Faturalar — Şube Özeti",
                ozet_rows=ozet,
                detay_fn=_kes_detay,
                tablo_tipi="faturalar",
                userid=uid,
                yil=yil,
                parent=self,
            )

        elif key == "gelen_fatura":
            ozet = detay_service.get_fatura_sube_ozet(uid, mno, yil, "gider")
            def _gel_detay(sube_adi, _u=uid, _m=mno, _y=yil):
                return detay_service.get_fatura_detay_by_sube(_u, _m, _y, "gider", sube_adi)
            dlg = DetayDialog(
                baslik="Gelen Faturalar — Şube Özeti",
                ozet_rows=ozet,
                detay_fn=_gel_detay,
                tablo_tipi="faturalar",
                userid=uid,
                yil=yil,
                parent=self,
            )

        elif key == "kurum_odemeleri":
            dlg = KurumOdemeDialog(mno, yil, parent=self)
            dlg.exec()
            return

        elif key == "kredi_karti":
            dlg = KrediKartiDialog(uid, yil, parent=self)
            dlg.exec()
            return

        elif key == "maas_kira_smm":
            dlg = MaasKiraSmmDialog(self._userid, str(self._musterino), yil=self._yil, parent=self)
            dlg.exec()
            return

        elif key == "fiziksel_pos":
            dlg = FizikselPosDialog(self._userid, parent=self)
            dlg.exec()
            return

        elif key == "sanal_pos":
            dlg = SanalPosDialog(self._userid, str(self._musterino), self._yil, parent=self)
            dlg.exec()
            return

        else:
            # Henüz uygulanmamış kartlar için basit mesaj
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(
                self, key,
                f"'{key}' kartı için detay görünümü yakında eklenecek."
            )
            return

        dlg.exec()

    def _on_yil_changed(self):
        combo_val = self.yil_combo.currentData()
        if combo_val is None:
            return   # combo henüz hazır değil, erken tetiklenmeden çık
        self._yil = combo_val
        self.banner.setText(f"FİNANS DURUM BİLGİSİ ( {self._yil} )")
        # İlk tarih DateEdit'i seçilen yılın 1 Ocak'ına güncelle
        try:
            from PyQt6.QtCore import QDate
            self.ilk_tarih.setDate(QDate(self._yil, 1, 1))
        except Exception:
            pass
        # Kartları yükleniyor moduna al
        for card in self._cards.values():
            card.set_value("Yükleniyor...")
        self._load_data()


# ─────────────────────────────────────────────────────────────────────────────
# Kurum Ödemeleri Detay Dialog
# PHP: nakitAkimParametreAjaxGider.php + gider_veriler.js DataTable
# ─────────────────────────────────────────────────────────────────────────────

class KurumOdemeDialog(QDialog):
    """
    Kurum Ödemeleri kartına tıklandığında açılan detay tablosu.
    PHP gider_veriler.js DataTable ile birebir aynı sütun yapısı.
    Düzeltmeler:
      1) Yazılar daima siyah (hover dahil)
      2) DateEdit picker — ilkTarih 01/01/YIL, sonTarih bugün
      3) Ay seçince tarihler otomatik dolar (Şubat→01.02/28.02)
      4) Satıra tıklandığında BeyannamePreviewDialog açılır
    """

    HESAP_ACIKLAMA = {
        "770.01": "770.01 — Vergi Giderleri (SGK / KDV / Gelir Vergisi)",
        "730.08": "730.08 — İşçilik / Müşavirlik Giderleri",
    }
    BEYANNAME_TUR = {
        "770.01": "KDV / SGK Beyannamesi",
        "730.08": "Muhtasar ve Prim Hizmet Beyannamesi",
    }

    AY_ADLARI = [
        "Hepsi", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
        "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık",
    ]

    SUTUNLAR = [
        ("Beyanname Türü",  "hesapKodu",     200),
        ("Ünvan",           "unvan",         140),
        ("Vergi No",        "vergiNo",       105),
        ("İlk Tarih",       "ilkTarih",       95),
        ("Son Tarih",       "sonTarih",       90),
        ("Sözleşme No",     "sozlesmeNo",    100),
        ("Sözl. Tarih",     "sozlesmeTarih",  90),
        ("Tutar",           "tutar",         115),
    ]

    _CBS = (
        "QComboBox{background:white;border:1.5px solid #cbd5e1;"
        "border-radius:6px;padding:0 8px;font-size:12px;color:#1e293b;}"
        "QComboBox:focus{border-color:#162C47;}"
        "QComboBox::drop-down{border:none;width:18px;}"
    )
    _DES = (
        "QDateEdit{background:white;border:1.5px solid #cbd5e1;"
        "border-radius:6px;padding:0 6px;font-size:12px;color:#1e293b;}"
        "QDateEdit:focus{border-color:#162C47;}"
        "QDateEdit::drop-down{border:none;width:18px;}"
    )

    def __init__(self, musterino: int, yil: int, parent=None):
        super().__init__(parent)
        self._musterino = musterino
        self._yil       = yil
        self._rows: list[dict] = []
        self._ay_degisiyor = False   # döngü koruması

        self.setWindowTitle("Kurum Ödemeleri — Detay")
        self.setMinimumSize(1060, 620)
        self.resize(1180, 700)
        self._setup_ui()
        self._load()

    # ── UI ───────────────────────────────────────────────────────────────────

    def _setup_ui(self):
        import calendar
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 14, 16, 14)
        root.setSpacing(8)

        # ── Başlık ──
        baslik = QLabel("📋  Kurum Ödemeleri — Gider Parametreleri")
        baslik.setStyleSheet(
            "font-size:15px;font-weight:700;color:#162C47;"
            "padding-bottom:4px;"
        )
        root.addWidget(baslik)

        # ── Filtre çubuğu ──
        bar = QHBoxLayout()
        bar.setSpacing(8)

        # Ay seçici
        bar.addWidget(self._lbl("Ay:"))
        self._ay_cb = QComboBox()
        self._ay_cb.setFixedSize(100, 32)
        self._ay_cb.setStyleSheet(self._CBS)
        for i, a in enumerate(self.AY_ADLARI):
            self._ay_cb.addItem(a, i)
        self._ay_cb.currentIndexChanged.connect(self._on_ay_change)
        bar.addWidget(self._ay_cb)

        # İlk Tarih DateEdit
        bar.addWidget(self._lbl("İlk Tarih:"))
        self._ilk_de = QDateEdit()
        self._ilk_de.setCalendarPopup(True)
        self._ilk_de.setDisplayFormat("dd.MM.yyyy")
        self._ilk_de.setFixedSize(120, 32)
        self._ilk_de.setStyleSheet(self._DES)
        self._ilk_de.setDate(QDate(self._yil, 1, 1))    # 01/01/YIL
        bar.addWidget(self._ilk_de)

        # Son Tarih DateEdit
        bar.addWidget(self._lbl("Son Tarih:"))
        self._son_de = QDateEdit()
        self._son_de.setCalendarPopup(True)
        self._son_de.setDisplayFormat("dd.MM.yyyy")
        self._son_de.setFixedSize(120, 32)
        self._son_de.setStyleSheet(self._DES)
        self._son_de.setDate(QDate.currentDate())
        bar.addWidget(self._son_de)

        bar.addStretch()

        # Filtrele butonu
        self._filtre_btn = QPushButton("🔍  Filtrele")
        self._filtre_btn.setFixedSize(110, 32)
        self._filtre_btn.setStyleSheet(
            "QPushButton{background:#162C47;color:white;border:none;"
            "border-radius:7px;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#1e3a5f;}"
        )
        self._filtre_btn.clicked.connect(self._load)
        bar.addWidget(self._filtre_btn)
        root.addLayout(bar)

        # ── Özet bant ──
        self._ozet_lbl = QLabel("")
        self._ozet_lbl.setFixedHeight(28)
        self._ozet_lbl.setStyleSheet(
            "background:#f0f9ff;border:1px solid #bae6fd;border-radius:6px;"
            "padding:0 10px;font-size:12px;color:#0c4a6e;"
        )
        root.addWidget(self._ozet_lbl)

        # ── Bilgi notu ──
        not_lbl = QLabel("💡  Satıra tıklayarak beyanname önizlemesini açabilirsiniz.")
        not_lbl.setStyleSheet("font-size:11px;color:#64748b;")
        root.addWidget(not_lbl)

        # ── Tablo ──
        self._tablo = QTableWidget()
        self._tablo.setColumnCount(len(self.SUTUNLAR))
        self._tablo.setHorizontalHeaderLabels([s[0] for s in self.SUTUNLAR])
        self._tablo.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tablo.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._tablo.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._tablo.setAlternatingRowColors(True)
        self._tablo.verticalHeader().setVisible(False)
        self._tablo.setSortingEnabled(True)
        self._tablo.horizontalHeader().setStretchLastSection(True)
        self._tablo.setMouseTracking(True)
        for i, (_, _, w) in enumerate(self.SUTUNLAR):
            self._tablo.setColumnWidth(i, w)
        # 1) Yazılar DAİMA siyah — hover ve seçim dahil
        self._tablo.setStyleSheet("""
            QTableWidget {
                background: white;
                gridline-color: #e2e8f0;
                font-size: 12px;
                color: #1e293b;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
            }
            QTableWidget::item {
                color: #1e293b;
                padding: 4px 6px;
            }
            QTableWidget::item:hover {
                background: #eff6ff;
                color: #1e293b;
            }
            QTableWidget::item:selected {
                background: #dbeafe;
                color: #1e293b;
            }
            QTableWidget::item:alternate {
                background: #f8fafc;
                color: #1e293b;
            }
            QHeaderView::section {
                background: #1e293b;
                color: white;
                font-weight: 700;
                font-size: 11px;
                padding: 6px 4px;
                border: none;
                border-right: 1px solid #334155;
            }
        """)
        self._tablo.cellClicked.connect(self._on_satir_tikla)
        root.addWidget(self._tablo)

        # ── Kapat ──
        kapat = QPushButton("✕  Kapat")
        kapat.setFixedSize(110, 32)
        kapat.setStyleSheet(
            "QPushButton{background:#64748b;color:white;border:none;"
            "border-radius:7px;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#475569;}"
        )
        kapat.clicked.connect(self.accept)
        bot = QHBoxLayout()
        bot.addStretch()
        bot.addWidget(kapat)
        root.addLayout(bot)

    # ── Yardımcılar ──────────────────────────────────────────────────────────

    @staticmethod
    def _lbl(txt: str) -> QLabel:
        l = QLabel(txt)
        l.setStyleSheet("font-size:12px;color:#475569;font-weight:600;")
        return l

    @staticmethod
    def _fmt_goster(t) -> str:
        """YYYYMMDD → DD.MM.YYYY"""
        s = str(t) if t else ""
        if len(s) == 8 and s.isdigit():
            return f"{s[6:8]}.{s[4:6]}.{s[0:4]}"
        return s

    # ── Ay seçince tarihler otomatik dolar ───────────────────────────────────

    def _on_ay_change(self, idx: int):
        """
        Şubat seçilirse → ilkTarih=01.02.YYYY, sonTarih=28.02.YYYY
        Hepsi seçilirse → ilkTarih=01.01.YYYY, sonTarih=bugün
        """
        import calendar
        if self._ay_degisiyor:
            return
        self._ay_degisiyor = True
        ay = self._ay_cb.currentData()   # 0=hepsi, 1=Ocak…12=Aralık
        yil = self._yil
        if ay and ay > 0:
            son_gun = calendar.monthrange(yil, ay)[1]
            self._ilk_de.setDate(QDate(yil, ay, 1))
            self._son_de.setDate(QDate(yil, ay, son_gun))
        else:
            self._ilk_de.setDate(QDate(yil, 1, 1))
            self._son_de.setDate(QDate.currentDate())
        self._ay_degisiyor = False

    # ── Veri yükleme ─────────────────────────────────────────────────────────

    def _load(self):
        from services.detay_service import get_kurum_odemeleri_detay_tarih
        ilk = self._ilk_de.date()
        son = self._son_de.date()
        ilk_str = f"{ilk.year()}{ilk.month():02d}{ilk.day():02d}"
        son_str = f"{son.year()}{son.month():02d}{son.day():02d}"
        self._rows = get_kurum_odemeleri_detay_tarih(
            self._musterino, ilk_str, son_str
        )
        self._doldur()

    def _doldur(self):
        from PyQt6.QtGui import QColor, QFont
        self._tablo.setSortingEnabled(False)
        self._tablo.setRowCount(0)

        toplam = 0.0
        for row in self._rows:
            ri = self._tablo.rowCount()
            self._tablo.insertRow(ri)
            self._tablo.setRowHeight(ri, 30)

            kod      = row.get("hesapKodu", "")
            beyan_t  = self.BEYANNAME_TUR.get(kod, self.HESAP_ACIKLAMA.get(kod, kod))
            unvan    = row.get("unvan", "") or "-"
            vergino  = row.get("vergiNo", "") or ""
            ilkT     = self._fmt_goster(row.get("ilkTarih", ""))
            sonT     = self._fmt_goster(row.get("sonTarih", ""))
            sozno    = row.get("sozlesmeNo", "") or ""
            soztarih = self._fmt_goster(row.get("sozlesmeTarih", ""))
            tutar    = float(row.get("tutar") or 0)
            toplam  += tutar

            degerler = [beyan_t, unvan, vergino, ilkT, sonT, sozno, soztarih, None]
            for ci, val in enumerate(degerler):
                if ci == 7:
                    # Tutar — sağa yasla, teal renk ama siyah hover için
                    it = QTableWidgetItem(f"{tutar:,.2f}")
                    it.setTextAlignment(
                        Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
                    )
                    it.setForeground(QColor("#0f766e"))
                    it.setFont(QFont("", -1, QFont.Weight.Bold))
                else:
                    it = QTableWidgetItem(str(val))
                    it.setTextAlignment(
                        Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
                    )
                    it.setForeground(QColor("#1e293b"))
                # satır indeksini UserRole olarak sakla (preview için)
                it.setData(Qt.ItemDataRole.UserRole, ri)
                self._tablo.setItem(ri, ci, it)

        self._tablo.setSortingEnabled(True)

        # Özet
        ilk_txt = self._ilk_de.date().toString("dd.MM.yyyy")
        son_txt = self._son_de.date().toString("dd.MM.yyyy")
        self._ozet_lbl.setText(
            f"📅 {ilk_txt} — {son_txt}  │  "
            f"<b>{len(self._rows)}</b> kayıt  │  "
            f"Toplam: <b>{toplam:,.2f} ₺</b>"
        )
        self._ozet_lbl.setTextFormat(Qt.TextFormat.RichText)

    # ── Satır tıklama → Beyanname preview ───────────────────────────────────

    def _on_satir_tikla(self, row: int, _col: int):
        if row < 0 or row >= len(self._rows):
            return
        veri = self._rows[row]
        dlg = BeyannamePreviewDialog(veri, self._musterino, self)
        dlg.exec()


# ─────────────────────────────────────────────────────────────────────────────
# Beyanname Önizleme Dialog
# ─────────────────────────────────────────────────────────────────────────────

class BeyannamePreviewDialog(QDialog):
    """
    Satıra tıklandığında açılan PDF önizleme dialog'u.
    - Moy'dan gerçek zamanlı beyanname_listeleri sorgusu yapar
    - İlgili dönemde bulunan beyannameleri listeler
    - Seçilen beyanname'nin PDF'ini beyanname_gib.Belge_Data'dan çeker
    - PDF'i geçici dosyaya yazıp sistem PDF görüntüleyicisiyle açar
    """

    BEYANNAME_TUR = {
        "770.01": "KDV / SGK Beyannamesi",
        "730.08": "Muhtasar ve Prim Hizmet Beyannamesi",
    }
    BEYANNAME_ICON = {"770.01": "🏛️", "730.08": "📄"}
    BEYANNAME_RENK = {"770.01": "#1d4ed8", "730.08": "#7c3aed"}

    BELGE_TUR_ADI = {
        "KDV1":     "KDV Beyannamesi (1.Tür)",
        "KDV2":     "KDV Beyannamesi (2.Tür)",
        "MUHSGK":   "Muhtasar ve SGK Beyannamesi",
        "KGECICI":  "Kurumlar Vergisi Geçici Beyan",
        "KURUMLAR": "Kurumlar Vergisi Beyannamesi",
        "LEVHA":    "Levha Beyannamesi",
        "MUHTAR":   "Muhtasar Beyanname",
    }

    def __init__(self, veri: dict, musterino: int, parent=None):
        super().__init__(parent)
        self._veri      = veri
        self._musterino = musterino
        self._beyanlar: list[dict] = []
        self._tmp_path: str | None = None

        kod = veri.get("hesapKodu", "")
        self._renk = self.BEYANNAME_RENK.get(kod, "#162C47")
        self.setWindowTitle("Beyanname PDF Önizleme")
        self.setMinimumSize(860, 640)
        self.resize(960, 720)
        self._setup_ui()
        self._load_beyanlar()

    @staticmethod
    def _fmt(t) -> str:
        s = str(t) if t else ""
        if len(s) == 8 and s.isdigit():
            return f"{s[6:8]}.{s[4:6]}.{s[0:4]}"
        return s or "—"

    def _donem_str(self) -> str:
        t = str(self._veri.get("ilkTarih", "") or "")
        if len(t) == 8 and t.isdigit():
            ay_map = {"01":"Ocak","02":"Şubat","03":"Mart","04":"Nisan",
                      "05":"Mayıs","06":"Haziran","07":"Temmuz","08":"Ağustos",
                      "09":"Eylül","10":"Ekim","11":"Kasım","12":"Aralık"}
            return f"{ay_map.get(t[4:6], t[4:6])} {t[0:4]}"
        return t or "—"

    # ── UI ───────────────────────────────────────────────────────────────────

    def _setup_ui(self):
        from PyQt6.QtGui import QColor
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        kod  = self._veri.get("hesapKodu", "")
        tur  = self.BEYANNAME_TUR.get(kod, "Beyanname")
        icon = self.BEYANNAME_ICON.get(kod, "📋")
        renk = self._renk

        # ── Başlık bandı ──
        hdr = QFrame()
        hdr.setFixedHeight(76)
        hdr.setStyleSheet(
            f"background:qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            f"stop:0 {renk},stop:1 #0f172a);"
        )
        h = QHBoxLayout(hdr)
        h.setContentsMargins(20, 0, 20, 0)

        left = QVBoxLayout()
        t1 = QLabel(f"{icon}  {tur}")
        t1.setStyleSheet("color:white;font-size:15px;font-weight:700;background:transparent;")
        t2 = QLabel(f"📅  Dönem: {self._donem_str()}   •   Tutar: {float(self._veri.get('tutar') or 0):,.2f} ₺")
        t2.setStyleSheet("color:rgba(255,255,255,0.82);font-size:12px;background:transparent;")
        left.addWidget(t1)
        left.addWidget(t2)
        h.addLayout(left, 1)
        root.addWidget(hdr)

        # ── İçerik: Sol liste + Sağ PDF alanı ──
        splitter_frame = QFrame()
        splitter_frame.setStyleSheet("background:#f8fafc;")
        sf_lay = QHBoxLayout(splitter_frame)
        sf_lay.setContentsMargins(10, 10, 10, 10)
        sf_lay.setSpacing(10)

        # Sol: Beyanname listesi
        sol = QFrame()
        sol.setFixedWidth(290)
        sol.setStyleSheet(
            "background:white;border-radius:8px;"
            "border:1px solid #e2e8f0;"
        )
        sol_lay = QVBoxLayout(sol)
        sol_lay.setContentsMargins(0, 0, 0, 0)
        sol_lay.setSpacing(0)

        sol_baslik = QLabel("📋  İlgili Beyannameler")
        sol_baslik.setFixedHeight(36)
        sol_baslik.setStyleSheet(
            f"background:{renk};color:white;font-size:12px;font-weight:700;"
            "padding:0 10px;border-radius:8px 8px 0 0;"
        )
        sol_lay.addWidget(sol_baslik)

        self._liste = QTableWidget()
        self._liste.setColumnCount(3)
        self._liste.setHorizontalHeaderLabels(["Tür", "Dönem", "Onay"])
        self._liste.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._liste.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._liste.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._liste.verticalHeader().setVisible(False)
        self._liste.setShowGrid(False)
        self._liste.horizontalHeader().setStretchLastSection(True)
        self._liste.setColumnWidth(0, 90)
        self._liste.setColumnWidth(1, 80)
        self._liste.setStyleSheet("""
            QTableWidget { background:white; font-size:11px; border:none; color:#1e293b; }
            QTableWidget::item { padding:4px 6px; color:#1e293b; }
            QTableWidget::item:hover { background:#eff6ff; color:#1e293b; }
            QTableWidget::item:selected { background:#dbeafe; color:#1e293b; }
            QHeaderView::section {
                background:#f1f5f9; color:#475569; font-weight:700;
                font-size:10px; padding:4px; border:none;
                border-bottom:1px solid #e2e8f0;
            }
        """)
        self._liste.cellClicked.connect(self._on_beyanname_sec)
        sol_lay.addWidget(self._liste, 1)

        self._yukleniyor_lbl = QLabel("🔄  Moy'dan yükleniyor...")
        self._yukleniyor_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._yukleniyor_lbl.setStyleSheet("color:#64748b;font-size:11px;padding:10px;")
        sol_lay.addWidget(self._yukleniyor_lbl)
        self._yukleniyor_lbl.hide()

        sf_lay.addWidget(sol)

        # Sağ: PDF alanı
        sag = QFrame()
        sag.setStyleSheet(
            "background:white;border-radius:8px;"
            "border:1px solid #e2e8f0;"
        )
        sag_lay = QVBoxLayout(sag)
        sag_lay.setContentsMargins(0, 0, 0, 0)

        sag_baslik_frame = QFrame()
        sag_baslik_frame.setFixedHeight(36)
        sag_baslik_frame.setStyleSheet(
            "background:#1e293b;border-radius:8px 8px 0 0;"
        )
        sb_lay = QHBoxLayout(sag_baslik_frame)
        sb_lay.setContentsMargins(12, 0, 12, 0)
        self._pdf_baslik = QLabel("PDF Önizleme")
        self._pdf_baslik.setStyleSheet(
            "color:white;font-size:12px;font-weight:700;"
        )
        sb_lay.addWidget(self._pdf_baslik)
        sb_lay.addStretch()

        self._pdf_ac_btn = QPushButton("📂  PDF'i Aç")
        self._pdf_ac_btn.setFixedSize(110, 26)
        self._pdf_ac_btn.setStyleSheet(
            "QPushButton{background:#0f766e;color:white;border:none;"
            "border-radius:5px;font-size:11px;font-weight:600;}"
            "QPushButton:hover{background:#0d9488;}"
            "QPushButton:disabled{background:#334155;color:#64748b;}"
        )
        self._pdf_ac_btn.setEnabled(False)
        self._pdf_ac_btn.clicked.connect(self._pdf_ac)
        sb_lay.addWidget(self._pdf_ac_btn)
        sag_lay.addWidget(sag_baslik_frame)

        # PDF içerik alanı — metin tabanlı önizleme
        self._pdf_alan = QLabel()
        self._pdf_alan.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._pdf_alan.setWordWrap(True)
        self._pdf_alan.setStyleSheet(
            "color:#94a3b8;font-size:13px;padding:20px;"
            "background:white;"
        )
        self._pdf_alan.setText(
            "⬅️  Soldaki listeden bir beyanname seçin\n"
            "PDF önizlemesi burada gösterilecektir."
        )

        # PDF Image görüntüleme için QLabel (PDF → PNG dönüşümü)
        self._pdf_img_scroll = QScrollArea()
        self._pdf_img_scroll.setWidgetResizable(True)
        self._pdf_img_scroll.setStyleSheet(
            "QScrollArea{border:none;background:white;}"
        )
        self._pdf_img_container = QLabel()
        self._pdf_img_container.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignHCenter)
        self._pdf_img_container.setStyleSheet("background:white;padding:8px;")
        self._pdf_img_scroll.setWidget(self._pdf_img_container)
        self._pdf_img_scroll.hide()

        sag_lay.addWidget(self._pdf_alan, 1)
        sag_lay.addWidget(self._pdf_img_scroll, 1)
        sf_lay.addWidget(sag, 1)

        root.addWidget(splitter_frame, 1)

        # ── Alt butonlar ──
        alt = QFrame()
        alt.setFixedHeight(48)
        alt.setStyleSheet("background:white;border-top:1px solid #e2e8f0;")
        a = QHBoxLayout(alt)
        a.setContentsMargins(16, 0, 16, 0)
        a.addStretch()
        kapat = QPushButton("✕  Kapat")
        kapat.setFixedSize(110, 32)
        kapat.setStyleSheet(
            "QPushButton{background:#64748b;color:white;border:none;"
            "border-radius:7px;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#475569;}"
        )
        kapat.clicked.connect(self.accept)
        a.addWidget(kapat)
        root.addWidget(alt)

    # ── Beyanname listesini yükle ─────────────────────────────────────────────

    def _load_beyanlar(self):
        ilk_tarih = self._veri.get("ilkTarih", "") or ""
        if not ilk_tarih:
            self._liste_bos_goster("İlk tarih bilgisi yok.")
            return
        self._yukleniyor_lbl.show()
        from services.moy_service import get_beyanname_listesi
        try:
            self._beyanlar = get_beyanname_listesi(self._musterino, ilk_tarih)
        except Exception as e:
            self._beyanlar = []
        self._yukleniyor_lbl.hide()
        self._liste_doldur()

    def _liste_doldur(self):
        self._liste.setRowCount(0)
        if not self._beyanlar:
            self._liste_bos_goster("Bu dönem için beyanname bulunamadı.")
            return

        for b in self._beyanlar:
            ri = self._liste.rowCount()
            self._liste.insertRow(ri)
            self._liste.setRowHeight(ri, 26)

            tur_adi = self.BELGE_TUR_ADI.get(b["belge_turu"], b["belge_turu"])
            onay = b["onay_tarihi"]
            # YYYYMMDDHHMMSS → DD.MM.YYYY
            if len(onay) >= 8 and onay[:8].isdigit():
                onay = f"{onay[6:8]}.{onay[4:6]}.{onay[0:4]}"

            for ci, txt in enumerate([tur_adi, b["donem_adi"], onay]):
                it = QTableWidgetItem(str(txt))
                it.setTextAlignment(
                    Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
                )
                from PyQt6.QtGui import QColor
                it.setForeground(QColor("#1e293b"))
                self._liste.setItem(ri, ci, it)

    def _liste_bos_goster(self, mesaj: str):
        self._pdf_alan.setText(f"ℹ️  {mesaj}")
        self._pdf_alan.show()
        self._pdf_img_scroll.hide()

    # ── Beyanname seçildi → PDF çek ──────────────────────────────────────────

    def _on_beyanname_sec(self, row: int, _col: int):
        if row < 0 or row >= len(self._beyanlar):
            return
        b = self._beyanlar[row]
        tur_adi = self.BELGE_TUR_ADI.get(b["belge_turu"], b["belge_turu"])
        self._pdf_baslik.setText(f"{tur_adi}  —  {b['donem_adi']} {b['donem_no']}")
        self._pdf_alan.setText("⏳  PDF yükleniyor...")
        self._pdf_alan.show()
        self._pdf_img_scroll.hide()
        self._pdf_ac_btn.setEnabled(False)

        from services.moy_service import get_beyanname_pdf_bytes
        pdf_bytes = get_beyanname_pdf_bytes(self._musterino, b["kayit_no"])

        if not pdf_bytes:
            self._pdf_alan.setText(
                "⚠️  Bu beyanname için PDF verisi bulunamadı.\n"
                "(beyanname_gib tablosunda kayıt yok)"
            )
            return

        # Geçici PDF dosyasına yaz (Workspace İçi)
        import tempfile, os
        workspace_tmp = os.path.expanduser("~/NakitAkim/data/tmp")
        os.makedirs(workspace_tmp, exist_ok=True)
        tmp = tempfile.NamedTemporaryFile(
            suffix=".pdf",
            prefix=f"beyanname_{b['belge_turu']}_{b['donem_no']}_",
            dir=workspace_tmp,
            delete=False
        )
        tmp.write(pdf_bytes)
        tmp.close()
        self._tmp_path = tmp.name
        self._pdf_ac_btn.setEnabled(True)

        # PDF'i resme dönüştürüp önizle (pymupdf varsa)
        self._pdf_goruntule(pdf_bytes)

    def _pdf_goruntule(self, pdf_bytes: bytes):
        """PDF'i resim olarak göster — fitz (PyMuPDF) kullanır."""
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            toplam = doc.page_count

            from PyQt6.QtGui import QImage, QPixmap
            from PyQt6.QtCore import QByteArray
            import io

            # Tüm sayfaları birleştirerek tek uzun görüntü oluştur
            pixmaplar = []
            toplam_y = 0
            maks_x = 0

            for pg_no in range(min(toplam, 10)):  # max 10 sayfa
                pg = doc[pg_no]
                mat = fitz.Matrix(1.8, 1.8)  # 1.8x zoom — yeterli kalite
                clip = pg.get_pixmap(matrix=mat)
                img_data = clip.tobytes("png")
                pm = QPixmap()
                pm.loadFromData(img_data)
                pixmaplar.append(pm)
                toplam_y += pm.height() + 4
                maks_x = max(maks_x, pm.width())

            doc.close()

            if pixmaplar:
                # Tek pixmap olarak birleştir
                from PyQt6.QtGui import QPainter
                combined = QPixmap(maks_x, toplam_y)
                combined.fill()
                p = QPainter(combined)
                y = 0
                for pm in pixmaplar:
                    p.drawPixmap(0, y, pm)
                    y += pm.height() + 4
                p.end()

                self._pdf_img_container.setPixmap(combined)
                self._pdf_img_container.setFixedSize(combined.size())
                self._pdf_alan.hide()
                self._pdf_img_scroll.show()
            else:
                self._pdf_alan.setText("⚠️  PDF sayfası okunamadı.")

        except ImportError:
            # PyMuPDF yüklü değil — PDF bilgisi göster
            kb = len(pdf_bytes) // 1024
            self._pdf_alan.setText(
                f"📄  PDF indirme başarılı ({kb} KB)\n\n"
                f"PDF önizlemesi için PyMuPDF gerekli:\n"
                f"  pip install pymupdf\n\n"
                f"'📂 PDF'i Aç' butonu ile görüntüleyebilirsiniz."
            )
        except Exception as e:
            self._pdf_alan.setText(f"⚠️  PDF görüntülenemedi:\n{e}")

    # ── PDF'i sistem görüntüleyicisinde aç ───────────────────────────────────

    def _pdf_ac(self):
        if not self._tmp_path:
            return
        import subprocess, sys
        try:
            if sys.platform == "darwin":
                subprocess.Popen(["open", self._tmp_path])
            elif sys.platform == "win32":
                import os; os.startfile(self._tmp_path)
            else:
                subprocess.Popen(["xdg-open", self._tmp_path])
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Hata", f"PDF açılamadı: {e}")

    def closeEvent(self, event):
        # Geçici dosyayı temizle
        if self._tmp_path:
            import os
            try:
                os.unlink(self._tmp_path)
            except Exception:
                pass
        super().closeEvent(event)


# ─────────────────────────────────────────────────────────────────────────────
# Kredi Kartı Detay Dialog
# PHP: admin.php → Kredi Kartları kartı → modal (kart özet + ekstre tablo)
# ─────────────────────────────────────────────────────────────────────────────

class KrediKartiDialog(QDialog):
    """
    PHP admin.php 'Kredi Kartları' modalının PyQt6 karşılığı.

    Yapı (admin.php ile birebir):
    - Sol panel : Kart özet listesi (Banka / Kart bazında GROUP BY toplam)
    - Sağ panel: Seçilen karta ait ekstre satırları
    - Tarih filtresi + Filtrele butonu
    - Özet bantı: Kayıt sayısı + Toplam tutar
    """

    _TBL_STYLE = """
        QTableWidget {
            background: white;
            gridline-color: #e2e8f0;
            font-size: 12px;
            color: #1e293b;
            border: 1px solid #bfdbfe;
            border-radius: 8px;
        }
        QTableWidget::item {
            color: #1e293b;
            padding: 4px 6px;
        }
        QTableWidget::item:hover {
            background: #eff6ff;
            color: #1e293b;
        }
        QTableWidget::item:selected {
            background: #bfdbfe;
            color: #1e293b;
        }
        QTableWidget::item:alternate {
            background: #f0f7ff;
            color: #1e293b;
        }
        QHeaderView::section {
            background: #1e40af;
            color: white;
            font-weight: 700;
            font-size: 11px;
            padding: 6px 4px;
            border: none;
            border-right: 1px solid #1d4ed8;
        }
    """

    _CBS = (
        "QComboBox{background:white;border:1.5px solid #93c5fd;"
        "border-radius:6px;padding:0 8px;font-size:12px;color:#1e293b;}"
        "QComboBox::drop-down{border:none;width:18px;}"
    )
    _DES = (
        "QDateEdit{background:white;border:1.5px solid #93c5fd;"
        "border-radius:6px;padding:0 6px;font-size:12px;color:#1e293b;}"
        "QDateEdit::drop-down{border:none;width:18px;}"
    )

    def __init__(self, userid: int, yil: int, parent=None):
        super().__init__(parent)
        self._userid    = userid
        self._yil       = yil
        self._ozet_rows: list[dict] = []
        self._ekstre_rows: list[dict] = []
        self._secili_banka: str = ""

        self.setWindowTitle("💳  Kredi Kartları — Detay")
        self.setMinimumSize(1160, 660)
        self.resize(1280, 740)
        self._setup_ui()
        self._load_ozet()

    # ── UI İnşası ─────────────────────────────────────────────

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Başlık bantı ──
        hdr = QFrame()
        hdr.setFixedHeight(72)
        hdr.setStyleSheet(
            "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "stop:0 #1e40af,stop:0.5 #2563eb,stop:1 #0f172a);"
        )
        h = QHBoxLayout(hdr)
        h.setContentsMargins(20, 0, 20, 0)
        ic = QLabel("💳")
        ic.setStyleSheet("font-size:26px;background:transparent;")
        h.addWidget(ic)
        left = QVBoxLayout()
        t1 = QLabel("Kredi Kartları")
        t1.setStyleSheet("color:white;font-size:16px;font-weight:700;background:transparent;")
        t2 = QLabel(f"📅  {self._yil} yılı ekstre verileri  •  SQLite kaydından")
        t2.setStyleSheet("color:rgba(255,255,255,.80);font-size:12px;background:transparent;")
        left.addWidget(t1)
        left.addWidget(t2)
        h.addLayout(left, 1)
        root.addWidget(hdr)

        # ── İçerik: Sol kart özet + Sağ ekstre ──
        body = QFrame()
        body.setStyleSheet("background:#f0f7ff;")
        body_lay = QHBoxLayout(body)
        body_lay.setContentsMargins(12, 12, 12, 12)
        body_lay.setSpacing(12)

        # ── Sol Panel: Kart Özet Listesi ──
        sol = QFrame()
        sol.setFixedWidth(310)
        sol.setStyleSheet(
            "background:white;border-radius:10px;"
            "border:1.5px solid #bfdbfe;"
        )
        sol_lay = QVBoxLayout(sol)
        sol_lay.setContentsMargins(0, 0, 0, 0)
        sol_lay.setSpacing(0)

        sol_hdr = QLabel("💳  Kayıtlı Kartlar")
        sol_hdr.setFixedHeight(38)
        sol_hdr.setStyleSheet(
            "background:#1e40af;color:white;font-size:13px;font-weight:700;"
            "padding:0 12px;border-radius:9px 9px 0 0;"
        )
        sol_lay.addWidget(sol_hdr)

        self._ozet_tbl = QTableWidget()
        self._ozet_tbl.setColumnCount(4)
        self._ozet_tbl.setHorizontalHeaderLabels(["Kart / Banka", "Kayıt", "Harcama (₺)", "Net (₺)"])
        self._ozet_tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._ozet_tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._ozet_tbl.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._ozet_tbl.verticalHeader().setVisible(False)
        self._ozet_tbl.setAlternatingRowColors(True)
        self._ozet_tbl.setShowGrid(False)
        self._ozet_tbl.horizontalHeader().setStretchLastSection(True)
        self._ozet_tbl.setColumnWidth(0, 130)
        self._ozet_tbl.setColumnWidth(1, 42)
        self._ozet_tbl.setColumnWidth(2, 90)
        self._ozet_tbl.setStyleSheet(self._TBL_STYLE)
        self._ozet_tbl.cellClicked.connect(self._on_kart_sec)
        sol_lay.addWidget(self._ozet_tbl, 1)

        self._ozet_toplam_lbl = QLabel("")
        self._ozet_toplam_lbl.setFixedHeight(28)
        self._ozet_toplam_lbl.setStyleSheet(
            "background:#dbeafe;border-top:1px solid #bfdbfe;"
            "padding:0 10px;font-size:11px;color:#1e40af;font-weight:600;"
        )
        sol_lay.addWidget(self._ozet_toplam_lbl)
        body_lay.addWidget(sol)

        # ── Sağ Panel: Ekstre Tablosu ──
        sag = QFrame()
        sag.setStyleSheet(
            "background:white;border-radius:10px;"
            "border:1.5px solid #bfdbfe;"
        )
        sag_lay = QVBoxLayout(sag)
        sag_lay.setContentsMargins(0, 0, 0, 0)
        sag_lay.setSpacing(0)

        # Sağ başlık
        sag_hdr_frame = QFrame()
        sag_hdr_frame.setFixedHeight(38)
        sag_hdr_frame.setStyleSheet(
            "background:#2563eb;border-radius:9px 9px 0 0;"
        )
        sag_hdr_lay = QHBoxLayout(sag_hdr_frame)
        sag_hdr_lay.setContentsMargins(12, 0, 12, 0)
        self._sag_baslik = QLabel("📊  Kart seçin →")
        self._sag_baslik.setStyleSheet(
            "color:white;font-size:13px;font-weight:700;"
        )
        sag_hdr_lay.addWidget(self._sag_baslik, 1)
        sag_lay.addWidget(sag_hdr_frame)

        # Ay seçici + Filtre çubuğu
        filtre_frame = QFrame()
        filtre_frame.setFixedHeight(44)
        filtre_frame.setStyleSheet(
            "background:#eff6ff;border-bottom:1px solid #bfdbfe;"
        )
        filtre_lay = QHBoxLayout(filtre_frame)
        filtre_lay.setContentsMargins(10, 0, 10, 0)
        filtre_lay.setSpacing(8)

        # Ay combo
        filtre_lay.addWidget(self._lbl("Ay:"))
        self._ay_cb = QComboBox()
        self._ay_cb.setFixedSize(90, 30)
        self._ay_cb.setStyleSheet(self._CBS)
        AY_ADLARI = [
            "Hepsi", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
            "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık",
        ]
        for i, ad in enumerate(AY_ADLARI):
            self._ay_cb.addItem(ad, i)
        self._ay_cb.currentIndexChanged.connect(self._on_ay_degis)
        filtre_lay.addWidget(self._ay_cb)

        filtre_lay.addWidget(self._lbl("İlk Tarih:"))
        self._ilk_de = QDateEdit()
        self._ilk_de.setCalendarPopup(True)
        self._ilk_de.setDisplayFormat("dd.MM.yyyy")
        self._ilk_de.setFixedSize(120, 30)
        self._ilk_de.setStyleSheet(self._DES)
        self._ilk_de.setDate(QDate(self._yil, 1, 1))
        filtre_lay.addWidget(self._ilk_de)

        filtre_lay.addWidget(self._lbl("Son Tarih:"))
        self._son_de = QDateEdit()
        self._son_de.setCalendarPopup(True)
        self._son_de.setDisplayFormat("dd.MM.yyyy")
        self._son_de.setFixedSize(120, 30)
        self._son_de.setStyleSheet(self._DES)
        self._son_de.setDate(QDate.currentDate())
        filtre_lay.addWidget(self._son_de)

        filtre_lay.addStretch()

        self._filtre_btn = QPushButton("🔍  Filtrele")
        self._filtre_btn.setFixedSize(100, 30)
        self._filtre_btn.setStyleSheet(
            "QPushButton{background:#2563eb;color:white;border:none;"
            "border-radius:7px;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#1d4ed8;}"
            "QPushButton:disabled{background:#cbd5e1;}"
        )
        self._filtre_btn.setEnabled(False)
        self._filtre_btn.clicked.connect(self._load_ekstre)
        filtre_lay.addWidget(self._filtre_btn)
        sag_lay.addWidget(filtre_frame)

        # Özet bantı
        self._ekstre_ozet_lbl = QLabel("💳  Sol panelden bir kart seçin")
        self._ekstre_ozet_lbl.setFixedHeight(26)
        self._ekstre_ozet_lbl.setStyleSheet(
            "background:#dbeafe;border-bottom:1px solid #bfdbfe;"
            "padding:0 10px;font-size:11px;color:#1e40af;"
        )
        self._ekstre_ozet_lbl.setTextFormat(Qt.TextFormat.RichText)
        sag_lay.addWidget(self._ekstre_ozet_lbl)

        # Ekstre tablosu
        SUTUNLAR = [
            ("Tarih",       80),
            ("Açıklama",   320),
            ("Tutar",        110),
            ("Hesap Kodu",   90),
            ("Banka / Kart", 200),
        ]
        self._ekstre_tbl = QTableWidget()
        self._ekstre_tbl.setColumnCount(len(SUTUNLAR))
        self._ekstre_tbl.setHorizontalHeaderLabels([s[0] for s in SUTUNLAR])
        self._ekstre_tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._ekstre_tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._ekstre_tbl.setAlternatingRowColors(True)
        self._ekstre_tbl.verticalHeader().setVisible(False)
        self._ekstre_tbl.setSortingEnabled(True)
        self._ekstre_tbl.horizontalHeader().setStretchLastSection(True)
        for i, (_, w) in enumerate(SUTUNLAR):
            self._ekstre_tbl.setColumnWidth(i, w)
        self._ekstre_tbl.setStyleSheet(self._TBL_STYLE)
        sag_lay.addWidget(self._ekstre_tbl, 1)

        body_lay.addWidget(sag, 1)
        root.addWidget(body, 1)

        # ── Alt bar ──
        alt = QFrame()
        alt.setFixedHeight(48)
        alt.setStyleSheet("background:white;border-top:1px solid #e2e8f0;")
        a = QHBoxLayout(alt)
        a.setContentsMargins(16, 0, 16, 0)
        a.addStretch()
        kapat = QPushButton("✕  Kapat")
        kapat.setFixedSize(110, 32)
        kapat.setStyleSheet(
            "QPushButton{background:#64748b;color:white;border:none;"
            "border-radius:7px;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#475569;}"
        )
        kapat.clicked.connect(self.accept)
        a.addWidget(kapat)
        root.addWidget(alt)

    # ── Yardımcılar ───────────────────────────────────────────

    @staticmethod
    def _lbl(txt: str) -> QLabel:
        l = QLabel(txt)
        l.setStyleSheet("font-size:12px;color:#1e40af;font-weight:600;")
        return l

    # ── Ay Değişimi ────────────────────────────────────────────

    def _on_ay_degis(self):
        """Ay seçilince ilk/son tarihi otomatik doldurur, kart seçiliyse filtreler."""
        import calendar as _cal
        ay = self._ay_cb.currentData()   # 0=Hepsi, 1=Ocak ... 12=Aralık
        yil = self._yil
        if ay and ay > 0:
            son_gun = _cal.monthrange(yil, ay)[1]
            self._ilk_de.setDate(QDate(yil, ay, 1))
            self._son_de.setDate(QDate(yil, ay, son_gun))
        else:
            self._ilk_de.setDate(QDate(yil, 1, 1))
            self._son_de.setDate(QDate.currentDate())
        if self._secili_banka:
            self._load_ekstre()

    # ── Kart Özet Yükleme ─────────────────────────────────────────

    def _load_ozet(self):
        from services.dashboard_service import get_kredi_karti_kart_ozet
        self._ozet_rows = get_kredi_karti_kart_ozet(self._userid, self._yil)
        self._ozet_tbl.setSortingEnabled(False)
        self._ozet_tbl.setRowCount(0)

        from PyQt6.QtGui import QColor, QFont as QF
        toplam_borc  = 0.0
        toplam_odeme = 0.0
        toplam_net   = 0.0

        for row in self._ozet_rows:
            ri = self._ozet_tbl.rowCount()
            self._ozet_tbl.insertRow(ri)
            self._ozet_tbl.setRowHeight(ri, 28)

            banka  = row.get("Banka", "")
            kayit  = int(row.get("kayit_sayisi", 0))
            borc   = float(row.get("borc",  0))
            odeme  = float(row.get("odeme", 0))
            net    = float(row.get("net",   0))
            toplam_borc  += borc
            toplam_odeme += odeme
            toplam_net   += net

            banka_kisa = banka[:22] + ("…" if len(banka) > 22 else "")

            it0 = QTableWidgetItem(banka_kisa)
            it0.setToolTip(banka)
            it0.setForeground(QColor("#1e293b"))
            it0.setData(Qt.ItemDataRole.UserRole, ri)

            it1 = QTableWidgetItem(str(kayit))
            it1.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            it1.setForeground(QColor("#374151"))

            it2 = QTableWidgetItem(f"{borc:,.0f}")
            it2.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            it2.setForeground(QColor("#1d4ed8"))
            it2.setFont(QF("", -1, QF.Weight.Bold))

            net_clr = "#dc2626" if net < 0 else ("#059669" if net > 0 else "#6b7280")
            it3 = QTableWidgetItem(f"{net:,.0f}")
            it3.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            it3.setForeground(QColor(net_clr))
            it3.setFont(QF("", -1, QF.Weight.Bold))

            for ci, it in enumerate([it0, it1, it2, it3]):
                self._ozet_tbl.setItem(ri, ci, it)

        self._ozet_tbl.setSortingEnabled(True)
        self._ozet_toplam_lbl.setText(
            f"Harcama: {toplam_borc:,.0f} ₺  "
            f"Ödeme: {abs(toplam_odeme):,.0f} ₺  "
            f"Net: {toplam_net:,.0f} ₺"
        )

    # ── Kart Seçimi ────────────────────────────────────────────

    def _on_kart_sec(self, row: int, _col: int):
        if row < 0 or row >= len(self._ozet_rows):
            return
        row_data = self._ozet_rows[row]
        if row_data.get("Banka") == "sanal_pos":
            dlg = SanalPosDialog(self._userid, self._musterino, self._yil, self)
            dlg.exec()
        else:
            self._secili_banka = row_data.get("Banka", "")
            banka_kisa = self._secili_banka[:40] + ("…" if len(self._secili_banka) > 40 else "")
            self._sag_baslik.setText(f"💳  {banka_kisa}")
            self._filtre_btn.setEnabled(True)
            self._load_ekstre()

    # ── Ekstre Yükleme ─────────────────────────────────────────

    def _load_ekstre(self):
        if not self._secili_banka:
            return

        from services.dashboard_service import get_kredi_karti_ekstre_detay
        ilk = self._ilk_de.date()
        son = self._son_de.date()
        ilk_str = f"{ilk.day():02d}.{ilk.month():02d}.{ilk.year()}"
        son_str = f"{son.day():02d}.{son.month():02d}.{son.year()}"

        self._ekstre_rows = get_kredi_karti_ekstre_detay(
            userid=self._userid,
            banka=self._secili_banka,
            yil=self._yil,
            ilk_tarih=ilk_str,
            son_tarih=son_str,
        )
        self._doldur_ekstre(ilk_str, son_str)

    def _doldur_ekstre(self, ilk_str: str, son_str: str):
        from PyQt6.QtGui import QColor, QFont as QF
        self._ekstre_tbl.setSortingEnabled(False)
        self._ekstre_tbl.setRowCount(0)

        toplam_borc  = 0.0
        toplam_odeme = 0.0

        for row in self._ekstre_rows:
            ri = self._ekstre_tbl.rowCount()
            self._ekstre_tbl.insertRow(ri)
            self._ekstre_tbl.setRowHeight(ri, 26)

            tarih      = row.get("tarih", "")
            aciklama   = row.get("aciklama", "") or ""
            tutar_val  = float(row.get("alinan_tutar1") or 0)
            hesap_kodu = row.get("hesapKodu", "") or ""
            banka      = row.get("Banka", "") or ""
            tur        = row.get("tur", "borc")  # 'borc' | 'odeme'

            if tutar_val < 0:
                toplam_odeme += tutar_val
            else:
                toplam_borc  += tutar_val

            # Satır rengi: ödeme satırları kırmızı tonı, harcama satırları normal
            is_odeme = tutar_val < 0
            row_bg   = "#fff1f2" if is_odeme else None   # çok hafif kırmızı
            tutar_clr = "#dc2626" if is_odeme else "#1d4ed8"
            tutar_txt = f"{tutar_val:,.2f}"  # negatif işaretli kalacak

            ack_kisa = aciklama[:80] + ("…" if len(aciklama) > 80 else "")

            vals = [tarih, ack_kisa, tutar_txt, hesap_kodu, banka]
            aligns = [
                Qt.AlignmentFlag.AlignCenter,
                Qt.AlignmentFlag.AlignLeft,
                Qt.AlignmentFlag.AlignRight,
                Qt.AlignmentFlag.AlignCenter,
                Qt.AlignmentFlag.AlignLeft,
            ]
            colors = ["#374151", "#1e293b", tutar_clr, "#6b7280", "#374151"]
            bolds  = [False, False, True, False, False]

            for ci, (val, aln, clr, bold) in enumerate(zip(vals, aligns, colors, bolds)):
                it = QTableWidgetItem(str(val))
                if ci == 1:
                    it.setToolTip(aciklama)
                it.setTextAlignment(aln | Qt.AlignmentFlag.AlignVCenter)
                it.setForeground(QColor(clr))
                if bold:
                    it.setFont(QF("", -1, QF.Weight.Bold))
                # Ödeme satırı arka planı
                if row_bg:
                    it.setBackground(QColor(row_bg))
                self._ekstre_tbl.setItem(ri, ci, it)

        self._ekstre_tbl.setSortingEnabled(True)

        net = toplam_borc + toplam_odeme
        self._ekstre_ozet_lbl.setText(
            f"📅 {ilk_str} — {son_str}  │  "
            f"<b>{len(self._ekstre_rows)}</b> kayıt  │  "
            f"Harcama: <b style='color:#1d4ed8'>{toplam_borc:,.2f} ₺</b>  "
            f"Ödeme: <b style='color:#dc2626'>{abs(toplam_odeme):,.2f} ₺</b>  "
            f"Net: <b>{net:,.2f} ₺</b>"
        )


# ─────────────────────────────────────────────────────────────────────────────
# Sanal Pos Hareketleri Dialog
# PHP: lib/panelparcalari/admin/admin.php  → #sanalPosModal
#      sabit/js/admin_dashboard.js         → spHareketleriYukle()
# 14 sütunlu DataTable + İşlem/Ödeme/Fark özet bantları
# ─────────────────────────────────────────────────────────────────────────────

class SanalPosDialog(QDialog):
    """
    PHP admin.php #sanalPosModal + admin_dashboard.js spHareketleriYukle()
    → PyQt6 karşılığı.

    Üst bölüm : Başlık + '📦 PayTR Veritabanı' badge + Tarih filtresi
               + İşlem / Ödeme / Fark toplam bantları
    Alt bölüm : 14 sütunlu hareket tablosu (paytr SQLite tablosu)
    """

    SUTUNLAR = [
        ("İşlem Tarihi",    "islemtarihi",    120, Qt.AlignmentFlag.AlignCenter),
        ("Sipariş No",      "siparisno",      140, Qt.AlignmentFlag.AlignLeft),
        ("İşlem Tutarı",   "islemtutari",    110, Qt.AlignmentFlag.AlignRight),
        ("Ödeme Tutarı",   "odemetutari",    110, Qt.AlignmentFlag.AlignRight),
        ("Kur",            "kur",             55,  Qt.AlignmentFlag.AlignCenter),
        ("Mağaza No",      "magazano",        90,  Qt.AlignmentFlag.AlignCenter),
        ("Net Tutar",      "nettutar",        100, Qt.AlignmentFlag.AlignRight),
        ("Kesinti Tutarı", "kesintitutari",   110, Qt.AlignmentFlag.AlignRight),
        ("Kesinti Oranı",  "kesintiorani",     90, Qt.AlignmentFlag.AlignCenter),
        ("Kart Markası",   "kartmarkasi",     100, Qt.AlignmentFlag.AlignCenter),
        ("Kart No",        "kartno",          110, Qt.AlignmentFlag.AlignCenter),
        ("Ödeme Tipi",     "odemetipi",       100, Qt.AlignmentFlag.AlignCenter),
        ("Kart Tipi",      "karttipi",         90, Qt.AlignmentFlag.AlignCenter),
        ("Taksit Sayısı", "taksitsayisi",      80, Qt.AlignmentFlag.AlignCenter),
    ]

    def __init__(self, userid: int, musterino: str, yil: int, parent=None):
        super().__init__(parent)
        self._userid    = userid
        self._musterino = str(musterino)
        self._yil       = yil
        self._rows: list[dict] = []
        self._toplam_islem = 0.0
        self._toplam_odeme = 0.0

        self.setWindowTitle("💳  Sanal Pos Hareketleri — PayTR")
        self.setMinimumSize(1280, 740)
        self.resize(1380, 800)
        self._setup_ui()
        self._load()

    # ── UI İnşası ──────────────────────────────────────────────────────────

    def _setup_ui(self):
        from datetime import date
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Üst bant (gri arka plan — PHP'deki #f8f9fa) ──────────────────
        top = QFrame()
        top.setStyleSheet("background:#f8f9fa;border-bottom:1px solid #dee2e6;")
        top_lay = QVBoxLayout(top)
        top_lay.setContentsMargins(16, 12, 16, 12)
        top_lay.setSpacing(8)

        # Başlık satırı
        baslik_row = QHBoxLayout()
        baslik_lbl = QLabel("Sanal Pos Hareketleri")
        baslik_lbl.setStyleSheet(
            "font-size:16px;font-weight:700;color:#212529;"
        )
        baslik_row.addWidget(baslik_lbl)

        # Kaynak badge — PHP: #spKaynakBadge = '📦 PayTR Veritabanı'
        self._badge_lbl = QLabel("📦 PayTR Veritabanı")
        self._badge_lbl.setStyleSheet(
            "background:#212121;color:#fff;font-size:11px;font-weight:600;"
            "border-radius:10px;padding:2px 10px;"
        )
        self._badge_lbl.setFixedHeight(22)
        baslik_row.addWidget(self._badge_lbl)
        baslik_row.addStretch()
        top_lay.addLayout(baslik_row)

        # Filtre satırı
        filtre_row = QHBoxLayout()
        filtre_row.setSpacing(10)

        _DE = (
            "QDateEdit{background:white;border:1px solid #ced4da;"
            "border-radius:4px;padding:3px 6px;font-size:12px;color:#212529;}"
            "QDateEdit::drop-down{border:none;}"
        )
        today = date.today()
        jan1  = date(today.year, 1, 1)

        filtre_row.addWidget(self._lbl("İlk Tarih:"))
        self._ilk_de = QDateEdit()
        self._ilk_de.setCalendarPopup(True)
        self._ilk_de.setDisplayFormat("dd.MM.yyyy")
        self._ilk_de.setFixedHeight(30)
        self._ilk_de.setFixedWidth(120)
        self._ilk_de.setDate(QDate(jan1.year, jan1.month, jan1.day))
        self._ilk_de.setStyleSheet(_DE)
        filtre_row.addWidget(self._ilk_de)

        filtre_row.addWidget(self._lbl("Son Tarih:"))
        self._son_de = QDateEdit()
        self._son_de.setCalendarPopup(True)
        self._son_de.setDisplayFormat("dd.MM.yyyy")
        self._son_de.setFixedHeight(30)
        self._son_de.setFixedWidth(120)
        self._son_de.setDate(QDate(today.year, today.month, today.day))
        self._son_de.setStyleSheet(_DE)
        filtre_row.addWidget(self._son_de)

        self._listele_btn = QPushButton("Listele")
        self._listele_btn.setFixedHeight(30)
        self._listele_btn.setFixedWidth(80)
        self._listele_btn.setStyleSheet(
            "QPushButton{background:#212121;color:white;border:none;"
            "border-radius:4px;font-size:12px;font-weight:700;}"
            "QPushButton:hover{background:#343a40;}"
        )
        self._listele_btn.clicked.connect(self._load)
        filtre_row.addWidget(self._listele_btn)

        self._excel_btn = QPushButton("📥 Excel İndir")
        self._excel_btn.setFixedHeight(30)
        self._excel_btn.setFixedWidth(110)
        self._excel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._excel_btn.setStyleSheet(
            "QPushButton{background:#10b981;color:white;border:none;"
            "border-radius:4px;font-size:12px;font-weight:700;}"
            "QPushButton:hover{background:#059669;}"
        )
        self._excel_btn.clicked.connect(self._export_excel)
        filtre_row.addWidget(self._excel_btn)

        filtre_row.addStretch()
        top_lay.addLayout(filtre_row)
        root.addWidget(top)

        # ── Özet bant (İşlem / Ödeme / Fark) ─────────────────────────────
        # PHP: #spIslemToplam / #spOdemeToplam / #spFarkToplam
        ozet = QFrame()
        ozet.setFixedHeight(64)
        ozet.setStyleSheet(
            "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "stop:0 #1a1a2e,stop:0.5 #16213e,stop:1 #0f3460);"
        )
        oz_lay = QHBoxLayout(ozet)
        oz_lay.setContentsMargins(20, 0, 20, 0)
        oz_lay.setSpacing(40)

        def _blok(label: str, attr: str, clr: str):
            col = QVBoxLayout()
            col.setSpacing(1)
            lbl_top = QLabel(label)
            lbl_top.setStyleSheet(
                "font-size:10px;color:rgba(255,255,255,.65);background:transparent;"
            )
            lbl_val = QLabel("-")
            lbl_val.setStyleSheet(
                f"font-size:18px;font-weight:700;color:{clr};background:transparent;"
            )
            col.addWidget(lbl_top)
            col.addWidget(lbl_val)
            setattr(self, attr, lbl_val)
            return col

        oz_lay.addLayout(_blok("İşlem Toplamı",  "_islem_lbl", "#dc3545"))
        oz_lay.addLayout(_blok("Ödeme Toplamı",  "_odeme_lbl", "#28a745"))
        oz_lay.addLayout(_blok("Fark",            "_fark_lbl",  "#ffffff"))
        oz_lay.addStretch()

        self._kayit_chip = QLabel("")
        self._kayit_chip.setStyleSheet(
            "background:rgba(255,255,255,.12);color:white;font-size:11px;"
            "font-weight:600;border-radius:10px;padding:2px 10px;"
        )
        oz_lay.addWidget(self._kayit_chip, 0, Qt.AlignmentFlag.AlignVCenter)
        root.addWidget(ozet)

        # ── Tablo ─────────────────────────────────────────────────────────
        self._tbl = QTableWidget()
        self._tbl.setColumnCount(len(self.SUTUNLAR))
        self._tbl.setHorizontalHeaderLabels([s[0] for s in self.SUTUNLAR])
        self._tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._tbl.setAlternatingRowColors(True)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.setSortingEnabled(True)
        hdr = self._tbl.horizontalHeader()
        hdr.setStretchLastSection(False)
        for i, (_, _, w, _) in enumerate(self.SUTUNLAR):
            self._tbl.setColumnWidth(i, w)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self._tbl.setStyleSheet("""
            QTableWidget {
                background: white;
                gridline-color: #e2e8f0;
                font-size: 12px;
                color: #1e293b;
                border: none;
            }
            QTableWidget::item {
                color: #1e293b;
                padding: 3px 6px;
            }
            QTableWidget::item:hover {
                background: #f0f9ff;
                color: #1e293b;
            }
            QTableWidget::item:selected {
                background: #dbeafe;
                color: #1e293b;
            }
            QTableWidget::item:alternate {
                background: #f8fafc;
            }
            QHeaderView::section {
                background: #212121;
                color: white;
                font-weight: 700;
                font-size: 11px;
                padding: 6px 4px;
                border: none;
                border-right: 1px solid #374151;
            }
        """)
        root.addWidget(self._tbl, 1)

        # ── Alt bar ───────────────────────────────────────────────────────
        alt = QFrame()
        alt.setFixedHeight(48)
        alt.setStyleSheet(
            "background:#f8f9fa;border-top:1px solid #dee2e6;"
        )
        a = QHBoxLayout(alt)
        a.setContentsMargins(16, 0, 16, 0)
        self._durum_lbl = QLabel("")
        self._durum_lbl.setStyleSheet("font-size:11px;color:#6c757d;")
        a.addWidget(self._durum_lbl)
        a.addStretch()
        kapat = QPushButton("Kapat")
        kapat.setFixedSize(110, 32)
        kapat.setStyleSheet(
            "QPushButton{background:#dc3545;color:white;border:none;"
            "border-radius:5px;font-size:13px;font-weight:700;}"
            "QPushButton:hover{background:#c82333;}"
        )
        kapat.clicked.connect(self.accept)
        a.addWidget(kapat)
        root.addWidget(alt)

    # ── Yardımcılar ───────────────────────────────────────────────────────

    @staticmethod
    def _lbl(txt: str) -> QLabel:
        l = QLabel(txt)
        l.setStyleSheet("font-size:12px;color:#495057;font-weight:600;")
        return l

    # ── Veri Yükleme ──────────────────────────────────────────────────────

    def _load(self):
        """Tarih aralığına göre paytr tablosundan satırları çeker."""
        from services.paytr_service import get_sanal_pos_hareketleri_db

        self._listele_btn.setEnabled(False)
        self._durum_lbl.setText("⏳  Yükleniyor...")
        self._islem_lbl.setText("Yükleniyor...")
        self._odeme_lbl.setText("...")
        self._fark_lbl.setText("...")

        ilk_qd = self._ilk_de.date()
        son_qd = self._son_de.date()
        ilk_str = f"{ilk_qd.year():04d}-{ilk_qd.month():02d}-{ilk_qd.day():02d}"
        son_str = f"{son_qd.year():04d}-{son_qd.month():02d}-{son_qd.day():02d}"

        result = get_sanal_pos_hareketleri_db(
            self._userid, self._musterino, ilk_str, son_str
        )

        self._listele_btn.setEnabled(True)

        if not result.get("success"):
            self._islem_lbl.setText("Hata")
            self._odeme_lbl.setText("-")
            self._fark_lbl.setText("-")
            self._durum_lbl.setText(f"❌  {result.get('message', 'Bilinmeyen hata')}")
            return

        self._rows = result.get("data", [])
        self._toplam_islem = result.get("toplam_islem", 0.0)
        self._toplam_odeme = result.get("toplam_odeme", 0.0)
        self._doldur(result)

    def _doldur(self, result: dict):
        from PyQt6.QtGui import QColor, QFont as QF
        import re

        # Özet bantları
        toplam_islem = self._toplam_islem
        toplam_odeme = self._toplam_odeme
        fark         = toplam_odeme - toplam_islem

        self._islem_lbl.setText(result.get("toplam_islem_fmt", "-"))
        self._odeme_lbl.setText(result.get("toplam_odeme_fmt", "-"))
        fark_fmt = result.get("toplam_fark_fmt", "-")
        self._fark_lbl.setText(fark_fmt)
        self._fark_lbl.setStyleSheet(
            f"font-size:18px;font-weight:700;background:transparent;"
            f"color:{'#28a745' if fark >= 0 else '#dc3545'};"
        )

        kayit_sayisi = result.get("kayit_sayisi", len(self._rows))
        self._kayit_chip.setText(f"📅 {kayit_sayisi:,} kayıt")

        ilk_qd = self._ilk_de.date()
        son_qd = self._son_de.date()
        self._durum_lbl.setText(
            f"📅 {ilk_qd.toString('dd.MM.yyyy')} — {son_qd.toString('dd.MM.yyyy')}"
            f"   │   {kayit_sayisi:,} kayıt   │   "
            f"İşlem: {result.get('toplam_islem_fmt','-')}  "
            f"Ödeme: {result.get('toplam_odeme_fmt','-')}  "
            f"Fark: {fark_fmt}"
        )

        # Tarih formatlama (DD.MM.YYYY)
        def _tarih_fmt(v) -> str:
            if not v:
                return ""
            s = str(v).strip()
            m = re.match(r'^(\d{1,2})\.(\d{1,2})\.(\d{4})', s)
            if m:
                return f"{m.group(1).zfill(2)}.{m.group(2).zfill(2)}.{m.group(3)}"
            m2 = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
            if m2:
                return f"{m2.group(3)}.{m2.group(2)}.{m2.group(1)}"
            return s

        # Tabloyu doldur
        self._tbl.setSortingEnabled(False)
        self._tbl.setRowCount(0)

        for row in self._rows:
            ri = self._tbl.rowCount()
            self._tbl.insertRow(ri)
            self._tbl.setRowHeight(ri, 24)

            for ci, (_, field, _, align) in enumerate(self.SUTUNLAR):
                raw = row.get(field, "")

                # Sayısal sütunlar — kırmızı/yeşil renklendirme (PHP DataTable render)
                if field in ("islemtutari", "odemetutari", "nettutar", "kesintitutari"):
                    try:
                        val_f = float(raw or 0)
                    except (ValueError, TypeError):
                        val_f = 0.0
                    txt = f"{val_f:,.2f}"
                    clr = (
                        "#dc3545" if field in ("islemtutari", "kesintitutari")
                        else "#28a745"
                    )
                    it = QTableWidgetItem(txt)
                    it.setForeground(QColor(clr))
                    it.setFont(QF("", -1, QF.Weight.Bold))
                    it.setData(Qt.ItemDataRole.UserRole, val_f)  # sayısal sıralama

                elif field == "islemtarihi":
                    txt = _tarih_fmt(raw)
                    it  = QTableWidgetItem(txt)
                    it.setForeground(QColor("#374151"))

                else:
                    txt = str(raw) if raw is not None else ""
                    it  = QTableWidgetItem(txt)
                    it.setForeground(QColor("#374151"))

                it.setTextAlignment(align | Qt.AlignmentFlag.AlignVCenter)
                self._tbl.setItem(ri, ci, it)

        self._tbl.setSortingEnabled(True)

    def _export_excel(self):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import re
        import datetime

        path, _ = QFileDialog.getSaveFileName(
            self, "Excel Olarak Kaydet", "sanal_pos_hareketleri.xlsx", "Excel Files (*.xlsx)"
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sanal Pos"

            # Excel Stylings
            font_title = Font(name="Segoe UI", size=14, bold=True, color="212121")
            font_subtitle = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            font_data = Font(name="Segoe UI", size=10, color="1F2937")
            font_total = Font(name="Segoe UI", size=11, bold=True, color="212121")

            fill_header = PatternFill(start_color="212121", end_color="212121", fill_type="solid") # Dark
            fill_total = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")  # Light Gray

            border_thin = Border(
                left=Side(style="thin", color="E5E7EB"),
                right=Side(style="thin", color="E5E7EB"),
                top=Side(style="thin", color="E5E7EB"),
                bottom=Side(style="thin", color="E5E7EB")
            )
            border_total = Border(
                top=Side(style="thin", color="D1D5DB"),
                bottom=Side(style="double", color="212121")
            )

            # 1. Rapor Başlık Bloğu
            ws.append(["Sanal Pos Hareketleri Raporu (PayTR)"])
            ws.cell(row=1, column=1).font = font_title

            ilk_qd = self._ilk_de.date()
            son_qd = self._son_de.date()
            tarih_araligi = f"Tarih Aralığı: {ilk_qd.toString('dd.MM.yyyy')} — {son_qd.toString('dd.MM.yyyy')}"
            ws.append([tarih_araligi])
            ws.cell(row=2, column=1).font = font_subtitle

            ws.append([]) # Boşluk

            # Headers
            headers = [col[0] for col in self.SUTUNLAR]
            ws.append(headers)

            header_row_idx = 4
            for col_idx in range(len(headers)):
                cell = ws.cell(row=header_row_idx, column=col_idx+1)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_thin

            # Sum tracker for columns
            column_sums = {col: 0.0 for col in range(len(self.SUTUNLAR))}

            # Rows writing
            current_row_idx = 5
            for row in range(self._tbl.rowCount()):
                if self._tbl.isRowHidden(row):
                    continue

                row_data = []
                for col in range(self._tbl.columnCount()):
                    item = self._tbl.item(row, col)
                    cell_text = item.text() if item else ""
                    field_name = self.SUTUNLAR[col][1]

                    # Check if it is an amount column
                    if field_name in ("islemtutari", "odemetutari", "nettutar", "kesintitutari"):
                        val_num = 0.0
                        if item:
                            val_data = item.data(Qt.ItemDataRole.UserRole)
                            if val_data is not None:
                                try:
                                    val_num = float(val_data)
                                except (ValueError, TypeError):
                                    val_num = 0.0
                            else:
                                # Fallback parse
                                s = cell_text.replace("₺", "").replace("TL", "").replace("$", "").replace("€", "").replace("+", "").strip()
                                s = s.replace(".", "").replace(",", ".")
                                try:
                                    val_num = float(s)
                                except ValueError:
                                    val_num = 0.0
                        row_data.append(val_num)
                        column_sums[col] += val_num
                    else:
                        row_data.append(cell_text)

                ws.append(row_data)

                # Style active data row
                for col in range(len(row_data)):
                    cell = ws.cell(row=current_row_idx, column=col+1)
                    cell.font = font_data
                    cell.border = border_thin
                    field_name = self.SUTUNLAR[col][1]

                    if field_name in ("islemtutari", "odemetutari", "nettutar", "kesintitutari"):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        # Format is exactly 0.00 with dot decimal separator, no currency symbol, no thousands separator
                        cell.number_format = '0.00'
                    elif field_name == "islemtarihi":
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                current_row_idx += 1

            # 2. Dynamic GENEL TOPLAM Row
            summary_row = []
            for col in range(self._tbl.columnCount()):
                field_name = self.SUTUNLAR[col][1]
                if col == 0:
                    summary_row.append("GENEL TOPLAM")
                elif field_name in ("islemtutari", "odemetutari", "nettutar", "kesintitutari"):
                    summary_row.append(column_sums[col])
                else:
                    summary_row.append("")

            ws.append(summary_row)

            # Style summary row
            for col in range(len(summary_row)):
                cell = ws.cell(row=current_row_idx, column=col+1)
                cell.font = font_total
                cell.fill = fill_total
                cell.border = border_total
                field_name = self.SUTUNLAR[col][1]

                if col == 0:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif field_name in ("islemtutari", "odemetutari", "nettutar", "kesintitutari"):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    # Amount columns use '.' decimal separator, no currency symbol, no thousands separator
                    cell.number_format = '0.00'

            # 3. Auto Width Adjustment
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row in (1, 2, 3): # Skip title headers for width
                        continue
                    if cell.value is not None:
                        if isinstance(cell.value, float):
                            val_str = f"{cell.value:.2f}"
                        else:
                            val_str = str(cell.value)
                        max_len = max(max_len, len(val_str))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            # Set gridlines visible
            ws.views.sheetView[0].showGridLines = True

            wb.save(path)

            msg = QMessageBox(self)
            msg.setWindowTitle("Başarılı")
            msg.setText("Sanal Pos Hareketleri Excel raporu başarıyla kaydedildi!")
            msg.setStyleSheet("""
                QMessageBox { background-color: white; }
                QLabel { color: #1F2937; font-size: 13px; font-weight: 600; min-width: 280px; min-height: 40px; }
                QPushButton { background-color: #212121; color: white; border: none; border-radius: 6px; padding: 6px 18px; font-size: 11px; font-weight: bold; }
                QPushButton:hover { background-color: #343a40; }
            """)
            msg.exec()

        except Exception as e:
            import traceback
            err_msg = f"Excel kaydedilirken hata oluştu:\n{e}\n\nDetay:\n{traceback.format_exc()}"
            print(err_msg)

            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Icon.Critical)
            msg.setWindowTitle("Hata")
            msg.setText(f"Excel raporu oluşturulurken beklenmedik hata oluştu:\n{e}")
            msg.setStyleSheet("""
                QMessageBox { background-color: white; }
                QLabel { color: #DC2626; font-size: 12px; font-weight: 600; min-width: 240px; }
                QPushButton { background-color: #DC2626; color: white; border: none; border-radius: 6px; padding: 6px 16px; }
            """)
            msg.exec()

# ─────────────────────────────────────────────────────────────────────────────
# Fiziksel Pos Hareketleri Dialog
# PHP: lib/panelparcalari/admin/admin.php  → #fizikselPosModal
#      sabit/js/admin_dashboard.js         → fpHareketleriYukle()
# 9 sütunlu DataTable + İşlem/İşyeri Ücreti/Net Tutar özet bantları
# ─────────────────────────────────────────────────────────────────────────────

class FizikselPosDialog(QDialog):
    """
    PHP admin.php #fizikselPosModal + admin_dashboard.js fpHareketleriYukle()
    → PyQt6 karşılığı.

    Üst bölüm : Başlık + '🏪 Womsis Yerel DB' badge + Tarih filtresi
    Özet bant : İşlem Toplamı (kırmızı) | İşyeri Ücreti (turuncu) | Net Tutar (yeşil)
    Alt bölüm : 9 sütunlu hareket tablosu (womsi_pos SQLite tablosu)
    """

    SUTUNLAR = [
        ("İşyeri No",           "isyerino",          100, Qt.AlignmentFlag.AlignLeft),
        ("Cari Hesap",          "carihesap",          140, Qt.AlignmentFlag.AlignLeft),
        ("Hesaba Geçiş Tarihi", "hesabagecistarihi",  130, Qt.AlignmentFlag.AlignCenter),
        ("İşlem Tutarı",        "islemtutari",        110, Qt.AlignmentFlag.AlignRight),
        ("İşlem Tarihi",        "islemtarihi",        110, Qt.AlignmentFlag.AlignCenter),
        ("POS No",              "posno",               90, Qt.AlignmentFlag.AlignCenter),
        ("İşyeri Ücreti",       "isyeritutar",        120, Qt.AlignmentFlag.AlignRight),
        ("Net Tutar",           "nettutar",           100, Qt.AlignmentFlag.AlignRight),
        ("Brand",               "brand",               90, Qt.AlignmentFlag.AlignCenter),
    ]

    def __init__(self, userid: int, parent=None):
        super().__init__(parent)
        self._userid = userid
        self._rows: list[dict] = []
        self.setWindowTitle("🏪  Fiziksel Pos Hareketleri — Womsis")
        self.setMinimumSize(1200, 720)
        self.resize(1340, 780)
        self._setup_ui()
        self._load()

    def _setup_ui(self):
        from datetime import date
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Üst bant
        top = QFrame()
        top.setStyleSheet("background:#f8f9fa;border-bottom:1px solid #dee2e6;")
        tl = QVBoxLayout(top)
        tl.setContentsMargins(16, 12, 16, 12)
        tl.setSpacing(8)

        br = QHBoxLayout()
        bl = QLabel("Fiziksel Pos Hareketleri")
        bl.setStyleSheet("font-size:16px;font-weight:700;color:#212529;")
        br.addWidget(bl)
        badge = QLabel("🏪 Womsis Yerel DB")
        badge.setStyleSheet(
            "background:#1a3a5c;color:#fff;font-size:11px;font-weight:600;"
            "border-radius:10px;padding:2px 10px;"
        )
        badge.setFixedHeight(22)
        br.addWidget(badge)
        br.addStretch()
        tl.addLayout(br)

        fr = QHBoxLayout()
        fr.setSpacing(10)
        _DE = (
            "QDateEdit{background:white;border:1px solid #ced4da;"
            "border-radius:4px;padding:3px 6px;font-size:12px;color:#212529;}"
            "QDateEdit::drop-down{border:none;}"
        )
        today = date.today()
        jan1 = date(today.year, 1, 1)

        fr.addWidget(self._lbl("İlk Tarih:"))
        self._ilk_de = QDateEdit()
        self._ilk_de.setCalendarPopup(True)
        self._ilk_de.setDisplayFormat("dd.MM.yyyy")
        self._ilk_de.setFixedHeight(30)
        self._ilk_de.setFixedWidth(120)
        self._ilk_de.setDate(QDate(jan1.year, jan1.month, jan1.day))
        self._ilk_de.setStyleSheet(_DE)
        fr.addWidget(self._ilk_de)

        fr.addWidget(self._lbl("Son Tarih:"))
        self._son_de = QDateEdit()
        self._son_de.setCalendarPopup(True)
        self._son_de.setDisplayFormat("dd.MM.yyyy")
        self._son_de.setFixedHeight(30)
        self._son_de.setFixedWidth(120)
        self._son_de.setDate(QDate(today.year, today.month, today.day))
        self._son_de.setStyleSheet(_DE)
        fr.addWidget(self._son_de)

        self._listele_btn = QPushButton("Listele")
        self._listele_btn.setFixedHeight(30)
        self._listele_btn.setFixedWidth(80)
        self._listele_btn.setStyleSheet(
            "QPushButton{background:#1a3a5c;color:white;border:none;"
            "border-radius:4px;font-size:12px;font-weight:700;}"
            "QPushButton:hover{background:#0d2137;}"
        )
        self._listele_btn.clicked.connect(self._load)
        fr.addWidget(self._listele_btn)
        fr.addStretch()
        tl.addLayout(fr)
        root.addWidget(top)

        # Özet bant — PHP: #fpIslemToplam / #fpIsyeriToplam / #fpNetToplam
        ozet = QFrame()
        ozet.setFixedHeight(64)
        ozet.setStyleSheet(
            "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "stop:0 #1a3a5c,stop:0.5 #0d2137,stop:1 #061422);"
        )
        oz = QHBoxLayout(ozet)
        oz.setContentsMargins(20, 0, 20, 0)
        oz.setSpacing(40)

        def _blok(lbl: str, attr: str, clr: str):
            col = QVBoxLayout()
            col.setSpacing(1)
            h = QLabel(lbl)
            h.setStyleSheet("font-size:10px;color:rgba(255,255,255,.65);background:transparent;")
            v = QLabel("-")
            v.setStyleSheet(f"font-size:18px;font-weight:700;color:{clr};background:transparent;")
            col.addWidget(h)
            col.addWidget(v)
            setattr(self, attr, v)
            return col

        oz.addLayout(_blok("İşlem Toplamı",  "_islem_lbl",  "#dc3545"))
        oz.addLayout(_blok("İşyeri Ücreti",  "_isyeri_lbl", "#e67e22"))
        oz.addLayout(_blok("Net Tutar",      "_net_lbl",    "#28a745"))
        oz.addStretch()

        self._kayit_chip = QLabel("")
        self._kayit_chip.setStyleSheet(
            "background:rgba(255,255,255,.12);color:white;font-size:11px;"
            "font-weight:600;border-radius:10px;padding:2px 10px;"
        )
        oz.addWidget(self._kayit_chip, 0, Qt.AlignmentFlag.AlignVCenter)
        root.addWidget(ozet)

        # Tablo
        self._tbl = QTableWidget()
        self._tbl.setColumnCount(len(self.SUTUNLAR))
        self._tbl.setHorizontalHeaderLabels([s[0] for s in self.SUTUNLAR])
        self._tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._tbl.setAlternatingRowColors(True)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.setSortingEnabled(True)
        hdr = self._tbl.horizontalHeader()
        hdr.setStretchLastSection(False)
        for i, (_, _, w, _) in enumerate(self.SUTUNLAR):
            self._tbl.setColumnWidth(i, w)
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self._tbl.setStyleSheet("""
            QTableWidget { background:white; gridline-color:#e2e8f0;
                font-size:12px; color:#1e293b; border:none; }
            QTableWidget::item { color:#1e293b; padding:3px 6px; }
            QTableWidget::item:hover { background:#f0f9ff; color:#1e293b; }
            QTableWidget::item:selected { background:#dbeafe; color:#1e293b; }
            QTableWidget::item:alternate { background:#f8fafc; }
            QHeaderView::section { background:#1a3a5c; color:white;
                font-weight:700; font-size:11px; padding:6px 4px;
                border:none; border-right:1px solid #0d2137; }
        """)
        root.addWidget(self._tbl, 1)

        # Alt bar
        alt = QFrame()
        alt.setFixedHeight(48)
        alt.setStyleSheet("background:#f8f9fa;border-top:1px solid #dee2e6;")
        a = QHBoxLayout(alt)
        a.setContentsMargins(16, 0, 16, 0)
        self._durum_lbl = QLabel("")
        self._durum_lbl.setStyleSheet("font-size:11px;color:#6c757d;")
        a.addWidget(self._durum_lbl)
        a.addStretch()
        kapat = QPushButton("Kapat")
        kapat.setFixedSize(110, 32)
        kapat.setStyleSheet(
            "QPushButton{background:#dc3545;color:white;border:none;"
            "border-radius:5px;font-size:13px;font-weight:700;}"
            "QPushButton:hover{background:#c82333;}"
        )
        kapat.clicked.connect(self.accept)
        a.addWidget(kapat)
        root.addWidget(alt)

    @staticmethod
    def _lbl(txt: str) -> QLabel:
        l = QLabel(txt)
        l.setStyleSheet("font-size:12px;color:#495057;font-weight:600;")
        return l

    def _load(self):
        """PHP: fpHareketleriYukle() — womsi_pos tablosundan tarih filtreli sorgu."""
        from services.fiziksel_pos_service import get_hareketler
        self._listele_btn.setEnabled(False)
        self._durum_lbl.setText("⏳  Yükleniyor...")
        self._islem_lbl.setText("Yükleniyor...")
        self._isyeri_lbl.setText("...")
        self._net_lbl.setText("...")

        ilk_qd = self._ilk_de.date()
        son_qd = self._son_de.date()
        ilk_str = f"{ilk_qd.year():04d}-{ilk_qd.month():02d}-{ilk_qd.day():02d}"
        son_str = f"{son_qd.year():04d}-{son_qd.month():02d}-{son_qd.day():02d}"

        result = get_hareketler(self._userid, ilk_str, son_str)
        self._listele_btn.setEnabled(True)

        if not result.get("success"):
            self._islem_lbl.setText("Hata")
            self._isyeri_lbl.setText("-")
            self._net_lbl.setText("-")
            self._durum_lbl.setText(f"❌  {result.get('message', 'Bilinmeyen hata')}")
            return

        self._rows = result.get("data", [])
        self._doldur(result)

    def _doldur(self, result: dict):
        from PyQt6.QtGui import QColor, QFont as QF
        # PHP: #fpIslemToplam / #fpIsyeriToplam / #fpNetToplam
        self._islem_lbl.setText(result.get("toplam_islem_fmt",  "-"))
        self._isyeri_lbl.setText(result.get("toplam_isyeri_fmt", "-"))
        self._net_lbl.setText(result.get("toplam_net_fmt",    "-"))

        kayit = result.get("kayit_sayisi", len(self._rows))
        self._kayit_chip.setText(f"📅 {kayit:,} kayıt")

        ilk_qd = self._ilk_de.date()
        son_qd = self._son_de.date()
        self._durum_lbl.setText(
            f"📅 {ilk_qd.toString('dd.MM.yyyy')} — {son_qd.toString('dd.MM.yyyy')}"
            f"   │   {kayit:,} kayıt   │   "
            f"İşlem: {result.get('toplam_islem_fmt','-')}  "
            f"İşyeri: {result.get('toplam_isyeri_fmt','-')}  "
            f"Net: {result.get('toplam_net_fmt','-')}"
        )

        self._tbl.setSortingEnabled(False)
        self._tbl.setRowCount(0)

        for row in self._rows:
            ri = self._tbl.rowCount()
            self._tbl.insertRow(ri)
            self._tbl.setRowHeight(ri, 24)
            for ci, (_, field, _, align) in enumerate(self.SUTUNLAR):
                raw = row.get(field, "")
                if field == "islemtutari":
                    clr, bold = "#dc3545", True
                elif field == "isyeritutar":
                    clr, bold = "#e67e22", True
                elif field == "nettutar":
                    clr, bold = "#28a745", True
                else:
                    clr, bold = "#374151", False

                if field in ("islemtutari", "isyeritutar", "nettutar"):
                    try:
                        val_f = float(raw or 0)
                    except (ValueError, TypeError):
                        val_f = 0.0
                    it = QTableWidgetItem(f"{val_f:,.2f}")
                    it.setData(Qt.ItemDataRole.UserRole, val_f)
                else:
                    it = QTableWidgetItem(str(raw) if raw is not None else "")

                it.setForeground(QColor(clr))
                if bold:
                    it.setFont(QF("", -1, QF.Weight.Bold))
                it.setTextAlignment(align | Qt.AlignmentFlag.AlignVCenter)
                self._tbl.setItem(ri, ci, it)

        self._tbl.setSortingEnabled(True)

        if not self._rows:
            self._tbl.insertRow(0)
            self._tbl.setRowHeight(0, 48)
            empty = QTableWidgetItem(
                "Bu tarih aralığında fiziksel POS hareketi bulunamadı."
            )
            empty.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            empty.setForeground(QColor("#6c757d"))
            self._tbl.setItem(0, 0, empty)
            self._tbl.setSpan(0, 0, 1, len(self.SUTUNLAR))

# ─────────────────────────────────────────────────────────────────────────────
# Maaş Kira Smm Dialog
# PHP: admin_dashboard.js → vmModalAc('Maaş Kira Smm') + vmBuildTable()
#      ajax/ayarlar/vergiMuhtasarGetir.php
# 5 sütunlu DataTable:
#   Dönem | Açıklama | Gayri Resmi Tutar (turuncu) |
#   Vergi Kesinti Tutarı (kırmızı) | Fark (yeşil)
# ─────────────────────────────────────────────────────────────────────────────

class MaasKiraSmmDialog(QDialog):
    """
    PHP admin.php #vergilerModal (maas_kira_smmm modu) + vmBuildTable()
    → PyQt6 karşılığı.

    Üst bölüm : Başlık + Dönem filtresi
    Özet bant : Gayri Resmi (turuncu) | Kesintiler (kırmızı) | Fark (yeşil)
    Alt bölüm : 5 sütunlu hareket tablosu (VergiMuhtasar SQLite)
    """

    SUTUNLAR = [
        # (başlık, alan_adı, genişlik, hizalama, renk)
        ("Dönem",                "donem",        100, Qt.AlignmentFlag.AlignLeft,    None),
        ("Açıklama",             "ack",          200, Qt.AlignmentFlag.AlignLeft,    None),
        ("Gayri Resmi Tutar",    "gaytutar",     140, Qt.AlignmentFlag.AlignRight, "#ff9800"),
        ("Vergi Kesinti Tutarı", "vergkestutar", 140, Qt.AlignmentFlag.AlignRight, "#dc3545"),
        ("Fark",                 "fark",         130, Qt.AlignmentFlag.AlignRight, "#28a745"),
    ]

    def __init__(self, userid: int, musterino: str = None, yil: int = None, parent=None):
        super().__init__(parent)
        self._userid    = userid
        self._musterino = musterino
        self._yil       = yil
        self._rows: list[dict] = []
        self._all_rows: list[dict] = []

        title_yil = f" ({yil})" if yil else ""
        self.setWindowTitle(f"💼  Maaş Kira Smm — Vergi Muhtasar{title_yil}")
        self.setMinimumSize(1000, 660)
        self.resize(1160, 720)
        self._setup_ui()
        self._load()

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Üst bant ─────────────────────────────────────────────────────────
        top = QFrame()
        top.setStyleSheet("background:#f8f9fa;border-bottom:1px solid #dee2e6;")
        tl = QVBoxLayout(top)
        tl.setContentsMargins(16, 12, 16, 12)
        tl.setSpacing(8)

        br = QHBoxLayout()
        bl = QLabel("Maaş Kira Smm — Vergi Muhtasar")
        bl.setStyleSheet("font-size:16px;font-weight:700;color:#212529;")
        br.addWidget(bl)
        badge = QLabel("💼 Personel · Kira · Müşavirlik")
        badge.setStyleSheet(
            "background:#1a3a5c;color:#fff;font-size:11px;font-weight:600;"
            "border-radius:10px;padding:2px 12px;"
        )
        badge.setFixedHeight(22)
        br.addWidget(badge)
        br.addStretch()
        tl.addLayout(br)

        # Dönem + Açıklama filtresi
        fr = QHBoxLayout()
        fr.setSpacing(10)
        _CB = (
            "QComboBox{background:white;border:1px solid #ced4da;border-radius:4px;"
            "padding:3px 8px;font-size:12px;color:#212529;min-height:28px;}"
        )
        fr.addWidget(self._lbl("Dönem:"))
        self._donem_cb = QComboBox()
        self._donem_cb.setFixedWidth(120)
        self._donem_cb.setStyleSheet(_CB)
        self._donem_cb.addItem("Tümü", "")
        fr.addWidget(self._donem_cb)

        fr.addWidget(self._lbl("Açıklama:"))
        self._ack_cb = QComboBox()
        self._ack_cb.setFixedWidth(180)
        self._ack_cb.setStyleSheet(_CB)
        self._ack_cb.addItem("Tümü", "")
        fr.addWidget(self._ack_cb)

        self._donem_cb.currentIndexChanged.connect(self._on_filter_change)
        self._ack_cb.currentIndexChanged.connect(self._on_ack_filter)
        fr.addStretch()
        tl.addLayout(fr)
        root.addWidget(top)

        # ── Özet bant — PHP: #vGayriResmiToplam / #vKesintilerToplam / #vFarkToplam
        ozet = QFrame()
        ozet.setFixedHeight(64)
        ozet.setStyleSheet(
            "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "stop:0 #1a3a5c,stop:0.5 #0d2137,stop:1 #061422);"
        )
        oz = QHBoxLayout(ozet)
        oz.setContentsMargins(20, 0, 20, 0)
        oz.setSpacing(40)

        def _blok(lbl: str, attr: str, clr: str):
            col = QVBoxLayout()
            col.setSpacing(1)
            h = QLabel(lbl)
            h.setStyleSheet("font-size:10px;color:rgba(255,255,255,.65);background:transparent;")
            v = QLabel("-")
            v.setStyleSheet(f"font-size:18px;font-weight:700;color:{clr};background:transparent;")
            col.addWidget(h)
            col.addWidget(v)
            setattr(self, attr, v)
            return col

        oz.addLayout(_blok("Gayri Resmi Toplam", "_gay_lbl",  "#ff9800"))   # PHP: #vGayriResmiToplam
        oz.addLayout(_blok("Kesintiler Toplam",  "_verg_lbl", "#dc3545"))   # PHP: #vKesintilerToplam
        oz.addLayout(_blok("Fark",               "_fark_lbl", "#28a745"))   # PHP: #vFarkToplam
        oz.addStretch()

        self._kayit_chip = QLabel("")
        self._kayit_chip.setStyleSheet(
            "background:rgba(255,255,255,.12);color:white;font-size:11px;"
            "font-weight:600;border-radius:10px;padding:2px 10px;"
        )
        oz.addWidget(self._kayit_chip, 0, Qt.AlignmentFlag.AlignVCenter)
        root.addWidget(ozet)

        # ── Tablo ────────────────────────────────────────────────────────────
        self._tbl = QTableWidget()
        self._tbl.setColumnCount(len(self.SUTUNLAR))
        self._tbl.setHorizontalHeaderLabels([s[0] for s in self.SUTUNLAR])
        self._tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._tbl.setAlternatingRowColors(True)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.setSortingEnabled(True)
        hdr = self._tbl.horizontalHeader()
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        for i, (_, _, w, _, _) in enumerate(self.SUTUNLAR):
            if i != 1:
                self._tbl.setColumnWidth(i, w)
        self._tbl.setStyleSheet("""
            QTableWidget { background:white; gridline-color:#e2e8f0;
                font-size:12px; color:#1e293b; border:none; }
            QTableWidget::item { color:#1e293b; padding:3px 6px; }
            QTableWidget::item:hover { background:#fef9f0; color:#1e293b; }
            QTableWidget::item:selected { background:#fef3c7; color:#1e293b; }
            QTableWidget::item:alternate { background:#f8fafc; }
            QHeaderView::section { background:#1a3a5c; color:white;
                font-weight:700; font-size:11px; padding:6px 4px;
                border:none; border-right:1px solid #0d2137; }
        """)
        root.addWidget(self._tbl, 1)

        # ── Alt bar ───────────────────────────────────────────────────────────
        alt = QFrame()
        alt.setFixedHeight(48)
        alt.setStyleSheet("background:#f8f9fa;border-top:1px solid #dee2e6;")
        a = QHBoxLayout(alt)
        a.setContentsMargins(16, 0, 16, 0)
        self._durum_lbl = QLabel("")
        self._durum_lbl.setStyleSheet("font-size:11px;color:#6c757d;")
        a.addWidget(self._durum_lbl)
        a.addStretch()
        kapat = QPushButton("Kapat")
        kapat.setFixedSize(110, 32)
        kapat.setStyleSheet(
            "QPushButton{background:#dc3545;color:white;border:none;"
            "border-radius:5px;font-size:13px;font-weight:700;}"
            "QPushButton:hover{background:#c82333;}"
        )
        kapat.clicked.connect(self.accept)
        a.addWidget(kapat)
        root.addWidget(alt)

    @staticmethod
    def _lbl(txt: str) -> QLabel:
        l = QLabel(txt)
        l.setStyleSheet("font-size:12px;color:#495057;font-weight:600;")
        return l

    # ── Veri yükleme ─────────────────────────────────────────────────────────

    def _load(self):
        """PHP: loadVergilerData() → vergiMuhtasarGetir.php"""
        from services.vergi_muhtasar_service import get_vergi_muhtasar
        donem = self._donem_cb.currentData() or ""
        result = get_vergi_muhtasar(self._userid, musterino=self._musterino,
                                    donem=donem, yil=self._yil)
        if not result.get("success"):
            self._durum_lbl.setText(f"❌ {result.get('message', 'Hata')}")
            return
        self._all_rows = result.get("data", [])
        self._doldur_filtreler(result.get("donemler", []),
                               list({r["ack"] for r in self._all_rows if r.get("ack")}))
        self._rows = self._all_rows
        self._doldur(result)

    def _doldur_filtreler(self, donemler: list[str], acklar: list[str]):
        """PHP: vmDoldurFiltreler() — combobox'ları doldur"""
        self._donem_cb.blockSignals(True)
        self._ack_cb.blockSignals(True)
        cur_don = self._donem_cb.currentData()
        cur_ack = self._ack_cb.currentData()
        self._donem_cb.clear()
        self._donem_cb.addItem("Tümü", "")
        for d in sorted(donemler):
            self._donem_cb.addItem(d, d)
        idx = self._donem_cb.findData(cur_don)
        if idx >= 0:
            self._donem_cb.setCurrentIndex(idx)
        self._ack_cb.clear()
        self._ack_cb.addItem("Tümü", "")
        for a in sorted(acklar):
            self._ack_cb.addItem(a, a)
        idx = self._ack_cb.findData(cur_ack)
        if idx >= 0:
            self._ack_cb.setCurrentIndex(idx)
        self._donem_cb.blockSignals(False)
        self._ack_cb.blockSignals(False)

    def _on_filter_change(self):
        """PHP: $('#vDonemFilter').on('change', loadVergilerData)"""
        self._load()

    def _on_ack_filter(self):
        """PHP: $('#vAckFilter').on('change', ...) — client-side filtre"""
        ack_val = self._ack_cb.currentData() or ""
        if ack_val:
            self._rows = [r for r in self._all_rows if r.get("ack") == ack_val]
        else:
            self._rows = self._all_rows
        self._doldur_from_rows()

    def _doldur(self, result: dict):
        """Servis sonucundan tabloyu doldur."""
        self._rows = result.get("data", [])
        self._doldur_from_rows()

    def _doldur_from_rows(self):
        """PHP: vmBuildTable() + vmToplamGuncelle()"""
        from PyQt6.QtGui import QColor, QFont as QF

        gay_top = verg_top = fark_top = 0.0
        for r in self._rows:
            gay  = float(r.get("gaytutar")     or 0)
            verg = float(r.get("vergkestutar") or 0)
            fark = gay - verg
            gay_top  += gay
            verg_top += verg
            fark_top += abs(fark)

        def _fmt(v): return f"{abs(v):,.2f} ₺".replace(",", "X").replace(".", ",").replace("X", ".")

        # PHP: #vGayriResmiToplam / #vKesintilerToplam / #vFarkToplam
        self._gay_lbl.setText(_fmt(gay_top))
        self._verg_lbl.setText(_fmt(verg_top))
        self._fark_lbl.setText(f"+{_fmt(fark_top)}")

        kayit = len(self._rows)
        self._kayit_chip.setText(f"📊 {kayit} kayıt")
        self._durum_lbl.setText(
            f"{kayit} kayıt  │  "
            f"Gayri Resmi: {_fmt(gay_top)}  "
            f"Kesinti: {_fmt(verg_top)}  "
            f"Fark: +{_fmt(fark_top)}"
        )

        self._tbl.setSortingEnabled(False)
        self._tbl.setRowCount(0)

        for row in self._rows:
            ri = self._tbl.rowCount()
            self._tbl.insertRow(ri)
            self._tbl.setRowHeight(ri, 28)

            for ci, (_, field, _, align, clr) in enumerate(self.SUTUNLAR):
                raw = row.get(field, "")

                if field in ("gaytutar", "vergkestutar", "fark"):
                    try:
                        val_f = float(raw or 0)
                    except (ValueError, TypeError):
                        val_f = 0.0
                    # Fark: abs değer + yeşil (PHP: Math.abs(d))
                    if field == "fark":
                        it = QTableWidgetItem(f"+{_fmt(abs(val_f))}")
                    else:
                        it = QTableWidgetItem(_fmt(val_f))
                    it.setData(Qt.ItemDataRole.UserRole, val_f)
                    it.setFont(QF("", -1, QF.Weight.Bold))
                    it.setForeground(QColor(clr or "#374151"))
                else:
                    it = QTableWidgetItem(str(raw) if raw is not None else "")
                    it.setForeground(QColor("#374151"))

                it.setTextAlignment(align | Qt.AlignmentFlag.AlignVCenter)
                self._tbl.setItem(ri, ci, it)

        self._tbl.setSortingEnabled(True)

        if not self._rows:
            self._tbl.insertRow(0)
            self._tbl.setRowHeight(0, 48)
            empty = QTableWidgetItem("Bu filtrede Vergi Muhtasar kaydı bulunamadı.")
            empty.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            empty.setForeground(QColor("#6c757d"))
            self._tbl.setItem(0, 0, empty)
            self._tbl.setSpan(0, 0, 1, len(self.SUTUNLAR))
