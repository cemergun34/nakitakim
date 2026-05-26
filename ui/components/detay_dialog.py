"""
KPI Detay Diyaloğu — kartlara tıklandığında açılır.
İki katmanlı:
  1. Şube / kategori bazlı özet kartları
  2. Seçilen şube / kategorinin işlem listesi (tablo)
"""
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QScrollArea, QWidget, QFrame, QGridLayout, QTableWidget,
    QTableWidgetItem, QHeaderView, QSizePolicy, QStackedWidget,
    QApplication, QLineEdit, QComboBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QSize
from PyQt6.QtGui import QFont, QColor, QBrush, QPainter, QLinearGradient

from ui.theme import COLORS, CARD_RADIUS
from utils.format import fmt_para
import tempfile
import os
import http.server
import socketserver
import threading
import webbrowser
import re

# ─── SORTABLE TABLE ITEM ──────────────────────────────────────────────────────

class SortableTableWidgetItem(QTableWidgetItem):
    """Numerik veya tarihsel değerleri doğru sıralamak için özel tablo hücresi."""
    def __init__(self, text, sort_val):
        super().__init__(text)
        self.sort_val = sort_val

    def __lt__(self, other):
        if isinstance(other, SortableTableWidgetItem):
            try:
                return self.sort_val < other.sort_val
            except TypeError:
                return str(self.sort_val) < str(other.sort_val)
        return super().__lt__(other)

# ─── GLOBAL EFATURA PREVIEW HTTP SUNUCUSU ───────────────────────────────────

class DetayGlobalServer:
    """E-Fatura XML + XSLT dosyalarını CORS engelleri olmadan sunan hafif yerel HTTP sunucusu."""
    _server_thread = None
    _server_port = None
    _temp_dir = None
    
    @classmethod
    def get_server_url(cls, filename: str) -> tuple[str, str]:
        if cls._server_thread is None:
            cls._temp_dir = tempfile.mkdtemp(prefix="efatura_global_")
            
            import socket
            s = socket.socket()
            s.bind(('', 0))
            cls._server_port = s.getsockname()[1]
            s.close()
            
            class Handler(http.server.SimpleHTTPRequestHandler):
                def __init__(self, *args, **kwargs):
                    super().__init__(*args, directory=cls._temp_dir, **kwargs)
                def log_message(self, format, *args):
                    pass
                    
            def run_server():
                try:
                    with socketserver.TCPServer(("", cls._server_port), Handler) as httpd:
                        httpd.serve_forever()
                except Exception:
                    pass
                    
            cls._server_thread = threading.Thread(target=run_server, daemon=True)
            cls._server_thread.start()
            
        return f"http://localhost:{cls._server_port}/{filename}", cls._temp_dir

# ─── KPI DETAY KARTI ──────────────────────────────────────────────────────────

class DetayKarti(QFrame):
    """Şube özet kartı (gradient, tıklanabilir)."""
    clicked = pyqtSignal(dict)  # row dict emit eder

    RENK_PALETI = [
        ("#10B981", "#059669"),
        ("#3B82F6", "#1D4ED8"),
        ("#8B5CF6", "#6D28D9"),
        ("#EC4899", "#DB2777"),
        ("#F59E0B", "#D97706"),
        ("#6B7280", "#4B5563"),
        ("#EA580C", "#C2410C"),
        ("#0EA5E9", "#0284C7"),
    ]

    def __init__(self, row: dict, index: int, gelir_field="toplam_gelir",
                 gider_field="toplam_gider", tutar_field=None, parent=None):
        super().__init__(parent)
        self._row = row
        self._hovered = False
        c1, c2 = self.RENK_PALETI[index % len(self.RENK_PALETI)]
        self._c1, self._c2 = c1, c2
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setFixedSize(240, 120)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(4)

        sube = str(row.get("sube_adi", "Bilinmeyen"))
        lbl_name = QLabel(sube)
        lbl_name.setStyleSheet("color: rgba(255,255,255,0.9); font-size: 12px; font-weight: 600; background: transparent;")
        lbl_name.setWordWrap(True)
        layout.addWidget(lbl_name)

        # Tutar
        if tutar_field and row.get(tutar_field) is not None:
            tutar = float(row.get(tutar_field) or 0)
            lbl_tutar = QLabel(fmt_para(tutar))
        else:
            gelir = float(row.get(gelir_field) or 0)
            gider = float(row.get(gider_field) or 0)
            net = gelir - gider
            sign = "+" if net >= 0 else "-"
            color = "white" if net >= 0 else "#FCA5A5"
            lbl_tutar = QLabel(f"{sign}{fmt_para(abs(net))}")
            lbl_tutar.setStyleSheet(f"color: {color}; font-size: 19px; font-weight: 700; background: transparent;")
            layout.addWidget(lbl_tutar)

            kayit = row.get("kayit_sayisi", 0)
            lbl_sub = QLabel(f"{kayit} işlem  •  Tıkla → detay")
            lbl_sub.setStyleSheet("color: rgba(255,255,255,0.7); font-size: 10px; background: transparent;")
            layout.addWidget(lbl_sub)
            layout.addStretch()
            return

        lbl_tutar.setStyleSheet("color: white; font-size: 19px; font-weight: 700; background: transparent;")
        layout.addWidget(lbl_tutar)
        kayit = row.get("kayit_sayisi", 0)
        lbl_sub = QLabel(f"{kayit} kayıt  •  Tıkla → detay")
        lbl_sub.setStyleSheet("color: rgba(255,255,255,0.7); font-size: 10px; background: transparent;")
        layout.addWidget(lbl_sub)
        layout.addStretch()

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        rect = self.rect()
        g = QLinearGradient(0, 0, rect.width(), rect.height())
        g.setColorAt(0, QColor(self._c1))
        g.setColorAt(1, QColor(self._c2))
        p.setBrush(QBrush(g))
        p.setPen(Qt.PenStyle.NoPen)
        r = rect.adjusted(0, -3, 0, 0) if self._hovered else rect
        p.drawRoundedRect(r, CARD_RADIUS, CARD_RADIUS)
        # dekoratif daire
        p.setBrush(QColor(255, 255, 255, 20))
        p.drawEllipse(rect.right() - 50, rect.bottom() - 50, 80, 80)
        p.end()
        super().paintEvent(event)

    def enterEvent(self, e):
        self._hovered = True; self.update()
    def leaveEvent(self, e):
        self._hovered = False; self.update()
    def mousePressEvent(self, e):
        if e.button() == Qt.MouseButton.LeftButton:
            self.clicked.emit(self._row)


# ─── İŞLEM TABLOSU ────────────────────────────────────────────────────────────

class IslemTablosu(QTableWidget):
    """Tüm kart tipleri için ortak işlem listesi tablosu."""

    KOLON_SETLERI = {
        "hareketler": [
            ("Tarih", "tarih"), ("Açıklama", "aciklama"),
            ("Şube", "sube_adi"), ("Tutar", "tutar"),
            ("Gelir/Gider", "gelirGider"), ("Teslim Şekli", "teslim_sekli"),
            ("Fatura No", "faturaNo"),
        ],
        "genel_hesap": [
            ("Tarih", "tarih"), ("Açıklama", "aciklama"),
            ("Şube", "sube_adi"), ("Gelir", "gelir"), ("Gider", "gider"),
            ("Teslim Şekli", "teslim_sekli"), ("Ödeme Şekli", "odeme_sekli"),
            ("Kategori", "kategori"),
        ],
        "faturalar": [
            ("Tarih", "tarih"), ("Ünvan", "unvan"),
            ("Fatura No", "faturano"), ("Vergi No", "vergino"),
            ("Tutar", "toplam"), ("Fatura Modu", "faturaMod"),
            ("Form No", "formNo"), ("Kaynak", "kaynak"),
        ],
    }

    def __init__(self, tablo_tipi: str = "hareketler", parent=None):
        super().__init__(parent)
        self._tablo_tipi = tablo_tipi
        self.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.setAlternatingRowColors(True)
        self.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.verticalHeader().setVisible(False)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.horizontalHeader().setStretchLastSection(True)
        self.setStyleSheet(f"""
            QTableWidget {{
                background: white;
                alternate-background-color: #F8FAFF;
                gridline-color: {COLORS['border']};
                font-size: 11px;
                border: 1.5px solid {COLORS['border']};
                border-radius: 8px;
            }}
            QTableWidget::item {{ padding: 5px 8px; color: {COLORS['text_primary']}; }}
            QTableWidget::item:selected {{ background: #DBEAFE; color: #1E40AF; }}
            QHeaderView::section {{
                background: {COLORS['table_header']};
                color: {COLORS['text_secondary']};
                font-size: 10px; font-weight: 600;
                padding: 6px;
                border: none;
                border-bottom: 1px solid {COLORS['border']};
                border-right: 1px solid {COLORS['border']};
            }}
        """)
        cols = self.KOLON_SETLERI.get(tablo_tipi, self.KOLON_SETLERI["hareketler"])
        self.setColumnCount(len(cols))
        self.setHorizontalHeaderLabels([c[0] for c in cols])
        self.setSortingEnabled(True)
        self.cellClicked.connect(self._on_cell_clicked)

    def load_rows(self, rows: list[dict]):
        self.setSortingEnabled(False)
        cols = self.KOLON_SETLERI.get(self._tablo_tipi, self.KOLON_SETLERI["hareketler"])
        self.setRowCount(len(rows))
        PARA_COLS = {"tutar", "gelir", "gider", "toplam"}
        for r_idx, row in enumerate(rows):
            for c_idx, (_, key) in enumerate(cols):
                val = row.get(key)
                if key in PARA_COLS and val is not None:
                    try:
                        text = fmt_para(float(val))
                    except Exception:
                        text = str(val) if val is not None else "-"
                else:
                    text = str(val) if val is not None else "-"

                # Sıralama değeri oluştur
                if key in PARA_COLS:
                    try:
                        sort_val = float(val) if val is not None else 0.0
                    except Exception:
                        sort_val = 0.0
                elif key == "tarih" and val:
                    import datetime as dt
                    try:
                        sort_val = dt.datetime.strptime(str(val).strip(), "%Y-%m-%d")
                    except Exception:
                        try:
                            sort_val = dt.datetime.strptime(str(val).strip(), "%d.%m.%Y")
                        except Exception:
                            sort_val = str(val)
                else:
                    sort_val = str(val).lower() if val is not None else ""

                item = SortableTableWidgetItem(text, sort_val)

                if key in PARA_COLS:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    if key in ("gelir", "tutar") and val and float(val) > 0:
                        item.setForeground(QBrush(QColor("#059669")))
                    elif key == "gider" and val and float(val) > 0:
                        item.setForeground(QBrush(QColor("#DC2626")))
                
                # Fatura No link özelliği
                if key in ("faturano", "faturaNo") and val and val != "-":
                    item.setForeground(QBrush(QColor("#2563EB")))
                    font = item.font()
                    font.setUnderline(True)
                    item.setFont(font)
                    item.setToolTip("E-Fatura Ön İzlemesini Görmek İçin Tıklayın")
                    item.setData(Qt.ItemDataRole.UserRole, row)

                self.setItem(r_idx, c_idx, item)
        self.resizeRowsToContents()
        self.setSortingEnabled(True)

    def _on_cell_clicked(self, row_idx, col_idx):
        cols = self.KOLON_SETLERI.get(self._tablo_tipi, [])
        if col_idx < len(cols):
            _, key = cols[col_idx]
            if key in ("faturano", "faturaNo"):
                item = self.item(row_idx, col_idx)
                if item:
                    row_data = item.data(Qt.ItemDataRole.UserRole)
                    if row_data:
                        self._show_fatura_preview(row_data)

    def _show_fatura_preview(self, row_data):
        """
        XSLT transform ile orijinal e-fatura görünümünü
        QWebEngineView (Chromium motoru) içinde gösterir.
        """
        import os, base64, tempfile
        import xml.etree.ElementTree as ET
        from PyQt6.QtWidgets import (
            QMessageBox, QDialog, QVBoxLayout, QHBoxLayout,
            QLabel, QPushButton
        )
        from PyQt6.QtCore import QUrl
        from PyQt6.QtWebEngineWidgets import QWebEngineView
        from PyQt6.QtWebEngineCore import QWebEngineSettings

        xml_path  = row_data.get("xml_dosya")
        fatura_no = row_data.get("faturano") or row_data.get("faturaNo") or "?"

        # ── 1. Dosya kontrolü ────────────────────────────────────────────────
        if not xml_path or not os.path.exists(str(xml_path)):
            QMessageBox.information(
                self,
                "Fatura Dosyası Yok",
                f"Bu fatura ({fatura_no}) sisteme XML olarak yüklenmemiş.\n"
                "Yalnızca XML ile aktarılan faturalar önizlenebilir."
            )
            return

        # ── 2. XML oku ───────────────────────────────────────────────────────
        try:
            with open(xml_path, "rb") as fh:
                xml_bytes = fh.read()
        except Exception as exc:
            QMessageBox.critical(self, "Hata", f"XML okunamadı:\n{exc}")
            return

        # ── 3. XSLT bul + dönüştür ──────────────────────────────────────────
        html_bytes = None
        xslt_filename = "style.xslt"
        try:
            from lxml import etree as let
            _NS = {
                "cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
                "cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
            }
            xml_root = let.fromstring(xml_bytes)
            hits = xml_root.xpath(
                ".//cac:AdditionalDocumentReference"
                "[cbc:DocumentType=\'XSLT\']"
                "/cac:Attachment/cbc:EmbeddedDocumentBinaryObject",
                namespaces=_NS,
            )
            if hits and hits[0].text:
                xslt_raw  = base64.b64decode(hits[0].text.strip())
                xslt_root = let.fromstring(xslt_raw)
                transform = let.XSLT(xslt_root)
                result    = transform(xml_root)
                html_bytes = bytes(result)
                # XSLT dosya adını da al (href olarak kullanacağız)
                fname = hits[0].attrib.get("filename", "style.xslt")
                xslt_filename = fname if fname else "style.xslt"
        except Exception as exc:
            print(f"[EFatura] XSLT dönüşüm hatası: {exc}")

        # ── 4. Fallback: XSLT yoksa kendi HTML şablonumuzu yaz ──────────────
        if not html_bytes:
            NS = {
                "cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
                "cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
            }
            root = ET.fromstring(xml_bytes)

            def tx(el, tag, ns="cbc"):
                e = el.find(f"{ns}:{tag}", NS) if el is not None else None
                return (e.text or "").strip() if e is not None else ""

            fno   = tx(root, "ID") or fatura_no
            tarih = tx(root, "IssueDate")

            sup  = root.find("cac:AccountingSupplierParty/cac:Party", NS)
            pn   = sup.find("cac:PartyName", NS) if sup else None
            sup_unvan = tx(pn, "Name") if pn else ""
            sup_vkn = ""
            if sup:
                for pid in sup.findall("cac:PartyIdentification", NS):
                    ie = pid.find("cbc:ID", NS)
                    if ie is not None and ie.attrib.get("schemeID") == "VKN":
                        sup_vkn = (ie.text or "").strip()

            cus  = root.find("cac:AccountingCustomerParty/cac:Party", NS)
            pnc  = cus.find("cac:PartyName", NS) if cus else None
            cus_unvan = tx(pnc, "Name") if pnc else ""
            cus_vkn = ""
            if cus:
                for pid in cus.findall("cac:PartyIdentification", NS):
                    ie = pid.find("cbc:ID", NS)
                    if ie is not None and ie.attrib.get("schemeID") == "VKN":
                        cus_vkn = (ie.text or "").strip()

            lmt     = root.find("cac:LegalMonetaryTotal", NS)
            payable = tx(lmt, "PayableAmount")

            def fn(v):
                try:    return f"{float(v):,.2f} TL"
                except: return v or "-"

            rows_html = ""
            for i, line in enumerate(root.findall("cac:InvoiceLine", NS)):
                item  = line.find("cac:Item", NS)
                urun  = tx(item, "Name") if item else "-"
                qty   = tx(line, "InvoicedQuantity")
                price = line.find("cac:Price", NS)
                bp    = tx(price, "PriceAmount") if price else ""
                le    = tx(line, "LineExtensionAmount")
                bg    = "#ffffff" if i % 2 == 0 else "#f8faff"
                rows_html += (
                    f"<tr style=\'background:{bg}\'>"
                    f"<td style=\'padding:8px\'>{urun}</td>"
                    f"<td style=\'padding:8px;text-align:center\'>{qty}</td>"
                    f"<td style=\'padding:8px;text-align:right\'>{fn(bp)}</td>"
                    f"<td style=\'padding:8px;text-align:right;font-weight:bold\'>{fn(le)}</td>"
                    f"</tr>"
                )

            html_str = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><style>
body{{font-family:Arial,sans-serif;font-size:13px;color:#1f2937;background:#f0f4ff;margin:0;padding:20px}}
.wrap{{max-width:900px;margin:0 auto}}
.card{{background:white;border-radius:8px;padding:20px;margin-bottom:14px;border:1px solid #e5e7eb}}
h1{{font-size:17px;color:#1e3a8a;margin:0 0 6px}} h2{{font-size:13px;color:#374151;margin:10px 0 6px;border-bottom:1px solid #e5e7eb;padding-bottom:4px}}
.label{{font-size:10px;color:#6b7280;text-transform:uppercase;font-weight:700}}
.val{{font-size:13px;margin-top:2px}}
.two{{display:flex;gap:20px}}.two>div{{flex:1}}
table{{width:100%;border-collapse:collapse}}
th{{background:#1e3a8a;color:white;padding:8px;text-align:left;font-size:11px}}
.total{{font-size:15px;font-weight:700;color:#059669;text-align:right;padding:10px 0}}
</style></head><body><div class="wrap">
<div class="card"><h1>Fatura No: {fno}</h1>
<div class="two">
  <div><div class="label">Tarih</div><div class="val">{tarih}</div></div>
  <div><div class="label">Tedarikçi</div><div class="val"><b>{sup_unvan}</b> (VKN: {sup_vkn})</div></div>
  <div><div class="label">Alıcı</div><div class="val"><b>{cus_unvan}</b> (VKN: {cus_vkn})</div></div>
</div></div>
<div class="card"><h2>Kalemler</h2>
<table><tr><th>Açıklama</th><th>Miktar</th><th style="text-align:right">Birim Fiyat</th><th style="text-align:right">Toplam</th></tr>
{rows_html}</table>
<div class="total">Ödenecek: {fn(payable)}</div>
</div></div></body></html>"""
            html_bytes = html_str.encode("utf-8")

        # ── 5. Geçici dizine yaz (yerel dosya sistemi üzerinden aç) ─────────
        tmp_dir  = tempfile.mkdtemp(prefix="efatura_preview_")
        html_file = os.path.join(tmp_dir, "fatura.html")
        with open(html_file, "wb") as fh:
            fh.write(html_bytes)

        # ── 6. QWebEngineView ile QDialog içinde göster ──────────────────────
        dialog = QDialog(self)
        dialog.setWindowTitle(f"E-Fatura Ön İzleme  —  {fatura_no}")
        dialog.setMinimumSize(1000, 820)
        dialog.resize(1100, 900)
        dialog.setStyleSheet("background:#1e3a8a;")

        dlg_layout = QVBoxLayout(dialog)
        dlg_layout.setContentsMargins(0, 0, 0, 0)
        dlg_layout.setSpacing(0)

        # Header
        hdr = QLabel()
        hdr.setFixedHeight(48)
        hdr.setText(f"  📄  E-Fatura Ön İzleme  —  {fatura_no}")
        hdr.setStyleSheet(
            "QLabel { background:#2563EB; color:white; font-size:13px; font-weight:700; padding-left:12px; }"
        )

        close_btn = QPushButton("✕  Kapat")
        close_btn.setFixedSize(100, 34)
        close_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        close_btn.setStyleSheet(
            "QPushButton{background:rgba(255,255,255,.22);color:white;"
            "border:none;border-radius:7px;font-size:13px;font-weight:700;}"
            "QPushButton:hover{background:rgba(255,255,255,.4);}"
        )
        close_btn.clicked.connect(dialog.reject)

        top = QHBoxLayout()
        top.setContentsMargins(0, 0, 12, 0)
        top.setSpacing(0)
        top.addWidget(hdr, 1)
        top.addWidget(close_btn)

        hdr_widget = __import__("PyQt6.QtWidgets", fromlist=["QWidget"]).QWidget()
        hdr_widget.setFixedHeight(48)
        hdr_widget.setStyleSheet("background:#2563EB;")
        hdr_widget.setLayout(top)
        dlg_layout.addWidget(hdr_widget)

        # WebEngine
        web = QWebEngineView()
        web.settings().setAttribute(QWebEngineSettings.WebAttribute.JavascriptEnabled, True)
        web.settings().setAttribute(QWebEngineSettings.WebAttribute.LocalContentCanAccessFileUrls, True)
        web.load(QUrl.fromLocalFile(html_file))
        dlg_layout.addWidget(web, 1)

        dialog.exec()


# ─── ANA DETAY DİYALOĞU ──────────────────────────────────────────────────────

class DetayDialog(QDialog):
    """
    2 aşamalı diyalog:
    Sayfa 0 → Şube özet kartları
    Sayfa 1 → Seçilen şubenin işlem listesi
    """

    def __init__(self, baslik: str, ozet_rows: list[dict],
                 detay_fn,          # callable(sube_adi) → list[dict]
                 tablo_tipi: str = "hareketler",
                 gelir_field="toplam_gelir", gider_field="toplam_gider",
                 tutar_field=None,
                 direct_detay: bool = False,   # True → özet atla, direkt liste aç
                 parent=None):
        super().__init__(parent)
        self.setWindowTitle(baslik)
        self.setStyleSheet(f"background: {COLORS['bg']};")

        self._detay_fn = detay_fn
        self._tablo_tipi = tablo_tipi
        self._baslik = baslik
        self._direct_detay = direct_detay

        # Kutucukların tam kapladığı pencere boyutunu dinamik hesapla
        cols_per_row = 4
        card_w, card_h = 240, 120
        spacing = 16
        margin_x, margin_y = 24, 20
        header_h = 56

        num_cards = len(ozet_rows)
        actual_cols = min(num_cards, cols_per_row) if num_cards > 0 else 1
        actual_rows = (num_cards + cols_per_row - 1) // cols_per_row if num_cards > 0 else 1

        self._ozet_width = actual_cols * card_w + (actual_cols - 1) * spacing + margin_x * 2
        self._ozet_height = header_h + actual_rows * card_h + (actual_rows - 1) * spacing + margin_y * 2
        self.resize(self._ozet_width, self._ozet_height)

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Header ────────────────────────────────────────────────────────────
        header = QFrame()
        header.setFixedHeight(56)
        header.setStyleSheet(f"""
            QFrame {{ background: {COLORS['btn_primary']}; }}
        """)
        hl = QHBoxLayout(header)
        hl.setContentsMargins(20, 0, 20, 0)

        self._back_btn = QPushButton("← Geri")
        self._back_btn.setFixedHeight(34)
        self._back_btn.setFixedWidth(90)
        self._back_btn.setStyleSheet(f"""
            QPushButton {{
                background: rgba(255,255,255,0.2);
                color: white; border: none;
                border-radius: 8px; font-size: 12px; font-weight: 600;
            }}
            QPushButton:hover {{ background: rgba(255,255,255,0.35); }}
        """)
        self._back_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._back_btn.hide()
        self._back_btn.clicked.connect(self._show_ozet)
        hl.addWidget(self._back_btn)

        self._title_lbl = QLabel(baslik)
        self._title_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._title_lbl.setStyleSheet("color: white; font-size: 16px; font-weight: 700;")
        hl.addWidget(self._title_lbl, 1)

        close_btn = QPushButton("✕")
        close_btn.setFixedSize(34, 34)
        close_btn.setStyleSheet("""
            QPushButton { background: rgba(255,255,255,0.2); color: white;
                          border: none; border-radius: 8px; font-size: 16px; }
            QPushButton:hover { background: rgba(255,255,255,0.4); }
        """)
        close_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        close_btn.clicked.connect(self.accept)
        hl.addWidget(close_btn)
        root.addWidget(header)

        # ── Stacked ───────────────────────────────────────────────────────────
        self._stack = QStackedWidget()
        root.addWidget(self._stack)

        # ── Sayfa 0: Özet kartlar ─────────────────────────────────────────────
        ozet_page = QScrollArea()
        ozet_page.setWidgetResizable(True)
        ozet_page.setFrameShape(QFrame.Shape.NoFrame)
        ozet_page.setStyleSheet("background: transparent;")

        ozet_inner = QWidget()
        ozet_inner.setStyleSheet("background: transparent;")
        ozet_grid = QGridLayout(ozet_inner)
        ozet_grid.setContentsMargins(24, 20, 24, 20)
        ozet_grid.setSpacing(16)

        cols_per_row = 4
        for idx, row in enumerate(ozet_rows):
            kart = DetayKarti(row, idx,
                              gelir_field=gelir_field,
                              gider_field=gider_field,
                              tutar_field=tutar_field)
            kart.clicked.connect(self._on_kart_clicked)
            r, c = divmod(idx, cols_per_row)
            ozet_grid.addWidget(kart, r, c)

        # Eğer boş sütunlar varsa, doldurucu ekle
        remainder = len(ozet_rows) % cols_per_row
        if remainder:
            for i in range(cols_per_row - remainder):
                spacer = QWidget()
                spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
                r_last = len(ozet_rows) // cols_per_row
                ozet_grid.addWidget(spacer, r_last, remainder + i)

        ozet_page.setWidget(ozet_inner)
        self._stack.addWidget(ozet_page)

        # ── Sayfa 1: İşlem listesi ────────────────────────────────────────────
        detay_page = QWidget()
        detay_page.setStyleSheet("background: transparent;")
        dp_layout = QVBoxLayout(detay_page)
        dp_layout.setContentsMargins(16, 12, 16, 12)
        dp_layout.setSpacing(8)

        # Bilgi Barları (Pills) Üst Bölüm
        top_row = QHBoxLayout()
        top_row.setSpacing(10)
        
        # 1. Başlık ve Şube Adı
        self._detay_sube_lbl = QLabel("")
        self._detay_sube_lbl.setStyleSheet(f"""
            color: {COLORS['text_primary']};
            font-size: 13px;
            font-weight: 700;
            padding: 8px 12px;
            background: #EFF6FF;
            border-radius: 8px;
            border: 1px solid #BFDBFE;
        """)
        top_row.addWidget(self._detay_sube_lbl)
        
        # 2. Gelir Pill
        self._detay_gelir_lbl = QLabel("")
        self._detay_gelir_lbl.setStyleSheet(f"""
            color: #047857;
            font-size: 13px;
            font-weight: 700;
            padding: 8px 12px;
            background: #D1FAE5;
            border-radius: 8px;
            border: 1px solid #A7F3D0;
        """)
        top_row.addWidget(self._detay_gelir_lbl)
        
        # 3. Gider Pill
        self._detay_gider_lbl = QLabel("")
        self._detay_gider_lbl.setStyleSheet(f"""
            color: #B91C1C;
            font-size: 13px;
            font-weight: 700;
            padding: 8px 12px;
            background: #FEE2E2;
            border-radius: 8px;
            border: 1px solid #FCA5A5;
        """)
        top_row.addWidget(self._detay_gider_lbl)
        
        # 4. Net Pill
        self._detay_net_lbl = QLabel("")
        top_row.addWidget(self._detay_net_lbl)
        top_row.addStretch()

        self._excel_btn = QPushButton("📥 Excel İndir")
        self._excel_btn.setFixedHeight(38)
        self._excel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._excel_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLORS['btn_excel'] if 'btn_excel' in COLORS else '#059669'};
                color: white;
                font-size: 12px;
                font-weight: 700;
                border: none;
                border-radius: 8px;
                padding: 0 16px;
                letter-spacing: 0.5px;
            }}
            QPushButton:hover {{ background: #047857; }}
        """)
        self._excel_btn.clicked.connect(self._export_excel)
        top_row.addWidget(self._excel_btn)
        dp_layout.addLayout(top_row)

        # Ay ve Yıl Bazında Filtreleme Layout
        date_filter_row = QHBoxLayout()
        date_filter_row.setSpacing(10)
        
        lbl_date_filter = QLabel("📅 Ay/Yıl Filtresi:")
        lbl_date_filter.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 11px; font-weight: 600;")
        date_filter_row.addWidget(lbl_date_filter)
        
        # Ay ComboBox
        self.combo_ay = QComboBox()
        self.combo_ay.setFixedHeight(28)
        self.combo_ay.setFixedWidth(120)
        self.combo_ay.setStyleSheet(self._combo_style())
        self.combo_ay.addItem("Tüm Aylar", None)
        aylar = [
            ("Ocak", 1), ("Şubat", 2), ("Mart", 3), ("Nisan", 4),
            ("Mayıs", 5), ("Haziran", 6), ("Temmuz", 7), ("Ağustos", 8),
            ("Eylül", 9), ("Ekim", 10), ("Kasım", 11), ("Aralık", 12)
        ]
        for name, code in aylar:
            self.combo_ay.addItem(name, code)
        self.combo_ay.currentIndexChanged.connect(self._apply_filters)
        date_filter_row.addWidget(self.combo_ay)
        
        # Yıl ComboBox
        import datetime
        current_year = datetime.datetime.now().year

        self.combo_yil = QComboBox()
        self.combo_yil.setFixedHeight(28)
        self.combo_yil.setFixedWidth(120)
        self.combo_yil.setStyleSheet(self._combo_style())
        self.combo_yil.addItem("Tüm Yıllar", None)
        for y in range(current_year + 1, current_year - 5, -1):
            self.combo_yil.addItem(str(y), int(y))
        # Mevcut yılı varsayılan seç
        idx = self.combo_yil.findData(current_year)
        if idx >= 0:
            self.combo_yil.setCurrentIndex(idx)
        self.combo_yil.currentIndexChanged.connect(self._apply_filters)
        date_filter_row.addWidget(self.combo_yil)
        
        date_filter_row.addStretch()
        dp_layout.addLayout(date_filter_row)

        # Kolon Filtreleri Satırı
        self._filter_container = QWidget()
        filter_layout = QHBoxLayout(self._filter_container)
        filter_layout.setContentsMargins(0, 4, 0, 4)
        filter_layout.setSpacing(8)
        
        self.filters = []
        cols = IslemTablosu.KOLON_SETLERI.get(tablo_tipi, [])
        for c_name, _ in cols:
            le = QLineEdit()
            le.setPlaceholderText(f"🔍 {c_name}...")
            le.setFixedHeight(28)
            le.setStyleSheet(f"""
                QLineEdit {{
                    background: white;
                    border: 1px solid {COLORS['border']};
                    border-radius: 6px;
                    padding: 0 8px;
                    font-size: 11px;
                    color: {COLORS['text_primary']};
                }}
                QLineEdit:focus {{ border-color: {COLORS['btn_primary']}; }}
            """)
            le.textChanged.connect(self._apply_filters)
            filter_layout.addWidget(le)
            self.filters.append(le)
            
        dp_layout.addWidget(self._filter_container)

        self._tablo = IslemTablosu(tablo_tipi)
        dp_layout.addWidget(self._tablo, 1)

        self._stack.addWidget(detay_page)

        # Direct mod: özet atla, direkt liste yükle
        if self._direct_detay:
            self._back_btn.hide()
            try:
                rows = self._detay_fn(None)
            except Exception:
                rows = []
            self._detay_sube_lbl.setText(f"📄 {baslik}  — {len(rows)} kayıt")
            self._detay_gelir_lbl.setText("")
            self._detay_gider_lbl.setText("")
            self._detay_net_lbl.setText("")
            self._tablo.load_rows(rows)
            self._stack.setCurrentIndex(1)
            self.resize(1200, 720)

    def _combo_style(self) -> str:
        return f"""
            QComboBox {{
                background: white;
                border: 1px solid {COLORS['border']};
                border-radius: 6px;
                padding: 0 8px;
                font-size: 11px;
                font-weight: 600;
                color: #1F2937;
            }}
            QComboBox:focus {{ border-color: {COLORS['btn_primary']}; }}
            QComboBox::drop-down {{ border: none; }}
            QComboBox QAbstractItemView {{
                background-color: white;
                color: #1F2937;
                selection-background-color: #DBEAFE;
                selection-color: #1E40AF;
                border: 1px solid {COLORS['border']};
            }}
        """

    def _parse_date(self, date_str: str):
        """
        Tarih stringinden (ay, yil) tuple döndürür.
        Desteklenen formatlar:
          YYYY-MM-DD   2026-03-15
          DD.MM.YYYY   15.03.2026
          DD/MM/YYYY   15/03/2026
          YYYY.MM.DD   2026.03.15
          ISO+time     2026-03-15T10:00:00
        """
        import datetime as dt
        if not date_str or date_str.strip() == "-":
            return None, None
        s = date_str.strip()[:10]   # sadece ilk 10 karakter (tarih kısmı)
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y.%m.%d"):
            try:
                d = dt.datetime.strptime(s, fmt)
                return d.month, d.year
            except ValueError:
                continue
        return None, None


    def _on_kart_clicked(self, row: dict):
        sube_adi = row.get("sube_adi", "")
        kayit = row.get("kayit_sayisi", 0)
        gelir = float(row.get("toplam_gelir") or row.get("toplam_tutar") or row.get("toplam_gider") or 0)
        gider = float(row.get("toplam_gider") or 0)
        net = gelir - gider

        self._detay_sube_lbl.setText(f"📍 {sube_adi}  ({kayit} kayıt)")
        self._detay_gelir_lbl.setText(f"🟢 Gelir: {fmt_para(gelir)}")
        self._detay_gider_lbl.setText(f"🔴 Gider: {fmt_para(gider)}")
        
        if net >= 0:
            self._detay_net_lbl.setText(f"💎 Net: +{fmt_para(net)}")
            self._detay_net_lbl.setStyleSheet(f"""
                color: #1E40AF;
                font-size: 13px;
                font-weight: 700;
                padding: 8px 12px;
                background: #DBEAFE;
                border-radius: 8px;
                border: 1px solid #BFDBFE;
            """)
        else:
            self._detay_net_lbl.setText(f"⚠️ Net: -{fmt_para(abs(net))}")
            self._detay_net_lbl.setStyleSheet(f"""
                color: #B91C1C;
                font-size: 13px;
                font-weight: 700;
                padding: 8px 12px;
                background: #FEE2E2;
                border-radius: 8px;
                border: 1px solid #FCA5A5;
            """)

        # Filtreleri temizle
        for le in self.filters:
            le.blockSignals(True)
            le.clear()
            le.blockSignals(False)

        # Ay/Yıl filtrelerini sıfırla
        self.combo_ay.blockSignals(True)
        self.combo_ay.setCurrentIndex(0)
        self.combo_ay.blockSignals(False)

        self.combo_yil.blockSignals(True)
        self.combo_yil.setCurrentIndex(0)
        self.combo_yil.blockSignals(False)

        # Veriyi yükle
        try:
            rows = self._detay_fn(sube_adi)
        except Exception as e:
            rows = []
            self._detay_info.setText(f"Hata: {e}")

        self._tablo.load_rows(rows)
        self._back_btn.show()
        self._title_lbl.setText(f"{self._baslik}  →  {sube_adi}")
        self._stack.setCurrentIndex(1)
        self.resize(1100, 680)

    def _show_ozet(self):
        self._back_btn.hide()
        self._title_lbl.setText(self._baslik)
        self._stack.setCurrentIndex(0)
        self.resize(self._ozet_width, self._ozet_height)

    def _apply_filters(self):
        selected_ay = self.combo_ay.currentData()       # int or None
        selected_yil = self.combo_yil.currentData()     # int or None

        for row_idx in range(self._tablo.rowCount()):
            match = True
            
            # 1. Ay ve Yıl Seçimi ile Tarih Filtresi (Tarih kolonu her zaman 0. kolondur)
            if (selected_ay is not None) or (selected_yil is not None):
                tarih_item = self._tablo.item(row_idx, 0)
                if tarih_item:
                    ay, yil = self._parse_date(tarih_item.text())
                    if selected_ay is not None and ay != selected_ay:
                        match = False
                    if selected_yil is not None and yil != selected_yil:
                        match = False
                else:
                    match = False

            # 2. Kolon Bazlı Metin Aramaları
            if match:
                for col_idx, le in enumerate(self.filters):
                    txt = le.text().strip().lower()
                    if not txt:
                        continue
                    item = self._tablo.item(row_idx, col_idx)
                    cell_txt = item.text().lower() if item else ""
                    if txt not in cell_txt:
                        match = False
                        break
                        
            self._tablo.setRowHidden(row_idx, not match)

    def _export_excel(self):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import re
        import datetime
        
        # Dosya adını temizle (özel karakterleri ve ok işaretini kaldır)
        title_text = self._title_lbl.text()
        clean_title = re.sub(r'[^\w\s\-]', '_', title_text)
        clean_title = re.sub(r'\s+', ' ', clean_title).strip()
        
        path, _ = QFileDialog.getSaveFileName(
            self, "Excel Olarak Kaydet", f"{clean_title}.xlsx", "Excel Files (*.xlsx)"
        )
        if not path:
            return
            
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Detay Listesi"
            
            # Excel Stylings
            font_title = Font(name="Segoe UI", size=14, bold=True, color="1E3A8A")
            font_subtitle = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            font_data = Font(name="Segoe UI", size=10, color="1F2937")
            font_total = Font(name="Segoe UI", size=11, bold=True, color="1E3A8A")
            
            fill_header = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid") # Soft Blue
            fill_total = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")  # Extremely Soft Blue
            
            border_thin = Border(
                left=Side(style="thin", color="E5E7EB"),
                right=Side(style="thin", color="E5E7EB"),
                top=Side(style="thin", color="E5E7EB"),
                bottom=Side(style="thin", color="E5E7EB")
            )
            border_total = Border(
                top=Side(style="thin", color="93C5FD"),
                bottom=Side(style="double", color="1E3A8A")
            )
            
            # 1. Rapor Başlık Bloğu
            ws.append([title_text])
            ws.cell(row=1, column=1).font = font_title
            
            rapor_tarih = f"Raporlama Tarihi: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
            ws.append([rapor_tarih])
            ws.cell(row=2, column=1).font = font_subtitle
            
            ws.append([]) # Boşluk
            
            # Headers
            headers = []
            for col in range(self._tablo.columnCount()):
                item = self._tablo.horizontalHeaderItem(col)
                headers.append(item.text() if item else f"Kolon {col+1}")
            ws.append(headers)
            
            header_row_idx = 4
            for col_idx in range(len(headers)):
                cell = ws.cell(row=header_row_idx, column=col_idx+1)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_thin
            
            # Helper to clean financial strings to pure float
            def clean_amount(val_str: str):
                if not val_str or val_str.strip() in ("-", ""):
                    return 0.0
                s = val_str.replace("₺", "").replace("TL", "").replace("$", "").replace("€", "").replace("+", "").strip()
                s = s.replace(".", "").replace(",", ".")
                try:
                    return float(s)
                except ValueError:
                    return 0.0
            
            # Sum tracker for columns
            column_sums = {col: 0.0 for col in range(self._tablo.columnCount())}
            
            # Rows writing
            current_row_idx = 5
            for row in range(self._tablo.rowCount()):
                if self._tablo.isRowHidden(row):
                    continue
                
                row_data = []
                for col in range(self._tablo.columnCount()):
                    header_text = headers[col]
                    item = self._tablo.item(row, col)
                    cell_text = item.text() if item else ""
                    
                    if header_text in ("Tutar", "Gelir", "Gider"):
                        val_num = clean_amount(cell_text)
                        row_data.append(val_num)
                        column_sums[col] += val_num
                    else:
                        row_data.append(cell_text)
                        
                ws.append(row_data)
                
                # Style active data row
                for col in range(len(row_data)):
                    cell = ws.cell(row=current_row_idx, column=col+1)
                    cell.font = font_data
                    cell.border = border_thin
                    
                    if headers[col] in ("Tutar", "Gelir", "Gider"):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '#,##0.00'
                    elif headers[col] == "Tarih":
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        
                current_row_idx += 1
            
            # 2. Dynamic GENEL TOPLAM Row
            summary_row = []
            for col in range(self._tablo.columnCount()):
                if col == 0:
                    summary_row.append("GENEL TOPLAM")
                elif headers[col] in ("Tutar", "Gelir", "Gider"):
                    summary_row.append(column_sums[col])
                else:
                    summary_row.append("")
                    
            ws.append(summary_row)
            
            # Style summary row
            for col in range(len(summary_row)):
                cell = ws.cell(row=current_row_idx, column=col+1)
                cell.font = font_total
                cell.fill = fill_total
                cell.border = border_total
                
                if col == 0:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif headers[col] in ("Tutar", "Gelir", "Gider"):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '#,##0.00'
            
            # 3. Auto Width Adjustment
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row in (1, 2, 3): # Skip title headers for width
                        continue
                    if cell.value is not None:
                        # If it is float formatted, round to string format
                        if isinstance(cell.value, float):
                            val_str = f"{cell.value:,.2f}"
                        else:
                            val_str = str(cell.value)
                        max_len = max(max_len, len(val_str))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
            # Set gridlines visible
            ws.views.sheetView[0].showGridLines = True
            
            wb.save(path)
            
            # Premium styled success popup
            msg = QMessageBox(self)
            msg.setWindowTitle("Başarılı")
            msg.setText("Excel raporu ve genel toplamlar başarıyla şekillendirilip kaydedildi!")
            msg.setStyleSheet("""
                QMessageBox { background-color: white; }
                QLabel { color: #1F2937; font-size: 13px; font-weight: 600; min-width: 280px; min-height: 40px; }
                QPushButton { background-color: #2563EB; color: white; border: none; border-radius: 6px; padding: 6px 18px; font-size: 11px; font-weight: bold; }
                QPushButton:hover { background-color: #1D4ED8; }
            """)
            msg.exec()
            
        except Exception as e:
            import traceback
            err_msg = f"Excel kaydedilirken hata oluştu:\n{e}\n\nDetay:\n{traceback.format_exc()}"
            print(err_msg)
            
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Icon.Critical)
            msg.setWindowTitle("Hata")
            msg.setText(f"Excel raporu oluşturulurken beklenmedik hata oluştu:\n{e}")
            msg.setStyleSheet("""
                QMessageBox { background-color: white; }
                QLabel { color: #DC2626; font-size: 12px; font-weight: 600; min-width: 240px; }
                QPushButton { background-color: #DC2626; color: white; border: none; border-radius: 6px; padding: 6px 16px; }
            """)
            msg.exec()
