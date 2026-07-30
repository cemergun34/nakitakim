"""
Dashboard ekranı — PHP admin_panel.php / admin.php'nin PyQt6 karşılığı.
12 KPI kartı + Yıllık Gelir-Gider grafik.
"""
from __future__ import annotations
from datetime import datetime
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QScrollArea,
    QGridLayout, QFrame, QPushButton, QComboBox, QSizePolicy,
    QDateEdit, QSpacerItem, QDialog, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView, QLineEdit
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QDate
from PyQt6.QtGui import QFont

from ui.theme import COLORS, FONTS
from ui.components.kpi_card import KPICard
from ui.components.detay_dialog import DetayDialog
from ui.components.gelir_gider_chart import GelirGiderChart
from services.dashboard_service import get_all_dashboard_data
from services import detay_service
from utils.format import fmt_para, fmt_signed


class PaytrKPICard(KPICard):
    """
    PHP admin.php #btnSanalPosHareketleri kartının PyQt6 karşılığı.

    Layout (PHP HTML birebir):
      H3  : 'Sanal Pos Paytr'  (üst sol)
      div : Fark değeri  (balance — yeşil badge)
      row : [İşlem  |  Ödeme]  (iki sütun)
      son : 'Son güncelleme: DD.MM.YYYY'  (alt sağ opak)
    """

    def __init__(self, click_cb=None, parent=None):
        super().__init__(
            title="Sanal Pos Paytr",
            value="Yükleniyor...",
            color="#212121",
            color2="#000000",
            click_cb=click_cb,
            parent=parent,
        )
        # setFixedHeight üst sınıftan geliyor; biraz daha yüksek yap
        self.setFixedHeight(160)

        # Üst sınıfın 'value_lbl' Fark değerini gösterir (yeşil renk)
        self.value_lbl.setStyleSheet(
            "color: rgb(170,255,204); font-size: 20px; font-weight: 700;"
            " background: transparent; letter-spacing: -0.5px;"
        )
        # Title rengi beyaz
        self.title_lbl.setStyleSheet(
            "color: #ffffff; font-size: 12px; font-weight: 600;"
            " background: transparent;"
        )

        # Ekstra satır: [İşlem | Ödeme]
        from PyQt6.QtWidgets import QHBoxLayout, QVBoxLayout, QLabel
        row_w = QHBoxLayout()
        row_w.setSpacing(4)

        def _blok(key: str, lbl: str, clr: str) -> QLabel:
            col = QVBoxLayout()
            col.setSpacing(0)
            hdr = QLabel(lbl)
            hdr.setAlignment(Qt.AlignmentFlag.AlignCenter)
            hdr.setStyleSheet(
                "color:rgba(255,255,255,.70);font-size:11px;background:transparent;"
            )
            val = QLabel("-")
            val.setAlignment(Qt.AlignmentFlag.AlignCenter)
            val.setStyleSheet(
                f"color:{clr};font-size:12px;font-weight:600;background:transparent;"
            )
            col.addWidget(hdr)
            col.addWidget(val)
            setattr(self, key, val)
            row_w.addLayout(col)
            return val

        _blok("_islem_lbl", "İşlem",  "#ffe0a0")
        _blok("_odeme_lbl", "Ödeme",  "#d0ffd0")

        # Son güncelleme etiketi
        self._son_guncelleme_lbl = QLabel("")
        self._son_guncelleme_lbl.setStyleSheet(
            "color:rgba(255,255,255,.80);font-size:10px;background:transparent;"
        )

        # Mevcut layout'a ekle
        lay = self.layout()
        lay.addLayout(row_w)
        lay.addWidget(self._son_guncelleme_lbl)

    def set_paytr(
        self,
        fark_fmt: str,
        islem_fmt: str,
        odeme_fmt: str,
        son_guncelleme: str = "",
    ):
        """
        PHP'deki:
          #paytrToplamBadge   → fark_fmt
          #paytrDashIslem     → islem_fmt
          #paytrDashOdeme     → odeme_fmt
          #paytrSonGuncelleme → son_guncelleme
        """
        self.value_lbl.setText(fark_fmt)
        self._islem_lbl.setText(islem_fmt)
        self._odeme_lbl.setText(odeme_fmt)
        self._son_guncelleme_lbl.setText(son_guncelleme)


class FizikselPosKPICard(KPICard):
    """
    PHP admin.php #btnFizikselPosHareketleri kartının PyQt6 karşılığı.

    Layout:
      H3  : 'Fiziksel Pos Womsis'
      div : Net Tutar  (balance — mint yeşil)
      row : [İşlem Tutarı | Net Tutar]  (sarı / yeşil)
      son : 'Yerel tablodan anlık veriler'

    Renk: #1a3a5c → #0d2137 (PHP: linear-gradient(135deg, #1a3a5c, #0d2137))
    """

    def __init__(self, click_cb=None, parent=None):
        super().__init__(
            title="Fiziksel Pos Womsis",
            value="₺0,00",
            color="#1a3a5c",
            color2="#0d2137",
            click_cb=click_cb,
            parent=parent,
        )
        self.setFixedHeight(160)

        # Net Tutar badge — PHP: #womsisToplamBadge  color: rgb(170,255,204)
        self.value_lbl.setStyleSheet(
            "color: rgb(170,255,204); font-size:20px; font-weight:700;"
            " background:transparent; letter-spacing:-0.5px;"
        )
        self.title_lbl.setStyleSheet(
            "color:#ffffff; font-size:12px; font-weight:600; background:transparent;"
        )

        from PyQt6.QtWidgets import QHBoxLayout as _HBL, QVBoxLayout as _VBL, QLabel as _LBL
        row_w = _HBL()
        row_w.setSpacing(4)

        def _blok(key: str, lbl: str, clr: str):
            col = _VBL()
            col.setSpacing(0)
            hdr = _LBL(lbl)
            hdr.setAlignment(Qt.AlignmentFlag.AlignCenter)
            hdr.setStyleSheet("color:rgba(255,255,255,.70);font-size:11px;background:transparent;")
            val = _LBL("-")
            val.setAlignment(Qt.AlignmentFlag.AlignCenter)
            val.setStyleSheet(f"color:{clr};font-size:12px;font-weight:600;background:transparent;")
            col.addWidget(hdr)
            col.addWidget(val)
            setattr(self, key, val)
            row_w.addLayout(col)

        _blok("_islem_lbl", "İşlem",  "#ffe0a0")   # PHP: #womsisDashIslem
        _blok("_odeme_lbl", "Net",    "#d0ffd0")   # PHP: #womsisDashOdeme

        self._son_lbl = _LBL("Yerel tablodan anlık veriler")
        self._son_lbl.setStyleSheet(
            "color:rgba(255,255,255,.80);font-size:10px;background:transparent;"
        )

        lay = self.layout()
        lay.addLayout(row_w)
        lay.addWidget(self._son_lbl)

    def set_fiziksel_pos(
        self,
        net_fmt: str,
        islem_fmt: str,
        odeme_fmt: str,
        alt_yazi: str = "Yerel tablodan anlık veriler",
    ):
        """
        PHP:
          #womsisToplamBadge  → net_fmt
          #womsisDashIslem    → islem_fmt
          #womsisDashOdeme    → odeme_fmt
          #womsisSonGuncelleme→ alt_yazi
        """
        self.value_lbl.setText(net_fmt)
        self._islem_lbl.setText(islem_fmt)
        self._odeme_lbl.setText(odeme_fmt)
        self._son_lbl.setText(alt_yazi)


class BankalaBakiyeKPICard(KPICard):
    """
    Bankalar Bakiye kartı — PaytrKPICard ile aynı pattern.

    Layout:
      H3  : 'Bankalar Bakiye'
      div : Net Tutar  (balance — mint yeşil badge)
      row : [Gelir | Gider]  (iki sütun, sarı / kırmızı)
      son : 'womsis_banka · tüm dönem'
    Renk: #374151 → #1F2937
    """

    def __init__(self, click_cb=None, parent=None):
        super().__init__(
            title="Bankalar Bakiye",
            value="₺0,00",
            color="#374151",
            color2="#1F2937",
            click_cb=click_cb,
            parent=parent,
        )
        self.setFixedHeight(160)

        # Net badge — mint yeşil
        self.value_lbl.setStyleSheet(
            "color: rgb(170,255,204); font-size:20px; font-weight:700;"
            " background:transparent; letter-spacing:-0.5px;"
        )
        self.title_lbl.setStyleSheet(
            "color:#ffffff; font-size:12px; font-weight:600; background:transparent;"
        )

        from PyQt6.QtWidgets import QHBoxLayout as _HBL, QVBoxLayout as _VBL, QLabel as _LBL
        row_w = _HBL()
        row_w.setSpacing(4)

        def _blok(key: str, lbl: str, clr: str):
            col = _VBL()
            col.setSpacing(0)
            hdr = _LBL(lbl)
            hdr.setAlignment(Qt.AlignmentFlag.AlignCenter)
            hdr.setStyleSheet("color:rgba(255,255,255,.70);font-size:11px;background:transparent;")
            val = _LBL("-")
            val.setAlignment(Qt.AlignmentFlag.AlignCenter)
            val.setStyleSheet(f"color:{clr};font-size:12px;font-weight:600;background:transparent;")
            col.addWidget(hdr)
            col.addWidget(val)
            setattr(self, key, val)
            row_w.addLayout(col)

        _blok("_gelir_lbl", "Gelir",  "#ffe0a0")   # sarı
        _blok("_gider_lbl", "Gider",  "#ffb3b3")   # açık kırmızı

        self._alt_lbl = _LBL("womsis_banka · tüm dönem")
        self._alt_lbl.setStyleSheet(
            "color:rgba(255,255,255,.80);font-size:10px;background:transparent;"
        )

        lay = self.layout()
        lay.addLayout(row_w)
        lay.addWidget(self._alt_lbl)

    def set_bankalar(self, net_fmt: str, gelir_fmt: str, gider_fmt: str,
                     alt_yazi: str = "womsis_banka · tüm dönem"):
        """Net (büyük badge) + Gelir | Gider satırını doldur."""
        self.value_lbl.setText(net_fmt)
        self._gelir_lbl.setText(gelir_fmt)
        self._gider_lbl.setText(gider_fmt)
        self._alt_lbl.setText(alt_yazi)

class KrediKartiKPICard(KPICard):
    """
    Kredi Kartları kartı — BankalaBakiyeKPICard ile aynı pattern.
    """
    def __init__(self, click_cb=None, parent=None):
        super().__init__(
            title="Kredi Kartları",
            value="₺0,00",
            color="#D97706",
            color2="#B45309",
            click_cb=click_cb,
            parent=parent,
        )
        self.setFixedHeight(160)

        # Net badge — mint yeşil
        self.value_lbl.setStyleSheet(
            "color: rgb(170,255,204); font-size:20px; font-weight:700;"
            " background:transparent; letter-spacing:-0.5px;"
        )
        self.title_lbl.setStyleSheet(
            "color:#ffffff; font-size:12px; font-weight:600; background:transparent;"
        )

        from PyQt6.QtWidgets import QHBoxLayout as _HBL, QVBoxLayout as _VBL, QLabel as _LBL
        row_w = _HBL()
        row_w.setSpacing(4)

        def _blok(key: str, lbl: str, clr: str):
            col = _VBL()
            col.setSpacing(0)
            hdr = _LBL(lbl)
            hdr.setAlignment(Qt.AlignmentFlag.AlignCenter)
            hdr.setStyleSheet("color:rgba(255,255,255,.70);font-size:11px;background:transparent;")
            val = _LBL("-")
            val.setAlignment(Qt.AlignmentFlag.AlignCenter)
            val.setStyleSheet(f"color:{clr};font-size:12px;font-weight:600;background:transparent;")
            col.addWidget(hdr)
            col.addWidget(val)
            setattr(self, key, val)
            row_w.addLayout(col)

        _blok("_harcama_lbl", "Harcama", "#ffb3b3")   # açık kırmızı
        _blok("_odeme_lbl", "Ödeme",   "#ffe0a0")   # sarı

        self._alt_lbl = _LBL("kredikartidata · tüm dönem")
        self._alt_lbl.setStyleSheet(
            "color:rgba(255,255,255,.80);font-size:10px;background:transparent;"
        )

        lay = self.layout()
        lay.addLayout(row_w)
        lay.addWidget(self._alt_lbl)

    def set_kredi_karti(self, net_fmt: str, harcama_fmt: str, odeme_fmt: str,
                     alt_yazi: str = "kredikartidata · tüm dönem"):
        self.value_lbl.setText(net_fmt)
        self._harcama_lbl.setText(harcama_fmt)
        self._odeme_lbl.setText(odeme_fmt)
        self._alt_lbl.setText(alt_yazi)


class DashboardLoader(QThread):
    """Arka planda veri yükler — UI donmaması için."""
    data_ready = pyqtSignal(dict)
    error = pyqtSignal(str)

    def __init__(self, userid: int, musterino: int, yil: int,
                 ilk_tarih: str = None, son_tarih: str = None):
        super().__init__()
        self.userid     = userid
        self.musterino  = musterino
        self.yil        = yil
        self.ilk_tarih  = ilk_tarih
        self.son_tarih  = son_tarih

    def run(self):
        try:
            data = get_all_dashboard_data(
                self.userid, self.musterino, self.yil,
                ilk_tarih=self.ilk_tarih,
                son_tarih=self.son_tarih,
            )
            self.data_ready.emit(data)
        except Exception as e:
            self.error.emit(str(e))



# ─────────────────────────────────────────────────────────────────────────────
# ExcelOzetBuilder — Her sheet için kayıt/gelir/gider/fark izleyici
# _export_excel metodunda kullanılır.
# ─────────────────────────────────────────────────────────────────────────────

class ExcelOzetBuilder:
    """
    Excel export sırasında her sheet için istatistik toplar ve
    ÖZET sheet'ini openpyxl Workbook'a yazar.

    Kullanım:
        ozet = ExcelOzetBuilder()
        ozet.kaydet("Nakit Kasa", kayit=100, gelir=5000.0, gider=2000.0)
        ozet.yaz(wb, baslik="Dashboard 2026", simdi="08.07.2026 14:00")

    Her sheet için desteklenen sütunlar:
        Sheet Adı | Kayıt | Gelir (₺) | Gider (₺) | Fark (₺)
    Tutar bilgisi olmayan sheetler için Gelir/Gider/Fark boş bırakılır.
    """

    # Her sheet için hangi para sütununu hangi gelir/gider anahtarıyla eşle
    _GELIR_HEADERS = {"Gelir (₺)", "İşlem Tutarı (₺)", "Ödeme Tutarı (₺)",
                      "Net Tutar (₺)", "Brüt Tutar (₺)", "Tutar (₺)",
                      "Gayri Resmi Tutar (₺)"}
    _GIDER_HEADERS = {"Gider (₺)", "Komisyon (₺)", "Kesinti Tutarı (₺)",
                      "Vergi Kesinti (₺)"}

    def __init__(self):
        # {sheet_adı: {"kayit": int, "gelir": float, "gider": float}}
        self._data: dict[str, dict] = {}

    # ── Public API ───────────────────────────────────────────────────────────

    def kaydet(self, sheet_adi: str, kayit: int = 0,
               gelir: float | None = None,
               gider: float | None = None) -> None:
        """Sheet istatistiğini kaydet."""
        self._data[sheet_adi] = {
            "kayit": kayit,
            "gelir": gelir,   # None → bu sheet için gelir yok
            "gider": gider,   # None → bu sheet için gider yok
        }

    def kayit_sayisi(self, sheet_adi: str) -> int:
        return self._data.get(sheet_adi, {}).get("kayit", 0)

    def toplam_kayit(self) -> int:
        return sum(d["kayit"] for d in self._data.values())

    def ozet_str(self) -> str:
        """Başarı popup'ı için tek satır özet."""
        satirlar = []
        for ad, d in self._data.items():
            gelir = d["gelir"]
            gider = d["gider"]
            if gelir is not None:
                net = (gelir or 0) - (gider or 0)
                satirlar.append(
                    f"  • {ad}: {d['kayit']:,} kayıt | "
                    f"G:{gelir:,.0f}₺ / G:{gider or 0:,.0f}₺ | Net:{net:+,.0f}₺"
                )
            else:
                satirlar.append(f"  • {ad}: {d['kayit']:,} kayıt")
        return "\n".join(satirlar)

    def yaz(self, wb, baslik: str, simdi: str,
            hdr_color: str = "1E3A8A",
            monthly_data: list = None) -> None:
        """
        Workbook'a index=0 olarak ÖZET sheet'ini ekler.
        Sütunlar: Sheet Adı | Kayıt | Gelir (₺) | Gider (₺) | Fark (₺)
        Tutar bilgisi olmayan satırlarda para sütunları boş kalır.
        monthly_data: [{'ay': int, 'toplam_gelir': float, 'toplam_gider': float}, ...]
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        ws = wb.create_sheet(title="ÖZET", index=0)

        # ── Stil setleri ──
        def _f(bold=False, size=10, color="FF1F2937", italic=False):
            return Font(name="Segoe UI", size=size, bold=bold,
                        color=color, italic=italic)

        def _fill(hex_color: str):
            return PatternFill(start_color=hex_color, end_color=hex_color,
                               fill_type="solid")

        def _border(light="FFE5E7EB"):
            s = Side(style="thin", color=light)
            return Border(left=s, right=s, top=s, bottom=s)

        border_total = Border(
            top=Side(style="thin", color="FF94A3B8"),
            bottom=Side(style="double", color=hdr_color),
        )

        # ── Başlık blokları ──
        ws.append([baslik])
        ws.cell(row=1, column=1).font = _f(bold=True, size=16, color=hdr_color)
        ws.merge_cells("A1:E1")

        ws.append([f"Oluşturulma: {simdi}"])
        ws.cell(row=2, column=1).font = _f(size=9, color="FF4B5563", italic=True)
        ws.append([])  # boşluk

        # ── Başlık satırı ──
        HEADERS = ["Sheet Adı", "Kayıt", "Gelir (₺)", "Gider (₺)", "Fark (₺)"]
        ws.append(HEADERS)
        for ci, hdr in enumerate(HEADERS, 1):
            cell = ws.cell(row=4, column=ci)
            cell.font      = _f(bold=True, size=10, color="FFFFFFFF")
            cell.fill      = _fill(hdr_color)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _border()

        # ── Veri satırları ──
        r = 5
        toplam_gelir = toplam_gider = 0.0
        has_gelir = False

        for sheet_adi, d in self._data.items():
            kayit  = d.get("kayit", 0)
            gelir  = d.get("gelir")
            gider  = d.get("gider") or 0.0

            row_vals = [sheet_adi, kayit, None, None, None]
            if gelir is not None:
                has_gelir = True
                fark = gelir - gider
                row_vals[2] = gelir
                row_vals[3] = gider
                row_vals[4] = fark
                toplam_gelir += gelir
                toplam_gider += gider

            # Satır arka plan: gelir/gider farka göre hafif ton
            if gelir is not None:
                fark_val = gelir - gider
                if fark_val > 0:
                    row_fill = _fill("F0FDF4")   # hafif yeşil
                elif fark_val < 0:
                    row_fill = _fill("FFF1F2")   # hafif kırmızı
                else:
                    row_fill = _fill("F8FAFC")   # nötr
            else:
                row_fill = _fill("F8FAFC")

            ws.append(row_vals)
            for ci, val in enumerate(row_vals, 1):
                cell = ws.cell(row=r, column=ci)
                cell.fill   = row_fill
                cell.border = _border()
                if ci == 1:
                    cell.font      = _f(bold=False)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif ci == 2:
                    cell.font         = _f(bold=True, color="FF1E3A8A")
                    cell.alignment    = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "#,##0"
                elif ci in (3, 4, 5):
                    cell.font         = _f(bold=(ci == 5), color="FF047857" if ci == 5 and (val or 0) >= 0 else "FFDC2626")
                    cell.alignment    = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "#,##0.00"
                    if val is not None and ci == 5:
                        cell.font = _f(bold=True,
                                       color="FF047857" if val >= 0 else "FFDC2626")
            r += 1

        # ── TOPLAM satırı ──
        if has_gelir:
            t_row = ["GENEL TOPLAM",
                     self.toplam_kayit(),
                     toplam_gelir,
                     toplam_gider,
                     toplam_gelir - toplam_gider]
            ws.append(t_row)
            for ci, val in enumerate(t_row, 1):
                cell = ws.cell(row=r, column=ci)
                cell.font   = _f(bold=True, size=11, color=hdr_color)
                cell.fill   = _fill("EFF6FF")
                cell.border = border_total
                if ci == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if ci == 2:
                        cell.number_format = "#,##0"
                    else:
                        cell.number_format = "#,##0.00"
                        if ci == 5:
                            net_val = toplam_gelir - toplam_gider
                            cell.font = _f(bold=True, size=11,
                                           color="FF047857" if net_val >= 0 else "FFDC2626")
            r += 1

        # ── Sütun genişlikleri ──
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 16

        ws.freeze_panes = "A5"
        ws.sheet_view.showGridLines = False

        # ── AYLIK GELİR-GİDER TABLOSU + GRAFİK ──────────────────────────────
        if monthly_data:
            try:
                from openpyxl.chart import BarChart, Reference
                from openpyxl.chart.series import SeriesLabel

                AY_ADLARI = ["Oca", "Şub", "Mar", "Nis", "May", "Haz",
                             "Tem", "Ağu", "Eyl", "Eki", "Kas", "Ara"]

                # Boş satır bırak, sonra tablo başlığı
                r += 1  # boşluk
                tablo_baslik_row = r
                ws.cell(row=r, column=1, value="Aylık Gelir-Gider Karşılaştırması")
                ws.cell(row=r, column=1).font  = _f(bold=True, size=12, color=hdr_color)
                ws.cell(row=r, column=1).fill  = _fill("EFF6FF")
                ws.merge_cells(f"A{r}:E{r}")
                r += 1

                # Tablo başlık satırı
                tablo_hdr_row = r
                for ci, hdr in enumerate(["Ay", "Gelir (₺)", "Gider (₺)", "Net (₺)"], 1):
                    cell = ws.cell(row=r, column=ci, value=hdr)
                    cell.font      = _f(bold=True, size=10, color="FFFFFFFF")
                    cell.fill      = _fill(hdr_color)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border    = _border()
                r += 1

                # Ay verilerini 12 ay olarak normalize et
                ay_map = {int(m.get("ay", 0)): m for m in monthly_data}
                tablo_veri_baslangic = r

                for ay_no in range(1, 13):
                    m = ay_map.get(ay_no, {})
                    gelir_v = float(m.get("toplam_gelir") or 0)
                    gider_v = float(m.get("toplam_gider") or 0)
                    net_v   = gelir_v - gider_v
                    ay_adi  = AY_ADLARI[ay_no - 1]

                    ws.cell(row=r, column=1, value=ay_adi)
                    ws.cell(row=r, column=1).font      = _f(bold=True)
                    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=r, column=1).border    = _border()
                    ws.cell(row=r, column=1).fill      = _fill("F8FAFC")

                    for ci, (val, is_net) in enumerate([(gelir_v, False), (gider_v, False), (net_v, True)], 2):
                        cell = ws.cell(row=r, column=ci, value=val)
                        cell.number_format = "#,##0.00"
                        cell.alignment     = Alignment(horizontal="right", vertical="center")
                        cell.border        = _border()
                        if is_net:
                            cell.font = _f(bold=True,
                                           color="FF047857" if val >= 0 else "FFDC2626")
                            cell.fill = _fill("F0FDF4" if val >= 0 else "FFF1F2")
                        else:
                            cell.font = _f(color="FF065F46" if ci == 2 else "FF991B1B")
                            cell.fill = _fill("F0FDF4" if ci == 2 else "FFF1F2")
                    r += 1

                tablo_veri_bitis = r - 1

                # Sütun genişlikleri
                ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 10)
                ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 18)
                ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width, 18)
                ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width, 18)

                # ── BarChart oluştur ──
                chart = BarChart()
                chart.type        = "col"          # dikey bar
                chart.grouping    = "clustered"    # yanyana
                chart.title       = "Aylık Gelir / Gider Karşılaştırması"
                chart.y_axis.title = "Tutar (₺)"
                chart.x_axis.title = "Ay"
                chart.style       = 10
                chart.width       = 22             # cm
                chart.height      = 12             # cm

                # Gelir serisi (kolon B)
                gelir_ref = Reference(ws,
                    min_col=2, max_col=2,
                    min_row=tablo_veri_baslangic,
                    max_row=tablo_veri_bitis)
                chart.add_data(gelir_ref)
                chart.series[0].title       = SeriesLabel(v="Gelir")
                chart.series[0].graphicalProperties.solidFill  = "10B981"
                chart.series[0].graphicalProperties.line.solidFill = "059669"

                # Gider serisi (kolon C)
                gider_ref = Reference(ws,
                    min_col=3, max_col=3,
                    min_row=tablo_veri_baslangic,
                    max_row=tablo_veri_bitis)
                chart.add_data(gider_ref)
                chart.series[1].title       = SeriesLabel(v="Gider")
                chart.series[1].graphicalProperties.solidFill  = "EF4444"
                chart.series[1].graphicalProperties.line.solidFill = "DC2626"

                # X ekseni etiketleri (Oca..Ara)
                cats = Reference(ws,
                    min_col=1, max_col=1,
                    min_row=tablo_veri_baslangic,
                    max_row=tablo_veri_bitis)
                chart.set_categories(cats)

                # Grafiği pivot tablonun hemen altına yap
                grafik_anchor = f"A{r + 1}"
                ws.add_chart(chart, grafik_anchor)

            except Exception as _chart_err:
                # Grafik oluşturulamazsa sessizce atla, tablo kalır
                ws.cell(row=r, column=1,
                        value=f"(Grafik oluşturulamadı: {_chart_err})")
                ws.cell(row=r, column=1).font = _f(size=9, color="FFDC2626", italic=True)



class DashboardScreen(QWidget):
    """
    Ana dashboard ekranı.
    12 KPI kartını 2 satır × 6 sütun grid içinde gösterir.
    """

    def __init__(self, user: dict, parent=None):
        super().__init__(parent)
        self._user = user
        self._userid = user.get("GercekUserId", user.get("Kayitno", 1))
        self._musterino = user.get("musterino", user.get("GercekUserId", 1))
        self._yil = datetime.now().year
        self._loader = None
        self._cards: dict[str, KPICard] = {}
        self._setup_ui()
        self._load_data()

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(20)

        # ── Başlık çubuğu ──────────────────────────────────────────────────────
        header = QHBoxLayout()
        title = QLabel("Dashboard")
        title.setStyleSheet(f"""
            color: {COLORS['text_primary']};
            font-size: 24px;
            font-weight: 700;
            letter-spacing: 1px;
        """)
        header.addWidget(title)
        header.addStretch()
        # Şirket / kullanıcı
        sirket = self._user.get("Kurum_Durumu", "IQ Finans")
        kullanici = self._user.get("Adi", "")
        profil = QLabel(f"🏢 {sirket}  •  👤 {kullanici}")
        profil.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 12px; margin-right: 15px;")
        header.addWidget(profil)

        root.addLayout(header)

        # ── Bilgi banner ────────────────────────────────────────────────────────
        self.banner = QLabel("FİNANS DURUM BİLGİSİ")
        self.banner.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.banner.setFixedHeight(38)
        self.banner.setStyleSheet(f"""
            background: {COLORS['text_primary']};
            color: white;
            font-size: 13px;
            font-weight: 700;
            letter-spacing: 3px;
            border-radius: 10px;
            padding: 0 20px;
        """)
        root.addWidget(self.banner)

        # Banner altı araç çubuğu (Yıl Seçimi sağda)
        banner_sub = QHBoxLayout()
        banner_sub.addStretch()

        self.yil_combo = QComboBox()
        current_year = datetime.now().year
        for y in range(current_year + 1, current_year - 6, -1):
            self.yil_combo.addItem(str(y), y)
        self.yil_combo.setCurrentText(str(self._yil))
        self.yil_combo.setFixedHeight(32)
        self.yil_combo.setFixedWidth(120)
        self.yil_combo.setStyleSheet(f"""
            QComboBox {{
                background: white;
                border: 1.5px solid {COLORS['border']};
                border-radius: 8px;
                padding: 0 10px;
                font-size: 12px;
                font-weight: 600;
                color: {COLORS['text_primary']};
            }}
            QComboBox:focus {{
                border-color: {COLORS['btn_primary']};
            }}
            QComboBox::drop-down {{ border: none; }}
        """)
        self.yil_combo.currentIndexChanged.connect(self._on_yil_changed)
        banner_sub.addWidget(self.yil_combo)
        root.addLayout(banner_sub)

        # ── Scroll alan ─────────────────────────────────────────────────────────
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setStyleSheet("QScrollArea { background: transparent; }")

        container = QWidget()
        container.setStyleSheet("background: transparent;")
        vbox = QVBoxLayout(container)
        vbox.setContentsMargins(0, 0, 0, 0)
        vbox.setSpacing(20)

        # ── KPI Grid (2 × 6) ───────────────────────────────────────────────────
        grid = QGridLayout()
        grid.setSpacing(16)

        KPI_DEFS = [
            # (key,              başlık,              renk1,                renk2,        satır, sütun)
            ("nakit_kasa_gelir", "Nakit Kasa Gelir",  COLORS["green"],      "#059669",    0, 0),
            ("nakit_kasa_odeme", "Nakit Kasa Ödeme",  COLORS["teal"],       "#0284C7",    0, 1),
            ("kesilen_fatura",   "Kesilen Fatura",     COLORS["purple"],     "#6D28D9",    0, 2),
            ("gelen_fatura",     "Gelen Fatura",       COLORS["pink"],       "#DB2777",    0, 3),
            ("gider_pusulasi",   "Gider Pusulası",     "#16A34A",            "#15803D",    0, 4),
            ("kurum_odemeleri",  "Kurum Ödemeleri",    COLORS["dark_blue"],  "#162C47",    0, 5),
            ("maas_kira_smm",    "Maaş Kira SMM",      "#1a3a5c",            "#0d2137",    1, 0),
            ("bankalar_bakiye",  "Bankalar Bakiye",    COLORS["grey"],       "#4B5563",    1, 1),
            ("sanal_pos",        "Sanal Pos",          "#111827",            "#030712",    1, 2),
            ("fiziksel_pos",     "Fiziksel Pos",       "#1F2937",            "#111827",    1, 3),
            ("kredi_karti",      "Kredi Kartları",     "#D97706",            "#B45309",    1, 4),
            ("genel_hesap",      "Genel Hesap Tablosu","#EA580C",            "#C2410C",    1, 5),
        ]

        for key, title, c1, c2, row, col in KPI_DEFS:
            if key == "sanal_pos":
                card = PaytrKPICard(
                    click_cb=lambda k=key: self._on_card_click(k),
                )
            elif key == "fiziksel_pos":
                card = FizikselPosKPICard(
                    click_cb=lambda k=key: self._on_card_click(k),
                )
            elif key == "bankalar_bakiye":
                card = BankalaBakiyeKPICard(
                    click_cb=lambda k=key: self._on_card_click(k),
                )
            elif key == "kredi_karti":
                card = KrediKartiKPICard(
                    click_cb=lambda k=key: self._on_card_click(k),
                )
            else:
                card = KPICard(
                    title=title,
                    value="Yükleniyor...",
                    color=c1,
                    color2=c2,
                    click_cb=lambda k=key: self._on_card_click(k),
                )
            card.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            self._cards[key] = card
            grid.addWidget(card, row, col)

        vbox.addLayout(grid)

        # ── Tarih filtresi + Göster butonu + Excel ──────────────────────────────
        filter_bar = QHBoxLayout()
        filter_bar.setSpacing(10)

        _DE_STYLE = """
            QDateEdit {
                background: #1A1A1A;
                border: 1.5px solid #555;
                border-radius: 8px;
                padding: 0 10px;
                font-size: 12px;
                color: white;
                font-weight: 600;
            }
            QDateEdit:focus { border-color: #2563EB; }
            QDateEdit::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 26px;
                border-left: 1.5px solid #555;
                border-top-right-radius: 8px;
                border-bottom-right-radius: 8px;
                background: #2A2A2A;
            }
            QDateEdit::down-arrow { image: none; width: 10px; height: 10px; }
        """

        ilk_lbl = QLabel("📅 İlk Tarih:")
        ilk_lbl.setStyleSheet("color: white; font-size: 12px; font-weight: 600;")
        filter_bar.addWidget(ilk_lbl)

        self.ilk_tarih = QDateEdit()
        self.ilk_tarih.setCalendarPopup(True)
        self.ilk_tarih.setDate(QDate(self._yil, 1, 1))
        self.ilk_tarih.setDisplayFormat("dd.MM.yyyy")
        self.ilk_tarih.setFixedHeight(34)
        self.ilk_tarih.setFixedWidth(135)
        self.ilk_tarih.setStyleSheet(_DE_STYLE)
        filter_bar.addWidget(self.ilk_tarih)

        son_lbl = QLabel("📅 Son Tarih:")
        son_lbl.setStyleSheet("color: white; font-size: 12px; font-weight: 600;")
        filter_bar.addWidget(son_lbl)

        self.son_tarih = QDateEdit()
        self.son_tarih.setCalendarPopup(True)
        self.son_tarih.setDate(QDate.currentDate())
        self.son_tarih.setDisplayFormat("dd.MM.yyyy")
        self.son_tarih.setFixedHeight(34)
        self.son_tarih.setFixedWidth(135)
        self.son_tarih.setStyleSheet(_DE_STYLE)
        filter_bar.addWidget(self.son_tarih)

        # —— Göster butonu
        self.goster_btn = QPushButton("▶  GÖSTER")
        self.goster_btn.setFixedHeight(34)
        self.goster_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.goster_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #2563EB, stop:1 #1d4ed8);
                color: white;
                font-size: 12px;
                font-weight: 700;
                border: none;
                border-radius: 8px;
                padding: 0 18px;
                letter-spacing: 1px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #1d4ed8, stop:1 #1e40af);
            }
            QPushButton:pressed { background: #1e40af; }
        """)
        self.goster_btn.clicked.connect(self._on_goster)
        filter_bar.addWidget(self.goster_btn)

        filter_bar.addStretch()

        self.excel_btn = QPushButton("📥 EXCEL İNDİR")
        self.excel_btn.setFixedHeight(34)
        self.excel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.excel_btn.setStyleSheet("""
            QPushButton {
                background: #000000;
                color: white;
                font-size: 12px;
                font-weight: 700;
                border: 1.5px solid #444;
                border-radius: 8px;
                padding: 0 18px;
                letter-spacing: 1px;
            }
            QPushButton:hover { background: #1A1A1A; border-color: #2563EB; }
        """)
        self.excel_btn.clicked.connect(self._export_excel)
        filter_bar.addWidget(self.excel_btn)

        vbox.addLayout(filter_bar)

        # ── Grafik ─────────────────────────────────────────────────────────────
        self.chart = GelirGiderChart()
        vbox.addWidget(self.chart)

        scroll.setWidget(container)
        root.addWidget(scroll)

    def _input_style(self) -> str:
        return f"""
            QDateEdit {{
                background: white;
                border: 1.5px solid {COLORS['border']};
                border-radius: 8px;
                padding: 0 10px;
                font-size: 12px;
                color: {COLORS['text_primary']};
            }}
        """

    def _load_data(self):
        if self._loader and self._loader.isRunning():
            return

        ilk_str = self.ilk_tarih.date().toString("yyyy-MM-dd")
        son_str  = self.son_tarih.date().toString("yyyy-MM-dd")

        self._loader = DashboardLoader(
            self._userid, self._musterino, self._yil,
            ilk_tarih=ilk_str, son_tarih=son_str,
        )
        self._loader.data_ready.connect(self._on_data_ready)
        self._loader.error.connect(self._on_error)
        self._loader.start()

    def _on_goster(self):
        """Göster butonuna basılınca kartları ve grafiği tarih aralığına göre yeniler."""
        ilk = self.ilk_tarih.date().toString("dd.MM.yyyy")
        son = self.son_tarih.date().toString("dd.MM.yyyy")
        self.banner.setText(f"FİNANS DURUM BİLGİSİ  —  📅 {ilk}  /  {son}")
        for card in self._cards.values():
            card.set_value("Yükleniyor...")
        self._load_data()

    def refresh(self):
        """Sol menüden sayfaya geçince veriler yeniden yüklenir."""
        self._load_data()

    def _on_data_ready(self, data: dict):
        c = self._cards
        
        # Grafik verilerini güncelle
        self.chart.load_data(data.get("monthly_chart", []))

        # Nakit Kasa Gelir
        kasa = data.get("nakit_kasa", {})
        c["nakit_kasa_gelir"].set_value(
            fmt_para(kasa.get("gelir", 0)),
            f"Gelir: {fmt_para(kasa.get('gelir', 0))}  Gider: {fmt_para(kasa.get('gider', 0))}"
        )

        # Nakit Kasa Ödeme
        gider = data.get("gider", {})
        c["nakit_kasa_odeme"].set_value(
            fmt_para(gider.get("gider", 0)),
            "genel_hesap_hareketleri · gider"
        )

        # Kesilen / Gelen Fatura
        fat = data.get("faturalar", {})
        c["kesilen_fatura"].set_value(
            fmt_para(fat.get("kesilen", 0)),
            "Detaylar için tıklayın..."
        )
        c["gelen_fatura"].set_value(
            fmt_para(fat.get("gelen", 0)),
            "Detaylar için tıklayın..."
        )

        # Gider Pusulası
        gp = data.get("gider_pusulasi", {})
        c["gider_pusulasi"].set_value(
            fmt_para(gp.get("gider", 0)),
            "Parça Alımı (Cihaz)"
        )

        # Kurum Ödemeleri
        kurum = data.get("kurum_odemeleri", {})
        kurum_moy = kurum.get("moy", 0)
        kurum_alt = "Vergi ödeme detayları"
        if kurum_moy > 0:
            kurum_alt += f"  •  Moy: {fmt_para(kurum_moy)}"
        c["kurum_odemeleri"].set_value(
            fmt_para(kurum.get("toplam", 0)),
            kurum_alt
        )

        # Maaş Kira SMM — PHP: #dashMaasKiraSmmmToplam
        # Formül: SUM(ABS(gaytutar - vergkestutar))  her satır için
        from services.vergi_muhtasar_service import get_dashboard_toplam as _mks_toplam
        mks_r = _mks_toplam(self._userid, musterino=str(self._musterino))
        if mks_r.get("success") and mks_r["fark_toplam"] > 0:
            mks_val_str = mks_r["fark_toplam_fmt"]   # '10.853.717,00 ₺'
        else:
            mks = data.get("maas_kira_smm", {})
            mks_val_str = fmt_para(mks.get("toplam", 0))
        c["maas_kira_smm"].set_value(
            mks_val_str,
            "Personel, Kira ve Müşavirlik Giderleri"
        )

        # Bankalar Bakiye (womsis_banka)
        banka = data.get("bankalar", {})
        b_gelir = banka.get("gelir", 0)
        b_gider = banka.get("gider", 0)
        b_net   = banka.get("net",   0)
        banka_card = self._cards.get("bankalar_bakiye")
        if isinstance(banka_card, BankalaBakiyeKPICard):
            banka_card.set_bankalar(
                net_fmt=fmt_para(b_net),
                gelir_fmt=fmt_para(b_gelir),
                gider_fmt=fmt_para(b_gider),
                alt_yazi=f"{banka.get('kayit', 0):,} hareket · tüm dönem",
            )
        else:
            banka_card.set_value(
                fmt_para(b_net),
                f"Gelir: {fmt_para(b_gelir)}  Gider: {fmt_para(b_gider)}"
            )

        # Sanal Pos (PayTR) — PHP: #paytrToplamBadge / #paytrDashIslem / #paytrDashOdeme / #paytrSonGuncelleme
        pos = data.get("sanal_pos", {})
        pos_card = self._cards.get("sanal_pos")
        if isinstance(pos_card, PaytrKPICard):
            # Fark = odeme - islem
            islem_val = pos.get("islem", 0)
            odeme_val = pos.get("odeme", 0)
            try:
                fark_val = float(odeme_val) - float(islem_val)
            except (TypeError, ValueError):
                fark_val = 0.0
            fark_sign = "+" if fark_val >= 0 else "-"
            fark_abs = abs(fark_val)
            fark_fmt = (
                f"{fark_sign}₺"
                + f"{fark_abs:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            )
            islem_fmt = fmt_para(islem_val)
            odeme_fmt = fmt_para(odeme_val)
            son = pos.get("son_guncelleme", "")
            pos_card.set_paytr(fark_fmt, islem_fmt, odeme_fmt, son)
        else:
            pos_card.set_value(
                fmt_para(pos.get("islem", 0)),
                f"İşlem: {fmt_para(pos.get('islem', 0))}  Ödeme: {fmt_para(pos.get('odeme', 0))}"
            )

        # Fiziksel Pos (Womsis) — PHP: #womsisToplamBadge / #womsisDashIslem / #womsisDashOdeme
        fp_card = self._cards.get("fiziksel_pos")
        if isinstance(fp_card, FizikselPosKPICard):
            from services.fiziksel_pos_service import get_dashboard_ozet
            fp = get_dashboard_ozet(self._userid, self._musterino)
            fp_card.set_fiziksel_pos(
                net_fmt=fp.get("toplam_net_fmt",   "₺0,00"),
                islem_fmt=fp.get("toplam_islem_fmt", "₺0,00"),
                odeme_fmt=fp.get("toplam_net_fmt",   "₺0,00"),
                alt_yazi="Yerel tablodan anlık veriler",
            )
        else:
            if fp_card:
                fp_card.set_value("₺0,00", "Yerel tablodan anlık veriler")


        # Kredi Kartları
        kk = data.get("kredi_karti", {})
        borc  = kk.get("borc", 0)
        odeme = kk.get("odeme", 0)
        net   = kk.get("net",   0)
        c["kredi_karti"].set_kredi_karti(
            net_fmt=fmt_para(net),
            harcama_fmt=fmt_para(borc),
            odeme_fmt=fmt_para(abs(odeme)),
            alt_yazi="kredikartidata"
        )

        # Genel Hesap
        gh = data.get("genel_hesap", {})
        c["genel_hesap"].set_value(
            fmt_para(gh.get("net", 0)),
            "genel_hesap_hareketleri · genelHesap"
        )

    def _on_error(self, msg: str):
        self.banner.setText(f"⚠  Hata: {msg}")
        self.banner.setStyleSheet(f"""
            background: {COLORS['red']};
            color: white;
            font-size: 12px;
            font-weight: 600;
            border-radius: 10px;
            padding: 0 20px;
        """)
        for card in self._cards.values():
            card.set_value("Hata")

    # ─── KART TIKLAMA ────────────────────────────────────────────────────────

    def _on_card_click(self, key: str):
        """KPI kartına tıklandığında ilgili detay diyaloğunu açar."""
        uid = self._userid
        mno = self._musterino
        yil = self._yil

        if key in ("nakit_kasa_gelir", "nakit_kasa_odeme"):
            ozet = detay_service.get_nakit_kasa_sube_ozet(uid, mno, yil)
            def detay_fn(sube_adi):
                return detay_service.get_nakit_kasa_detay(uid, mno, yil, sube_adi=sube_adi)
            dlg = DetayDialog(
                baslik="Nakit Kasa — Şube Özeti",
                ozet_rows=ozet,
                detay_fn=detay_fn,
                tablo_tipi="genel_hesap",
                parent=self,
            )
        elif key == "gider_pusulasi":
            ozet = detay_service.get_gider_pusulasi_sube_ozet(uid, mno, yil)
            def detay_fn(sube_adi):
                return detay_service.get_gider_pusulasi_detay(uid, mno, yil, sube_adi=sube_adi)
            dlg = DetayDialog(
                baslik="Gider Pusulası — Şube Özeti",
                ozet_rows=ozet,
                detay_fn=detay_fn,
                tablo_tipi="genel_hesap",
                userid=uid,
                parent=self,
            )

        elif key in ("genel_hesap",):
            ozet = detay_service.get_genel_hesap_sube_ozet(uid, mno, yil)
            def detay_fn(sube_adi):
                return detay_service.get_genel_hesap_detay(uid, mno, yil, sube_adi=sube_adi)
            dlg = DetayDialog(
                baslik="Genel Hesap Tablosu — Şube Özeti",
                ozet_rows=ozet,
                detay_fn=detay_fn,
                tablo_tipi="genel_hesap",
                userid=uid,
                parent=self,
            )

        elif key == "kesilen_fatura":
            ozet = detay_service.get_fatura_sube_ozet(uid, mno, yil, "gelir")
            def _kes_detay(sube_adi, _u=uid, _m=mno, _y=yil):
                return detay_service.get_fatura_detay_by_sube(_u, _m, _y, "gelir", sube_adi)
            dlg = DetayDialog(
                baslik="Kesilen Faturalar — Şube Özeti",
                ozet_rows=ozet,
                detay_fn=_kes_detay,
                tablo_tipi="faturalar",
                userid=uid,
                parent=self,
            )

        elif key == "gelen_fatura":
            ozet = detay_service.get_fatura_sube_ozet(uid, mno, yil, "gider")
            def _gel_detay(sube_adi, _u=uid, _m=mno, _y=yil):
                return detay_service.get_fatura_detay_by_sube(_u, _m, _y, "gider", sube_adi)
            dlg = DetayDialog(
                baslik="Gelen Faturalar — Şube Özeti",
                ozet_rows=ozet,
                detay_fn=_gel_detay,
                tablo_tipi="faturalar",
                userid=uid,
                parent=self,
            )

        elif key == "kurum_odemeleri":
            dlg = KurumOdemeDialog(mno, yil, parent=self)
            dlg.exec()
            return

        elif key == "kredi_karti":
            dlg = KrediKartiDialog(uid, self._musterino, yil, parent=self)
            dlg.exec()
            return

        elif key == "maas_kira_smm":
            dlg = MaasKiraSmmDialog(self._userid, str(self._musterino), parent=self)
            dlg.exec()
            return

        elif key == "fiziksel_pos":
            dlg = FizikselPosDialog(self._userid, self._musterino, parent=self)
            dlg.exec()
            return

        elif key == "sanal_pos":
            dlg = SanalPosDialog(self._userid, str(self._musterino), self._yil, parent=self)
            dlg.exec()
            return

        elif key == "bankalar_bakiye":
            dlg = BankalaBakiyeDialog(self._userid, self._musterino, parent=self)
            dlg.exec()
            return

        else:
            # Henüz uygulanmamış kartlar için basit mesaj
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(
                self, key,
                f"'{key}' kartı için detay görünümü yakında eklenecek."
            )
            return

        dlg.exec()

    def _on_yil_changed(self):
        self._yil = self.yil_combo.currentData()
        self.banner.setText(f"FİNANS DURUM BİLGİSİ ( {self._yil} )")
        # Kartları yükleniyor moduna al
        for card in self._cards.values():
            card.set_value("Yükleniyor...")
        self._load_data()



    def _export_excel(self):
        """
        Dashboard'daki tüm kartların verilerini TEK Excel dosyasına,
        her kart için ayrı bir sheet olarak aktarır.
        Şube filtresi UYGULANMAZ — tüm veri çekilir.
        """
        from PyQt6.QtWidgets import QFileDialog, QMessageBox, QProgressDialog
        from PyQt6.QtCore import Qt as QtC
        import datetime
        import traceback
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(self, "Eksik Kütüphane",
                                 "openpyxl yüklenmemiş.\nTerminalde: pip install openpyxl")
            return
    
        yil        = self._yil
        uid        = self._userid
        mno        = self._musterino
        ilk        = self.ilk_tarih.date()
        son        = self.son_tarih.date()
        ilk_str    = ilk.toString("yyyy-MM-dd")
        son_str    = son.toString("yyyy-MM-dd")
        ilk_goster = ilk.toString("dd.MM.yyyy")
        son_goster = son.toString("dd.MM.yyyy")
        simdi      = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    
        dosya_adi = f"dashboard_rapor_{yil}_{ilk_str}_{son_str}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Excel Olarak Kaydet", dosya_adi, "Excel Files (*.xlsx)"
        )
        if not path:
            return
    
        # ── İlerleme dialogu ──
        prog = QProgressDialog("Excel hazırlanıyor...", None, 0, 13, self)
        prog.setWindowTitle("Excel Export")
        prog.setWindowModality(QtC.WindowModality.WindowModal)
        prog.setMinimumWidth(340)
        prog.setStyleSheet("""
            QProgressDialog {
                background-color: #0A0A0A;
                color: white;
                border: 1px solid #333;
                border-radius: 10px;
            }
            QLabel {
                color: white;
                font-size: 13px;
                font-weight: 600;
                background: transparent;
            }
            QProgressBar {
                background-color: #1F1F1F;
                border: 1px solid #444;
                border-radius: 6px;
                height: 14px;
                text-align: center;
                color: white;
                font-size: 11px;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2563EB, stop:1 #7C3AED);
                border-radius: 5px;
            }
        """)
        prog.show()
    
        try:
            from db.database import get_connection
            from db.db_compat import yr, numeric_cast
            from services import detay_service
    
            wb = openpyxl.Workbook()
            wb.remove(wb.active)   # default boş sheet'i sil
    
            def get_val(r, key):
                try:
                    if key in r:
                        return r[key]
                    kl = key.lower()
                    if kl in r:
                        return r[kl]
                    return None
                except (TypeError, AttributeError):
                    return None

            def _fmt_goster_yyyymmdd(t) -> str:
                s = str(t) if t else ""
                if len(s) == 8 and s.isdigit():
                    return f"{s[6:8]}.{s[4:6]}.{s[0:4]}"
                return s

            # ─────────────────────────────────────────────────────────────────
            # YARDIMCI FONKSİYONLAR
            # ─────────────────────────────────────────────────────────────────
    
            def _stiller(hdr_color: str):
                """Renk koduna göre stil seti döndürür."""
                return dict(
                    font_title    = Font(name="Segoe UI", size=13, bold=True,  color=hdr_color),
                    font_subtitle = Font(name="Segoe UI", size=9,  italic=True, color="FF4B5563"),
                    font_header   = Font(name="Segoe UI", size=10, bold=True,  color="FFFFFFFF"),
                    font_data     = Font(name="Segoe UI", size=10, color="FF1F2937"),
                    font_total    = Font(name="Segoe UI", size=10, bold=True,  color=hdr_color),
                    fill_header   = PatternFill(start_color=hdr_color, end_color=hdr_color, fill_type="solid"),
                    fill_gelir    = PatternFill(start_color="FFDCFCE7", end_color="FFDCFCE7", fill_type="solid"),
                    fill_gider    = PatternFill(start_color="FFFEE2E2", end_color="FFFEE2E2", fill_type="solid"),
                    fill_total    = PatternFill(start_color="FFF1F5F9", end_color="FFF1F5F9", fill_type="solid"),
                    border_thin   = Border(
                        left=Side(style="thin", color="FFE5E7EB"), right=Side(style="thin", color="FFE5E7EB"),
                        top=Side(style="thin",  color="FFE5E7EB"), bottom=Side(style="thin", color="FFE5E7EB"),
                    ),
                    border_total  = Border(
                        top=Side(style="thin", color="FF94A3B8"),
                        bottom=Side(style="double", color=hdr_color),
                    ),
                )
    
            def _yeni_sheet(title: str, baslik: str, alt_yazi: str,
                            headers: list[str], rows_data: list[dict],
                            para_sutunlar: set[str], hdr_color: str,
                            para_prefix: str = ""):
                """
                Yeni sheet oluşturur ve doldurur.
                para_sutunlar: sütun başlığındaki para birimi içeren isimler.
                """
                # ── 1. Toplam para değerlerini önceden hesapla ──
                _gelir_total = 0.0
                _gider_total = 0.0
                _has_para = False
                for _r in rows_data:
                    _rd = dict(_r) if not isinstance(_r, dict) else _r
                    for _hdr in para_sutunlar:
                        _val = float(get_val(_rd, _hdr) or 0)
                        # Tutar sütunlarında: pozitif = gelir, negatif = gider
                        if "tutar" in _hdr.lower() and "gider" not in _hdr.lower() and "kesinti" not in _hdr.lower() and "komisyon" not in _hdr.lower():
                            if _val >= 0:
                                _gelir_total += _val
                            else:
                                _gider_total += abs(_val)
                        elif "gider" in _hdr.lower() or "kesinti" in _hdr.lower() or "komisyon" in _hdr.lower():
                            _gider_total += abs(_val)
                        else:
                            _gelir_total += _val
                        _has_para = True

                ws = wb.create_sheet(title=title)
                s  = _stiller(hdr_color)
    
                # ── Başlık blokları ──
                ws.append([baslik])
                ws.cell(row=1, column=1).font = s["font_title"]

                # ── Gelir / Gider / Kalan özet satırı (İlk satırdan sonra açılan satır) ──
                if _has_para:
                    _kalan = _gelir_total - _gider_total
                    summary_text = f"Gelir Toplamı: {_gelir_total:,.2f} ₺   │   Gider Toplamı: {_gider_total:,.2f} ₺   │   Kalan: {_kalan:+,.2f} ₺"
                    ws.append([summary_text])
                    summary_cell = ws.cell(row=2, column=1)
                    summary_cell.font = Font(name="Segoe UI", size=10, bold=True, color="FF1E3A8A")
                else:
                    ws.append([]) # boş satır
                    
                ws.append([alt_yazi])
                ws.cell(row=3, column=1).font = s["font_subtitle"]
                ws.append([])   # boşluk satırı
    
                # ── Sütun başlıkları ──
                ws.append(headers)
                for ci, hdr in enumerate(headers, 1):
                    cell = ws.cell(row=5, column=ci)
                    cell.font      = s["font_header"]
                    cell.fill      = s["fill_header"]
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border    = s["border_thin"]
    
                # ── Veri satırları ──
                toplam: dict[str, float] = {h: 0.0 for h in para_sutunlar}
                cur = 6
                for row in rows_data:
                    r = dict(row) if not isinstance(row, dict) else row
                    row_data = []
                    for hdr in headers:
                        val = get_val(r, hdr)
                        if hdr in para_sutunlar:
                            fval = float(val or 0)
                            row_data.append(fval)
                            toplam[hdr] = toplam.get(hdr, 0.0) + fval
                        else:
                            row_data.append(str(val or ""))
                    ws.append(row_data)
    
                    for ci, hdr in enumerate(headers, 1):
                        cell = ws.cell(row=cur, column=ci)
                        cell.font   = s["font_data"]
                        cell.border = s["border_thin"]
                        if hdr in para_sutunlar:
                            fval = float(row_data[ci - 1] or 0)
                            cell.alignment     = Alignment(horizontal="right", vertical="center")
                            cell.number_format = "#,##0.00"
                            if "gelir" in hdr.lower() and fval > 0:
                                cell.fill = s["fill_gelir"]
                            elif "gider" in hdr.lower() and fval > 0:
                                cell.fill = s["fill_gider"]
                            elif "tutar" in hdr.lower() or "toplam" in hdr.lower():
                                cell.fill = s["fill_gelir"]
                        elif ci == 1:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                    cur += 1
    
                # ── GENEL TOPLAM satırı ──
                if para_sutunlar:
                    t_row = {h: "" for h in headers}
                    t_row[headers[0]] = "GENEL TOPLAM"
                    for h in para_sutunlar:
                        t_row[h] = toplam.get(h, 0.0)
                    ws.append([t_row[h] for h in headers])
                    for ci, hdr in enumerate(headers, 1):
                        cell = ws.cell(row=cur, column=ci)
                        cell.font   = s["font_total"]
                        cell.fill   = s["fill_total"]
                        cell.border = s["border_total"]
                        if hdr in para_sutunlar:
                            cell.alignment     = Alignment(horizontal="right", vertical="center")
                            cell.number_format = "#,##0.00"
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
    
                # ── Otomatik sütun genişliği ──
                for col in ws.columns:
                    max_len    = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.row <= 4:
                            continue
                        if cell.value is not None:
                            s_val = f"{cell.value:,.2f}" if isinstance(cell.value, float) else str(cell.value)
                            max_len = max(max_len, len(s_val))
                    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
    
                # ── AutoFilter (Filtre Özelliği) ──
                if rows_data:
                    last_col_letter = get_column_letter(len(headers))
                    ws.auto_filter.ref = f"A5:{last_col_letter}{cur - 1}"
    
                ws.freeze_panes = "A6"
                return (len(rows_data),
                        _gelir_total if _has_para else None,
                        _gider_total if _has_para else None)
    
            def _rows_to_dicts(rows):
                import json
                res = []
                for r in rows:
                    d = dict(r) if not isinstance(r, dict) else dict(r)
                    if "fatura" in d:
                        try:
                            meta = json.loads(d["fatura"] or "{}")
                            d["Açıklama"] = meta.get("aciklama", "")
                        except Exception:
                            d["Açıklama"] = ""
                    res.append(d)
                return res
    
            ozet = ExcelOzetBuilder()
            conn = get_connection()
    
            # ═══════════════════════════════════════════════════════════════════
            # 1. NAKİT KASA HAREKETLERİ (kasa kaynağı, tüm yıl)
            # ═══════════════════════════════════════════════════════════════════
            prog.setLabelText("1/12 Nakit Kasa Hareketleri...")
            prog.setValue(0)
            rows = conn.execute(f"""
                SELECT tarih_date AS Tarih, form_id AS "Form No",
                       COALESCE(sube,'(Şubesiz)') AS Şube,
                       COALESCE(aciklama,'') AS Açıklama,
                       COALESCE(kategori,'') AS Kategori,
                       COALESCE(teslim_sekli,'') AS "Teslim Şekli",
                       COALESCE(odeme_sekli,'') AS "Ödeme Şekli",
                       COALESCE({numeric_cast('gelir')},0) AS "Gelir (₺)",
                       COALESCE({numeric_cast('gider')},0) AS "Gider (₺)",
                       COALESCE(nerden_geliyor,'') AS Kaynak
                FROM genel_hesap_hareketleri
                WHERE userid=? AND musteri_no=? AND nerden_geliyor='kasa'
                  AND tarih_date >= ? AND tarih_date <= ?
                ORDER BY tarih_date ASC, id ASC
            """, (uid, mno, ilk_str, son_str)).fetchall()
            _n, _g, _gd = _yeni_sheet(
                title="Nakit Kasa",
                baslik=f"Nakit Kasa Hareketleri — {ilk_goster} / {son_goster}",
                alt_yazi=f"Kayıt: {len(rows):,}  •  {simdi}",
                headers=["Tarih","Form No","Şube","Açıklama","Kategori",
                         "Teslim Şekli","Ödeme Şekli","Gelir (₺)","Gider (₺)","Kaynak"],
                rows_data=_rows_to_dicts(rows),
                para_sutunlar={"Gelir (₺)", "Gider (₺)"},
                hdr_color="FF059669",
            )
            ozet.kaydet("Nakit Kasa", kayit=_n, gelir=_g, gider=_gd)
    
            # ═══════════════════════════════════════════════════════════════════
            # 2. KESİLEN FATURALAR
            # ═══════════════════════════════════════════════════════════════════
            prog.setLabelText("2/12 Kesilen Faturalar...")
            prog.setValue(1)
            _mod_col  = "gelirGiderMod" if not hasattr(conn, "server_version") else "gelirgidermod"
            _fmod_col = "faturaMod"     if not hasattr(conn, "server_version") else "faturamod"
            _fno_col  = "formNo"        if not hasattr(conn, "server_version") else "formno"
            _ykl_col  = "yuklenmeTarihi" if not hasattr(conn, "server_version") else "yuklenmetarihi"
            from db.db_compat import left4
            rows = conn.execute(f"""
                SELECT f.tarih AS Tarih, f.unvan AS Ünvan,
                       f.{_fno_col} AS "Form No",
                       f.vergino AS "Vergi No",
                       f.faturano AS "Fatura No",
                       CAST(f.toplam AS REAL) AS "Toplam (₺)",
                       COALESCE(
                           (SELECT MIN(sube) FROM genel_hesap_hareketleri 
                            WHERE form_id = f.{_fno_col} AND userid = f.userid AND musteri_no = ?),
                           '(Şubesiz)'
                       ) AS Şube,
                       f.kaynak AS Kaynak,
                       f.fatura AS fatura
                FROM faturalar f
                WHERE f.userid=? AND {left4('f.tarih')}=?
                  AND f.{_mod_col}='gelir'
                ORDER BY f.tarih DESC, f.id DESC
            """, (mno, uid, str(yil))).fetchall()
            _n, _g, _gd = _yeni_sheet(
                title="Kesilen Faturalar",
                baslik=f"Kesilen Faturalar — {yil}",
                alt_yazi=f"Kayıt: {len(rows):,}  •  {simdi}",
                headers=["Tarih","Ünvan","Form No","Açıklama","Vergi No","Fatura No","Toplam (₺)","Şube","Kaynak"],
                rows_data=_rows_to_dicts(rows),
                para_sutunlar={"Toplam (₺)"},
                hdr_color="FF6D28D9",
            )
            ozet.kaydet("Kesilen Faturalar", kayit=_n, gelir=_g, gider=_gd)
    
            # ═══════════════════════════════════════════════════════════════════
            # 3. GELEN FATURALAR
            # ═══════════════════════════════════════════════════════════════════
            prog.setLabelText("3/12 Gelen Faturalar...")
            prog.setValue(2)
            rows = conn.execute(f"""
                SELECT f.tarih AS Tarih, f.unvan AS Ünvan, f.vergino AS "Vergi No",
                       f.faturano AS "Fatura No",
                       CAST(f.toplam AS REAL) AS "Toplam (₺)",
                       COALESCE(
                           (SELECT MIN(sube) FROM genel_hesap_hareketleri 
                            WHERE form_id = f.{_fno_col} AND userid = f.userid AND musteri_no = ?),
                           '(Şubesiz)'
                       ) AS Şube,
                       f.kaynak AS Kaynak,
                       f.fatura AS fatura
                FROM faturalar f
                WHERE f.userid=? AND {left4('f.tarih')}=?
                  AND f.{_mod_col}='gider'
                ORDER BY f.tarih DESC, f.id DESC
            """, (mno, uid, str(yil))).fetchall()
            _n, _g, _gd = _yeni_sheet(
                title="Gelen Faturalar",
                baslik=f"Gelen Faturalar — {yil}",
                alt_yazi=f"Kayıt: {len(rows):,}  •  {simdi}",
                headers=["Tarih","Ünvan","Açıklama","Vergi No","Fatura No","Toplam (₺)","Şube","Kaynak"],
                rows_data=_rows_to_dicts(rows),
                para_sutunlar={"Toplam (₺)"},
                hdr_color="FFDB2777",
            )
            ozet.kaydet("Gelen Faturalar", kayit=_n, gelir=_g, gider=_gd)
    
            # ═══════════════════════════════════════════════════════════════════
            # 4. GİDER PUSULASI
            # ═══════════════════════════════════════════════════════════════════
            prog.setLabelText("4/12 Gider Pusulası...")
            prog.setValue(3)
            rows = conn.execute(f"""
                SELECT tarih_date AS Tarih, form_id AS "Form No",
                       COALESCE(sube,'(Şubesiz)') AS Şube,
                       COALESCE(aciklama,'') AS Açıklama,
                       COALESCE(teslim_sekli,'') AS "Teslim Şekli",
                       COALESCE(odeme_sekli,'') AS "Ödeme Şekli",
                       COALESCE({numeric_cast('gelir')},0) AS "Gelir (₺)",
                       COALESCE({numeric_cast('gider')},0) AS "Gider (₺)",
                       COALESCE(kategori,'') AS Kategori,
                       COALESCE(nerden_geliyor,'') AS Kaynak
                FROM genel_hesap_hareketleri
                WHERE userid=? AND musteri_no=?
                  AND tarih_date >= ? AND tarih_date <= ?
                  AND teslim_sekli LIKE '%Parça Alımı (Cihaz)%'
                ORDER BY tarih_date ASC, id ASC
            """, (uid, mno, ilk_str, son_str)).fetchall()
            _n, _g, _gd = _yeni_sheet(
                title="Gider Pusulası",
                baslik=f"Gider Pusulası — {ilk_goster} / {son_goster}",
                alt_yazi=f"Teslim: Parça Alımı (Cihaz)  •  Kayıt: {len(rows):,}  •  {simdi}",
                headers=["Tarih","Form No","Şube","Açıklama","Teslim Şekli",
                         "Ödeme Şekli","Gelir (₺)","Gider (₺)","Kategori","Kaynak"],
                rows_data=_rows_to_dicts(rows),
                para_sutunlar={"Gelir (₺)", "Gider (₺)"},
                hdr_color="FF16A34A",
            )
            ozet.kaydet("Gider Pusulası", kayit=_n, gelir=_g, gider=_gd)
    
            # ═══════════════════════════════════════════════════════════════════
            # 5. KURUM ÖDEMELERİ
            # ═══════════════════════════════════════════════════════════════════
            prog.setLabelText("5/12 Kurum Ödemeleri...")
            prog.setValue(4)
            from services.detay_service import get_kurum_odemeleri_detay_tarih
            ilk_yyyymmdd = ilk.toString("yyyyMMdd")
            son_yyyymmdd = son.toString("yyyyMMdd")
            
            BEYANNAME_TUR = {
                "770.01": "KDV / SGK Beyannamesi",
                "730.08": "Muhtasar ve Prim Hizmet Beyannamesi",
            }
            HESAP_ACIKLAMA = {
                "770.01": "770.01 — Vergi Giderleri (SGK / KDV / Gelir Vergisi)",
                "730.08": "730.08 — İşçilik / Müşavirlik Giderleri",
            }
            
            kurum_rows_dialog, _toplam = get_kurum_odemeleri_detay_tarih(mno, ilk_yyyymmdd, son_yyyymmdd)
            
            # Yerel önbellekten beyanname bilgilerini çek
            from services.moy_service import get_local_beyannameler
            BELGE_TUR_ADI = {
                "KDV1":     "KDV Beyannamesi (1.Tür)",
                "KDV2":     "KDV Beyannamesi (2.Tür)",
                "MUHSGK":   "SGK Tahakkuk Fişi (5510)",
                "KGECICI":  "Kurumlar Vg. Geçici",
                "KURUMLAR": "Kurumlar Vergisi",
                "LEVHA":    "Levha Beyannamesi",
                "MUHTAR":   "Muhtasar Beyanname",
            }
            
            kurum_export_rows = []
            for r in kurum_rows_dialog:
                kod     = get_val(r, "hesapKodu")
                ilk_t   = get_val(r, "ilkTarih") or ""
                son_t   = get_val(r, "sonTarih") or ""
                soz_no  = get_val(r, "sozlesmeNo") or ""
                soz_tar = get_val(r, "sozlesmeTarih") or ""
                
                # Beyanname verisini yerel tablodan eşleştir
                beyanlar = get_local_beyannameler(mno, ilk_t, kod) if ilk_t else []
                byn = beyanlar[0] if beyanlar else None
                
                beyan_t  = BEYANNAME_TUR.get(kod, HESAP_ACIKLAMA.get(kod, kod))
                
                # Boş alanları beyanname verisinden doldur
                if byn:
                    if not son_t:
                        son_t = byn.get("beyan_tarih_2", "") or ""
                    if not soz_tar:
                        soz_tar = byn.get("onay_tarihi", "") or ""
                    _raw_bt = byn.get("belge_turu", "")
                    if _raw_bt == "MUHSGK" and kod == "730.08":
                        byn_belge = "Muhtasar ve Prim Hizmet Beyannamesi"
                    else:
                        byn_belge = BELGE_TUR_ADI.get(_raw_bt, _raw_bt or "")
                    byn_donem  = byn.get("donem_adi", "") or ""
                    byn_durum  = byn.get("belge_durumu", "") or ""
                else:
                    byn_belge = ""
                    byn_donem = ""
                    byn_durum = ""
                
                kurum_export_rows.append({
                    "Beyanname Türü":  beyan_t,
                    "Ünvan":           get_val(r, "unvan") or "-",
                    "Vergi No":        get_val(r, "vergiNo") or "",
                    "İlk Tarih":       _fmt_goster_yyyymmdd(ilk_t),
                    "Son Tarih":       _fmt_goster_yyyymmdd(son_t),
                    "Sözleşme No":     soz_no,
                    "Sözl. Tarih":     _fmt_goster_yyyymmdd(soz_tar),
                    "Tutar (₺)":       float(get_val(r, "tutar") or 0),
                    "Belge Türü":      byn_belge,
                    "Dönem":           byn_donem,
                    "Belge Durum":     byn_durum,
                })
                
            _n, _g, _gd = _yeni_sheet(
                title="Kurum Ödemeleri",
                baslik=f"Kurum Ödemeleri — {ilk_goster} / {son_goster}",
                alt_yazi=f"Kayıt: {len(kurum_export_rows):,}  •  {simdi}",
                headers=["Beyanname Türü", "Ünvan", "Vergi No", "İlk Tarih", "Son Tarih", "Sözleşme No", "Sözl. Tarih", "Tutar (₺)", "Belge Türü", "Dönem", "Belge Durum"],
                rows_data=kurum_export_rows,
                para_sutunlar={"Tutar (₺)"},
                hdr_color="FF1E3A8A",
            )
            ozet.kaydet("Kurum Ödemeleri", kayit=_n, gelir=_g, gider=_gd)
    
            # ═══════════════════════════════════════════════════════════════════
            # 6. MAAŞ KİRA SMM
            # ═══════════════════════════════════════════════════════════════════
            prog.setLabelText("6/12 Maaş Kira SMM...")
            prog.setValue(5)
            try:
                from services.vergi_muhtasar_service import get_vergi_muhtasar
                muhtasar_res = get_vergi_muhtasar(uid, musterino=str(mno), yil=yil)
                muhtasar_rows = muhtasar_res.get("data", [])
            except Exception:
                muhtasar_rows = []
                
            muhtasar_export_rows = []
            for r in muhtasar_rows:
                gay = float(get_val(r, "gaytutar") or 0)
                verg = float(get_val(r, "vergkestutar") or 0)
                fark = gay - verg
                muhtasar_export_rows.append({
                    "Dönem": get_val(r, "donem") or "",
                    "Açıklama": get_val(r, "ack") or "",
                    "Gayri Resmi Tutar": gay,
                    "Vergi Kesinti Tutarı": verg,
                    "Fark": fark
                })
                
            _n, _g, _gd = _yeni_sheet(
                title="Maas Kira SMM",
                baslik=f"Maaş / Kira / SMM — {yil}",
                alt_yazi=f"Kayıt: {len(muhtasar_export_rows):,}  •  {simdi}",
                headers=["Dönem", "Açıklama", "Gayri Resmi Tutar", "Vergi Kesinti Tutarı", "Fark"],
                rows_data=muhtasar_export_rows,
                para_sutunlar={"Gayri Resmi Tutar", "Vergi Kesinti Tutarı", "Fark"},
                hdr_color="FF1a3a5c",
            )
            ozet.kaydet("Maaş Kira SMM", kayit=_n, gelir=_g, gider=_gd)
    
            # ═══════════════════════════════════════════════════════════════════
            # 7. BANKALAR BAKİYE
            # ═══════════════════════════════════════════════════════════════════
            prog.setLabelText("7/12 Bankalar Bakiye...")
            prog.setValue(6)
            rows = conn.execute("""
                SELECT
                    tarih,
                    COALESCE(sube, '(Şubesiz)') AS sube,
                    COALESCE(aciklama, '')       AS aciklama,
                    COALESCE(gelirgider, '')     AS gelirgider,
                    CAST(tutar AS REAL)          AS tutar,
                    COALESCE(faturaunvan, '')    AS faturaunvan,
                    COALESCE(kaynak, '')         AS kaynak,
                    COALESCE(iban, '')           AS iban,
                    COALESCE(hesap_turu, '')     AS hesap_turu,
                    COALESCE(dekont_no, '')      AS dekont_no,
                    CAST(bakiye AS REAL)         AS bakiye
                FROM womsis_banka
                WHERE musterino = ?
                ORDER BY tarih DESC, id DESC
            """, (self._musterino,)).fetchall()
            
            banka_export_rows = []
            for r in rows:
                raw_tutar = float(get_val(r, "tutar") or 0)
                gelirgider = get_val(r, "gelirgider") or ""
                # Gider ise tutarı negatif yap
                signed_tutar = -abs(raw_tutar) if gelirgider == "gider" else abs(raw_tutar)
                raw_bakiye = get_val(r, "bakiye")
                banka_export_rows.append({
                    "Tarih": get_val(r, "tarih") or "",
                    "Banka / Şube": get_val(r, "sube") or "(Şubesiz)",
                    "Açıklama": get_val(r, "aciklama") or "",
                    "Tür": gelirgider,
                    "Tutar (₺)": signed_tutar,
                    "Bakiye (₺)": float(raw_bakiye) if raw_bakiye is not None else "",
                    "Karşı Taraf": get_val(r, "faturaunvan") or "",
                    "Kaynak": get_val(r, "kaynak") or "",
                    "IBAN": get_val(r, "iban") or "",
                    "Hesap Türü": get_val(r, "hesap_turu") or "",
                    "Dekont No": get_val(r, "dekont_no") or "",
                })
                
            _n, _g, _gd = _yeni_sheet(
                title="Bankalar",
                baslik=f"Banka Hareketleri — Tüm Dönem",
                alt_yazi=f"Kayıt: {len(banka_export_rows):,}  •  {simdi}",
                headers=["Tarih", "Banka / Şube", "Açıklama", "Tür", "Tutar (₺)", "Bakiye (₺)", "Karşı Taraf", "Kaynak", "IBAN", "Hesap Türü", "Dekont No"],
                rows_data=banka_export_rows,
                para_sutunlar={"Tutar (₺)"},
                hdr_color="FF374151",
            )
            ozet.kaydet("Bankalar", kayit=_n, gelir=_g, gider=_gd)
    
            # ═══════════════════════════════════════════════════════════════════
            # 8. SANAL POS
            # ═══════════════════════════════════════════════════════════════════
            prog.setLabelText("8/12 Sanal Pos...")
            prog.setValue(7)
            try:
                from services.paytr_service import get_sanal_pos_hareketleri_db
                sp_res = get_sanal_pos_hareketleri_db(uid, str(mno), ilk_str, son_str)
                sp_raw = sp_res.get("data", [])
            except Exception:
                sp_raw = []
                
            sp_export_rows = []
            for r in sp_raw:
                sp_export_rows.append({
                    "İşlem Tarihi": get_val(r, "islemtarihi") or "",
                    "Sipariş No": get_val(r, "siparisno") or "",
                    "İşlem Tutarı": float(get_val(r, "islemtutari") or 0),
                    "Ödeme Tutarı": float(get_val(r, "odemetutari") or 0),
                    "Kur": get_val(r, "kur") or "",
                    "Mağaza No": get_val(r, "magazano") or "",
                    "Net Tutar": float(get_val(r, "nettutar") or 0),
                    "Kesinti Tutarı": float(get_val(r, "kesintitutari") or 0),
                    "Kesinti Oranı": get_val(r, "kesintiorani") or "",
                    "Kart Markası": get_val(r, "kartmarkasi") or "",
                    "Kart No": get_val(r, "kartno") or "",
                    "Ödeme Tipi": get_val(r, "odemetipi") or "",
                    "Kart Tipi": get_val(r, "karttipi") or "",
                    "Taksit Sayısı": get_val(r, "taksitsayisi") or ""
                })
                
            _n, _g, _gd = _yeni_sheet(
                title="Sanal POS",
                baslik=f"Sanal POS Hareketleri — {ilk_goster} / {son_goster}",
                alt_yazi=f"Kayıt: {len(sp_export_rows):,}  •  {simdi}",
                headers=["İşlem Tarihi", "Sipariş No", "İşlem Tutarı", "Ödeme Tutarı", "Kur", 
                         "Mağaza No", "Net Tutar", "Kesinti Tutarı", "Kesinti Oranı", 
                         "Kart Markası", "Kart No", "Ödeme Tipi", "Kart Tipi", "Taksit Sayısı"],
                rows_data=sp_export_rows,
                para_sutunlar={"İşlem Tutarı", "Ödeme Tutarı", "Net Tutar", "Kesinti Tutarı"},
                hdr_color="FF111827",
            )
            ozet.kaydet("Sanal POS", kayit=_n, gelir=_g, gider=_gd)
    
            # ═══════════════════════════════════════════════════════════════════
            # 9. FİZİKSEL POS
            # ═══════════════════════════════════════════════════════════════════
            prog.setLabelText("9/12 Fiziksel Pos...")
            prog.setValue(8)
            rows = conn.execute("""
                SELECT islemTarihi                      AS "Tarih",
                       cariHesap                       AS "Banka",
                       posNo                           AS "Terminal No",
                       CAST(netTutar AS REAL)          AS "Net Tutar (₺)",
                       CAST(isyeriUcretiTutar AS REAL) AS "Komisyon (₺)",
                       CAST(islemTutari AS REAL)       AS "Brüt Tutar (₺)",
                       islemTipi                       AS "Ödeme Türü",
                       brand                           AS "Kart Markası"
                FROM womsi_pos
                WHERE userid=?
                ORDER BY islemTarihi DESC, id DESC
            """, (uid,)).fetchall()
            _n, _g, _gd = _yeni_sheet(
                title="Fiziksel POS",
                baslik=f"Fiziksel POS Hareketleri — Tüm Dönem",
                alt_yazi=f"Kayıt: {len(rows):,}  •  {simdi}",
                headers=["Tarih","Banka","Terminal No","Net Tutar (₺)",
                         "Komisyon (₺)","Brüt Tutar (₺)","Ödeme Türü","Kart Markası"],
                rows_data=_rows_to_dicts(rows),
                para_sutunlar={"Net Tutar (₺)","Komisyon (₺)","Brüt Tutar (₺)"},
                hdr_color="FF1F2937",
            )
            ozet.kaydet("Fiziksel POS", kayit=_n, gelir=_g, gider=_gd)
    
            # ═══════════════════════════════════════════════════════════════════
            # 10. KREDİ KARTLARI
            # ═══════════════════════════════════════════════════════════════════
            prog.setLabelText("10/11 Kredi Kartları...")
            prog.setValue(9)
            try:
                from db.db_compat import right4 as _r4
                kk_raw = conn.execute(f"""
                    SELECT
                        tarih                          AS "Tarih",
                        aciklama                       AS "Açıklama",
                        Banka                          AS "Kart / Banka",
                        hesapKodu                      AS "Hesap Kodu",
                        CAST(alinan_tutar1 AS REAL)    AS "Tutar (₺)",
                        CASE WHEN alinan_tutar1 < 0
                             THEN 'Ödeme' ELSE 'Borç' END AS "Tür"
                    FROM kredikartidata
                    WHERE userid = ?
                      AND {_r4('tarih')} = ?
                    ORDER BY tarih ASC, id ASC
                """, (str(uid), str(yil))).fetchall()
                kk_rows = _rows_to_dicts(kk_raw)
            except Exception as _kk_e:
                print(f"Kredi kartı export hata: {_kk_e}")
                kk_rows = []
            _n, _g, _gd = _yeni_sheet(
                title="Kredi Kartları",
                baslik=f"Kredi Kartı Hareketleri — {yil}",
                alt_yazi=f"Kayıt: {len(kk_rows):,}  •  {simdi}",
                headers=["Tarih", "Açıklama", "Kart / Banka",
                         "Hesap Kodu", "Tutar (₺)", "Tür"],
                rows_data=kk_rows,
                para_sutunlar={"Tutar (₺)"},
                hdr_color="FFD97706",
            )
            ozet.kaydet("Kredi Kartları", kayit=_n, gelir=_g, gider=_gd)
    
            # ═══════════════════════════════════════════════════════════════════
            # 11. GENEL HESAP TABLOSU
            # ═══════════════════════════════════════════════════════════════════
            prog.setLabelText("11/12 Genel Hesap Tablosu...")
            prog.setValue(10)
            rows = conn.execute(f"""
                SELECT tarih_date AS Tarih, form_id AS "Form No",
                       COALESCE(sube,'(Şubesiz)') AS Şube,
                       COALESCE(aciklama,'') AS Açıklama,
                       COALESCE(kategori,'') AS Kategori,
                       COALESCE(teslim_sekli,'') AS "Teslim Şekli",
                       COALESCE(odeme_sekli,'') AS "Ödeme Şekli",
                       COALESCE({numeric_cast('gelir')},0) AS "Gelir (₺)",
                       COALESCE({numeric_cast('gider')},0) AS "Gider (₺)",
                       COALESCE(nerden_geliyor,'') AS Kaynak
                FROM genel_hesap_hareketleri
                WHERE userid=? AND musteri_no=? AND nerden_geliyor='genelHesap'
                  AND tarih_date >= ? AND tarih_date <= ?
                ORDER BY tarih_date ASC, id ASC
            """, (uid, mno, ilk_str, son_str)).fetchall()
            _n, _g, _gd = _yeni_sheet(
                title="Genel Hesap",
                baslik=f"Genel Hesap Tablosu — {ilk_goster} / {son_goster}",
                alt_yazi=f"Kayıt: {len(rows):,}  •  {simdi}",
                headers=["Tarih","Form No","Şube","Açıklama","Kategori",
                         "Teslim Şekli","Ödeme Şekli","Gelir (₺)","Gider (₺)","Kaynak"],
                rows_data=_rows_to_dicts(rows),
                para_sutunlar={"Gelir (₺)", "Gider (₺)"},
                hdr_color="FFEA580C",
            )
            ozet.kaydet("Genel Hesap", kayit=_n, gelir=_g, gider=_gd)

            # ═══════════════════════════════════════════════════════════════════
            # 12. GENELDE VAR KESİLENDE YOK
            # ═══════════════════════════════════════════════════════════════════
            prog.setLabelText("12/13 Genelde Var Kesilende Yok...")
            prog.setValue(11)

            # Form numarası olan ama kesilen faturası bulunmayan kayıtları doğrudan SQL ile çek
            _gvky_filtered = conn.execute(f"""
                SELECT tarih_date AS Tarih, form_id AS "Form No",
                       COALESCE(sube,'(Şubesiz)') AS Şube,
                       COALESCE(aciklama,'') AS Açıklama,
                       COALESCE(kategori,'') AS Kategori,
                       COALESCE(teslim_sekli,'') AS "Teslim Şekli",
                       COALESCE(odeme_sekli,'') AS "Ödeme Şekli",
                       COALESCE({numeric_cast('gelir')},0) AS "Gelir (₺)",
                       COALESCE({numeric_cast('gider')},0) AS "Gider (₺)",
                       COALESCE(nerden_geliyor,'') AS Kaynak
                FROM genel_hesap_hareketleri
                WHERE userid=? AND musteri_no=? AND nerden_geliyor='genelHesap'
                  AND tarih_date >= ? AND tarih_date <= ?
                  AND form_id IS NOT NULL AND TRIM(form_id) != ''
                  AND form_id NOT IN (
                      SELECT {_fno_col} FROM faturalar
                      WHERE userid=? AND {_mod_col}='gelir'
                        AND {_fno_col} IS NOT NULL AND TRIM({_fno_col}) != ''
                  )
                ORDER BY tarih_date ASC, id ASC
            """, (uid, mno, ilk_str, son_str, uid)).fetchall()
            _gvky_filtered = [dict(r) if not isinstance(r, dict) else r for r in _gvky_filtered]

            # ── Sheet oluştur ──
            _gvky_hdrs = ["Tarih","Form No","Şube","Açıklama","Kategori",
                          "Teslim Şekli","Ödeme Şekli","Gelir (₺)","Gider (₺)","Kaynak"]
            _gvky_para = {"Gelir (₺)", "Gider (₺)"}
            _hdr_clr   = "FF7C3AED"  # mor
            _s         = _stiller(_hdr_clr)

            ws_gvky = wb.create_sheet(title="Genelde Var Kesilende Yok")

            # Başlık
            ws_gvky.append([f"Genelde Var Kesilende Yok — {ilk_goster} / {son_goster}"])
            ws_gvky.cell(row=1, column=1).font = _s["font_title"]

            # Gelir/Gider özet satırı (önceden hesapla)
            _gvky_gelir_total = sum(float(_r.get("Gelir (₺)") or 0) for _r in _gvky_filtered)
            _gvky_gider_total = sum(float(_r.get("Gider (₺)") or 0) for _r in _gvky_filtered)
            _gvky_kalan = _gvky_gelir_total - _gvky_gider_total
            _summary_txt = (f"Gelir Toplamı: {_gvky_gelir_total:,.2f} ₺   │   "
                            f"Gider Toplamı: {_gvky_gider_total:,.2f} ₺   │   "
                            f"Kalan: {_gvky_kalan:+,.2f} ₺")
            ws_gvky.append([_summary_txt])
            ws_gvky.cell(row=2, column=1).font = Font(name="Segoe UI", size=10, bold=True, color="FF1E3A8A")

            ws_gvky.append([f"Fatura Kesilmemiş Kayıt: {len(_gvky_filtered):,}  •  {simdi}"])
            ws_gvky.cell(row=3, column=1).font = _s["font_subtitle"]
            ws_gvky.append([])  # boşluk

            # Sütun başlıkları (5. satır)
            ws_gvky.append(_gvky_hdrs)
            for _ci, _hdr in enumerate(_gvky_hdrs, 1):
                _cell = ws_gvky.cell(row=5, column=_ci)
                _cell.font      = _s["font_header"]
                _cell.fill      = _s["fill_header"]
                _cell.alignment = Alignment(horizontal="center", vertical="center")
                _cell.border    = _s["border_thin"]

            # ── Veri satırları — ay bazlı gruplama ──
            _cur_row   = 6
            _cur_ay    = None  # "YYYY-MM"
            _ay_gelir  = 0.0
            _ay_gider  = 0.0

            _AY_ADLARI = {1:"Ocak",2:"Şubat",3:"Mart",4:"Nisan",5:"Mayıs",6:"Haziran",
                          7:"Temmuz",8:"Ağustos",9:"Eylül",10:"Ekim",11:"Kasım",12:"Aralık"}

            def _yaz_ay_toplam(ay_str, ay_gelir, ay_gider):
                nonlocal _cur_row
                # ay_str örn: "2026-03"
                try:
                    _ay_no = int(ay_str.split("-")[1])
                    _ay_yil = ay_str.split("-")[0]
                    _ay_adi = f"── {_AY_ADLARI.get(_ay_no, ay_str)} {_ay_yil} Toplamı ──"
                except Exception:
                    _ay_adi = f"── {ay_str} Toplamı ──"
                _ay_kalan = ay_gelir - ay_gider
                _row_vals = []
                for _h in _gvky_hdrs:
                    if _h == "Tarih":
                        _row_vals.append(_ay_adi)
                    elif _h == "Gelir (₺)":
                        _row_vals.append(ay_gelir)
                    elif _h == "Gider (₺)":
                        _row_vals.append(ay_gider)
                    elif _h == "Açıklama":
                        _row_vals.append(f"Kalan: {_ay_kalan:+,.2f} ₺")
                    else:
                        _row_vals.append("")
                ws_gvky.append(_row_vals)
                for _ci2, _h2 in enumerate(_gvky_hdrs, 1):
                    _c2 = ws_gvky.cell(row=_cur_row, column=_ci2)
                    _c2.font   = Font(name="Segoe UI", size=10, bold=True, color="FFFFFFFF")
                    _c2.fill   = PatternFill(start_color=_hdr_clr, end_color=_hdr_clr, fill_type="solid")
                    _c2.border = _s["border_thin"]
                    if _h2 in _gvky_para:
                        _c2.alignment     = Alignment(horizontal="right", vertical="center")
                        _c2.number_format = "#,##0.00"
                    else:
                        _c2.alignment = Alignment(horizontal="left", vertical="center")
                _cur_row += 1

            for _rd in _gvky_filtered:
                _tarih_str = str(_rd.get("Tarih") or "")
                # "YYYY-MM" kısmını al
                _this_ay = _tarih_str[:7] if len(_tarih_str) >= 7 else ""

                # Ay değişti → önceki ayın toplamını yaz
                if _cur_ay is not None and _this_ay != _cur_ay:
                    _yaz_ay_toplam(_cur_ay, _ay_gelir, _ay_gider)
                    _ay_gelir = 0.0
                    _ay_gider = 0.0

                _cur_ay = _this_ay

                # Satır verisi
                _row_vals = []
                for _h in _gvky_hdrs:
                    _v = _rd.get(_h)
                    if _h in _gvky_para:
                        _fv = float(_v or 0)
                        _row_vals.append(_fv)
                        if _h == "Gelir (₺)":
                            _ay_gelir += _fv
                        else:
                            _ay_gider += _fv
                    else:
                        _row_vals.append(str(_v or ""))
                ws_gvky.append(_row_vals)

                for _ci3, _h3 in enumerate(_gvky_hdrs, 1):
                    _c3 = ws_gvky.cell(row=_cur_row, column=_ci3)
                    _c3.font   = _s["font_data"]
                    _c3.border = _s["border_thin"]
                    if _h3 in _gvky_para:
                        _fv3 = float(_row_vals[_ci3 - 1] or 0)
                        _c3.alignment     = Alignment(horizontal="right", vertical="center")
                        _c3.number_format = "#,##0.00"
                        if "gelir" in _h3.lower() and _fv3 > 0:
                            _c3.fill = _s["fill_gelir"]
                        elif "gider" in _h3.lower() and _fv3 > 0:
                            _c3.fill = _s["fill_gider"]
                    elif _ci3 == 1:
                        _c3.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        _c3.alignment = Alignment(horizontal="left", vertical="center")
                _cur_row += 1

            # Son ayın toplamı
            if _cur_ay is not None:
                _yaz_ay_toplam(_cur_ay, _ay_gelir, _ay_gider)

            # Genel toplam satırı
            _gt_vals = []
            for _h in _gvky_hdrs:
                if _h == "Tarih":
                    _gt_vals.append("GENEL TOPLAM")
                elif _h == "Gelir (₺)":
                    _gt_vals.append(_gvky_gelir_total)
                elif _h == "Gider (₺)":
                    _gt_vals.append(_gvky_gider_total)
                elif _h == "Açıklama":
                    _gt_vals.append(f"Kalan: {_gvky_kalan:+,.2f} ₺")
                else:
                    _gt_vals.append("")
            ws_gvky.append(_gt_vals)
            for _ci4, _h4 in enumerate(_gvky_hdrs, 1):
                _c4 = ws_gvky.cell(row=_cur_row, column=_ci4)
                _c4.font   = _s["font_total"]
                _c4.fill   = _s["fill_total"]
                _c4.border = _s["border_total"]
                if _h4 in _gvky_para:
                    _c4.alignment     = Alignment(horizontal="right", vertical="center")
                    _c4.number_format = "#,##0.00"
                else:
                    _c4.alignment = Alignment(horizontal="left", vertical="center")

            # Sütun genişlikleri
            for _col in ws_gvky.columns:
                _max_len   = 0
                _col_ltr   = get_column_letter(_col[0].column)
                for _cell in _col:
                    if _cell.row <= 4:
                        continue
                    if _cell.value is not None:
                        _sv = f"{_cell.value:,.2f}" if isinstance(_cell.value, float) else str(_cell.value)
                        _max_len = max(_max_len, len(_sv))
                ws_gvky.column_dimensions[_col_ltr].width = max(_max_len + 4, 12)

            # AutoFilter
            if _gvky_filtered:
                ws_gvky.auto_filter.ref = f"A5:{get_column_letter(len(_gvky_hdrs))}{_cur_row - 1}"
            ws_gvky.freeze_panes = "A6"

            ozet.kaydet("Genelde Var Kesilende Yok",
                        kayit=len(_gvky_filtered),
                        gelir=_gvky_gelir_total,
                        gider=_gvky_gider_total)

            # ═══════════════════════════════════════════════════════════════════
            # 13. ÖZET (ExcelOzetBuilder ile oluşturulur)
            # ═══════════════════════════════════════════════════════════════════
            prog.setLabelText("13/13 Özet sayfası...")
            prog.setValue(12)

            # Aylık gelir-gider verisi (grafik için)
            from db.db_compat import yr, mo
            try:
                monthly_rows = conn.execute(f"""
                    SELECT
                        {mo('tarih_date')} AS ay,
                        SUM(gelir) AS toplam_gelir,
                        SUM(gider) AS toplam_gider
                    FROM genel_hesap_hareketleri
                    WHERE userid=%s AND musteri_no=%s AND {yr('tarih_date')}=%s
                    GROUP BY {mo('tarih_date')}
                    ORDER BY {mo('tarih_date')}
                """, (uid, mno, str(yil))).fetchall()
                monthly_data = [{"ay": int(r["ay"]), "toplam_gelir": float(r["toplam_gelir"] or 0),
                                 "toplam_gider": float(r["toplam_gider"] or 0)} for r in monthly_rows]
            except Exception:
                monthly_data = []

            ozet.yaz(
                wb,
                baslik=f"Dashboard Raporu — {yil}  ({ilk_goster} / {son_goster})",
                simdi=simdi,
                hdr_color="FF1E3A8A",
                monthly_data=monthly_data,
            )
    
            conn.close()
    
            # ── Kaydet ──
            prog.setValue(12)
            wb.save(path)
            prog.close()
    
            msg = QMessageBox(self)
            msg.setWindowTitle("✅  Excel Hazır")
            msg.setText(
                f"Tüm kartlar başarıyla dışa aktarıldı!\n\n"
                f"📄  {path}\n"
                f"📋  Toplam {ozet.toplam_kayit():,} kayıt  •  {len(ozet._data)} sheet\n"
                f"📅  {ilk_goster} — {son_goster}\n\n"
                + ozet.ozet_str()
            )
            msg.setStyleSheet("""
                QMessageBox {
                    background-color: #0A0A0A;
                    border: 1px solid #333;
                    border-radius: 10px;
                }
                QLabel {
                    color: white;
                    font-size: 12px;
                    font-weight: 500;
                    min-width: 420px;
                    min-height: 40px;
                    background: transparent;
                    line-height: 1.6;
                }
                QPushButton {
                    background-color: #2563EB;
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 8px 24px;
                    font-size: 12px;
                    font-weight: bold;
                    min-width: 80px;
                }
                QPushButton:hover { background-color: #1D4ED8; }
            """)
            msg.exec()
    
        except Exception as exc:
            prog.close()
            print(traceback.format_exc())
            QMessageBox.critical(self, "Hata", f"Excel oluşturulurken hata:\n{exc}")



# ─────────────────────────────────────────────────────────────────────────────
# Kurum Ödemeleri Detay Dialog
# PHP: nakitAkimParametreAjaxGider.php + gider_veriler.js DataTable
# ─────────────────────────────────────────────────────────────────────────────

class KurumOdemeDialog(QDialog):
    """
    Kurum Ödemeleri kartına tıklandığında açılan DETAYLI tablo.
    Beyanname bilgileri (belge türü, dönem, onay tarihi, PDF) doğrudan
    tablo sütunlarında gösterilir — ayrı dialog açılmaz.
    """

    HESAP_ACIKLAMA = {
        "770.01": "770.01 — Vergi Giderleri (SGK / KDV / Gelir Vergisi)",
        "730.08": "730.08 — İşçilik / Müşavirlik Giderleri",
    }
    BEYANNAME_TUR = {
        "770.01": "KDV / SGK Beyannamesi",
        "730.08": "Muhtasar ve Prim Hizmet Beyannamesi",
    }

    AY_ADLARI = [
        "Hepsi", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
        "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık",
    ]

    SUTUNLAR = [
        ("Beyanname Türü",  "hesapKodu",     185),
        ("Ünvan",           "unvan",         130),
        ("Vergi No",        "vergiNo",       100),
        ("İlk Tarih",       "ilkTarih",       90),
        ("Son Tarih",       "sonTarih",       90),
        ("Sözleşme No",     "sozlesmeNo",     95),
        ("Sözl. Tarih",     "sozlesmeTarih",  85),
        ("Tutar",           "tutar",         110),
        ("Belge Türü",      "byn_belge_turu", 130),
        ("Dönem",           "byn_donem",     100),
        ("Onay Tarihi",     "byn_onay",      100),
        ("Belge Durum",     "byn_durum",      85),
        ("PDF",             "byn_pdf",        60),
    ]

    _CBS = (
        "QComboBox{background:white;border:1.5px solid #cbd5e1;"
        "border-radius:6px;padding:0 8px;font-size:12px;color:#1e293b;}"
        "QComboBox:focus{border-color:#162C47;}"
        "QComboBox::drop-down{border:none;width:18px;}"
    )
    _DES = (
        "QDateEdit{background:white;border:1.5px solid #cbd5e1;"
        "border-radius:6px;padding:0 6px;font-size:12px;color:#1e293b;}"
        "QDateEdit:focus{border-color:#162C47;}"
        "QDateEdit::drop-down{border:none;width:18px;}"
    )

    def __init__(self, musterino: int, yil: int, parent=None):
        super().__init__(parent)
        self._musterino    = musterino
        self._yil          = yil
        self._rows: list[dict] = []
        self._ay_degisiyor = False   # döngü koruması
        self._pdf_map: dict[int, bytes] = {}  # satır_no → pdf bytes

        self.setWindowTitle("Kurum Ödemeleri — Detay")
        self.setMinimumSize(1300, 660)
        self.resize(1480, 760)
        self._setup_ui()
        self._load()

    # ── UI ───────────────────────────────────────────────────────────────────

    def _setup_ui(self):
        import calendar
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 14, 16, 14)
        root.setSpacing(8)

        # ── Başlık ──
        baslik = QLabel("📋  Kurum Ödemeleri — Tüm Detaylar")
        baslik.setStyleSheet(
            "font-size:15px;font-weight:700;color:#162C47;"
            "padding-bottom:4px;"
        )
        root.addWidget(baslik)

        # ── Filtre çubuğu ──
        bar = QHBoxLayout()
        bar.setSpacing(8)

        # Ay seçici
        bar.addWidget(self._lbl("Ay:"))
        self._ay_cb = QComboBox()
        self._ay_cb.setFixedSize(100, 32)
        self._ay_cb.setStyleSheet(self._CBS)
        for i, a in enumerate(self.AY_ADLARI):
            self._ay_cb.addItem(a, i)
        self._ay_cb.currentIndexChanged.connect(self._on_ay_change)
        bar.addWidget(self._ay_cb)

        # İlk Tarih DateEdit
        bar.addWidget(self._lbl("İlk Tarih:"))
        self._ilk_de = QDateEdit()
        self._ilk_de.setCalendarPopup(True)
        self._ilk_de.setDisplayFormat("dd.MM.yyyy")
        self._ilk_de.setFixedSize(120, 32)
        self._ilk_de.setStyleSheet(self._DES)
        self._ilk_de.setDate(QDate(self._yil, 1, 1))
        bar.addWidget(self._ilk_de)

        # Son Tarih DateEdit
        bar.addWidget(self._lbl("Son Tarih:"))
        self._son_de = QDateEdit()
        self._son_de.setCalendarPopup(True)
        self._son_de.setDisplayFormat("dd.MM.yyyy")
        self._son_de.setFixedSize(120, 32)
        self._son_de.setStyleSheet(self._DES)
        self._son_de.setDate(QDate.currentDate())
        bar.addWidget(self._son_de)

        bar.addStretch()

        # ── Arama Kutusu ──
        bar.addWidget(self._lbl("Ara:"))
        self._ara_le = QLineEdit()
        self._ara_le.setPlaceholderText("Beyanname türü, unvan vb...")
        self._ara_le.setFixedSize(160, 32)
        self._ara_le.setStyleSheet(
            "QLineEdit{background:white;border:1.5px solid #cbd5e1;"
            "border-radius:6px;padding:0 8px;font-size:12px;color:#1e293b;}"
            "QLineEdit:focus{border-color:#162C47;}"
        )
        self._ara_le.textChanged.connect(self._doldur)
        bar.addWidget(self._ara_le)

        # Filtrele butonu
        self._filtre_btn = QPushButton("🔍  Tarihi Filtrele")
        self._filtre_btn.setFixedSize(130, 32)
        self._filtre_btn.setStyleSheet(
            "QPushButton{background:#162C47;color:white;border:none;"
            "border-radius:7px;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#1e3a5f;}"
        )
        self._filtre_btn.clicked.connect(self._load)
        bar.addWidget(self._filtre_btn)
        root.addLayout(bar)

        # ── Özet bant ──
        self._ozet_lbl = QLabel("")
        self._ozet_lbl.setFixedHeight(28)
        self._ozet_lbl.setStyleSheet(
            "background:#f0f9ff;border:1px solid #bae6fd;border-radius:6px;"
            "padding:0 10px;font-size:12px;color:#0c4a6e;"
        )
        root.addWidget(self._ozet_lbl)

        # ── Bilgi notu ──
        not_lbl = QLabel("💡  PDF sütunundaki 📄 butonuna tıklayarak beyanname PDF'ini açabilirsiniz.")
        not_lbl.setStyleSheet("font-size:11px;color:#64748b;")
        root.addWidget(not_lbl)

        # ── Tablo ──
        self._tablo = QTableWidget()
        self._tablo.setColumnCount(len(self.SUTUNLAR))
        self._tablo.setHorizontalHeaderLabels([s[0] for s in self.SUTUNLAR])
        self._tablo.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tablo.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._tablo.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._tablo.setAlternatingRowColors(True)
        self._tablo.verticalHeader().setVisible(False)
        self._tablo.setSortingEnabled(False)  # butonlar olduğu için sıralama kapalı
        self._tablo.horizontalHeader().setStretchLastSection(False)
        self._tablo.horizontalHeader().setSectionResizeMode(
            0, self._tablo.horizontalHeader().ResizeMode.Stretch
        )
        self._tablo.setMouseTracking(True)
        for i, (_, _, w) in enumerate(self.SUTUNLAR):
            self._tablo.setColumnWidth(i, w)
        self._tablo.setStyleSheet("""
            QTableWidget {
                background: white;
                gridline-color: #e2e8f0;
                font-size: 12px;
                color: #1e293b;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
            }
            QTableWidget::item {
                color: #1e293b;
                padding: 4px 6px;
            }
            QTableWidget::item:hover {
                background: #eff6ff;
                color: #1e293b;
            }
            QTableWidget::item:selected {
                background: #dbeafe;
                color: #1e293b;
            }
            QTableWidget::item:alternate {
                background: #f8fafc;
                color: #1e293b;
            }
            QHeaderView::section {
                background: #1e293b;
                color: white;
                font-weight: 700;
                font-size: 11px;
                padding: 6px 4px;
                border: none;
                border-right: 1px solid #334155;
            }
        """)
        root.addWidget(self._tablo)

        # ── Kapat ──
        kapat = QPushButton("✕  Kapat")
        kapat.setFixedSize(110, 32)
        kapat.setStyleSheet(
            "QPushButton{background:#64748b;color:white;border:none;"
            "border-radius:7px;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#475569;}"
        )
        kapat.clicked.connect(self.accept)
        bot = QHBoxLayout()
        bot.addStretch()
        bot.addWidget(kapat)
        root.addLayout(bot)

    # ── Yardımcılar ──────────────────────────────────────────────────────────

    @staticmethod
    def _lbl(txt: str) -> QLabel:
        l = QLabel(txt)
        l.setStyleSheet("font-size:12px;color:#475569;font-weight:600;")
        return l

    @staticmethod
    def _fmt_goster(t) -> str:
        """YYYYMMDD → DD.MM.YYYY"""
        s = str(t) if t else ""
        if len(s) == 8 and s.isdigit():
            return f"{s[6:8]}.{s[4:6]}.{s[0:4]}"
        return s

    # ── Ay seçince tarihler otomatik dolar ───────────────────────────────────

    def _on_ay_change(self, idx: int):
        """
        Şubat seçilirse → ilkTarih=01.02.YYYY, sonTarih=28.02.YYYY
        Hepsi seçilirse → ilkTarih=01.01.YYYY, sonTarih=bugün
        """
        import calendar
        if self._ay_degisiyor:
            return
        self._ay_degisiyor = True
        ay = self._ay_cb.currentData()   # 0=hepsi, 1=Ocak…12=Aralık
        yil = self._yil
        if ay and ay > 0:
            son_gun = calendar.monthrange(yil, ay)[1]
            self._ilk_de.setDate(QDate(yil, ay, 1))
            self._son_de.setDate(QDate(yil, ay, son_gun))
        else:
            self._ilk_de.setDate(QDate(yil, 1, 1))
            self._son_de.setDate(QDate.currentDate())
        self._ay_degisiyor = False

    # ── Veri yükleme ─────────────────────────────────────────────────────────

    def _load(self):
        from services.detay_service import get_kurum_odemeleri_detay_tarih
        from db.database import get_connection

        ilk = self._ilk_de.date()
        son = self._son_de.date()
        ilk_str = f"{ilk.year()}{ilk.month():02d}{ilk.day():02d}"
        son_str = f"{son.year()}{son.month():02d}{son.day():02d}"

        self._rows, self._sql_toplam = get_kurum_odemeleri_detay_tarih(
            self._musterino, ilk_str, son_str
        )
        self._pdf_map = {}

        # ── Beyanname bilgilerini önce byn_kayit_no ile çek ────────────────────
        # Her nakitakis_parametre satırında byn_kayit_no varsa
        # moy_beyannameler'den DIREKT çek (tarih eşleştirmesi gerekmez).
        # byn_kayit_no NULL ise yerel önbellekte tarih+kod ile fallback yap.
        from services.moy_service import get_local_beyannameler

        # moy_beyannameler tablosundan önce bu müşteriye ait tüm önbelleği çek
        beyan_cache: dict[int, dict] = {}   # kayit_no → beyanname satırı
        try:
            conn = get_connection()
            rows_b = conn.execute(
                """SELECT kayit_no, belge_tipi, belge_turu, donem_no, donem_adi,
                          onay_tarihi, belge_no, belge_durumu,
                          beyan_tarih_1, beyan_tarih_2,
                          sube_adi AS sgm_kodu, sube_alanlar, musteri_unvani
                   FROM moy_beyannameler
                   WHERE musteri_no = ?""",
                (self._musterino,)
            ).fetchall()
            conn.close()
            beyan_cache = {int(r["kayit_no"]): dict(r) for r in rows_b}
        except Exception:
            beyan_cache = {}

        self._enriched = []
        for row in self._rows:
            ilk_tarih = str(row.get("ilkTarih", row.get("ilktarih", "")))
            hkod = str(row.get("hesapKodu", row.get("hesapkodu", "")))
            byn_kno = row.get("byn_kayit_no")   # nakitakis_parametre kolonu

            byn = None
            if byn_kno and int(byn_kno) in beyan_cache:
                # ✅ Doğrudan ilişkilendirilmiş beyanname
                byn = beyan_cache[int(byn_kno)]
            elif ilk_tarih:
                # Fallback: eski tarih-tabanlı arama (byn_kayit_no yoksa)
                beyanlar = get_local_beyannameler(self._musterino, ilk_tarih, hkod)
                byn = beyanlar[0] if beyanlar else None

            self._enriched.append((row, byn))

        # ── Beyanname pair index: (donem_no, belge_turu, sube_adi) → {Byn/Thk: kayit_no} ─
        # Aynı dönem+tür için hem Byn hem Thk kaydı olabilir.
        # İndeks sayesinde her satırda her ikisini de bulabiliriz.
        pair_idx: dict[tuple, dict] = {}
        for kno, b in beyan_cache.items():
            key = (
                str(b.get("donem_no", "") or ""),
                str(b.get("belge_turu", "") or ""),
                str(b.get("sgm_kodu", "") or ""),   # sube_adi
            )
            if key not in pair_idx:
                pair_idx[key] = {}
            tipi = str(b.get("belge_tipi", "") or "")
            if tipi not in pair_idx[key]:
                pair_idx[key][tipi] = kno
        self._pair_idx  = pair_idx
        self._beyan_cache = beyan_cache

        self._doldur()


    def _doldur(self):
        from PyQt6.QtGui import QColor, QFont

        # Belge türü adları — PDF'teki başlıklarla uyumlu
        BELGE_TUR_ADI = {
            "KDV1":     "KDV Beyannamesi (1.Tür)",
            "KDV2":     "KDV Beyannamesi (2.Tür)",
            "MUHSGK":   "SGK Tahakkuk Fişi (5510)",
            "KGECICI":  "Kurumlar Vg. Geçici",
            "KURUMLAR": "Kurumlar Vergisi",
            "LEVHA":    "Levha Beyannamesi",
            "MUHTAR":   "Muhtasar Beyanname",
        }

        self._tablo.setRowCount(0)
        toplam = 0.0
        sayilan_tutarlar = set()  # (id, tutar) → toplama bir kez say
        
        q = self._ara_le.text().strip().lower() if hasattr(self, "_ara_le") else ""

        for ri_idx, (row, byn) in enumerate(self._enriched):
            kod      = row.get("hesapKodu", "")
            beyan_t  = self.BEYANNAME_TUR.get(kod, self.HESAP_ACIKLAMA.get(kod, kod))
            vergino  = row.get("vergiNo", "") or ""
            ilkT     = self._fmt_goster(row.get("ilkTarih", ""))
            sonT     = self._fmt_goster(row.get("sonTarih", ""))
            sozno    = row.get("sozlesmeNo", "") or ""
            soztarih = self._fmt_goster(row.get("sozlesmeTarih", ""))
            tutar    = float(row.get("tutar") or 0)

            # Son Tarih / Sözleşme Tarih → DB boşsa beyanname verilerinden doldur
            if byn:
                if not sonT:
                    sonT = self._fmt_goster(byn.get("beyan_tarih_2", ""))
                if not soztarih:
                    soztarih = self._fmt_goster(byn.get("onay_tarihi", ""))

            # Ünvan: 1) DB unvanı (moy_kaydet_veriler'den), 2) Moy musteri_unvani, 3) SGM
            db_unvan = str(row.get("unvan", "") or "")
            if byn:
                _mu = byn.get("musteri_unvani", "") or ""
                _sk = byn.get("sgm_kodu", "") or ""
                _sa = byn.get("sgm_adi",  "") or ""
                if db_unvan and db_unvan not in ("-", ""):
                    unvan = db_unvan
                elif _mu:
                    unvan = _mu
                elif _sk and _sa:
                    unvan = f"{_sk} - {_sa}"
                elif _sa:
                    unvan = _sa
                elif _sk:
                    unvan = _sk
                else:
                    unvan = "-"
            else:
                unvan = db_unvan if db_unvan and db_unvan not in ("-", "") else "-"

            if q:
                if q not in beyan_t.lower() and q not in unvan.lower() and q not in kod.lower():
                    continue

            ri = self._tablo.rowCount()
            self._tablo.insertRow(ri)
            self._tablo.setRowHeight(ri, 32)

            # Tutarı Python'da toplama — sadece sayım için
            row_key = (row.get("id", id(row)), tutar)
            sayilan_tutarlar.add(row_key)

            # Beyanname sütunları
            if byn:
                raw_belge_turu = byn.get("belge_turu", "")
                # MUHSGK belgesi hesap koduna göre farklı etiket alır:
                # 770.01 (SGK ödemesi) → Tahakkuk Fişi
                # 730.08 (Muhtasar)    → Muhtasar ve Prim Hizmet Beyannamesi
                if raw_belge_turu == "MUHSGK" and kod == "730.08":
                    byn_belge = "Muhtasar ve Prim Hizmet Beyannamesi"
                else:
                    byn_belge = BELGE_TUR_ADI.get(raw_belge_turu, raw_belge_turu)
                byn_donem = byn.get("donem_adi", "") or ""
                onay_raw  = byn.get("onay_tarihi", "") or ""
                if len(str(onay_raw)) >= 8 and str(onay_raw)[:8].isdigit():
                    s = str(onay_raw)
                    byn_onay = f"{s[6:8]}.{s[4:6]}.{s[0:4]}"
                else:
                    byn_onay = str(onay_raw)
                byn_durum  = byn.get("belge_durumu", "") or ""
                # Belge tipi etiketi: Thk (tahakkuk) olsa bile 'Byn' göster
                _btipi = str(byn.get("belge_tipi", "") or "")
                _btipi_etiket = {"Thk": "Byn", "Byn": "Byn",
                                  "Hiz": "Byn", "Blg": "Byn"}.get(_btipi, _btipi)
                if _btipi_etiket:
                    byn_durum = f"{_btipi_etiket}  {byn_durum}".strip()
                byn_kayit  = byn.get("kayit_no", None)
            else:
                byn_belge = "—"
                byn_donem = "—"
                byn_onay  = "—"
                byn_durum = "—"
                byn_kayit = None

            # Sütun verileri: 0-7 normal, 8-11 beyanname, 12 PDF butonu
            metin_sutunlar = [
                beyan_t, unvan, vergino, ilkT, sonT, sozno, soztarih,
                None,  # tutar (özel)
                byn_belge, byn_donem, byn_onay, byn_durum,
            ]

            for ci, val in enumerate(metin_sutunlar):
                if ci == 7:
                    it = QTableWidgetItem(f"{tutar:,.2f} ₺")
                    it.setTextAlignment(
                        Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
                    )
                    it.setForeground(QColor("#0f766e"))
                    it.setFont(QFont("", -1, QFont.Weight.Bold))
                elif ci in (8, 9):
                    # Belge türü ve dönem — mavi vurgu
                    it = QTableWidgetItem(str(val))
                    it.setTextAlignment(
                        Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
                    )
                    it.setForeground(QColor("#1d4ed8") if byn else QColor("#94a3b8"))
                elif ci == 10:
                    # Onay tarihi — yeşil
                    it = QTableWidgetItem(str(val))
                    it.setTextAlignment(
                        Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter
                    )
                    it.setForeground(QColor("#059669") if byn else QColor("#94a3b8"))
                elif ci == 11:
                    # Belge durumu
                    it = QTableWidgetItem(str(val))
                    it.setTextAlignment(
                        Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter
                    )
                    it.setForeground(QColor("#7c3aed") if byn else QColor("#94a3b8"))
                else:
                    it = QTableWidgetItem(str(val))
                    it.setTextAlignment(
                        Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
                    )
                    it.setForeground(QColor("#1e293b"))
                it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self._tablo.setItem(ri, ci, it)

            # ── PDF butonları (sütun 12) — Byn + Tah yan yana ───────────────
            ilk_tarih_row  = str(row.get("ilkTarih", "") or "")
            hesap_kodu_row = str(row.get("hesapKodu", "") or "")

            # Pair index'ten Byn ve Thk kayit_no'larını bul
            byn_kno_btn = None
            thk_kno_btn = None
            if byn:
                _pair_key = (
                    str(byn.get("donem_no", "") or ""),
                    str(byn.get("belge_turu", "") or ""),
                    str(byn.get("sgm_kodu", "") or ""),
                )
                _pair = getattr(self, "_pair_idx", {}).get(_pair_key, {})
                byn_kno_btn = _pair.get("Byn") or _pair.get("Hiz")
                thk_kno_btn = _pair.get("Thk") or _pair.get("Blg")
                # Fallback: mevcut byn_kayit uygun tipe ata
                if byn_kayit is not None:
                    _cur_tipi = str(byn.get("belge_tipi", "") or "")
                    if _cur_tipi in ("Thk", "Blg") and thk_kno_btn is None:
                        thk_kno_btn = byn_kayit
                    elif _cur_tipi in ("Byn", "Hiz") and byn_kno_btn is None:
                        byn_kno_btn = byn_kayit

            # Widget container
            btn_widget  = QWidget()
            btn_layout  = QHBoxLayout(btn_widget)
            btn_layout.setContentsMargins(2, 2, 2, 2)
            btn_layout.setSpacing(4)

            _BTN_STYLE_BYN = (
                "QPushButton{background:#0f766e;color:white;border:none;"
                "border-radius:5px;font-size:11px;font-weight:600;padding:0 6px;}"
                "QPushButton:hover{background:#0d9488;}"
            )
            _BTN_STYLE_TAH = (
                "QPushButton{background:#7c3aed;color:white;border:none;"
                "border-radius:5px;font-size:11px;font-weight:600;padding:0 6px;}"
                "QPushButton:hover{background:#6d28d9;}"
            )

            def _make_pdf_handler(kno, mno):
                def _h():
                    self._pdf_ac_kayit(kno, mno)
                return _h

            if byn_kno_btn is not None:
                b_btn = QPushButton("📄 Byn")
                b_btn.setFixedHeight(26)
                b_btn.setStyleSheet(_BTN_STYLE_BYN)
                b_btn.clicked.connect(_make_pdf_handler(byn_kno_btn, self._musterino))
                btn_layout.addWidget(b_btn)

            if thk_kno_btn is not None:
                t_btn = QPushButton("📄 Tah")
                t_btn.setFixedHeight(26)
                t_btn.setStyleSheet(_BTN_STYLE_TAH)
                t_btn.clicked.connect(_make_pdf_handler(thk_kno_btn, self._musterino))
                btn_layout.addWidget(t_btn)

            if byn_kno_btn is None and thk_kno_btn is None:
                no_pdf = QLabel("—")
                no_pdf.setAlignment(Qt.AlignmentFlag.AlignCenter)
                no_pdf.setStyleSheet("color:#94a3b8;font-size:11px;")
                btn_layout.addWidget(no_pdf)

            btn_layout.addStretch()
            self._tablo.setCellWidget(ri, 12, btn_widget)


        # Özet
        ilk_txt = self._ilk_de.date().toString("dd.MM.yyyy")
        son_txt = self._son_de.date().toString("dd.MM.yyyy")
        self._ozet_lbl.setText(
            f"📅 {ilk_txt} — {son_txt}  │  "
            f"<b>{len(self._rows)}</b> kayıt  │  "
            f"Toplam: <b>{self._sql_toplam:,.2f} ₺</b>"
        )
        self._ozet_lbl.setTextFormat(Qt.TextFormat.RichText)

    # ── PDF Aç (Kayıt No ile) ────────────────────────────────────────────────

    def _pdf_ac_kayit(self, kayit_no: int, musterino: int):
        from services.moy_service import get_beyanname_pdf_bytes
        import tempfile, os, subprocess, sys
        pdf_bytes = get_beyanname_pdf_bytes(musterino, kayit_no)
        if not pdf_bytes:
            from PyQt6.QtWidgets import QDialog, QVBoxLayout, QLabel, QPushButton as _PB
            dlg = QDialog(self)
            dlg.setWindowTitle("PDF Bulunamadı")
            dlg.setFixedSize(380, 160)
            dlg.setStyleSheet("""
                QDialog  { background:#1e293b; }
                QLabel   { color:#f1f5f9; font-size:13px; padding:8px; }
                QPushButton { background:#0f766e; color:white; border:none;
                              border-radius:6px; padding:6px 24px;
                              font-size:12px; font-weight:600; }
                QPushButton:hover { background:#0d9488; }
            """)
            lay = QVBoxLayout(dlg)
            lay.addWidget(QLabel(
                "⚠️  Bu ödeme için PDF verisi bulunamadı.\n\n"
                "Beyanname Moy'da kaydedilmemiş olabilir."
            ))
            btn = _PB("Kapat")
            btn.clicked.connect(dlg.accept)
            lay.addWidget(btn)
            dlg.exec()
            return
        workspace_tmp = os.path.expanduser("~/NakitAkim/data/tmp")
        os.makedirs(workspace_tmp, exist_ok=True)
        tmp = tempfile.NamedTemporaryFile(
            suffix=".pdf",
            prefix=f"beyanname_{kayit_no}_",
            dir=workspace_tmp,
            delete=False
        )
        tmp.write(pdf_bytes)
        tmp.close()
        try:
            if sys.platform == "darwin":
                subprocess.Popen(["open", tmp.name])
            elif sys.platform == "win32":
                os.startfile(tmp.name)
            else:
                subprocess.Popen(["xdg-open", tmp.name])
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Hata", f"PDF açılamadı: {e}")

    def _pdf_ac_lazy(self, ilk_tarih: str, hesap_kodu: str, musterino: int):
        """PDF butonuna tıklanınca o an Moy'dan beyanname listesini çeker."""
        from PyQt6.QtWidgets import QMessageBox, QInputDialog
        from services.moy_service import get_beyanname_listesi

        try:
            beyanlar = get_beyanname_listesi(musterino, ilk_tarih, hesap_kodu=hesap_kodu)
        except Exception as e:
            QMessageBox.warning(self, "Moy Bağlantı Hatası",
                f"Beyanname listesi alınamadı:\n{e}")
            return

        if not beyanlar:
            QMessageBox.information(self, "PDF Bulunamadı",
                "Bu tarihe ait beyanname bulunamadı.")
            return

        if len(beyanlar) == 1:
            self._pdf_ac_kayit(beyanlar[0]["kayit_no"], musterino)
        else:
            # Birden fazla beyanname — seçim sun
            secenekler = [
                f"{b['belge_turu']} — {b.get('donem_adi', '')} (Onay: {b.get('onay_tarihi', '')})"
                for b in beyanlar
            ]
            secim, ok = QInputDialog.getItem(
                self, "Beyanname Seç",
                "Bu tarihe ait birden fazla beyanname var. Hangisini açmak istiyorsunuz?",
                secenekler, 0, False
            )
            if ok and secim:
                idx = secenekler.index(secim)
                self._pdf_ac_kayit(beyanlar[idx]["kayit_no"], musterino)




# ─────────────────────────────────────────────────────────────────────────────
# Beyanname Önizleme Dialog
# ─────────────────────────────────────────────────────────────────────────────

class BeyannamePreviewDialog(QDialog):
    """
    Satıra tıklandığında açılan PDF önizleme dialog'u.
    - Moy'dan gerçek zamanlı beyanname_listeleri sorgusu yapar
    - İlgili dönemde bulunan beyannameleri listeler
    - Seçilen beyanname'nin PDF'ini beyanname_gib.Belge_Data'dan çeker
    - PDF'i geçici dosyaya yazıp sistem PDF görüntüleyicisiyle açar
    """

    BEYANNAME_TUR = {
        "770.01": "KDV / SGK Beyannamesi",
        "730.08": "Muhtasar ve Prim Hizmet Beyannamesi",
    }
    BEYANNAME_ICON = {"770.01": "🏛️", "730.08": "📄"}
    BEYANNAME_RENK = {"770.01": "#1d4ed8", "730.08": "#7c3aed"}

    BELGE_TUR_ADI = {
        "KDV1":     "KDV Beyannamesi (1.Tür)",
        "KDV2":     "KDV Beyannamesi (2.Tür)",
        "MUHSGK":   "SGK Tahakkuk Fişi (5510)",
        "KGECICI":  "Kurumlar Vergisi Geçici Beyan",
        "KURUMLAR": "Kurumlar Vergisi Beyannamesi",
        "LEVHA":    "Levha Beyannamesi",
        "MUHTAR":   "Muhtasar Beyanname",
    }

    def __init__(self, veri: dict, musterino: int, parent=None):
        super().__init__(parent)
        self._veri      = veri
        self._musterino = musterino
        self._beyanlar: list[dict] = []
        self._tmp_path: str | None = None

        kod = veri.get("hesapKodu", "")
        self._renk = self.BEYANNAME_RENK.get(kod, "#162C47")
        self.setWindowTitle("Beyanname PDF Önizleme")
        self.setMinimumSize(860, 640)
        self.resize(960, 720)
        self._setup_ui()
        self._load_beyanlar()

    @staticmethod
    def _fmt(t) -> str:
        s = str(t) if t else ""
        if len(s) == 8 and s.isdigit():
            return f"{s[6:8]}.{s[4:6]}.{s[0:4]}"
        return s or "—"

    def _donem_str(self) -> str:
        t = str(self._veri.get("ilkTarih", "") or "")
        if len(t) == 8 and t.isdigit():
            ay_map = {"01":"Ocak","02":"Şubat","03":"Mart","04":"Nisan",
                      "05":"Mayıs","06":"Haziran","07":"Temmuz","08":"Ağustos",
                      "09":"Eylül","10":"Ekim","11":"Kasım","12":"Aralık"}
            return f"{ay_map.get(t[4:6], t[4:6])} {t[0:4]}"
        return t or "—"

    # ── UI ───────────────────────────────────────────────────────────────────

    def _setup_ui(self):
        from PyQt6.QtGui import QColor
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        kod  = self._veri.get("hesapKodu", "")
        tur  = self.BEYANNAME_TUR.get(kod, "Beyanname")
        icon = self.BEYANNAME_ICON.get(kod, "📋")
        renk = self._renk

        # ── Başlık bandı ──
        hdr = QFrame()
        hdr.setFixedHeight(76)
        hdr.setStyleSheet(
            f"background:qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            f"stop:0 {renk},stop:1 #0f172a);"
        )
        h = QHBoxLayout(hdr)
        h.setContentsMargins(20, 0, 20, 0)

        left = QVBoxLayout()
        t1 = QLabel(f"{icon}  {tur}")
        t1.setStyleSheet("color:white;font-size:15px;font-weight:700;background:transparent;")
        t2 = QLabel(f"📅  Dönem: {self._donem_str()}   •   Tutar: {float(self._veri.get('tutar') or 0):,.2f} ₺")
        t2.setStyleSheet("color:rgba(255,255,255,0.82);font-size:12px;background:transparent;")
        left.addWidget(t1)
        left.addWidget(t2)
        h.addLayout(left, 1)
        root.addWidget(hdr)

        # ── İçerik: Sol liste + Sağ PDF alanı ──
        splitter_frame = QFrame()
        splitter_frame.setStyleSheet("background:#f8fafc;")
        sf_lay = QHBoxLayout(splitter_frame)
        sf_lay.setContentsMargins(10, 10, 10, 10)
        sf_lay.setSpacing(10)

        # Sol: Beyanname listesi
        sol = QFrame()
        sol.setFixedWidth(290)
        sol.setStyleSheet(
            "background:white;border-radius:8px;"
            "border:1px solid #e2e8f0;"
        )
        sol_lay = QVBoxLayout(sol)
        sol_lay.setContentsMargins(0, 0, 0, 0)
        sol_lay.setSpacing(0)

        sol_baslik = QLabel("📋  İlgili Beyannameler")
        sol_baslik.setFixedHeight(36)
        sol_baslik.setStyleSheet(
            f"background:{renk};color:white;font-size:12px;font-weight:700;"
            "padding:0 10px;border-radius:8px 8px 0 0;"
        )
        sol_lay.addWidget(sol_baslik)

        self._liste = QTableWidget()
        self._liste.setColumnCount(4)
        self._liste.setHorizontalHeaderLabels(["Tür", "Dönem", "Onay", "SGM"])
        self._liste.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._liste.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._liste.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._liste.verticalHeader().setVisible(False)
        self._liste.setShowGrid(False)
        self._liste.horizontalHeader().setStretchLastSection(True)
        self._liste.setColumnWidth(0, 90)
        self._liste.setColumnWidth(1, 80)
        self._liste.setStyleSheet("""
            QTableWidget { background:white; font-size:11px; border:none; color:#1e293b; }
            QTableWidget::item { padding:4px 6px; color:#1e293b; }
            QTableWidget::item:hover { background:#eff6ff; color:#1e293b; }
            QTableWidget::item:selected { background:#dbeafe; color:#1e293b; }
            QHeaderView::section {
                background:#f1f5f9; color:#475569; font-weight:700;
                font-size:10px; padding:4px; border:none;
                border-bottom:1px solid #e2e8f0;
            }
        """)
        self._liste.cellClicked.connect(self._on_beyanname_sec)
        sol_lay.addWidget(self._liste, 1)

        self._yukleniyor_lbl = QLabel("🔄  Moy'dan yükleniyor...")
        self._yukleniyor_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._yukleniyor_lbl.setStyleSheet("color:#64748b;font-size:11px;padding:10px;")
        sol_lay.addWidget(self._yukleniyor_lbl)
        self._yukleniyor_lbl.hide()

        sf_lay.addWidget(sol)

        # Sağ: PDF alanı
        sag = QFrame()
        sag.setStyleSheet(
            "background:white;border-radius:8px;"
            "border:1px solid #e2e8f0;"
        )
        sag_lay = QVBoxLayout(sag)
        sag_lay.setContentsMargins(0, 0, 0, 0)

        sag_baslik_frame = QFrame()
        sag_baslik_frame.setFixedHeight(36)
        sag_baslik_frame.setStyleSheet(
            "background:#1e293b;border-radius:8px 8px 0 0;"
        )
        sb_lay = QHBoxLayout(sag_baslik_frame)
        sb_lay.setContentsMargins(12, 0, 12, 0)
        self._pdf_baslik = QLabel("PDF Önizleme")
        self._pdf_baslik.setStyleSheet(
            "color:white;font-size:12px;font-weight:700;"
        )
        sb_lay.addWidget(self._pdf_baslik)
        sb_lay.addStretch()

        self._pdf_ac_btn = QPushButton("📂  PDF'i Aç")
        self._pdf_ac_btn.setFixedSize(110, 26)
        self._pdf_ac_btn.setStyleSheet(
            "QPushButton{background:#0f766e;color:white;border:none;"
            "border-radius:5px;font-size:11px;font-weight:600;}"
            "QPushButton:hover{background:#0d9488;}"
            "QPushButton:disabled{background:#334155;color:#64748b;}"
        )
        self._pdf_ac_btn.setEnabled(False)
        self._pdf_ac_btn.clicked.connect(self._pdf_ac)
        sb_lay.addWidget(self._pdf_ac_btn)
        sag_lay.addWidget(sag_baslik_frame)

        # PDF içerik alanı — metin tabanlı önizleme
        self._pdf_alan = QLabel()
        self._pdf_alan.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._pdf_alan.setWordWrap(True)
        self._pdf_alan.setStyleSheet(
            "color:#94a3b8;font-size:13px;padding:20px;"
            "background:white;"
        )
        self._pdf_alan.setText(
            "⬅️  Soldaki listeden bir beyanname seçin\n"
            "PDF önizlemesi burada gösterilecektir."
        )

        # PDF Image görüntüleme için QLabel (PDF → PNG dönüşümü)
        self._pdf_img_scroll = QScrollArea()
        self._pdf_img_scroll.setWidgetResizable(True)
        self._pdf_img_scroll.setStyleSheet(
            "QScrollArea{border:none;background:white;}"
        )
        self._pdf_img_container = QLabel()
        self._pdf_img_container.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignHCenter)
        self._pdf_img_container.setStyleSheet("background:white;padding:8px;")
        self._pdf_img_scroll.setWidget(self._pdf_img_container)
        self._pdf_img_scroll.hide()

        sag_lay.addWidget(self._pdf_alan, 1)
        sag_lay.addWidget(self._pdf_img_scroll, 1)
        sf_lay.addWidget(sag, 1)

        root.addWidget(splitter_frame, 1)

        # ── Alt butonlar ──
        alt = QFrame()
        alt.setFixedHeight(48)
        alt.setStyleSheet("background:white;border-top:1px solid #e2e8f0;")
        a = QHBoxLayout(alt)
        a.setContentsMargins(16, 0, 16, 0)
        a.addStretch()
        kapat = QPushButton("✕  Kapat")
        kapat.setFixedSize(110, 32)
        kapat.setStyleSheet(
            "QPushButton{background:#64748b;color:white;border:none;"
            "border-radius:7px;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#475569;}"
        )
        kapat.clicked.connect(self.accept)
        a.addWidget(kapat)
        root.addWidget(alt)

    # ── Beyanname listesini yükle ─────────────────────────────────────────────

    def _load_beyanlar(self):
        ilk_tarih   = self._veri.get("ilkTarih", "") or ""
        hesap_kodu  = self._veri.get("hesapKodu", "") or ""
        if not ilk_tarih:
            self._liste_bos_goster("İlk tarih bilgisi yok.")
            return

        # Yükleme dialogu
        from PyQt6.QtWidgets import QProgressDialog, QApplication
        from PyQt6.QtCore import Qt
        prog = QProgressDialog(
            "🔄  Beyanname bilgileri MöY'dan çekiliyor...\n"
            "Lütfen bekleyiniz.",
            None, 0, 0, self
        )
        prog.setWindowTitle("Lütfen Bekleyiniz")
        prog.setWindowModality(Qt.WindowModality.WindowModal)
        prog.setMinimumDuration(0)
        prog.setValue(0)
        QApplication.processEvents()

        self._yukleniyor_lbl.show()
        from services.moy_service import get_beyanname_listesi
        try:
            self._beyanlar = get_beyanname_listesi(
                self._musterino,
                ilk_tarih,
                hesap_kodu=hesap_kodu
            )
        except Exception:
            self._beyanlar = []
        prog.close()
        self._yukleniyor_lbl.hide()
        self._liste_doldur()

    def _liste_doldur(self):
        self._liste.setRowCount(0)
        if not self._beyanlar:
            self._liste_bos_goster("Bu dönem için beyanname bulunamadı.")
            return

        for b in self._beyanlar:
            ri = self._liste.rowCount()
            self._liste.insertRow(ri)
            self._liste.setRowHeight(ri, 26)

            tur_adi = self.BELGE_TUR_ADI.get(b["belge_turu"], b["belge_turu"])
            onay = b["onay_tarihi"]
            if len(onay) >= 8 and onay[:8].isdigit():
                onay = f"{onay[6:8]}.{onay[4:6]}.{onay[0:4]}"

            _sk = b.get("sgm_kodu", "") or ""
            _sa = b.get("sgm_adi",  "") or ""
            if _sk and _sa:
                sgm_txt = f"{_sk} - {_sa}"
            elif _sa:
                sgm_txt = _sa
            elif _sk:
                sgm_txt = _sk
            else:
                sgm_txt = "—"

            for ci, txt in enumerate([tur_adi, b["donem_adi"], onay, sgm_txt]):
                it = QTableWidgetItem(str(txt))
                it.setTextAlignment(
                    Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
                )
                from PyQt6.QtGui import QColor
                it.setForeground(QColor("#1e293b"))
                self._liste.setItem(ri, ci, it)

    def _liste_bos_goster(self, mesaj: str):
        self._pdf_alan.setText(f"ℹ️  {mesaj}")
        self._pdf_alan.show()
        self._pdf_img_scroll.hide()

    # ── Beyanname seçildi → PDF çek ──────────────────────────────────────────

    def _on_beyanname_sec(self, row: int, _col: int):
        if row < 0 or row >= len(self._beyanlar):
            return
        b = self._beyanlar[row]
        tur_adi = self.BELGE_TUR_ADI.get(b["belge_turu"], b["belge_turu"])
        _sk = b.get("sgm_kodu", "") or ""
        _sa = b.get("sgm_adi",  "") or ""
        if _sk and _sa:
            sgm_ek = f"  •  SGM: {_sk} – {_sa}"
        elif _sa:
            sgm_ek = f"  •  SGM: {_sa}"
        elif _sk:
            sgm_ek = f"  •  SGM: {_sk}"
        else:
            sgm_ek = ""
        self._pdf_baslik.setText(f"{tur_adi}  —  {b['donem_adi']} {b['donem_no']}{sgm_ek}")
        self._pdf_alan.setText("⏳  PDF yükleniyor...")
        self._pdf_alan.show()
        self._pdf_img_scroll.hide()
        self._pdf_ac_btn.setEnabled(False)

        from services.moy_service import get_beyanname_pdf_bytes
        pdf_bytes = get_beyanname_pdf_bytes(self._musterino, b["kayit_no"])

        if not pdf_bytes:
            self._pdf_alan.setText(
                "⚠️  Bu beyanname için PDF verisi bulunamadı.\n"
                "(beyanname_gib tablosunda kayıt yok)"
            )
            return

        # Geçici PDF dosyasına yaz (Workspace İçi)
        import tempfile, os
        workspace_tmp = os.path.expanduser("~/NakitAkim/data/tmp")
        os.makedirs(workspace_tmp, exist_ok=True)
        tmp = tempfile.NamedTemporaryFile(
            suffix=".pdf",
            prefix=f"beyanname_{b['belge_turu']}_{b['donem_no']}_",
            dir=workspace_tmp,
            delete=False
        )
        tmp.write(pdf_bytes)
        tmp.close()
        self._tmp_path = tmp.name
        self._pdf_ac_btn.setEnabled(True)

        # PDF'i resme dönüştürüp önizle (pymupdf varsa)
        self._pdf_goruntule(pdf_bytes)

    def _pdf_goruntule(self, pdf_bytes: bytes):
        """PDF'i resim olarak göster — fitz (PyMuPDF) kullanır."""
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            toplam = doc.page_count

            from PyQt6.QtGui import QImage, QPixmap
            from PyQt6.QtCore import QByteArray
            import io

            # Tüm sayfaları birleştirerek tek uzun görüntü oluştur
            pixmaplar = []
            toplam_y = 0
            maks_x = 0

            for pg_no in range(min(toplam, 10)):  # max 10 sayfa
                pg = doc[pg_no]
                mat = fitz.Matrix(1.8, 1.8)  # 1.8x zoom — yeterli kalite
                clip = pg.get_pixmap(matrix=mat)
                img_data = clip.tobytes("png")
                pm = QPixmap()
                pm.loadFromData(img_data)
                pixmaplar.append(pm)
                toplam_y += pm.height() + 4
                maks_x = max(maks_x, pm.width())

            doc.close()

            if pixmaplar:
                # Tek pixmap olarak birleştir
                from PyQt6.QtGui import QPainter
                combined = QPixmap(maks_x, toplam_y)
                combined.fill()
                p = QPainter(combined)
                y = 0
                for pm in pixmaplar:
                    p.drawPixmap(0, y, pm)
                    y += pm.height() + 4
                p.end()

                self._pdf_img_container.setPixmap(combined)
                self._pdf_img_container.setFixedSize(combined.size())
                self._pdf_alan.hide()
                self._pdf_img_scroll.show()
            else:
                self._pdf_alan.setText("⚠️  PDF sayfası okunamadı.")

        except ImportError:
            # PyMuPDF yüklü değil — PDF bilgisi göster
            kb = len(pdf_bytes) // 1024
            self._pdf_alan.setText(
                f"📄  PDF indirme başarılı ({kb} KB)\n\n"
                f"PDF önizlemesi için PyMuPDF gerekli:\n"
                f"  pip install pymupdf\n\n"
                f"'📂 PDF'i Aç' butonu ile görüntüleyebilirsiniz."
            )
        except Exception as e:
            self._pdf_alan.setText(f"⚠️  PDF görüntülenemedi:\n{e}")

    # ── PDF'i sistem görüntüleyicisinde aç ───────────────────────────────────

    def _pdf_ac(self):
        if not self._tmp_path:
            return
        import subprocess, sys
        try:
            if sys.platform == "darwin":
                subprocess.Popen(["open", self._tmp_path])
            elif sys.platform == "win32":
                import os; os.startfile(self._tmp_path)
            else:
                subprocess.Popen(["xdg-open", self._tmp_path])
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Hata", f"PDF açılamadı: {e}")

    def closeEvent(self, event):
        # Geçici dosyayı temizle
        if self._tmp_path:
            import os
            try:
                os.unlink(self._tmp_path)
            except Exception:
                pass
        super().closeEvent(event)


# ─────────────────────────────────────────────────────────────────────────────
# Bankalar Bakiye Dialog
# womsis_banka tablosu — şube/banka bazlı gelir/gider/net hareketler
# ─────────────────────────────────────────────────────────────────────────────

class BankalaBakiyeDialog(QDialog):
    """
    Bankalar Bakiye kartına tıklandığında açılan detay diyalogu.

    Yapı (eski PHP sistemiyle birebir — iki panelli):
    ┌──────────────────┬──────────────────────────────────────────────┐
    │  Sol panel       │  Sağ panel                                   │
    │  Banka / Şube    │  Seçilen bankaya ait hareket tablosu         │
    │  GROUP BY özet   │  Tarih | Açıklama | Tür | Tutar | Karşı Taraf│
    │  Hepsi + liste   │  Alt bant: Gelir / Gider / Net               │
    └──────────────────┴──────────────────────────────────────────────┘
    """

    _STYLE = """
        QDialog  { background: #F0F4F8; }
        QFrame#sol_panel {
            background: white;
            border: 1px solid #CBD5E1;
            border-radius: 10px;
        }
        QFrame#sag_panel {
            background: white;
            border: 1px solid #CBD5E1;
            border-radius: 10px;
        }
        QLabel#diyalog_baslik {
            font-size: 15px; font-weight: 800; color: #1E3A8A;
        }
        QLabel#panel_baslik {
            font-size: 12px; font-weight: 700; color: #374151;
            padding: 6px 10px;
            background: #EFF6FF;
            border-radius: 6px;
        }
        QLabel#ozet_bant {
            font-size: 11px; color: #374151;
            background: #F8FAFC;
            border: 1px solid #E2E8F0;
            border-radius: 6px;
            padding: 5px 10px;
        }
        QTableWidget {
            background: white;
            gridline-color: #F1F5F9;
            font-size: 11px;
            color: #1F2937;
            border: none;
        }
        QTableWidget::item { padding: 3px 6px; }
        QTableWidget::item:hover    { background: #EFF6FF; }
        QTableWidget::item:selected { background: #DBEAFE; color: #1E3A8A; }
        QTableWidget::item:alternate{ background: #F8FAFC; }
        QHeaderView::section {
            background: #1E3A8A; color: white;
            font-weight: 700; font-size: 10px;
            padding: 5px 6px; border: none;
            border-right: 1px solid #1D4ED8;
        }
        QPushButton#kapat_btn {
            background: #1E3A8A; color: white;
            border: none; border-radius: 6px;
            padding: 6px 22px; font-weight: 700; font-size: 11px;
        }
        QPushButton#kapat_btn:hover { background: #1D4ED8; }
        QPushButton#excel_btn {
            background: #059669; color: white;
            border: none; border-radius: 6px;
            padding: 6px 18px; font-weight: 700; font-size: 11px;
        }
        QPushButton#excel_btn:hover { background: #047857; }
    """

    _SOL_COLS = ["Banka / Şube", "Kayıt", "Gelir (₺)", "Gider (₺)", "Net (₺)"]
    _SAG_COLS = ["Tarih", "Açıklama", "Tür", "Tutar (₺)", "Karşı Taraf", "Kaynak"]

    def __init__(self, userid: int, musterino: int = 1, parent=None):
        super().__init__(parent)
        self._userid       = userid
        self._musterino    = musterino
        self._tum_data: list[dict] = []
        self._secili_sube: str | None = None

        self.setWindowTitle("🏦  Bankalar — Banka / Şube Detayı")
        self.setMinimumSize(1180, 660)
        self.resize(1300, 740)
        self.setStyleSheet(self._STYLE)
        self._build_ui()
        self._load_all()

    # ── UI ───────────────────────────────────────────────────────────────────

    def _build_ui(self):
        from PyQt6.QtWidgets import (
            QVBoxLayout, QHBoxLayout, QLabel, QFrame,
            QTableWidget, QHeaderView, QPushButton, QSplitter,
        )
        from PyQt6.QtCore import Qt

        root = QVBoxLayout(self)
        root.setContentsMargins(14, 12, 14, 12)
        root.setSpacing(10)

        hdr = QLabel("🏦  Bankalar — Şube Bazlı Hareket Tablosu")
        hdr.setObjectName("diyalog_baslik")
        root.addWidget(hdr)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setHandleWidth(6)

        # ─ Sol Panel ─────────────────────────────────────────────
        sol_frame = QFrame()
        sol_frame.setObjectName("sol_panel")
        sol_lay = QVBoxLayout(sol_frame)
        sol_lay.setContentsMargins(10, 10, 10, 10)
        sol_lay.setSpacing(6)

        sol_lbl = QLabel("🏛  Banka / Şube Listesi")
        sol_lbl.setObjectName("panel_baslik")
        sol_lay.addWidget(sol_lbl)

        self._sol_tablo = QTableWidget()
        self._sol_tablo.setColumnCount(len(self._SOL_COLS))
        self._sol_tablo.setHorizontalHeaderLabels(self._SOL_COLS)
        self._sol_tablo.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch
        )
        for ci in range(1, len(self._SOL_COLS)):
            self._sol_tablo.setColumnWidth(ci, 90)
        self._sol_tablo.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._sol_tablo.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._sol_tablo.setAlternatingRowColors(True)
        self._sol_tablo.verticalHeader().setVisible(False)
        self._sol_tablo.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self._sol_tablo.itemSelectionChanged.connect(self._on_sube_secildi)
        sol_lay.addWidget(self._sol_tablo)

        splitter.addWidget(sol_frame)

        # ─ Sağ Panel ─────────────────────────────────────────────
        sag_frame = QFrame()
        sag_frame.setObjectName("sag_panel")
        sag_lay = QVBoxLayout(sag_frame)
        sag_lay.setContentsMargins(10, 10, 10, 10)
        sag_lay.setSpacing(6)

        self._sag_baslik = QLabel("📋  Tüm Hareketler")
        self._sag_baslik.setObjectName("panel_baslik")
        sag_lay.addWidget(self._sag_baslik)

        self._sag_tablo = QTableWidget()
        self._sag_tablo.setColumnCount(len(self._SAG_COLS))
        self._sag_tablo.setHorizontalHeaderLabels(self._SAG_COLS)
        hh = self._sag_tablo.horizontalHeader()
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        for ci, w in enumerate([110, 0, 70, 120, 160, 100]):
            if w:
                self._sag_tablo.setColumnWidth(ci, w)
        self._sag_tablo.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._sag_tablo.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._sag_tablo.setAlternatingRowColors(True)
        self._sag_tablo.verticalHeader().setVisible(False)
        self._sag_tablo.setSortingEnabled(True)
        sag_lay.addWidget(self._sag_tablo)

        self._ozet_bant = QLabel("—")
        self._ozet_bant.setObjectName("ozet_bant")
        self._ozet_bant.setTextFormat(Qt.TextFormat.RichText)
        sag_lay.addWidget(self._ozet_bant)

        splitter.addWidget(sag_frame)
        splitter.setSizes([340, 860])
        root.addWidget(splitter, stretch=1)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        excel_btn = QPushButton("📥 Excel İndir")
        excel_btn.setObjectName("excel_btn")
        excel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        excel_btn.clicked.connect(self._export_excel)
        btn_row.addWidget(excel_btn)
        kapat = QPushButton("  Kapat  ")
        kapat.setObjectName("kapat_btn")
        kapat.clicked.connect(self.accept)
        btn_row.addWidget(kapat)
        root.addLayout(btn_row)

    # ── Veri ─────────────────────────────────────────────────────────────────

    def _load_all(self):
        from db.database import get_connection
        try:
            conn = get_connection()
            rows = conn.execute("""
                SELECT
                    tarih,
                    COALESCE(aciklama, '')       AS aciklama,
                    COALESCE(gelirgider, '')     AS gelirgider,
                    CAST(tutar AS REAL)          AS tutar,
                    COALESCE(sube, '(Şubesiz)') AS sube,
                    COALESCE(faturaunvan, '')    AS faturaunvan,
                    COALESCE(kaynak, '')         AS kaynak
                FROM womsis_banka
                WHERE musterino = ?
                ORDER BY tarih DESC, id DESC
            """, (self._musterino,)).fetchall()
            self._tum_data = [dict(r) if not isinstance(r, dict) else r for r in rows]
            conn.close()
        except Exception as exc:
            self._tum_data = []
            print(f"BankalaBakiyeDialog hata: {exc}")
        self._doldur_sol()
        self._doldur_sag(None)

    def _doldur_sol(self):
        from PyQt6.QtWidgets import QTableWidgetItem
        from PyQt6.QtCore import Qt
        from PyQt6.QtGui import QColor

        gruplar: dict[str, dict] = {}
        for r in self._tum_data:
            sube  = r.get("sube") or "(Şubesiz)"
            tur   = str(r.get("gelirgider", "")).lower()
            tutar = float(r.get("tutar") or 0)
            if sube not in gruplar:
                gruplar[sube] = {"kayit": 0, "gelir": 0.0, "gider": 0.0}
            gruplar[sube]["kayit"] += 1
            if tur == "gelir":
                gruplar[sube]["gelir"] += tutar
            else:
                gruplar[sube]["gider"] += tutar

        self._sol_tablo.setSortingEnabled(False)
        self._sol_tablo.setRowCount(0)

        # Hepsi satırı
        tg = sum(d["gelir"] for d in gruplar.values())
        tgd = sum(d["gider"] for d in gruplar.values())
        self._ekle_sol_satir("🔷  Hepsi", len(self._tum_data), tg, tgd,
                              bold=True, bg="#EFF6FF")

        # Şubeler (gelir büyükten küçüğe)
        for sube, d in sorted(gruplar.items(),
                               key=lambda x: x[1]["gelir"], reverse=True):
            self._ekle_sol_satir(sube, d["kayit"], d["gelir"], d["gider"])

        self._sol_tablo.setSortingEnabled(False)
        self._sol_tablo.selectRow(0)

    def _ekle_sol_satir(self, ad, kayit, gelir, gider, bold=False, bg=None):
        from PyQt6.QtWidgets import QTableWidgetItem
        from PyQt6.QtCore import Qt
        from PyQt6.QtGui import QColor

        net = gelir - gider
        ri  = self._sol_tablo.rowCount()
        self._sol_tablo.insertRow(ri)
        self._sol_tablo.setRowHeight(ri, 30)

        vals   = [ad, str(kayit), f"{gelir:,.2f}", f"{gider:,.2f}", f"{net:+,.2f}"]
        aligns = [Qt.AlignmentFlag.AlignLeft,   Qt.AlignmentFlag.AlignRight,
                  Qt.AlignmentFlag.AlignRight,  Qt.AlignmentFlag.AlignRight,
                  Qt.AlignmentFlag.AlignRight]
        colors = [None, None, "#047857", "#DC2626",
                  "#047857" if net >= 0 else "#DC2626"]

        for ci, (v, align, col) in enumerate(zip(vals, aligns, colors)):
            it = QTableWidgetItem(v)
            it.setTextAlignment(align | Qt.AlignmentFlag.AlignVCenter)
            if col:
                it.setForeground(QColor(col))
            if bold:
                f = it.font(); f.setBold(True); it.setFont(f)
            if bg:
                it.setBackground(QColor(bg))
            self._sol_tablo.setItem(ri, ci, it)

    def _on_sube_secildi(self):
        sel = self._sol_tablo.selectedItems()
        if not sel:
            return
        ad = self._sol_tablo.item(sel[0].row(), 0)
        if not ad:
            return
        metin = ad.text().replace("🔷  ", "").strip()
        sube  = None if metin == "Hepsi" else metin
        self._secili_sube = sube
        self._sag_baslik.setText(
            f"📋  {'Tüm Hareketler' if sube is None else sube}"
        )
        self._doldur_sag(sube)

    def _doldur_sag(self, sube: str | None):
        from PyQt6.QtWidgets import QTableWidgetItem
        from PyQt6.QtCore import Qt
        from PyQt6.QtGui import QColor

        rows = (self._tum_data if sube is None
                else [r for r in self._tum_data
                      if (r.get("sube") or "(Şubesiz)") == sube])

        self._sag_tablo.setSortingEnabled(False)
        self._sag_tablo.setRowCount(0)

        toplam_gelir = toplam_gider = 0.0
        for r in rows:
            tur   = str(r.get("gelirgider", "")).lower()
            tutar = float(r.get("tutar") or 0)
            if tur == "gelir":
                toplam_gelir += tutar
            else:
                toplam_gider += tutar

            ri = self._sag_tablo.rowCount()
            self._sag_tablo.insertRow(ri)
            self._sag_tablo.setRowHeight(ri, 26)

            vals   = [
                str(r.get("tarih", "") or ""),
                str(r.get("aciklama", "") or ""),
                str(r.get("gelirgider", "") or ""),
                f"{tutar:,.2f}",
                str(r.get("faturaunvan", "") or ""),
                str(r.get("kaynak", "") or ""),
            ]
            aligns = [Qt.AlignmentFlag.AlignLeft,   Qt.AlignmentFlag.AlignLeft,
                      Qt.AlignmentFlag.AlignCenter, Qt.AlignmentFlag.AlignRight,
                      Qt.AlignmentFlag.AlignLeft,   Qt.AlignmentFlag.AlignLeft]
            para_col = "#047857" if tur == "gelir" else "#DC2626"

            for ci, (v, align) in enumerate(zip(vals, aligns)):
                it = QTableWidgetItem(v)
                it.setTextAlignment(align | Qt.AlignmentFlag.AlignVCenter)
                if ci in (2, 3):
                    it.setForeground(QColor(para_col))
                self._sag_tablo.setItem(ri, ci, it)

        self._sag_tablo.setSortingEnabled(True)

        net = toplam_gelir - toplam_gider
        ns  = "+" if net >= 0 else ""
        nc  = "#047857" if net >= 0 else "#DC2626"
        self._ozet_bant.setText(
            f"<b>{len(rows):,}</b> kayıt  &nbsp;|&nbsp;  "
            f"Gelir: <b style='color:#047857'>{toplam_gelir:,.2f} ₺</b>  "
            f"Gider: <b style='color:#DC2626'>{toplam_gider:,.2f} ₺</b>  "
            f"Net: <b style='color:{nc}'>{ns}{net:,.2f} ₺</b>"
        )

    # ── Excel Export ──────────────────────────────────────────────────────────

    def _export_excel(self):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import datetime

        sube = self._secili_sube
        rows = (
            self._tum_data if sube is None
            else [r for r in self._tum_data
                  if (r.get("sube") or "(Şubesiz)") == sube]
        )

        dosya_adi = (
            f"banka_hareketleri_{sube.replace('/', '-')}.xlsx"
            if sube else "banka_hareketleri_tum.xlsx"
        )
        path, _ = QFileDialog.getSaveFileName(
            self, "Excel Olarak Kaydet", dosya_adi, "Excel Files (*.xlsx)"
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Banka Hareketleri"

            # Stiller
            font_title   = Font(name="Segoe UI", size=13, bold=True,  color="FF1E3A8A")
            font_sub     = Font(name="Segoe UI", size=9,  italic=True, color="FF4B5563")
            font_header  = Font(name="Segoe UI", size=10, bold=True,  color="FFFFFFFF")
            font_data    = Font(name="Segoe UI", size=10, color="FF1F2937")
            font_total   = Font(name="Segoe UI", size=10, bold=True,  color="FF1F2937")
            fill_header  = PatternFill(start_color="FF1E3A8A", end_color="FF1E3A8A", fill_type="solid")
            fill_total   = PatternFill(start_color="FFEFF6FF", end_color="FFEFF6FF", fill_type="solid")
            fill_gelir   = PatternFill(start_color="FFF0FDF4", end_color="FFF0FDF4", fill_type="solid")
            fill_gider   = PatternFill(start_color="FFFFF5F5", end_color="FFFFF5F5", fill_type="solid")
            border_thin  = Border(
                left=Side(style="thin",  color="FFE5E7EB"),
                right=Side(style="thin", color="FFE5E7EB"),
                top=Side(style="thin",   color="FFE5E7EB"),
                bottom=Side(style="thin",color="FFE5E7EB"),
            )

            simdi = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
            baslik = "Tüm Hareketler" if sube is None else sube

            # Başlık satırları
            ws.append([f"🏦 Bankalar — {baslik}"])
            ws.cell(1, 1).font = font_title

            ws.append([f"Oluşturulma: {simdi}  •  Kayıt: {len(rows):,}"])
            ws.cell(2, 1).font = font_sub

            ws.append([])  # boşluk

            # Kolon başlıkları
            headers = ["Tarih", "Açıklama", "Tür", "Tutar (₺)", "Karşı Taraf", "Kaynak", "Banka / Şube"]
            ws.append(headers)
            HDR_ROW = 4
            for ci, _ in enumerate(headers, start=1):
                cell = ws.cell(HDR_ROW, ci)
                cell.font      = font_header
                cell.fill      = fill_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = border_thin

            # Veri satırları
            toplam_gelir = toplam_gider = 0.0
            for r in rows:
                tur   = str(r.get("gelirgider", "") or "").lower()
                tutar = float(r.get("tutar") or 0)
                if tur == "gelir":
                    toplam_gelir += tutar
                else:
                    toplam_gider += tutar

                row_vals = [
                    str(r.get("tarih", "") or ""),
                    str(r.get("aciklama", "") or ""),
                    str(r.get("gelirgider", "") or ""),
                    tutar,
                    str(r.get("faturaunvan", "") or ""),
                    str(r.get("kaynak", "") or ""),
                    str(r.get("sube", "") or "(Şubesiz)"),
                ]
                ws.append(row_vals)
                ri = ws.max_row
                bg = fill_gelir if tur == "gelir" else fill_gider
                for ci in range(1, len(headers) + 1):
                    cell = ws.cell(ri, ci)
                    cell.font      = font_data
                    cell.border    = border_thin
                    cell.alignment = Alignment(vertical="center")
                    cell.fill      = bg
                    if ci == 4:  # Tutar
                        cell.number_format = '#,##0.00 ₺'
                        cell.alignment = Alignment(horizontal="right", vertical="center")

            # Toplam satırı
            net = toplam_gelir - toplam_gider
            ws.append([])
            tot_row = ws.max_row + 1
            ws.cell(tot_row, 1, "TOPLAM")
            ws.cell(tot_row, 3, "Gelir")
            ws.cell(tot_row, 4, toplam_gelir)
            ws.cell(tot_row + 1, 3, "Gider")
            ws.cell(tot_row + 1, 4, toplam_gider)
            ws.cell(tot_row + 2, 3, "Net")
            ws.cell(tot_row + 2, 4, net)
            for rr in [tot_row, tot_row + 1, tot_row + 2]:
                for ci in range(1, 8):
                    cell = ws.cell(rr, ci)
                    cell.font   = font_total
                    cell.fill   = fill_total
                    cell.border = border_thin
                    if ci == 4:
                        cell.number_format = '#,##0.00 ₺'
                        cell.alignment = Alignment(horizontal="right", vertical="center")

            # Kolon genişlikleri
            col_widths = [14, 42, 10, 16, 30, 14, 28]
            for ci, w in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(ci)].width = w

            # Satır yüksekliği (başlık)
            ws.row_dimensions[HDR_ROW].height = 22

            wb.save(path)

            msg = QMessageBox(self)
            msg.setWindowTitle("✅ Excel Hazır")
            msg.setText(f"Excel dosyası başarıyla kaydedildi:\n{path}")
            msg.setIcon(QMessageBox.Icon.Information)
            msg.exec()

        except Exception as exc:
            import traceback
            QMessageBox.critical(
                self, "Hata",
                f"Excel oluşturulurken hata:\n{exc}\n\n{traceback.format_exc()}"
            )


# ─────────────────────────────────────────────────────────────────────────────
# Kredi Kartı Detay Dialog
# PHP: admin.php → Kredi Kartları kartı → modal (kart özet + ekstre tablo)
# ─────────────────────────────────────────────────────────────────────────────

class KrediKartiDialog(QDialog):

    """
    PHP admin.php 'Kredi Kartları' modalının PyQt6 karşılığı.

    Yapı (admin.php ile birebir):
    - Sol panel : Kart özet listesi (Banka / Kart bazında GROUP BY toplam)
    - Sağ panel: Seçilen karta ait ekstre satırları
    - Tarih filtresi + Filtrele butonu
    - Özet bantı: Kayıt sayısı + Toplam tutar
    """

    _TBL_STYLE = """
        QTableWidget {
            background: white;
            gridline-color: #e2e8f0;
            font-size: 12px;
            color: #1e293b;
            border: 1px solid #bfdbfe;
            border-radius: 8px;
        }
        QTableWidget::item {
            color: #1e293b;
            padding: 4px 6px;
        }
        QTableWidget::item:hover {
            background: #eff6ff;
            color: #1e293b;
        }
        QTableWidget::item:selected {
            background: transparent;
            color: inherit;
        }
        QTableWidget::item:alternate {
            background: #f0f7ff;
            color: #1e293b;
        }
        QHeaderView::section {
            background: #1e40af;
            color: white;
            font-weight: 700;
            font-size: 11px;
            padding: 6px 4px;
            border: none;
            border-right: 1px solid #1d4ed8;
        }
    """

    # Sol panel (ozet_tbl) için sarı/siyah başlık stili
    _SOL_TBL_STYLE = """
        QTableWidget {
            background: white;
            gridline-color: #e2e8f0;
            font-size: 12px;
            color: #1e293b;
            border: 1px solid #fde047;
            border-radius: 8px;
            selection-background-color: #fde047;
            selection-color: #000000;
        }
        QTableWidget::item {
            color: #1e293b;
            padding: 4px 6px;
        }
        QTableWidget::item:hover {
            background: #fef9c3;
            color: #000;
        }
        QTableWidget::item:selected {
            background: #fde047;
            color: #000000;
        }
        QHeaderView::section {
            background: #fde047;
            color: #000000;
            font-weight: 700;
            font-size: 11px;
            padding: 6px 4px;
            border: none;
            border-right: 1px solid #f59e0b;
        }
    """

    _CBS = (
        "QComboBox{background:white;border:1.5px solid #93c5fd;"
        "border-radius:6px;padding:0 8px;font-size:12px;color:#1e293b;}"
        "QComboBox::drop-down{border:none;width:18px;}"
    )
    _DES = (
        "QDateEdit{"
        "  background:white;"
        "  border:1.5px solid #93c5fd;"
        "  border-radius:6px;"
        "  padding:0 6px;"
        "  font-size:12px;"
        "  color:#1e293b;"
        "}"
        "QDateEdit:focus{"
        "  border:1.5px solid #2563eb;"
        "  background:#eff6ff;"
        "}"
        "QDateEdit::drop-down{"
        "  subcontrol-origin:padding;"
        "  subcontrol-position:top right;"
        "  width:24px;"
        "  border-left:1.5px solid #93c5fd;"
        "  border-top-right-radius:6px;"
        "  border-bottom-right-radius:6px;"
        "  background:#dbeafe;"
        "}"
        "QDateEdit::down-arrow{image:none;width:10px;height:10px;}"
    )

    def __init__(self, userid: int, musterino: int, yil: int, parent=None):
        super().__init__(parent)
        self._userid    = userid
        self._musterino = musterino
        self._yil       = yil
        self._ozet_rows: list[dict] = []
        self._ekstre_rows: list[dict] = []
        self._secili_banka: str = ""
        self._secili_hesapkodu: str = ""
        self._secili_pdf_adi: str = ""
        self._secili_row_idx: int = -1  # Sarı boyama için

        self.setWindowTitle("💳  Kredi Kartları — Detay")
        self.setMinimumSize(1160, 660)
        self.resize(1280, 740)
        self._setup_ui()
        self._load_ozet()

    # ── UI İnşası ─────────────────────────────────────────────

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Başlık bantı ──
        hdr = QFrame()
        hdr.setFixedHeight(72)
        hdr.setStyleSheet(
            "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "stop:0 #1e40af,stop:0.5 #2563eb,stop:1 #0f172a);"
        )
        h = QHBoxLayout(hdr)
        h.setContentsMargins(20, 0, 20, 0)
        ic = QLabel("💳")
        ic.setStyleSheet("font-size:26px;background:transparent;")
        h.addWidget(ic)
        left = QVBoxLayout()
        t1 = QLabel("Kredi Kartları")
        t1.setStyleSheet("color:white;font-size:16px;font-weight:700;background:transparent;")
        t2 = QLabel(f"📅  {self._yil} yılı ekstre verileri  •  SQLite kaydından")
        t2.setStyleSheet("color:rgba(255,255,255,.80);font-size:12px;background:transparent;")
        left.addWidget(t1)
        left.addWidget(t2)
        h.addLayout(left, 1)
        root.addWidget(hdr)

        # ── İçerik: Sol kart özet + Sağ ekstre ──
        body = QFrame()
        body.setStyleSheet("background:#f0f7ff;")
        body_lay = QHBoxLayout(body)
        body_lay.setContentsMargins(12, 12, 12, 12)
        body_lay.setSpacing(12)

        # ── Sol Panel: Kart Özet Listesi ──
        sol = QFrame()
        sol.setFixedWidth(360)
        sol.setStyleSheet(
            "background:white;border-radius:10px;"
            "border:1.5px solid #fde047;"
        )
        sol_lay = QVBoxLayout(sol)
        sol_lay.setContentsMargins(0, 0, 0, 0)
        sol_lay.setSpacing(0)

        # Sol panel başlık + Hepsi butonu
        sol_hdr_row = QHBoxLayout()
        sol_hdr_row.setContentsMargins(0, 0, 0, 0)
        sol_hdr_row.setSpacing(0)
        sol_hdr = QLabel("💳  Ekstre Dosyaları")
        sol_hdr.setFixedHeight(38)
        sol_hdr.setStyleSheet(
            "background:#fde047;color:#000000;font-size:13px;font-weight:700;"
            "padding:0 12px;border-radius:9px 0 0 0;"
        )
        sol_hdr_row.addWidget(sol_hdr, 1)
        self._hepsi_btn = QPushButton("★ Hepsi")
        self._hepsi_btn.setFixedSize(72, 38)
        self._hepsi_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._hepsi_btn.setStyleSheet(
            "QPushButton{background:#1e293b;color:#fde047;font-size:12px;font-weight:700;"
            "border:none;border-radius:0 9px 0 0;}"
            "QPushButton:hover{background:#0f172a;}"
        )
        self._hepsi_btn.clicked.connect(self._on_hepsi)
        sol_hdr_row.addWidget(self._hepsi_btn)
        sol_hdr_frame = QFrame()
        sol_hdr_frame.setLayout(sol_hdr_row)
        sol_lay.addWidget(sol_hdr_frame)

        self._ozet_tbl = QTableWidget()
        self._ozet_tbl.setColumnCount(5)
        self._ozet_tbl.setHorizontalHeaderLabels(["Ekstre Dosyası", "Dönem", "Kayıt", "Harcama (₺)", "Net (₺)"])
        self._ozet_tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._ozet_tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._ozet_tbl.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self._ozet_tbl.verticalHeader().setVisible(False)
        self._ozet_tbl.setAlternatingRowColors(False)
        self._ozet_tbl.setShowGrid(False)
        self._ozet_tbl.horizontalHeader().setStretchLastSection(True)
        self._ozet_tbl.setColumnWidth(0, 110)
        self._ozet_tbl.setColumnWidth(1, 75)
        self._ozet_tbl.setColumnWidth(2, 38)
        self._ozet_tbl.setColumnWidth(3, 85)
        self._ozet_tbl.setStyleSheet(self._SOL_TBL_STYLE)
        self._ozet_tbl.itemSelectionChanged.connect(self._on_secim_degisti)
        sol_lay.addWidget(self._ozet_tbl, 1)

        self._ozet_toplam_lbl = QLabel("")
        self._ozet_toplam_lbl.setFixedHeight(28)
        self._ozet_toplam_lbl.setStyleSheet(
            "background:#fef9c3;border-top:1px solid #fde047;"
            "padding:0 10px;font-size:11px;color:#000000;font-weight:600;"
        )
        sol_lay.addWidget(self._ozet_toplam_lbl)
        body_lay.addWidget(sol)

        # ── Sağ Panel: Ekstre Tablosu ──
        sag = QFrame()
        sag.setStyleSheet(
            "background:white;border-radius:10px;"
            "border:1.5px solid #bfdbfe;"
        )
        sag_lay = QVBoxLayout(sag)
        sag_lay.setContentsMargins(0, 0, 0, 0)
        sag_lay.setSpacing(0)

        # Sağ başlık
        sag_hdr_frame = QFrame()
        sag_hdr_frame.setFixedHeight(38)
        sag_hdr_frame.setStyleSheet(
            "background:#2563eb;border-radius:9px 9px 0 0;"
        )
        sag_hdr_lay = QHBoxLayout(sag_hdr_frame)
        sag_hdr_lay.setContentsMargins(12, 0, 12, 0)
        self._sag_baslik = QLabel("📊  Kart seçin →")
        self._sag_baslik.setStyleSheet(
            "color:white;font-size:13px;font-weight:700;"
        )
        sag_hdr_lay.addWidget(self._sag_baslik, 1)
        sag_lay.addWidget(sag_hdr_frame)

        # Ay seçici + Filtre çubuğu
        filtre_frame = QFrame()
        filtre_frame.setFixedHeight(44)
        filtre_frame.setStyleSheet(
            "background:#eff6ff;border-bottom:1px solid #bfdbfe;"
        )
        filtre_lay = QHBoxLayout(filtre_frame)
        filtre_lay.setContentsMargins(10, 0, 10, 0)
        filtre_lay.setSpacing(8)

        # Ay combo
        filtre_lay.addWidget(self._lbl("Ay:"))
        self._ay_cb = QComboBox()
        self._ay_cb.setFixedSize(90, 30)
        self._ay_cb.setStyleSheet(self._CBS)
        AY_ADLARI = [
            "Hepsi", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
            "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık",
        ]
        for i, ad in enumerate(AY_ADLARI):
            self._ay_cb.addItem(ad, i)
        self._ay_cb.currentIndexChanged.connect(self._on_ay_degis)
        filtre_lay.addWidget(self._ay_cb)

        filtre_lay.addWidget(self._lbl("📅 İlk Tarih:"))
        self._ilk_de = QDateEdit()
        self._ilk_de.setCalendarPopup(True)
        self._ilk_de.setDisplayFormat("dd.MM.yyyy")
        self._ilk_de.setFixedSize(130, 30)
        self._ilk_de.setStyleSheet(self._DES)
        self._ilk_de.setDate(QDate(self._yil, 1, 1))
        filtre_lay.addWidget(self._ilk_de)

        filtre_lay.addWidget(self._lbl("📅 Son Tarih:"))
        self._son_de = QDateEdit()
        self._son_de.setCalendarPopup(True)
        self._son_de.setDisplayFormat("dd.MM.yyyy")
        self._son_de.setFixedSize(130, 30)
        self._son_de.setStyleSheet(self._DES)
        self._son_de.setDate(QDate.currentDate())
        filtre_lay.addWidget(self._son_de)

        filtre_lay.addStretch()

        self._filtre_btn = QPushButton("🔍  Filtrele")
        self._filtre_btn.setFixedSize(100, 30)
        self._filtre_btn.setStyleSheet(
            "QPushButton{background:#2563eb;color:white;border:none;"
            "border-radius:7px;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#1d4ed8;}"
            "QPushButton:disabled{background:#cbd5e1;}"
        )
        self._filtre_btn.setEnabled(False)
        self._filtre_btn.clicked.connect(self._load_ekstre)
        filtre_lay.addWidget(self._filtre_btn)

        self._excel_btn = QPushButton("📥 Excel İndir")
        self._excel_btn.setFixedSize(110, 30)
        self._excel_btn.setStyleSheet(
            "QPushButton{background:#10b981;color:white;border:none;"
            "border-radius:7px;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#059669;}"
        )
        self._excel_btn.clicked.connect(self._export_excel)
        filtre_lay.addWidget(self._excel_btn)

        sag_lay.addWidget(filtre_frame)

        # Özet bantı
        self._ekstre_ozet_lbl = QLabel("💳  Sol panelden bir kart seçin")
        self._ekstre_ozet_lbl.setFixedHeight(26)
        self._ekstre_ozet_lbl.setStyleSheet(
            "background:#dbeafe;border-bottom:1px solid #bfdbfe;"
            "padding:0 10px;font-size:11px;color:#1e40af;"
        )
        self._ekstre_ozet_lbl.setTextFormat(Qt.TextFormat.RichText)
        sag_lay.addWidget(self._ekstre_ozet_lbl)

        # Ekstre tablosu
        SUTUNLAR = [
            ("Tarih",       80),
            ("Açıklama",   320),
            ("Tutar",        110),
            ("Hesap Kodu",   90),
            ("Banka / Kart", 200),
        ]
        self._ekstre_tbl = QTableWidget()
        self._ekstre_tbl.setColumnCount(len(SUTUNLAR))
        self._ekstre_tbl.setHorizontalHeaderLabels([s[0] for s in SUTUNLAR])
        self._ekstre_tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._ekstre_tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._ekstre_tbl.setAlternatingRowColors(True)
        self._ekstre_tbl.verticalHeader().setVisible(False)
        self._ekstre_tbl.setSortingEnabled(True)
        self._ekstre_tbl.horizontalHeader().setStretchLastSection(True)
        for i, (_, w) in enumerate(SUTUNLAR):
            self._ekstre_tbl.setColumnWidth(i, w)
        self._ekstre_tbl.setStyleSheet(self._TBL_STYLE)
        self._ekstre_tbl.itemSelectionChanged.connect(self._hesapla_secili_toplam)
        sag_lay.addWidget(self._ekstre_tbl, 1)

        body_lay.addWidget(sag, 1)
        root.addWidget(body, 1)

        # ── Alt bar ──
        alt = QFrame()
        alt.setFixedHeight(48)
        alt.setStyleSheet("background:white;border-top:1px solid #e2e8f0;")
        a = QHBoxLayout(alt)
        a.setContentsMargins(16, 0, 16, 0)
        
        self._secim_toplam_lbl = QLabel("")
        self._secim_toplam_lbl.setStyleSheet("font-size:13px; font-weight:700; color:#166534; padding-left:10px;")
        a.addWidget(self._secim_toplam_lbl)
        
        a.addStretch()
        kapat = QPushButton("✕  Kapat")
        kapat.setFixedSize(110, 32)
        kapat.setStyleSheet(
            "QPushButton{background:#64748b;color:white;border:none;"
            "border-radius:7px;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#475569;}"
        )
        kapat.clicked.connect(self.accept)
        a.addWidget(kapat)
        root.addWidget(alt)

    # ── Yardımcılar ───────────────────────────────────────────

    @staticmethod
    def _lbl(txt: str) -> QLabel:
        l = QLabel(txt)
        l.setStyleSheet("font-size:12px;color:#1e40af;font-weight:600;")
        return l

    # ── Ay Değişimi ────────────────────────────────────────────

    def _on_ay_degis(self):
        """Ay seçilince ilk/son tarihi otomatik doldurur, kart seçiliyse filtreler."""
        import calendar as _cal
        ay = self._ay_cb.currentData()   # 0=Hepsi, 1=Ocak ... 12=Aralık
        yil = self._yil
        if ay and ay > 0:
            son_gun = _cal.monthrange(yil, ay)[1]
            self._ilk_de.setDate(QDate(yil, ay, 1))
            self._son_de.setDate(QDate(yil, ay, son_gun))
        else:
            self._ilk_de.setDate(QDate(yil, 1, 1))
            self._son_de.setDate(QDate.currentDate())
        if self._secili_banka:
            self._load_ekstre()

    # ── Kart Özet Yükleme ─────────────────────────────────────────

    def _load_pdf_listesi(self):
        """PDF Dönemleri sekmesini doldurur."""
        from services.dashboard_service import get_kredi_karti_pdf_listesi
        from PyQt6.QtGui import QColor, QFont as QF
        rows = get_kredi_karti_pdf_listesi(self._userid, self._musterino)
        self._pdf_tbl.setRowCount(0)
        toplam_kayit = 0
        toplam_tutar = 0.0
        for r in rows:
            pdf_adi  = str(r.get("pdf_adi", "") or "")
            kayit    = int(r.get("kayit_sayisi", 0) or 0)
            tutar    = float(r.get("toplam", 0) or 0)
            ilk      = str(r.get("ilk_tarih", "") or "")
            son      = str(r.get("son_tarih", "") or "")
            toplam_kayit += kayit
            toplam_tutar += tutar

            ri = self._pdf_tbl.rowCount()
            self._pdf_tbl.insertRow(ri)
            self._pdf_tbl.setRowHeight(ri, 26)

            # PDF adı
            it0 = QTableWidgetItem(pdf_adi)
            it0.setToolTip(f"{pdf_adi}\n{ilk} → {son}")
            it0.setForeground(QColor("#1e293b"))

            # Kayıt sayısı
            it1 = QTableWidgetItem(str(kayit))
            it1.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            it1.setForeground(QColor("#374151"))

            # Tutar
            it2 = QTableWidgetItem(f"{tutar:,.0f}")
            it2.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            it2.setForeground(QColor("#1d4ed8"))
            it2.setFont(QF("", -1, QF.Weight.Bold))

            for ci, it in enumerate([it0, it1, it2]):
                self._pdf_tbl.setItem(ri, ci, it)

        self._pdf_sayac_lbl.setText(
            f"📄 {len(rows)} PDF dosyası  •  Toplam {toplam_kayit} kayıt  •  {toplam_tutar:,.0f} ₺"
        )

    def _load_ozet(self):
        from services.dashboard_service import get_kredi_karti_kart_ozet
        self._ozet_rows = get_kredi_karti_kart_ozet(self._userid, self._musterino, self._yil)
        self._ozet_tbl.setSortingEnabled(False)
        self._ozet_tbl.setRowCount(0)

        from PyQt6.QtGui import QColor, QFont as QF
        toplam_borc  = 0.0
        toplam_odeme = 0.0
        toplam_net   = 0.0

        AY_ADI = {
            "01": "Oca", "02": "Şub", "03": "Mar", "04": "Nis",
            "05": "May", "06": "Haz", "07": "Tem", "08": "Ağu",
            "09": "Eyl", "10": "Eki", "11": "Kas", "12": "Ara",
        }

        for row in self._ozet_rows:
            ri = self._ozet_tbl.rowCount()
            self._ozet_tbl.insertRow(ri)
            self._ozet_tbl.setRowHeight(ri, 28)

            # PDF adı ile göster (banka ise fallback)
            pdf_adi    = row.get("pdf_adi") or row.get("banka", "") or ""
            banka      = row.get("banka", "") or ""
            kayit      = int(row.get("kayit_sayisi", 0))
            borc       = float(row.get("borc",  0))
            odeme      = float(row.get("odeme", 0))
            net        = float(row.get("net",   0))
            ilk_tarih  = str(row.get("ilk_tarih", "") or "")
            son_tarih  = str(row.get("son_tarih",  "") or "")
            toplam_borc  += borc
            toplam_odeme += odeme
            toplam_net   += net

            # Dönem: "Oca-Haz 2026" gibi
            def _ay_yil(t: str) -> str:
                """DD.MM.YYYY veya DD/MM/YYYY → Oca 2026"""
                try:
                    sep = "." if "." in t else "/"
                    parts = t.split(sep)
                    return AY_ADI.get(parts[1], parts[1]) + " " + parts[2]
                except Exception:
                    return t[:7] if t else ""

            ilk_str = _ay_yil(ilk_tarih)
            son_str = _ay_yil(son_tarih)
            donem   = ilk_str if ilk_str == son_str else f"{ilk_str}–{son_str}"

            pdf_kisa = pdf_adi[:22] + ("…" if len(pdf_adi) > 22 else "")

            it0 = QTableWidgetItem(pdf_kisa)
            it0.setToolTip(pdf_adi)
            it0.setForeground(QColor("#1e293b"))
            it0.setData(Qt.ItemDataRole.UserRole, ri)

            it1 = QTableWidgetItem(donem)
            it1.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            it1.setForeground(QColor("#6d28d9"))
            it1.setFont(QF("", -1, QF.Weight.Bold))
            it1.setToolTip(f"{ilk_tarih} → {son_tarih}")

            it2 = QTableWidgetItem(str(kayit))
            it2.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            it2.setForeground(QColor("#374151"))

            it3 = QTableWidgetItem(f"{borc:,.0f}")
            it3.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            it3.setForeground(QColor("#1d4ed8"))
            it3.setFont(QF("", -1, QF.Weight.Bold))

            net_clr = "#dc2626" if net < 0 else ("#059669" if net > 0 else "#6b7280")
            it4 = QTableWidgetItem(f"{net:,.0f}")
            it4.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            it4.setForeground(QColor(net_clr))
            it4.setFont(QF("", -1, QF.Weight.Bold))

            for ci, it in enumerate([it0, it1, it2, it3, it4]):
                self._ozet_tbl.setItem(ri, ci, it)

        self._ozet_tbl.setSortingEnabled(True)
        self._ozet_toplam_lbl.setText(
            f"Harcama: {toplam_borc:,.0f} ₺  "
            f"Ödeme: {abs(toplam_odeme):,.0f} ₺  "
            f"Net: {toplam_net:,.0f} ₺"
        )

    # ── Kart Seçimi ─────────────────────────────────────────────

    def _on_hepsi(self):
        """Tüm satırları seç → tüm PDF'lerin ekstresi birleşik görünsün."""
        # itemSelectionChanged defalarca tetiklenmesin diye sinyali blokla
        self._ozet_tbl.blockSignals(True)
        self._ozet_tbl.selectAll()
        self._ozet_tbl.blockSignals(False)
        # Tek seferde çağır
        self._on_secim_degisti()

    def _on_secim_degisti(self):
        """Seçili PDF satırlarının ekstre verilerini birleştirip gösterir."""
        from PyQt6.QtGui import QColor as QC
        selected_rows = sorted(set(idx.row() for idx in self._ozet_tbl.selectedIndexes()))
        if not selected_rows:
            return

        # Seçili → sarı/siyah, diğerleri → beyaz
        for ri in range(self._ozet_tbl.rowCount()):
            selected = ri in selected_rows
            for c in range(self._ozet_tbl.columnCount()):
                it = self._ozet_tbl.item(ri, c)
                if it:
                    if selected:
                        it.setBackground(QC("#fde047"))
                        it.setForeground(QC("#000000"))
                    else:
                        it.setBackground(QC("white"))
                        it.setForeground(QC("#1e293b"))

        # Seçili PDF'lerin ekstre verilerini topla
        from services.dashboard_service import get_kredi_karti_ekstre_detay
        ilk_str = self._ilk_de.date().toString("dd.MM.yyyy")
        son_str = self._son_de.date().toString("dd.MM.yyyy")

        birlesik: list = []
        etiketler = []
        for ri in selected_rows:
            if ri >= len(self._ozet_rows):
                continue
            row_data = self._ozet_rows[ri]
            pdf_adi = row_data.get("pdf_adi") or ""
            banka   = row_data.get("banka", "") or ""
            satirlar = get_kredi_karti_ekstre_detay(
                userid=self._userid, musterino=self._musterino,
                banka=banka, yil=self._yil,
                ilk_tarih=ilk_str, son_tarih=son_str,
                pdf_adi=pdf_adi or None,
            )
            birlesik.extend(satirlar)
            etiketler.append((pdf_adi or banka)[:30])

        self._ekstre_rows = birlesik
        label = ", ".join(etiketler[:3]) + ("…" if len(etiketler) > 3 else "")
        self._sag_baslik.setText(f"💳  {label}")
        self._filtre_btn.setEnabled(True)
        self._doldur_ekstre(ilk_str, son_str)

    def _on_kart_sec(self, row: int, _col: int):
        pass  # itemSelectionChanged ile yönetiliyor


        if self._secili_row_idx >= 0:
            for c in range(self._ozet_tbl.columnCount()):
                it = self._ozet_tbl.item(self._secili_row_idx, c)
                if it:
                    it.setBackground(QColor("white"))
                    it.setForeground(QColor("#1e293b"))
                    f = it.font()
                    f.setBold(False)
                    it.setFont(f)
        
        # Yeni seçili satırı SARI / Siyah yap
        self._secili_row_idx = row
        for c in range(self._ozet_tbl.columnCount()):
            it = self._ozet_tbl.item(row, c)
            if it:
                it.setBackground(QColor("#fde047"))
                it.setForeground(QColor("#000000"))
                f = it.font()
                f.setBold(True)
                it.setFont(f)
        
        row_data = self._ozet_rows[row]
        if row_data.get("banka") == "sanal_pos":
            dlg = SanalPosDialog(self._userid, self._musterino, self._yil, self)
            dlg.exec()
        else:
            self._secili_banka     = row_data.get("banka", "") or ""
            self._secili_hesapkodu = ""
            self._secili_pdf_adi   = row_data.get("pdf_adi") or ""
            label = self._secili_pdf_adi or self._secili_banka
            label_kisa = label[:45] + ("…" if len(label) > 45 else "")
            self._sag_baslik.setText(f"💳  {label_kisa}")
            self._filtre_btn.setEnabled(True)
            self._load_ekstre()

    # ── Ekstre Yükleme ─────────────────────────────────────────

    def _load_ekstre(self):
        if not self._secili_pdf_adi and not self._secili_banka:
            return

        # Tablo hemen temizlenir — kullanıcı eski veri görmesin
        self._ekstre_tbl.setRowCount(0)

        from services.dashboard_service import get_kredi_karti_ekstre_detay
        ilk_str = self._ilk_de.date().toString("dd.MM.yyyy")
        son_str = self._son_de.date().toString("dd.MM.yyyy")

        self._ekstre_rows = get_kredi_karti_ekstre_detay(
            userid=self._userid,
            musterino=self._musterino,
            banka=self._secili_banka,
            yil=self._yil,
            ilk_tarih=ilk_str,
            son_tarih=son_str,
            pdf_adi=self._secili_pdf_adi or None,
        )
        self._doldur_ekstre(ilk_str, son_str)

    def _hesapla_secili_toplam(self):
        from utils.format import fmt_para
        ranges = self._ekstre_tbl.selectedRanges()
        toplam = 0.0
        for r in ranges:
            for row in range(r.topRow(), r.bottomRow() + 1):
                item = self._ekstre_tbl.item(row, 2)
                if item:
                    # Item text looks like "1.234,56 ₺" or "-1.234,56 ₺"
                    # Table uses standard python formatting: f"{val:,.2f}" -> "1,234.56"
                    # So we just remove the comma to get "1234.56"
                    txt = item.text().replace(",", "")
                    try:
                        toplam += float(txt)
                    except ValueError:
                        pass
        
        if toplam != 0:
            self._secim_toplam_lbl.setText(f"Seçilen Toplam: {fmt_para(toplam)}")
        else:
            self._secim_toplam_lbl.setText("")

    def _export_excel(self):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        
        if self._ekstre_tbl.rowCount() == 0:
            QMessageBox.information(self, "Bilgi", "Dışa aktarılacak veri yok.")
            return
            
        dosya_adi, _ = QFileDialog.getSaveFileName(
            self, "Excel Olarak Kaydet", "kredi_karti_hareketleri.xlsx", "Excel Files (*.xlsx)"
        )
        if not dosya_adi:
            return
            
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            QMessageBox.critical(self, "Eksik Kütüphane", "openpyxl modülü eksik. Terminalde 'pip install openpyxl' yazarak yükleyebilirsiniz.")
            return
            
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Kredi Kartı Hareketleri"
            
            # Write headers
            headers = [self._ekstre_tbl.horizontalHeaderItem(i).text() for i in range(self._ekstre_tbl.columnCount())]
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
            
            # Write data
            for r in range(self._ekstre_tbl.rowCount()):
                row_data = []
                for c in range(self._ekstre_tbl.columnCount()):
                    it = self._ekstre_tbl.item(r, c)
                    val = it.text() if it else ""
                    
                    if c == 2 and val:
                        # Remove thousands separator to convert to float
                        txt = val.replace(",", "")
                        try:
                            val = float(txt)
                        except ValueError:
                            pass
                            
                    row_data.append(val)
                ws.append(row_data)
                
            # Tutar sütununu (C sütunu = 3. sütun) sayısal olarak formatla
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                
            wb.save(dosya_adi)
            QMessageBox.information(self, "Başarılı", f"Excel dosyası başarıyla kaydedildi:\n{dosya_adi}")
        except Exception as e:
            import traceback
            QMessageBox.critical(self, "Hata", f"Excel kaydedilemedi:\n{e}\n{traceback.format_exc()}")

    def _doldur_ekstre(self, ilk_str: str, son_str: str):
        from PyQt6.QtGui import QColor, QFont as QF
        self._ekstre_tbl.setSortingEnabled(False)
        self._ekstre_tbl.setRowCount(0)

        toplam_borc  = 0.0
        toplam_odeme = 0.0

        for row in self._ekstre_rows:
            ri = self._ekstre_tbl.rowCount()
            self._ekstre_tbl.insertRow(ri)
            self._ekstre_tbl.setRowHeight(ri, 26)

            tarih      = row.get("tarih", "")
            aciklama   = row.get("aciklama", "") or ""
            tutar_val  = float(row.get("alinan_tutar1") or 0)
            hesap_kodu = row.get("hesapKodu", "") or ""
            banka      = row.get("Banka", "") or ""
            tur        = row.get("tur", "borc")  # 'borc' | 'odeme'

            if tutar_val < 0:
                toplam_odeme += tutar_val
            else:
                toplam_borc  += tutar_val

            # Satır rengi: ödeme satırları kırmızı tonı, harcama satırları normal
            is_odeme = tutar_val < 0
            row_bg   = "#fff1f2" if is_odeme else None   # çok hafif kırmızı
            tutar_clr = "#dc2626" if is_odeme else "#1d4ed8"
            tutar_txt = f"{tutar_val:,.2f}"  # negatif işaretli kalacak

            ack_kisa = aciklama[:80] + ("…" if len(aciklama) > 80 else "")

            vals = [tarih, ack_kisa, tutar_txt, hesap_kodu, banka]
            aligns = [
                Qt.AlignmentFlag.AlignCenter,
                Qt.AlignmentFlag.AlignLeft,
                Qt.AlignmentFlag.AlignRight,
                Qt.AlignmentFlag.AlignCenter,
                Qt.AlignmentFlag.AlignLeft,
            ]
            colors = ["#374151", "#1e293b", tutar_clr, "#6b7280", "#374151"]
            bolds  = [False, False, True, False, False]

            for ci, (val, aln, clr, bold) in enumerate(zip(vals, aligns, colors, bolds)):
                it = QTableWidgetItem(str(val))
                if ci == 1:
                    it.setToolTip(aciklama)
                it.setTextAlignment(aln | Qt.AlignmentFlag.AlignVCenter)
                it.setForeground(QColor(clr))
                if bold:
                    it.setFont(QF("", -1, QF.Weight.Bold))
                # Ödeme satırı arka planı
                if row_bg:
                    it.setBackground(QColor(row_bg))
                self._ekstre_tbl.setItem(ri, ci, it)

        self._ekstre_tbl.setSortingEnabled(True)

        net = toplam_borc + toplam_odeme
        self._ekstre_ozet_lbl.setText(
            f"📅 {ilk_str} — {son_str}  │  "
            f"<b>{len(self._ekstre_rows)}</b> kayıt  │  "
            f"Harcama: <b style='color:#1d4ed8'>{toplam_borc:,.2f} ₺</b>  "
            f"Ödeme: <b style='color:#dc2626'>{abs(toplam_odeme):,.2f} ₺</b>  "
            f"Net: <b>{net:,.2f} ₺</b>"
        )


# ─────────────────────────────────────────────────────────────────────────────
# Sanal Pos Hareketleri Dialog
# PHP: lib/panelparcalari/admin/admin.php  → #sanalPosModal
#      sabit/js/admin_dashboard.js         → spHareketleriYukle()
# 14 sütunlu DataTable + İşlem/Ödeme/Fark özet bantları
# ─────────────────────────────────────────────────────────────────────────────

class SanalPosDialog(QDialog):
    """
    PHP admin.php #sanalPosModal + admin_dashboard.js spHareketleriYukle()
    → PyQt6 karşılığı.

    Üst bölüm : Başlık + '📦 PayTR Veritabanı' badge + Tarih filtresi
               + İşlem / Ödeme / Fark toplam bantları
    Alt bölüm : 14 sütunlu hareket tablosu (paytr SQLite tablosu)
    """

    SUTUNLAR = [
        ("İşlem Tarihi",    "islemtarihi",    120, Qt.AlignmentFlag.AlignCenter),
        ("Sipariş No",      "siparisno",      140, Qt.AlignmentFlag.AlignLeft),
        ("İşlem Tutarı",   "islemtutari",    110, Qt.AlignmentFlag.AlignRight),
        ("Ödeme Tutarı",   "odemetutari",    110, Qt.AlignmentFlag.AlignRight),
        ("Kur",            "kur",             55,  Qt.AlignmentFlag.AlignCenter),
        ("Mağaza No",      "magazano",        90,  Qt.AlignmentFlag.AlignCenter),
        ("Net Tutar",      "nettutar",        100, Qt.AlignmentFlag.AlignRight),
        ("Kesinti Tutarı", "kesintitutari",   110, Qt.AlignmentFlag.AlignRight),
        ("Kesinti Oranı",  "kesintiorani",     90, Qt.AlignmentFlag.AlignCenter),
        ("Kart Markası",   "kartmarkasi",     100, Qt.AlignmentFlag.AlignCenter),
        ("Kart No",        "kartno",          110, Qt.AlignmentFlag.AlignCenter),
        ("Ödeme Tipi",     "odemetipi",       100, Qt.AlignmentFlag.AlignCenter),
        ("Kart Tipi",      "karttipi",         90, Qt.AlignmentFlag.AlignCenter),
        ("Taksit Sayısı", "taksitsayisi",      80, Qt.AlignmentFlag.AlignCenter),
    ]

    def __init__(self, userid: int, musterino: str, yil: int, parent=None):
        super().__init__(parent)
        self._userid    = userid
        self._musterino = str(musterino)
        self._yil       = yil
        self._rows: list[dict] = []
        self._toplam_islem = 0.0
        self._toplam_odeme = 0.0

        self.setWindowTitle("💳  Sanal Pos Hareketleri — PayTR")
        self.setMinimumSize(1280, 740)
        self.resize(1380, 800)
        self._setup_ui()
        self._load()

    # ── UI İnşası ──────────────────────────────────────────────────────────

    def _setup_ui(self):
        from datetime import date
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Üst bant (gri arka plan — PHP'deki #f8f9fa) ──────────────────
        top = QFrame()
        top.setStyleSheet("background:#f8f9fa;border-bottom:1px solid #dee2e6;")
        top_lay = QVBoxLayout(top)
        top_lay.setContentsMargins(16, 12, 16, 12)
        top_lay.setSpacing(8)

        # Başlık satırı
        baslik_row = QHBoxLayout()
        baslik_lbl = QLabel("Sanal Pos Hareketleri")
        baslik_lbl.setStyleSheet(
            "font-size:16px;font-weight:700;color:#212529;"
        )
        baslik_row.addWidget(baslik_lbl)

        # Kaynak badge — PHP: #spKaynakBadge = '📦 PayTR Veritabanı'
        self._badge_lbl = QLabel("📦 PayTR Veritabanı")
        self._badge_lbl.setStyleSheet(
            "background:#212121;color:#fff;font-size:11px;font-weight:600;"
            "border-radius:10px;padding:2px 10px;"
        )
        self._badge_lbl.setFixedHeight(22)
        baslik_row.addWidget(self._badge_lbl)
        baslik_row.addStretch()
        top_lay.addLayout(baslik_row)

        # Filtre satırı
        filtre_row = QHBoxLayout()
        filtre_row.setSpacing(10)

        _DE = (
            "QDateEdit{"
            "  background:white;"
            "  border:1px solid #ced4da;"
            "  border-radius:4px;"
            "  padding:3px 6px;"
            "  font-size:12px;"
            "  color:#212529;"
            "}"
            "QDateEdit:focus{"
            "  border:1px solid #6366f1;"
            "  background:#f8f7ff;"
            "}"
            "QDateEdit::drop-down{"
            "  subcontrol-origin:padding;"
            "  subcontrol-position:top right;"
            "  width:24px;"
            "  border-left:1px solid #ced4da;"
            "  border-top-right-radius:4px;"
            "  border-bottom-right-radius:4px;"
            "  background:#f0f0f0;"
            "}"
            "QDateEdit::down-arrow{"
            "  image:none;"
            "  width:10px;height:10px;"
            "}"
        )
        today = date.today()
        jan1  = date(today.year, 1, 1)

        filtre_row.addWidget(self._lbl("📅 İlk Tarih:"))
        self._ilk_de = QDateEdit()
        self._ilk_de.setCalendarPopup(True)
        self._ilk_de.setDisplayFormat("dd.MM.yyyy")
        self._ilk_de.setFixedHeight(30)
        self._ilk_de.setFixedWidth(130)
        self._ilk_de.setDate(QDate(jan1.year, jan1.month, jan1.day))
        self._ilk_de.setStyleSheet(_DE)
        filtre_row.addWidget(self._ilk_de)

        filtre_row.addWidget(self._lbl("📅 Son Tarih:"))
        self._son_de = QDateEdit()
        self._son_de.setCalendarPopup(True)
        self._son_de.setDisplayFormat("dd.MM.yyyy")
        self._son_de.setFixedHeight(30)
        self._son_de.setFixedWidth(130)
        self._son_de.setDate(QDate(today.year, today.month, today.day))
        self._son_de.setStyleSheet(_DE)
        filtre_row.addWidget(self._son_de)

        self._listele_btn = QPushButton("Listele")
        self._listele_btn.setFixedHeight(30)
        self._listele_btn.setFixedWidth(80)
        self._listele_btn.setStyleSheet(
            "QPushButton{background:#212121;color:white;border:none;"
            "border-radius:4px;font-size:12px;font-weight:700;}"
            "QPushButton:hover{background:#343a40;}"
        )
        self._listele_btn.clicked.connect(self._load)
        filtre_row.addWidget(self._listele_btn)

        self._excel_btn = QPushButton("📥 Excel İndir")
        self._excel_btn.setFixedHeight(30)
        self._excel_btn.setFixedWidth(110)
        self._excel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._excel_btn.setStyleSheet(
            "QPushButton{background:#10b981;color:white;border:none;"
            "border-radius:4px;font-size:12px;font-weight:700;}"
            "QPushButton:hover{background:#059669;}"
        )
        self._excel_btn.clicked.connect(self._export_excel)
        filtre_row.addWidget(self._excel_btn)

        filtre_row.addStretch()
        top_lay.addLayout(filtre_row)
        root.addWidget(top)

        # ── Özet bant (İşlem / Ödeme / Fark) ─────────────────────────────
        # PHP: #spIslemToplam / #spOdemeToplam / #spFarkToplam
        ozet = QFrame()
        ozet.setFixedHeight(64)
        ozet.setStyleSheet(
            "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "stop:0 #1a1a2e,stop:0.5 #16213e,stop:1 #0f3460);"
        )
        oz_lay = QHBoxLayout(ozet)
        oz_lay.setContentsMargins(20, 0, 20, 0)
        oz_lay.setSpacing(40)

        def _blok(label: str, attr: str, clr: str):
            col = QVBoxLayout()
            col.setSpacing(1)
            lbl_top = QLabel(label)
            lbl_top.setStyleSheet(
                "font-size:10px;color:rgba(255,255,255,.65);background:transparent;"
            )
            lbl_val = QLabel("-")
            lbl_val.setStyleSheet(
                f"font-size:18px;font-weight:700;color:{clr};background:transparent;"
            )
            col.addWidget(lbl_top)
            col.addWidget(lbl_val)
            setattr(self, attr, lbl_val)
            return col

        oz_lay.addLayout(_blok("İşlem Toplamı",  "_islem_lbl", "#dc3545"))
        oz_lay.addLayout(_blok("Ödeme Toplamı",  "_odeme_lbl", "#28a745"))
        oz_lay.addLayout(_blok("Fark",            "_fark_lbl",  "#ffffff"))
        oz_lay.addStretch()

        self._kayit_chip = QLabel("")
        self._kayit_chip.setStyleSheet(
            "background:rgba(255,255,255,.12);color:white;font-size:11px;"
            "font-weight:600;border-radius:10px;padding:2px 10px;"
        )
        oz_lay.addWidget(self._kayit_chip, 0, Qt.AlignmentFlag.AlignVCenter)
        root.addWidget(ozet)

        # ── Tablo ─────────────────────────────────────────────────────────
        self._tbl = QTableWidget()
        self._tbl.setColumnCount(len(self.SUTUNLAR))
        self._tbl.setHorizontalHeaderLabels([s[0] for s in self.SUTUNLAR])
        self._tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._tbl.setAlternatingRowColors(True)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.setSortingEnabled(True)
        hdr = self._tbl.horizontalHeader()
        hdr.setStretchLastSection(False)
        for i, (_, _, w, _) in enumerate(self.SUTUNLAR):
            self._tbl.setColumnWidth(i, w)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self._tbl.setStyleSheet("""
            QTableWidget {
                background: white;
                gridline-color: #e2e8f0;
                font-size: 12px;
                color: #1e293b;
                border: none;
            }
            QTableWidget::item {
                color: #1e293b;
                padding: 3px 6px;
            }
            QTableWidget::item:hover {
                background: #f0f9ff;
                color: #1e293b;
            }
            QTableWidget::item:selected {
                background: #dbeafe;
                color: #1e293b;
            }
            QTableWidget::item:alternate {
                background: #f8fafc;
            }
            QHeaderView::section {
                background: #212121;
                color: white;
                font-weight: 700;
                font-size: 11px;
                padding: 6px 4px;
                border: none;
                border-right: 1px solid #374151;
            }
        """)
        root.addWidget(self._tbl, 1)

        # ── Alt bar ───────────────────────────────────────────────────────
        alt = QFrame()
        alt.setFixedHeight(48)
        alt.setStyleSheet(
            "background:#f8f9fa;border-top:1px solid #dee2e6;"
        )
        a = QHBoxLayout(alt)
        a.setContentsMargins(16, 0, 16, 0)
        self._durum_lbl = QLabel("")
        self._durum_lbl.setStyleSheet("font-size:11px;color:#6c757d;")
        a.addWidget(self._durum_lbl)
        a.addStretch()
        kapat = QPushButton("Kapat")
        kapat.setFixedSize(110, 32)
        kapat.setStyleSheet(
            "QPushButton{background:#dc3545;color:white;border:none;"
            "border-radius:5px;font-size:13px;font-weight:700;}"
            "QPushButton:hover{background:#c82333;}"
        )
        kapat.clicked.connect(self.accept)
        a.addWidget(kapat)
        root.addWidget(alt)

    # ── Yardımcılar ───────────────────────────────────────────────────────

    @staticmethod
    def _lbl(txt: str) -> QLabel:
        l = QLabel(txt)
        l.setStyleSheet("font-size:12px;color:#495057;font-weight:600;")
        return l

    # ── Veri Yükleme ──────────────────────────────────────────────────────

    def _load(self):
        """Tarih aralığına göre paytr tablosundan satırları çeker."""
        from services.paytr_service import get_sanal_pos_hareketleri_db

        self._listele_btn.setEnabled(False)
        self._durum_lbl.setText("⏳  Yükleniyor...")
        self._islem_lbl.setText("Yükleniyor...")
        self._odeme_lbl.setText("...")
        self._fark_lbl.setText("...")

        # Tablo hemen temizlenir — kullanıcı eski veri görmesin
        self._tbl.setRowCount(0)

        ilk_str = self._ilk_de.date().toString("yyyy-MM-dd")
        son_str = self._son_de.date().toString("yyyy-MM-dd")

        result = get_sanal_pos_hareketleri_db(
            self._userid, self._musterino, ilk_str, son_str
        )

        self._listele_btn.setEnabled(True)

        if not result.get("success"):
            self._islem_lbl.setText("Hata")
            self._odeme_lbl.setText("-")
            self._fark_lbl.setText("-")
            self._durum_lbl.setText(f"❌  {result.get('message', 'Bilinmeyen hata')}")
            return

        self._rows = result.get("data", [])
        self._toplam_islem = result.get("toplam_islem", 0.0)
        self._toplam_odeme = result.get("toplam_odeme", 0.0)
        self._doldur(result)

    def _doldur(self, result: dict):
        from PyQt6.QtGui import QColor, QFont as QF
        import re

        # Özet bantları
        toplam_islem = self._toplam_islem
        toplam_odeme = self._toplam_odeme
        fark         = toplam_odeme - toplam_islem

        self._islem_lbl.setText(result.get("toplam_islem_fmt", "-"))
        self._odeme_lbl.setText(result.get("toplam_odeme_fmt", "-"))
        fark_fmt = result.get("toplam_fark_fmt", "-")
        self._fark_lbl.setText(fark_fmt)
        self._fark_lbl.setStyleSheet(
            f"font-size:18px;font-weight:700;background:transparent;"
            f"color:{'#28a745' if fark >= 0 else '#dc3545'};"
        )

        kayit_sayisi = result.get("kayit_sayisi", len(self._rows))
        self._kayit_chip.setText(f"📅 {kayit_sayisi:,} kayıt")

        ilk_qd = self._ilk_de.date()
        son_qd = self._son_de.date()
        self._durum_lbl.setText(
            f"📅 {ilk_qd.toString('dd.MM.yyyy')} — {son_qd.toString('dd.MM.yyyy')}"
            f"   │   {kayit_sayisi:,} kayıt   │   "
            f"İşlem: {result.get('toplam_islem_fmt','-')}  "
            f"Ödeme: {result.get('toplam_odeme_fmt','-')}  "
            f"Fark: {fark_fmt}"
        )

        # Tarih formatlama (DD.MM.YYYY)
        def _tarih_fmt(v) -> str:
            if not v:
                return ""
            s = str(v).strip()
            m = re.match(r'^(\d{1,2})\.(\d{1,2})\.(\d{4})', s)
            if m:
                return f"{m.group(1).zfill(2)}.{m.group(2).zfill(2)}.{m.group(3)}"
            m2 = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
            if m2:
                return f"{m2.group(3)}.{m2.group(2)}.{m2.group(1)}"
            return s

        # Tabloyu doldur
        self._tbl.setSortingEnabled(False)
        self._tbl.setRowCount(0)

        for row in self._rows:
            ri = self._tbl.rowCount()
            self._tbl.insertRow(ri)
            self._tbl.setRowHeight(ri, 24)

            for ci, (_, field, _, align) in enumerate(self.SUTUNLAR):
                raw = row.get(field, "")

                # Sayısal sütunlar — kırmızı/yeşil renklendirme (PHP DataTable render)
                if field in ("islemtutari", "odemetutari", "nettutar", "kesintitutari"):
                    try:
                        val_f = float(raw or 0)
                    except (ValueError, TypeError):
                        val_f = 0.0
                    txt = f"{val_f:,.2f}"
                    clr = (
                        "#dc3545" if field in ("islemtutari", "kesintitutari")
                        else "#28a745"
                    )
                    it = QTableWidgetItem(txt)
                    it.setForeground(QColor(clr))
                    it.setFont(QF("", -1, QF.Weight.Bold))
                    it.setData(Qt.ItemDataRole.UserRole, val_f)  # sayısal sıralama

                elif field == "islemtarihi":
                    txt = _tarih_fmt(raw)
                    it  = QTableWidgetItem(txt)
                    it.setForeground(QColor("#374151"))

                else:
                    txt = str(raw) if raw is not None else ""
                    it  = QTableWidgetItem(txt)
                    it.setForeground(QColor("#374151"))

                it.setTextAlignment(align | Qt.AlignmentFlag.AlignVCenter)
                self._tbl.setItem(ri, ci, it)

        self._tbl.setSortingEnabled(True)

    def _export_excel(self):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import re
        import datetime

        path, _ = QFileDialog.getSaveFileName(
            self, "Excel Olarak Kaydet", "sanal_pos_hareketleri.xlsx", "Excel Files (*.xlsx)"
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sanal Pos"

            # Excel Stylings
            font_title = Font(name="Segoe UI", size=14, bold=True, color="FF212121")
            font_subtitle = Font(name="Segoe UI", size=9, italic=True, color="FF4B5563")
            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFFFF")
            font_data = Font(name="Segoe UI", size=10, color="FF1F2937")
            font_total = Font(name="Segoe UI", size=11, bold=True, color="FF212121")

            fill_header = PatternFill(start_color="FF212121", end_color="FF212121", fill_type="solid") # Dark
            fill_total = PatternFill(start_color="FFF3F4F6", end_color="FFF3F4F6", fill_type="solid")  # Light Gray

            border_thin = Border(
                left=Side(style="thin", color="FFE5E7EB"),
                right=Side(style="thin", color="FFE5E7EB"),
                top=Side(style="thin", color="FFE5E7EB"),
                bottom=Side(style="thin", color="FFE5E7EB")
            )
            border_total = Border(
                top=Side(style="thin", color="FFD1D5DB"),
                bottom=Side(style="double", color="FF212121")
            )

            # 1. Rapor Başlık Bloğu
            ws.append(["Sanal Pos Hareketleri Raporu (PayTR)"])
            ws.cell(row=1, column=1).font = font_title

            ilk_qd = self._ilk_de.date()
            son_qd = self._son_de.date()
            tarih_araligi = f"Tarih Aralığı: {ilk_qd.toString('dd.MM.yyyy')} — {son_qd.toString('dd.MM.yyyy')}"
            ws.append([tarih_araligi])
            ws.cell(row=2, column=1).font = font_subtitle

            ws.append([]) # Boşluk

            # Headers
            headers = [col[0] for col in self.SUTUNLAR]
            ws.append(headers)

            header_row_idx = 4
            for col_idx in range(len(headers)):
                cell = ws.cell(row=header_row_idx, column=col_idx+1)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_thin

            # Sum tracker for columns
            column_sums = {col: 0.0 for col in range(len(self.SUTUNLAR))}

            # Rows writing
            current_row_idx = 5
            for row in range(self._tbl.rowCount()):
                if self._tbl.isRowHidden(row):
                    continue

                row_data = []
                for col in range(self._tbl.columnCount()):
                    item = self._tbl.item(row, col)
                    cell_text = item.text() if item else ""
                    field_name = self.SUTUNLAR[col][1]

                    # Check if it is an amount column
                    if field_name in ("islemtutari", "odemetutari", "nettutar", "kesintitutari"):
                        val_num = 0.0
                        if item:
                            val_data = item.data(Qt.ItemDataRole.UserRole)
                            if val_data is not None:
                                try:
                                    val_num = float(val_data)
                                except (ValueError, TypeError):
                                    val_num = 0.0
                            else:
                                # Fallback parse
                                s = cell_text.replace("₺", "").replace("TL", "").replace("$", "").replace("€", "").replace("+", "").strip()
                                s = s.replace(".", "").replace(",", ".")
                                try:
                                    val_num = float(s)
                                except ValueError:
                                    val_num = 0.0
                        row_data.append(val_num)
                        column_sums[col] += val_num
                    else:
                        row_data.append(cell_text)

                ws.append(row_data)

                # Style active data row
                for col in range(len(row_data)):
                    cell = ws.cell(row=current_row_idx, column=col+1)
                    cell.font = font_data
                    cell.border = border_thin
                    field_name = self.SUTUNLAR[col][1]

                    if field_name in ("islemtutari", "odemetutari", "nettutar", "kesintitutari"):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        # Format is exactly 0.00 with dot decimal separator, no currency symbol, no thousands separator
                        cell.number_format = '0.00'
                    elif field_name == "islemtarihi":
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                current_row_idx += 1

            # 2. Dynamic GENEL TOPLAM Row
            summary_row = []
            for col in range(self._tbl.columnCount()):
                field_name = self.SUTUNLAR[col][1]
                if col == 0:
                    summary_row.append("GENEL TOPLAM")
                elif field_name in ("islemtutari", "odemetutari", "nettutar", "kesintitutari"):
                    summary_row.append(column_sums[col])
                else:
                    summary_row.append("")

            ws.append(summary_row)

            # Style summary row
            for col in range(len(summary_row)):
                cell = ws.cell(row=current_row_idx, column=col+1)
                cell.font = font_total
                cell.fill = fill_total
                cell.border = border_total
                field_name = self.SUTUNLAR[col][1]

                if col == 0:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif field_name in ("islemtutari", "odemetutari", "nettutar", "kesintitutari"):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    # Amount columns use '.' decimal separator, no currency symbol, no thousands separator
                    cell.number_format = '0.00'

            # 3. Auto Width Adjustment
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row in (1, 2, 3): # Skip title headers for width
                        continue
                    if cell.value is not None:
                        if isinstance(cell.value, float):
                            val_str = f"{cell.value:.2f}"
                        else:
                            val_str = str(cell.value)
                        max_len = max(max_len, len(val_str))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            # Set gridlines visible
            ws.views.sheetView[0].showGridLines = True

            wb.save(path)

            msg = QMessageBox(self)
            msg.setWindowTitle("Başarılı")
            msg.setText("Sanal Pos Hareketleri Excel raporu başarıyla kaydedildi!")
            msg.setStyleSheet("""
                QMessageBox { background-color: white; }
                QLabel { color: #1F2937; font-size: 13px; font-weight: 600; min-width: 280px; min-height: 40px; }
                QPushButton { background-color: #212121; color: white; border: none; border-radius: 6px; padding: 6px 18px; font-size: 11px; font-weight: bold; }
                QPushButton:hover { background-color: #343a40; }
            """)
            msg.exec()

        except Exception as e:
            import traceback
            err_msg = f"Excel kaydedilirken hata oluştu:\n{e}\n\nDetay:\n{traceback.format_exc()}"
            print(err_msg)

            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Icon.Critical)
            msg.setWindowTitle("Hata")
            msg.setText(f"Excel raporu oluşturulurken beklenmedik hata oluştu:\n{e}")
            msg.setStyleSheet("""
                QMessageBox { background-color: white; }
                QLabel { color: #DC2626; font-size: 12px; font-weight: 600; min-width: 240px; }
                QPushButton { background-color: #DC2626; color: white; border: none; border-radius: 6px; padding: 6px 16px; }
            """)
            msg.exec()

# ─────────────────────────────────────────────────────────────────────────────
# Fiziksel Pos Hareketleri Dialog
# PHP: lib/panelparcalari/admin/admin.php  → #fizikselPosModal
#      sabit/js/admin_dashboard.js         → fpHareketleriYukle()
# 9 sütunlu DataTable + İşlem/İşyeri Ücreti/Net Tutar özet bantları
# ─────────────────────────────────────────────────────────────────────────────

class FizikselPosDialog(QDialog):
    """
    PHP admin.php #fizikselPosModal + admin_dashboard.js fpHareketleriYukle()
    → PyQt6 karşılığı.

    Üst bölüm : Başlık + '🏪 Womsis Yerel DB' badge + Tarih filtresi
    Özet bant : İşlem Toplamı (kırmızı) | İşyeri Ücreti (turuncu) | Net Tutar (yeşil)
    Alt bölüm : 9 sütunlu hareket tablosu (womsi_pos SQLite tablosu)
    """

    SUTUNLAR = [
        ("İşyeri No",           "isyerino",          100, Qt.AlignmentFlag.AlignLeft),
        ("Cari Hesap",          "carihesap",          140, Qt.AlignmentFlag.AlignLeft),
        ("Hesaba Geçiş Tarihi", "hesabagecistarihi",  130, Qt.AlignmentFlag.AlignCenter),
        ("İşlem Tutarı",        "islemtutari",        110, Qt.AlignmentFlag.AlignRight),
        ("İşlem Tarihi",        "islemtarihi",        110, Qt.AlignmentFlag.AlignCenter),
        ("POS No",              "posno",               90, Qt.AlignmentFlag.AlignCenter),
        ("İşyeri Ücreti",       "isyeritutar",        120, Qt.AlignmentFlag.AlignRight),
        ("Net Tutar",           "nettutar",           100, Qt.AlignmentFlag.AlignRight),
        ("Brand",               "brand",               90, Qt.AlignmentFlag.AlignCenter),
    ]

    def __init__(self, userid: int, musterino: int, parent=None):
        super().__init__(parent)
        self._userid = userid
        self._musterino = musterino
        self._rows: list[dict] = []
        self.setWindowTitle("🏪  Fiziksel Pos Hareketleri — Womsis")
        self.setMinimumSize(1200, 720)
        self.resize(1340, 780)
        self._setup_ui()
        self._load()

    def _setup_ui(self):
        from datetime import date
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Üst bant
        top = QFrame()
        top.setStyleSheet("background:#f8f9fa;border-bottom:1px solid #dee2e6;")
        tl = QVBoxLayout(top)
        tl.setContentsMargins(16, 12, 16, 12)
        tl.setSpacing(8)

        br = QHBoxLayout()
        bl = QLabel("Fiziksel Pos Hareketleri")
        bl.setStyleSheet("font-size:16px;font-weight:700;color:#212529;")
        br.addWidget(bl)
        badge = QLabel("🏪 Womsis Yerel DB")
        badge.setStyleSheet(
            "background:#1a3a5c;color:#fff;font-size:11px;font-weight:600;"
            "border-radius:10px;padding:2px 10px;"
        )
        badge.setFixedHeight(22)
        br.addWidget(badge)
        br.addStretch()
        tl.addLayout(br)

        fr = QHBoxLayout()
        fr.setSpacing(10)
        _DE = (
            "QDateEdit{"
            "  background:white;"
            "  border:1px solid #ced4da;"
            "  border-radius:4px;"
            "  padding:3px 6px;"
            "  font-size:12px;"
            "  color:#212529;"
            "}"
            "QDateEdit:focus{"
            "  border:1px solid #1a3a5c;"
            "  background:#f0f4f8;"
            "}"
            "QDateEdit::drop-down{"
            "  subcontrol-origin:padding;"
            "  subcontrol-position:top right;"
            "  width:24px;"
            "  border-left:1px solid #ced4da;"
            "  border-top-right-radius:4px;"
            "  border-bottom-right-radius:4px;"
            "  background:#f0f0f0;"
            "}"
            "QDateEdit::down-arrow{image:none;width:10px;height:10px;}"
        )
        today = date.today()
        jan1 = date(today.year, 1, 1)

        fr.addWidget(self._lbl("📅 İlk Tarih:"))
        self._ilk_de = QDateEdit()
        self._ilk_de.setCalendarPopup(True)
        self._ilk_de.setDisplayFormat("dd.MM.yyyy")
        self._ilk_de.setFixedHeight(30)
        self._ilk_de.setFixedWidth(130)
        self._ilk_de.setDate(QDate(jan1.year, jan1.month, jan1.day))
        self._ilk_de.setStyleSheet(_DE)
        fr.addWidget(self._ilk_de)

        fr.addWidget(self._lbl("📅 Son Tarih:"))
        self._son_de = QDateEdit()
        self._son_de.setCalendarPopup(True)
        self._son_de.setDisplayFormat("dd.MM.yyyy")
        self._son_de.setFixedHeight(30)
        self._son_de.setFixedWidth(130)
        self._son_de.setDate(QDate(today.year, today.month, today.day))
        self._son_de.setStyleSheet(_DE)
        fr.addWidget(self._son_de)

        self._listele_btn = QPushButton("Listele")
        self._listele_btn.setFixedHeight(30)
        self._listele_btn.setFixedWidth(80)
        self._listele_btn.setStyleSheet(
            "QPushButton{background:#1a3a5c;color:white;border:none;"
            "border-radius:4px;font-size:12px;font-weight:700;}"
            "QPushButton:hover{background:#0d2137;}"
        )
        self._listele_btn.clicked.connect(self._load)
        fr.addWidget(self._listele_btn)
        fr.addStretch()
        tl.addLayout(fr)
        root.addWidget(top)

        # Özet bant — PHP: #fpIslemToplam / #fpIsyeriToplam / #fpNetToplam
        ozet = QFrame()
        ozet.setFixedHeight(64)
        ozet.setStyleSheet(
            "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "stop:0 #1a3a5c,stop:0.5 #0d2137,stop:1 #061422);"
        )
        oz = QHBoxLayout(ozet)
        oz.setContentsMargins(20, 0, 20, 0)
        oz.setSpacing(40)

        def _blok(lbl: str, attr: str, clr: str):
            col = QVBoxLayout()
            col.setSpacing(1)
            h = QLabel(lbl)
            h.setStyleSheet("font-size:10px;color:rgba(255,255,255,.65);background:transparent;")
            v = QLabel("-")
            v.setStyleSheet(f"font-size:18px;font-weight:700;color:{clr};background:transparent;")
            col.addWidget(h)
            col.addWidget(v)
            setattr(self, attr, v)
            return col

        oz.addLayout(_blok("İşlem Toplamı",  "_islem_lbl",  "#dc3545"))
        oz.addLayout(_blok("İşyeri Ücreti",  "_isyeri_lbl", "#e67e22"))
        oz.addLayout(_blok("Net Tutar",      "_net_lbl",    "#28a745"))
        oz.addStretch()

        self._kayit_chip = QLabel("")
        self._kayit_chip.setStyleSheet(
            "background:rgba(255,255,255,.12);color:white;font-size:11px;"
            "font-weight:600;border-radius:10px;padding:2px 10px;"
        )
        oz.addWidget(self._kayit_chip, 0, Qt.AlignmentFlag.AlignVCenter)
        root.addWidget(ozet)

        # Tablo
        self._tbl = QTableWidget()
        self._tbl.setColumnCount(len(self.SUTUNLAR))
        self._tbl.setHorizontalHeaderLabels([s[0] for s in self.SUTUNLAR])
        self._tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._tbl.setAlternatingRowColors(True)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.setSortingEnabled(True)
        hdr = self._tbl.horizontalHeader()
        hdr.setStretchLastSection(False)
        for i, (_, _, w, _) in enumerate(self.SUTUNLAR):
            self._tbl.setColumnWidth(i, w)
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self._tbl.setStyleSheet("""
            QTableWidget { background:white; gridline-color:#e2e8f0;
                font-size:12px; color:#1e293b; border:none; }
            QTableWidget::item { color:#1e293b; padding:3px 6px; }
            QTableWidget::item:hover { background:#f0f9ff; color:#1e293b; }
            QTableWidget::item:selected { background:#dbeafe; color:#1e293b; }
            QTableWidget::item:alternate { background:#f8fafc; }
            QHeaderView::section { background:#1a3a5c; color:white;
                font-weight:700; font-size:11px; padding:6px 4px;
                border:none; border-right:1px solid #0d2137; }
        """)
        root.addWidget(self._tbl, 1)

        # Alt bar
        alt = QFrame()
        alt.setFixedHeight(48)
        alt.setStyleSheet("background:#f8f9fa;border-top:1px solid #dee2e6;")
        a = QHBoxLayout(alt)
        a.setContentsMargins(16, 0, 16, 0)
        self._durum_lbl = QLabel("")
        self._durum_lbl.setStyleSheet("font-size:11px;color:#6c757d;")
        a.addWidget(self._durum_lbl)
        a.addStretch()
        kapat = QPushButton("Kapat")
        kapat.setFixedSize(110, 32)
        kapat.setStyleSheet(
            "QPushButton{background:#dc3545;color:white;border:none;"
            "border-radius:5px;font-size:13px;font-weight:700;}"
            "QPushButton:hover{background:#c82333;}"
        )
        kapat.clicked.connect(self.accept)
        a.addWidget(kapat)
        root.addWidget(alt)

    @staticmethod
    def _lbl(txt: str) -> QLabel:
        l = QLabel(txt)
        l.setStyleSheet("font-size:12px;color:#495057;font-weight:600;")
        return l

    def _load(self):
        """PHP: fpHareketleriYukle() — womsi_pos tablosundan tarih filtreli sorgu."""
        from services.fiziksel_pos_service import get_hareketler
        self._listele_btn.setEnabled(False)
        self._durum_lbl.setText("⏳  Yükleniyor...")
        self._islem_lbl.setText("Yükleniyor...")
        self._isyeri_lbl.setText("...")
        self._net_lbl.setText("...")

        # Tablo hemen temizlenir — kullanıcı eski veri görmesin
        self._tbl.setRowCount(0)

        ilk_str = self._ilk_de.date().toString("yyyy-MM-dd")
        son_str = self._son_de.date().toString("yyyy-MM-dd")

        result = get_hareketler(self._userid, self._musterino, ilk_str, son_str)
        self._listele_btn.setEnabled(True)

        if not result.get("success"):
            self._islem_lbl.setText("Hata")
            self._isyeri_lbl.setText("-")
            self._net_lbl.setText("-")
            self._durum_lbl.setText(f"❌  {result.get('message', 'Bilinmeyen hata')}")
            return

        self._rows = result.get("data", [])
        self._doldur(result)

    def _doldur(self, result: dict):
        from PyQt6.QtGui import QColor, QFont as QF
        # PHP: #fpIslemToplam / #fpIsyeriToplam / #fpNetToplam
        self._islem_lbl.setText(result.get("toplam_islem_fmt",  "-"))
        self._isyeri_lbl.setText(result.get("toplam_isyeri_fmt", "-"))
        self._net_lbl.setText(result.get("toplam_net_fmt",    "-"))

        kayit = result.get("kayit_sayisi", len(self._rows))
        self._kayit_chip.setText(f"📅 {kayit:,} kayıt")

        ilk_qd = self._ilk_de.date()
        son_qd = self._son_de.date()
        self._durum_lbl.setText(
            f"📅 {ilk_qd.toString('dd.MM.yyyy')} — {son_qd.toString('dd.MM.yyyy')}"
            f"   │   {kayit:,} kayıt   │   "
            f"İşlem: {result.get('toplam_islem_fmt','-')}  "
            f"İşyeri: {result.get('toplam_isyeri_fmt','-')}  "
            f"Net: {result.get('toplam_net_fmt','-')}"
        )

        self._tbl.setSortingEnabled(False)
        self._tbl.setRowCount(0)

        for row in self._rows:
            ri = self._tbl.rowCount()
            self._tbl.insertRow(ri)
            self._tbl.setRowHeight(ri, 24)
            for ci, (_, field, _, align) in enumerate(self.SUTUNLAR):
                raw = row.get(field, "")
                if field == "islemtutari":
                    clr, bold = "#dc3545", True
                elif field == "isyeritutar":
                    clr, bold = "#e67e22", True
                elif field == "nettutar":
                    clr, bold = "#28a745", True
                else:
                    clr, bold = "#374151", False

                if field in ("islemtutari", "isyeritutar", "nettutar"):
                    try:
                        val_f = float(raw or 0)
                    except (ValueError, TypeError):
                        val_f = 0.0
                    it = QTableWidgetItem(f"{val_f:,.2f}")
                    it.setData(Qt.ItemDataRole.UserRole, val_f)
                else:
                    it = QTableWidgetItem(str(raw) if raw is not None else "")

                it.setForeground(QColor(clr))
                if bold:
                    it.setFont(QF("", -1, QF.Weight.Bold))
                it.setTextAlignment(align | Qt.AlignmentFlag.AlignVCenter)
                self._tbl.setItem(ri, ci, it)

        self._tbl.setSortingEnabled(True)

        if not self._rows:
            self._tbl.insertRow(0)
            self._tbl.setRowHeight(0, 48)
            empty = QTableWidgetItem(
                "Bu tarih aralığında fiziksel POS hareketi bulunamadı."
            )
            empty.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            empty.setForeground(QColor("#6c757d"))
            self._tbl.setItem(0, 0, empty)
            self._tbl.setSpan(0, 0, 1, len(self.SUTUNLAR))

# ─────────────────────────────────────────────────────────────────────────────
# Maaş Kira Smm Dialog
# PHP: admin_dashboard.js → vmModalAc('Maaş Kira Smm') + vmBuildTable()
#      ajax/ayarlar/vergiMuhtasarGetir.php
# 5 sütunlu DataTable:
#   Dönem | Açıklama | Gayri Resmi Tutar (turuncu) |
#   Vergi Kesinti Tutarı (kırmızı) | Fark (yeşil)
# ─────────────────────────────────────────────────────────────────────────────

class MaasKiraSmmDialog(QDialog):
    """
    PHP admin.php #vergilerModal (maas_kira_smmm modu) + vmBuildTable()
    → PyQt6 karşılığı.

    Üst bölüm : Başlık + Dönem filtresi
    Özet bant : Gayri Resmi (turuncu) | Kesintiler (kırmızı) | Fark (yeşil)
    Alt bölüm : 5 sütunlu hareket tablosu (VergiMuhtasar SQLite)
    """

    SUTUNLAR = [
        # (başlık, alan_adı, genişlik, hizalama, renk)
        ("Dönem",                "donem",        100, Qt.AlignmentFlag.AlignLeft,    None),
        ("Açıklama",             "ack",          200, Qt.AlignmentFlag.AlignLeft,    None),
        ("Gayri Resmi Tutar",    "gaytutar",     140, Qt.AlignmentFlag.AlignRight, "#ff9800"),
        ("Vergi Kesinti Tutarı", "vergkestutar", 140, Qt.AlignmentFlag.AlignRight, "#dc3545"),
        ("Fark",                 "fark",         130, Qt.AlignmentFlag.AlignRight, "#28a745"),
    ]

    def __init__(self, userid: int, musterino: str = None, parent=None):
        super().__init__(parent)
        self._userid    = userid
        self._musterino = musterino
        self._rows: list[dict] = []
        self._all_rows: list[dict] = []

        self.setWindowTitle("💼  Maaş Kira Smm — Vergi Muhtasar")
        self.setMinimumSize(1000, 660)
        self.resize(1160, 720)
        self._setup_ui()
        self._load()

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Üst bant ─────────────────────────────────────────────────────────
        top = QFrame()
        top.setStyleSheet("background:#f8f9fa;border-bottom:1px solid #dee2e6;")
        tl = QVBoxLayout(top)
        tl.setContentsMargins(16, 12, 16, 12)
        tl.setSpacing(8)

        br = QHBoxLayout()
        bl = QLabel("Maaş Kira Smm — Vergi Muhtasar")
        bl.setStyleSheet("font-size:16px;font-weight:700;color:#212529;")
        br.addWidget(bl)
        badge = QLabel("💼 Personel · Kira · Müşavirlik")
        badge.setStyleSheet(
            "background:#1a3a5c;color:#fff;font-size:11px;font-weight:600;"
            "border-radius:10px;padding:2px 12px;"
        )
        badge.setFixedHeight(22)
        br.addWidget(badge)
        br.addStretch()
        tl.addLayout(br)

        # Dönem + Açıklama filtresi
        fr = QHBoxLayout()
        fr.setSpacing(10)
        _CB = (
            "QComboBox{background:white;border:1px solid #ced4da;border-radius:4px;"
            "padding:3px 8px;font-size:12px;color:#212529;min-height:28px;}"
        )
        fr.addWidget(self._lbl("Dönem:"))
        self._donem_cb = QComboBox()
        self._donem_cb.setFixedWidth(120)
        self._donem_cb.setStyleSheet(_CB)
        self._donem_cb.addItem("Tümü", "")
        fr.addWidget(self._donem_cb)

        fr.addWidget(self._lbl("Açıklama:"))
        self._ack_cb = QComboBox()
        self._ack_cb.setFixedWidth(180)
        self._ack_cb.setStyleSheet(_CB)
        self._ack_cb.addItem("Tümü", "")
        fr.addWidget(self._ack_cb)

        self._donem_cb.currentIndexChanged.connect(self._on_filter_change)
        self._ack_cb.currentIndexChanged.connect(self._on_ack_filter)
        
        self._excel_btn = QPushButton("📥 Excel İndir")
        self._excel_btn.setFixedHeight(30)
        self._excel_btn.setFixedWidth(110)
        self._excel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._excel_btn.setStyleSheet(
            "QPushButton{background:#10b981;color:white;border:none;"
            "border-radius:4px;font-size:12px;font-weight:700;}"
            "QPushButton:hover{background:#059669;}"
        )
        self._excel_btn.clicked.connect(self._export_excel)
        fr.addWidget(self._excel_btn)

        fr.addStretch()
        tl.addLayout(fr)
        root.addWidget(top)

        # ── Özet bant — PHP: #vGayriResmiToplam / #vKesintilerToplam / #vFarkToplam
        ozet = QFrame()
        ozet.setFixedHeight(64)
        ozet.setStyleSheet(
            "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "stop:0 #1a3a5c,stop:0.5 #0d2137,stop:1 #061422);"
        )
        oz = QHBoxLayout(ozet)
        oz.setContentsMargins(20, 0, 20, 0)
        oz.setSpacing(40)

        def _blok(lbl: str, attr: str, clr: str):
            col = QVBoxLayout()
            col.setSpacing(1)
            h = QLabel(lbl)
            h.setStyleSheet("font-size:10px;color:rgba(255,255,255,.65);background:transparent;")
            v = QLabel("-")
            v.setStyleSheet(f"font-size:18px;font-weight:700;color:{clr};background:transparent;")
            col.addWidget(h)
            col.addWidget(v)
            setattr(self, attr, v)
            return col

        oz.addLayout(_blok("Gayri Resmi Toplam", "_gay_lbl",  "#ff9800"))   # PHP: #vGayriResmiToplam
        oz.addLayout(_blok("Kesintiler Toplam",  "_verg_lbl", "#dc3545"))   # PHP: #vKesintilerToplam
        oz.addLayout(_blok("Fark",               "_fark_lbl", "#28a745"))   # PHP: #vFarkToplam
        oz.addStretch()

        self._kayit_chip = QLabel("")
        self._kayit_chip.setStyleSheet(
            "background:rgba(255,255,255,.12);color:white;font-size:11px;"
            "font-weight:600;border-radius:10px;padding:2px 10px;"
        )
        oz.addWidget(self._kayit_chip, 0, Qt.AlignmentFlag.AlignVCenter)
        root.addWidget(ozet)

        # ── Tablo ────────────────────────────────────────────────────────────
        self._tbl = QTableWidget()
        self._tbl.setColumnCount(len(self.SUTUNLAR))
        self._tbl.setHorizontalHeaderLabels([s[0] for s in self.SUTUNLAR])
        self._tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._tbl.setAlternatingRowColors(True)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.setSortingEnabled(True)
        hdr = self._tbl.horizontalHeader()
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        for i, (_, _, w, _, _) in enumerate(self.SUTUNLAR):
            if i != 1:
                self._tbl.setColumnWidth(i, w)
        self._tbl.setStyleSheet("""
            QTableWidget { background:white; gridline-color:#e2e8f0;
                font-size:12px; color:#1e293b; border:none; }
            QTableWidget::item { color:#1e293b; padding:3px 6px; }
            QTableWidget::item:hover { background:#fef9f0; color:#1e293b; }
            QTableWidget::item:selected { background:#fef3c7; color:#1e293b; }
            QTableWidget::item:alternate { background:#f8fafc; }
            QHeaderView::section { background:#1a3a5c; color:white;
                font-weight:700; font-size:11px; padding:6px 4px;
                border:none; border-right:1px solid #0d2137; }
        """)
        root.addWidget(self._tbl, 1)

        # ── Alt bar ───────────────────────────────────────────────────────────
        alt = QFrame()
        alt.setFixedHeight(48)
        alt.setStyleSheet("background:#f8f9fa;border-top:1px solid #dee2e6;")
        a = QHBoxLayout(alt)
        a.setContentsMargins(16, 0, 16, 0)
        self._durum_lbl = QLabel("")
        self._durum_lbl.setStyleSheet("font-size:11px;color:#6c757d;")
        a.addWidget(self._durum_lbl)
        a.addStretch()
        kapat = QPushButton("Kapat")
        kapat.setFixedSize(110, 32)
        kapat.setStyleSheet(
            "QPushButton{background:#dc3545;color:white;border:none;"
            "border-radius:5px;font-size:13px;font-weight:700;}"
            "QPushButton:hover{background:#c82333;}"
        )
        kapat.clicked.connect(self.accept)
        a.addWidget(kapat)
        root.addWidget(alt)

    @staticmethod
    def _lbl(txt: str) -> QLabel:
        l = QLabel(txt)
        l.setStyleSheet("font-size:12px;color:#495057;font-weight:600;")
        return l

    # ── Veri yükleme ─────────────────────────────────────────────────────────

    def _load(self):
        """PHP: loadVergilerData() → vergiMuhtasarGetir.php"""
        from services.vergi_muhtasar_service import get_vergi_muhtasar
        donem = self._donem_cb.currentData() or ""
        result = get_vergi_muhtasar(self._userid, musterino=self._musterino, donem=donem)
        if not result.get("success"):
            self._durum_lbl.setText(f"❌ {result.get('message', 'Hata')}")
            return
        self._all_rows = result.get("data", [])
        self._doldur_filtreler(result.get("donemler", []),
                               list({r["ack"] for r in self._all_rows if r.get("ack")}))
        self._rows = self._all_rows
        self._doldur(result)

    def _doldur_filtreler(self, donemler: list[str], acklar: list[str]):
        """PHP: vmDoldurFiltreler() — combobox'ları doldur"""
        self._donem_cb.blockSignals(True)
        self._ack_cb.blockSignals(True)
        cur_don = self._donem_cb.currentData()
        cur_ack = self._ack_cb.currentData()
        self._donem_cb.clear()
        self._donem_cb.addItem("Tümü", "")
        for d in sorted(donemler):
            self._donem_cb.addItem(d, d)
        idx = self._donem_cb.findData(cur_don)
        if idx >= 0:
            self._donem_cb.setCurrentIndex(idx)
        self._ack_cb.clear()
        self._ack_cb.addItem("Tümü", "")
        for a in sorted(acklar):
            self._ack_cb.addItem(a, a)
        idx = self._ack_cb.findData(cur_ack)
        if idx >= 0:
            self._ack_cb.setCurrentIndex(idx)
        self._donem_cb.blockSignals(False)
        self._ack_cb.blockSignals(False)

    def _on_filter_change(self):
        """PHP: $('#vDonemFilter').on('change', loadVergilerData)"""
        self._load()

    def _on_ack_filter(self):
        """PHP: $('#vAckFilter').on('change', ...) — client-side filtre"""
        ack_val = self._ack_cb.currentData() or ""
        if ack_val:
            self._rows = [r for r in self._all_rows if r.get("ack") == ack_val]
        else:
            self._rows = self._all_rows
        self._doldur_from_rows()

    def _doldur(self, result: dict):
        """Servis sonucundan tabloyu doldur."""
        self._rows = result.get("data", [])
        self._doldur_from_rows()

    def _doldur_from_rows(self):
        """PHP: vmBuildTable() + vmToplamGuncelle()"""
        from PyQt6.QtGui import QColor, QFont as QF

        gay_top = verg_top = fark_top = 0.0
        for r in self._rows:
            gay  = float(r.get("gaytutar")     or 0)
            verg = float(r.get("vergkestutar") or 0)
            fark = gay - verg
            gay_top  += gay
            verg_top += verg
            fark_top += abs(fark)

        def _fmt(v): return f"{abs(v):,.2f} ₺".replace(",", "X").replace(".", ",").replace("X", ".")

        # PHP: #vGayriResmiToplam / #vKesintilerToplam / #vFarkToplam
        self._gay_lbl.setText(_fmt(gay_top))
        self._verg_lbl.setText(_fmt(verg_top))
        self._fark_lbl.setText(f"+{_fmt(fark_top)}")

        kayit = len(self._rows)
        self._kayit_chip.setText(f"📊 {kayit} kayıt")
        self._durum_lbl.setText(
            f"{kayit} kayıt  │  "
            f"Gayri Resmi: {_fmt(gay_top)}  "
            f"Kesinti: {_fmt(verg_top)}  "
            f"Fark: +{_fmt(fark_top)}"
        )

        self._tbl.setSortingEnabled(False)
        self._tbl.setRowCount(0)

        for row in self._rows:
            ri = self._tbl.rowCount()
            self._tbl.insertRow(ri)
            self._tbl.setRowHeight(ri, 28)

            for ci, (_, field, _, align, clr) in enumerate(self.SUTUNLAR):
                raw = row.get(field, "")

                if field in ("gaytutar", "vergkestutar", "fark"):
                    try:
                        val_f = float(raw or 0)
                    except (ValueError, TypeError):
                        val_f = 0.0
                    # Fark: abs değer + yeşil (PHP: Math.abs(d))
                    if field == "fark":
                        it = QTableWidgetItem(f"+{_fmt(abs(val_f))}")
                    else:
                        it = QTableWidgetItem(_fmt(val_f))
                    it.setData(Qt.ItemDataRole.UserRole, val_f)
                    it.setFont(QF("", -1, QF.Weight.Bold))
                    it.setForeground(QColor(clr or "#374151"))
                else:
                    it = QTableWidgetItem(str(raw) if raw is not None else "")
                    it.setForeground(QColor("#374151"))

                it.setTextAlignment(align | Qt.AlignmentFlag.AlignVCenter)
                self._tbl.setItem(ri, ci, it)

        self._tbl.setSortingEnabled(True)

        if not self._rows:
            self._tbl.insertRow(0)
            self._tbl.setRowHeight(0, 48)
            empty = QTableWidgetItem("Bu filtrede Vergi Muhtasar kaydı bulunamadı.")
            empty.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            empty.setForeground(QColor("#6c757d"))
            self._tbl.setItem(0, 0, empty)
            self._tbl.setSpan(0, 0, 1, len(self.SUTUNLAR))

    def _export_excel(self):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        path, _ = QFileDialog.getSaveFileName(
            self, "Excel Olarak Kaydet", "maas_kira_smm_raporu.xlsx", "Excel Files (*.xlsx)"
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Maaş Kira SMM"

            # Excel Stylings
            font_title = Font(name="Segoe UI", size=14, bold=True, color="FF1A3A5C")
            font_subtitle = Font(name="Segoe UI", size=9, italic=True, color="FF4B5563")
            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFFFF")
            font_data = Font(name="Segoe UI", size=10, color="FF1F2937")
            font_total = Font(name="Segoe UI", size=11, bold=True, color="FF1A3A5C")

            fill_header = PatternFill(start_color="FF1A3A5C", end_color="FF1A3A5C", fill_type="solid")
            fill_total = PatternFill(start_color="FFF1F5F9", end_color="FFF1F5F9", fill_type="solid")

            border_thin = Border(
                left=Side(style="thin", color="FFE5E7EB"),
                right=Side(style="thin", color="FFE5E7EB"),
                top=Side(style="thin", color="FFE5E7EB"),
                bottom=Side(style="thin", color="FFE5E7EB")
            )
            border_total = Border(
                top=Side(style="thin", color="FF94A3B8"),
                bottom=Side(style="double", color="FF1A3A5C")
            )

            # 1. Title Block
            ws.append(["Maaş / Kira / SMM — Vergi Muhtasar Raporu"])
            ws.cell(row=1, column=1).font = font_title

            donem_val = self._donem_cb.currentText()
            ack_val = self._ack_cb.currentText()
            filtre_str = f"Filtreler - Dönem: {donem_val}, Açıklama: {ack_val}"
            ws.append([filtre_str])
            ws.cell(row=2, column=1).font = font_subtitle

            ws.append([]) # Blank row

            # Headers
            headers = [col[0] for col in self.SUTUNLAR]
            ws.append(headers)

            header_row_idx = 4
            for col_idx in range(len(headers)):
                cell = ws.cell(row=header_row_idx, column=col_idx+1)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_thin

            # Sum tracker for columns
            column_sums = {col: 0.0 for col in range(len(self.SUTUNLAR))}

            # Rows writing
            current_row_idx = 5
            for row in range(self._tbl.rowCount()):
                if self._tbl.isRowHidden(row):
                    continue

                # Check if there is an empty table item indicator
                item_first = self._tbl.item(row, 0)
                if item_first and "kaydı bulunamadı" in item_first.text():
                    continue

                row_data = []
                for col in range(self._tbl.columnCount()):
                    item = self._tbl.item(row, col)
                    cell_text = item.text() if item else ""
                    field_name = self.SUTUNLAR[col][1]

                    # Check if it is an amount column
                    if field_name in ("gaytutar", "vergkestutar", "fark"):
                        val_num = 0.0
                        if item:
                            val_data = item.data(Qt.ItemDataRole.UserRole)
                            if val_data is not None:
                                try:
                                    val_num = float(val_data)
                                except (ValueError, TypeError):
                                    val_num = 0.0
                            else:
                                # Fallback parse
                                s = cell_text.replace("₺", "").replace("TL", "").replace("+", "").replace("-", "").strip()
                                s = s.replace(".", "").replace(",", ".")
                                try:
                                    val_num = float(s)
                                except ValueError:
                                    val_num = 0.0
                        row_data.append(val_num)
                        column_sums[col] += val_num
                    else:
                        row_data.append(cell_text)

                ws.append(row_data)

                # Style active data row
                for col in range(len(row_data)):
                    cell = ws.cell(row=current_row_idx, column=col+1)
                    cell.font = font_data
                    cell.border = border_thin
                    field_name = self.SUTUNLAR[col][1]

                    if field_name in ("gaytutar", "vergkestutar", "fark"):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '#,##0.00'
                    elif field_name == "donem":
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                current_row_idx += 1

            # 2. Dynamic GENEL TOPLAM Row
            summary_row = []
            for col in range(self._tbl.columnCount()):
                field_name = self.SUTUNLAR[col][1]
                if col == 0:
                    summary_row.append("GENEL TOPLAM")
                elif field_name in ("gaytutar", "vergkestutar", "fark"):
                    summary_row.append(column_sums[col])
                else:
                    summary_row.append("")

            ws.append(summary_row)

            # Style summary row
            for col in range(len(summary_row)):
                cell = ws.cell(row=current_row_idx, column=col+1)
                cell.font = font_total
                cell.fill = fill_total
                cell.border = border_total
                field_name = self.SUTUNLAR[col][1]

                if col == 0:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif field_name in ("gaytutar", "vergkestutar", "fark"):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '#,##0.00'

            # 3. Auto Width Adjustment
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row in (1, 2, 3): # Skip title headers for width
                        continue
                    if cell.value is not None:
                        if isinstance(cell.value, float):
                            val_str = f"{cell.value:,.2f}"
                        else:
                            val_str = str(cell.value)
                        max_len = max(max_len, len(val_str))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            # Set gridlines visible
            ws.views.sheetView[0].showGridLines = True

            wb.save(path)

            msg = QMessageBox(self)
            msg.setWindowTitle("Başarılı")
            msg.setText("Maaş Kira SMM Excel raporu başarıyla kaydedildi!")
            msg.setStyleSheet("""
                QMessageBox { background-color: white; }
                QLabel { color: #1F2937; font-size: 13px; font-weight: 600; min-width: 280px; min-height: 40px; }
                QPushButton { background-color: #1a3a5c; color: white; border: none; border-radius: 6px; padding: 6px 18px; font-size: 11px; font-weight: bold; }
                QPushButton:hover { background-color: #0d2137; }
            """)
            msg.exec()

        except Exception as e:
            import traceback
            err_msg = f"Excel kaydedilirken hata oluştu:\n{e}\n\nDetay:\n{traceback.format_exc()}"
            print(err_msg)

            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Icon.Critical)
            msg.setWindowTitle("Hata")
            msg.setText(f"Excel raporu oluşturulurken beklenmedik hata oluştu:\n{e}")
            msg.setStyleSheet("""
                QMessageBox { background-color: white; }
                QLabel { color: #DC2626; font-size: 12px; font-weight: 600; min-width: 240px; }
                QPushButton { background-color: #DC2626; color: white; border: none; border-radius: 6px; padding: 6px 16px; }
            """)
            msg.exec()
