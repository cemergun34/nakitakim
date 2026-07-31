"""
Ayarlar Ekranı — PyQt6
PHP ayarlar.php → Eklentiler tab → E-Fatura Çek + VOMSİS API bölümlerinin karşılığı.
"""
from __future__ import annotations
import os
from pathlib import Path
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QComboBox, QFileDialog, QFrame, QScrollArea,
    QMessageBox, QProgressBar, QDialog, QDialogButtonBox,
    QLineEdit, QDateEdit, QSizePolicy, QGridLayout,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QStackedWidget
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QDate
from PyQt6.QtGui import QFont, QBrush, QColor

from ui.theme import COLORS
from services.efatura_service import (
    import_xml,
    get_alt_hesap_kodlari, get_cari_hesaplar
)
from db.database import get_connection

import datetime
CURRENT_YEAR = datetime.datetime.now().year


# ── Arka plan iş parçacığı ───────────────────────────────────────────────────

class ImportWorker(QThread):
    progress = pyqtSignal(int, int, str, str)  # (current, total, dosya, durum)
    finished = pyqtSignal(int, int, int, int, int)  # (basarili, atlandi, hatali, eslesmez, tip_filtresi)

    def __init__(self, paths: list[str], userid: int, musterino: int = 1):
        super().__init__()
        self.paths = paths
        self.userid = userid
        self.musterino = musterino

    def run(self):
        basarili = atlandi = hatali = eslesmez = tip_filtresi = 0
        for i, path in enumerate(self.paths):
            res = import_xml(path, self.userid, musterino=self.musterino)
            if res.get("skipped") and res.get("skip_reason") == "tip_filtresi":
                tip_filtresi += 1
                durum = "tip_filtresi"
            elif res.get("skipped"):
                atlandi += 1
                durum = "atlandi"
            elif res.get("no_match"):
                eslesmez += 1
                durum = "eslesmez"
            elif res["success"]:
                basarili += 1
                durum = "ok"
            else:
                hatali += 1
                durum = "hata"
            self.progress.emit(i + 1, len(self.paths), os.path.basename(path), durum)
        self.finished.emit(basarili, atlandi, hatali, eslesmez, tip_filtresi)



# ── E-Fatura Çek Kartı ───────────────────────────────────────────────────────

class EFaturaCard(QFrame):
    data_changed = pyqtSignal()   # Import bitince dashboard'u yenilemek için

    def __init__(self, userid: int, musterino: int = 1, parent=None):
        super().__init__(parent)
        self.userid = userid
        self.musterino = musterino
        self._worker: ImportWorker | None = None

        self.setStyleSheet(
            "QFrame{background:white;border:1.5px solid #e2e8f0;"
            "border-radius:14px;}"
        )
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(22, 20, 22, 20)
        root.setSpacing(14)

        # Başlık
        h = QHBoxLayout()
        ic = QLabel("🧾")
        ic.setStyleSheet("font-size:20px;")
        h.addWidget(ic)
        t = QLabel("E-Fatura Çek")
        t.setStyleSheet("font-size:14px;font-weight:700;color:#000000;letter-spacing:.5px;")
        h.addWidget(t)
        h.addStretch()
        root.addLayout(h)

        # Bilgi chip
        self._profil_chip = QLabel("")
        self._profil_chip.setWordWrap(True)
        self._profil_chip.setStyleSheet(
            "background:#e6edfa;color:#000000;border-radius:6px;"
            "padding:9px 12px;font-size:12px;border:none;"
        )
        root.addWidget(self._profil_chip)

        # Butonlar
        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)

        self.aktar_btn = QPushButton("\U0001f4c1  Klasör Seç & Tüm XML'leri Aktar")
        self.aktar_btn.setFixedHeight(38)
        self.aktar_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.aktar_btn.setStyleSheet(self._btn_style("#418def"))
        self.aktar_btn.clicked.connect(self._on_aktar)
        btn_row.addWidget(self.aktar_btn)

        btn_row.addStretch()
        root.addLayout(btn_row)

        # Progress bar
        self.progress = QProgressBar()
        self.progress.setFixedHeight(6)
        self.progress.setTextVisible(False)
        self.progress.hide()
        self.progress.setStyleSheet(
            "QProgressBar{background:#f1f5f9;border-radius:3px;border:none;}"
            "QProgressBar::chunk{background:#418def;border-radius:3px;}"
        )
        root.addWidget(self.progress)

        # Durum etiketi
        self.status_lbl = QLabel("")
        self.status_lbl.setStyleSheet("font-size:12px;color:#000000;")
        self.status_lbl.hide()
        root.addWidget(self.status_lbl)

        # aktar_btn olustuktan sonra profil durumunu guncelle
        self._update_profil_chip()

    def _btn_style(self, color: str) -> str:
        return (f"QPushButton{{background:{color};color:white;border:none;"
                f"border-radius:9px;font-size:13px;font-weight:600;"
                f"padding:0 18px;letter-spacing:.5px;}}"
                f"QPushButton:hover{{opacity:.85;background:{color}cc;}}"
                f"QPushButton:disabled{{background:#cbd5e1;color:#94a3b8;}}")

    def _update_profil_chip(self):
        from services.sirket_service import get_sirket_profili
        p = get_sirket_profili(self.userid)
        if p and (p.get("vergino") or p.get("tckn")):
            vkn = p.get("vergino") or p.get("tckn")
            self._profil_chip.setText(
                f"✅  Aktif Şirket: <b>{p.get('unvan','')}</b>  |  VKN: {vkn}  |  "
                "Gelen/Kesilen ayrımı otomatik yapılacak."
            )
            self._profil_chip.setStyleSheet(
                "background:#dcfce7;color:#166534;border-radius:6px;"
                "padding:9px 12px;font-size:12px;border:none;"
            )
            self.aktar_btn.setEnabled(True)
        else:
            self._profil_chip.setText(
                "⚠️  Şirket profili tanımlanmamış. "
                "Önce 'Hesap & Güvenlik' sekmesinden şirket bilgilerini kaydedin."
            )
            self._profil_chip.setStyleSheet(
                "background:#fef3c7;color:#92400e;border-radius:6px;"
                "padding:9px 12px;font-size:12px;border:none;"
            )
            self.aktar_btn.setEnabled(False)

    def refresh(self):
        self._update_profil_chip()

    def _on_aktar(self):
        from services.sirket_service import get_sirket_profili
        p = get_sirket_profili(self.userid)
        if not p or not (p.get("vergino") or p.get("tckn")):
            QMessageBox.warning(self, "Şirket Profili Yok",
                "Önce 'Hesap & Güvenlik' sekmesinden şirket profilinizi kaydedin.")
            return

        klasor = QFileDialog.getExistingDirectory(
            self, "XML Fatura Klasörü Seç", ""
        )
        if not klasor:
            return

        paths = [
            os.path.join(klasor, f)
            for f in os.listdir(klasor)
            if f.lower().endswith(".xml")
        ]

        if not paths:
            QMessageBox.information(
                self, "Dosya Bulunamadı",
                f"Seçilen klasörde hiç .xml dosyası yok:\n{klasor}"
            )
            return

        self.status_lbl.setText(f"📁  {len(paths)} XML dosyası bulundu, aktarılıyor...")
        self.status_lbl.show()
        self._start_import(paths)

    def _start_import(self, paths: list[str]):
        """İmport worker'ı başlatır — UI'yi bloklamaz."""
        self.aktar_btn.setEnabled(False)
        self.progress.setMaximum(len(paths))
        self.progress.setValue(0)
        self.progress.show()

        self._worker = ImportWorker(paths, self.userid, getattr(self, "musterino", 1))
        self._worker.progress.connect(self._on_progress)
        self._worker.finished.connect(self._on_import_done)
        self._worker.start()

    def _on_progress(self, current: int, total: int, dosya: str, durum: str):
        self.progress.setValue(current)
        icon = {
            "ok": "✔",
            "atlandi": "⏩",
            "hata": "✖",
            "eslesmez": "❓",
            "tip_filtresi": "⛔",
        }.get(durum, "")
        self.status_lbl.setText(f"{icon}  {dosya}  ({current}/{total})")

    def _on_import_done(self, basarili: int, atlandi: int, hatali: int,
                        eslesmez: int, tip_filtresi: int):
        self.progress.hide()
        self.aktar_btn.setEnabled(True)
        parts = []
        if basarili:
            parts.append(f"✔ {basarili} aktarıldı")
        if atlandi:
            parts.append(f"⏩ {atlandi} zaten mevcuttu")
        if tip_filtresi:
            parts.append(f"⛔ {tip_filtresi} iptal/iade atlandı")
        if eslesmez:
            parts.append(f"❓ {eslesmez} VKN eşleşmedi")
        if hatali:
            parts.append(f"✖ {hatali} hatalı")
        self.status_lbl.setText("  ·  ".join(parts) if parts else "Tamamlandı")

        # musterino alanı boş (NULL veya 0) olan tüm faturalara musterino=1 yaz
        try:
            conn = get_connection()
            conn.execute(
                "UPDATE faturalar SET musterino=1 WHERE userid=? AND (musterino IS NULL OR musterino=0)",
                (self.userid,)
            )
            conn.commit()
            conn.close()
        except Exception:
            pass

        # Yeni fatura aktarıldıysa dashboard'u yenile
        if basarili > 0:
            self.data_changed.emit()




# ── VOMSİS API Worker ────────────────────────────────────────────────────────

class VomsisTestWorker(QThread):
    """'Kontrol Et' butonu için arka plan iş parçacığı."""
    result = pyqtSignal(dict)   # {'success': bool, 'message': str, ...}

    def __init__(self, api_base: str, app_key: str, app_secret: str):
        super().__init__()
        self._api_base   = api_base
        self._app_key    = app_key
        self._app_secret = app_secret

    def run(self):
        from services.vomsis_service import vomsis_test_connection
        r = vomsis_test_connection(self._api_base, self._app_key, self._app_secret)
        self.result.emit(r)


class VomsisIsleWorker(QThread):
    """'VOMSİS İşle' butonu — seçilen tarih aralığında hareketleri çekip DB'e yazar."""
    progress = pyqtSignal(str)   # durum metni
    finished = pyqtSignal(dict)  # {'success': bool, 'message': str, 'count': int}

    def __init__(self, api_base: str, app_key: str, app_secret: str,
                 start_dt: datetime.datetime, end_dt: datetime.datetime,
                 userid: int, musterino: int = 1):
        super().__init__()
        self._api_base   = api_base
        self._app_key    = app_key
        self._app_secret = app_secret
        self._start_dt   = start_dt
        self._end_dt     = end_dt
        self._userid     = userid
        self._musterino  = musterino

    def run(self):
        from services.vomsis_service import (
            vomsis_authenticate, vomsis_get_all_transactions_chunked
        )

        self.progress.emit("🔑  Token alınıyor...")
        token, err_msg = vomsis_authenticate(self._api_base, self._app_key, self._app_secret)
        if not token:
            self.finished.emit({
                "success": False,
                "message": err_msg or "Token alınamadı. API bilgilerini kontrol edin.",
                "count": 0
            })
            return

        self.progress.emit("📡  Banka hareketleri çekiliyor...")
        txs = vomsis_get_all_transactions_chunked(
            self._api_base, token, self._start_dt, self._end_dt
        )

        count = 0
        conn = get_connection()
        try:
            for tx in txs:
                # VOMSİS hareket alanlarını hareketler tablosuna yaz
                tarih_raw = tx.get("date") or tx.get("processDate") or ""
                aciklama  = tx.get("description") or tx.get("explanation") or ""
                tutar_raw = tx.get("amount") or tx.get("tryAmount") or 0
                try:
                    tutar = round(float(str(tutar_raw).replace(",", ".")), 2)
                    tutar = -tutar
                except (ValueError, TypeError):
                    tutar = 0.0
                yon       = tx.get("direction") or tx.get("transactionDirection") or ""
                gelir_gider = "gelir" if str(yon).upper() in ("CREDIT", "ALACAK", "+") else "gider"
                borc   = tutar if gelir_gider == "gelir" else 0.0
                alacak = tutar if gelir_gider == "gider" else 0.0
                vomsis_key  = tx.get("id") or tx.get("transactionId") or ""
                banka_sube  = tx.get("bankName") or tx.get("accountName") or tx.get("bank") or ""
                karsi_unvan = tx.get("counterpartyName") or tx.get("receiverName") or tx.get("senderName") or ""

                # Aynı VOMSİS kaydı zaten varsa atla
                if vomsis_key:
                    exists = conn.execute(
                        "SELECT id FROM womsis_banka "
                        "WHERE womsiskey=? AND userid=? AND musterino=? LIMIT 1",
                        (str(vomsis_key), self._userid, self._musterino)
                    ).fetchone()
                    if exists:
                        continue

                conn.execute(
                    """INSERT INTO womsis_banka
                       (tarih, aciklama, gelirgider, tutar, kaynak, womsiskey,
                        userid, sube, faturaunvan, musterino)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (tarih_raw, aciklama, gelir_gider, tutar, "vomsis",
                     str(vomsis_key), self._userid, banka_sube, karsi_unvan,
                     self._musterino)
                )
                count += 1
                self.progress.emit(f"✔  {count} hareket işlendi...")
            conn.commit()
        except Exception as exc:
            conn.rollback()
            self.finished.emit({"success": False, "message": str(exc), "count": count})
            return
        finally:
            conn.close()

        self.finished.emit({
            "success": True,
            "message": f"{count} banka hareketi aktarıldı.",
            "count": count
        })


# ─────────────────────────────────────────────────────────────────────────────
# ── Manuel Toplu İşle — Worker Sınıfları  (PHP: btnTopluBankalarIsle, vb.)  ──
# ─────────────────────────────────────────────────────────────────────────────

class TopluBankalarIsleWorker(QThread):
    """
    PHP: #btnTopluBankalarIsle click → ajax/topluWomIsle.php
    VOMSİS banka hareketlerini 15 günlük batch'ler halinde çekip hareketler
    tablosuna yazar.  VomsisIsleWorker'dan farklı olarak tarihi dışarıdan alır
    ve tüm batch döngüsü boyunca ilerleme bildirir.
    """
    progress = pyqtSignal(str)    # durum metni
    batch_done = pyqtSignal(int, int)   # (tamamlanan_batch, toplam_batch)
    finished = pyqtSignal(dict)   # {'success': bool, 'message': str, 'count': int}

    def __init__(self, api_base: str, app_key: str, app_secret: str,
                 start_dt: datetime.datetime, end_dt: datetime.datetime,
                 userid: int, musterino: int = 1):
        super().__init__()
        self._api_base   = api_base
        self._app_key    = app_key
        self._app_secret = app_secret
        self._start_dt   = start_dt
        self._end_dt     = end_dt
        self._userid     = userid
        self._musterino  = musterino

    def run(self):
        try:
            from services.vomsis_service import (
                vomsis_authenticate, vomsis_get_all_transactions
            )
            self.progress.emit("🔑  VOMSİS token alınıyor...")
            token, err_msg = vomsis_authenticate(
                self._api_base, self._app_key, self._app_secret
            )
            if not token:
                self.finished.emit({
                    "success": False,
                    "message": err_msg or "Token alınamadı.",
                    "count": 0
                })
                return

            # PHP: 15 günlük batch döngüsü  (topluWomIsle.php benzeri)
            batches: list[tuple[datetime.datetime, datetime.datetime]] = []
            cur = self._start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
            while cur <= self._end_dt:
                batch_end = min(
                    cur + datetime.timedelta(days=14),
                    self._end_dt.replace(hour=23, minute=59, second=59)
                )
                batches.append((cur, batch_end))
                cur = batch_end + datetime.timedelta(seconds=1)
                cur = cur.replace(hour=0, minute=0, second=0, microsecond=0)

            total_batches  = len(batches)
            total_inserted = 0
            conn = get_connection()
            try:
                for i, (bs, be) in enumerate(batches):
                    begin_str = bs.strftime("%d-%m-%Y %H:%M:%S")
                    end_str   = be.strftime("%d-%m-%Y %H:%M:%S")
                    self.progress.emit(
                        f"📡  Parça {i+1}/{total_batches}: "
                        f"{bs.strftime('%d.%m.%Y')} → {be.strftime('%d.%m.%Y')} çekiliyor..."
                    )
                    txs = vomsis_get_all_transactions(
                        self._api_base, token, begin_str, end_str
                    )
                    for tx in txs:
                        tarih_raw  = tx.get("date") or tx.get("processDate") or ""
                        aciklama   = tx.get("description") or tx.get("explanation") or ""
                        tutar_raw  = tx.get("amount") or tx.get("tryAmount") or 0
                        try:
                            tutar = round(float(str(tutar_raw).replace(",", ".")), 2)
                            tutar = -tutar
                        except (ValueError, TypeError):
                            tutar = 0.0
                        yon       = tx.get("direction") or tx.get("transactionDirection") or ""
                        gelir_gider = "gelir" if str(yon).upper() in ("CREDIT", "ALACAK", "+") else "gider"
                        vomsis_key  = tx.get("id") or tx.get("transactionId") or ""
                        banka_sube  = tx.get("bankName") or tx.get("accountName") or tx.get("bank") or ""
                        karsi_unvan = tx.get("counterpartyName") or tx.get("receiverName") or tx.get("senderName") or ""

                        if vomsis_key:
                            exists = conn.execute(
                                "SELECT id FROM womsis_banka "
                                "WHERE womsiskey=? AND userid=? AND musterino=? LIMIT 1",
                                (str(vomsis_key), self._userid, self._musterino)
                            ).fetchone()
                            if exists:
                                continue

                        conn.execute(
                            """INSERT INTO womsis_banka
                               (tarih, aciklama, gelirgider, tutar, kaynak, womsiskey,
                                userid, sube, faturaunvan, musterino)
                               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                            (tarih_raw, aciklama, gelir_gider, tutar, "vomsis",
                             str(vomsis_key), self._userid, banka_sube, karsi_unvan,
                             self._musterino)
                        )
                        total_inserted += 1
                    conn.commit()
                    self.batch_done.emit(i + 1, total_batches)
            except Exception as exc:
                conn.rollback()
                self.finished.emit({"success": False, "message": str(exc), "count": total_inserted})
                return
            finally:
                conn.close()

            self.finished.emit({
                "success": True,
                "message": f"{total_inserted} banka hareketi aktarıldı ({total_batches} parça).",
                "count": total_inserted
            })
        except Exception as exc:
            self.finished.emit({"success": False, "message": f"Beklenmeyen hata: {exc}", "count": 0})


class WomsisPosIsleWorker(QThread):
    """
    PHP: #btnGiderlerIsle click → ajax/ayarlar/womsisPosIsle.php
    VOMSİS POS terminal hareketlerini çekip womsi_pos tablosuna yazar.
    PHP womsisPosIsle.php ile birebir alan eşlemesi yapılmıştır.
    """
    progress = pyqtSignal(str)
    finished = pyqtSignal(dict)

    def __init__(self, api_base: str, app_key: str, app_secret: str,
                 start_dt: datetime.datetime, end_dt: datetime.datetime,
                 userid: int, musterino: int = 1):
        super().__init__()
        self._api_base   = api_base
        self._app_key    = app_key
        self._app_secret = app_secret
        self._start_dt   = start_dt
        self._end_dt     = end_dt
        self._userid     = userid
        self._musterino  = musterino

    def run(self):
        try:
            from services.vomsis_service import (
                vomsis_authenticate, vomsis_get_terminals, vomsis_get_terminal_transactions
            )

            CHUNK_DAYS = 14  # PHP: $CHUNK_DAYS = 14

            self.progress.emit("🔑  VOMSİS POS token alınıyor...")
            token, err_msg = vomsis_authenticate(
                self._api_base, self._app_key, self._app_secret
            )
            if not token:
                self.finished.emit({
                    "success": False,
                    "message": err_msg or "Token alınamadı: VOMSİS sunucusuna ulaşılamadı.",
                    "count": 0
                })
                return

            self.progress.emit("📋  Terminal listesi alınıyor...")
            terminals = vomsis_get_terminals(self._api_base, token)
            if not terminals:
                self.finished.emit({
                    "success": False,
                    "message": "Hiç POS terminali bulunamadı.",
                    "count": 0
                })
                return

            # ── PHP: $CHUNK_DAYS = 14 ─────────────────────────────────────────
            # Tarih aralığını 14 günlük parçalara böl
            chunks = []
            current = self._start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
            end_norm = self._end_dt.replace(hour=23, minute=59, second=59)
            while current <= end_norm:
                chunk_end = min(
                    current + datetime.timedelta(days=CHUNK_DAYS - 1),
                    end_norm
                )
                chunks.append({
                    "begin": current.strftime("%d-%m-%Y %H:%M"),
                    "end":   chunk_end.strftime("%d-%m-%Y %H:%M"),
                })
                current = (chunk_end + datetime.timedelta(seconds=1)).replace(
                    hour=0, minute=0, second=0, microsecond=0
                )

            now_str    = datetime.datetime.now().strftime("%d/%m/%Y")
            conn       = get_connection()
            total_ins  = 0
            total_skip = 0

            try:
                for idx, terminal in enumerate(terminals):
                    # ── PHP: $station (terminal nesnesi) ────────────────────
                    tid          = terminal.get("id") or terminal.get("stationId") or ""
                    station_no   = str(terminal.get("station_no") or terminal.get("no") or tid)
                    workplace_no = str(terminal.get("workplace_no") or terminal.get("merchantNo") or "")
                    bank_title   = str(terminal.get("bank_title") or terminal.get("bank_name") or terminal.get("bankTitle") or "")

                    if not tid:
                        continue

                    self.progress.emit(
                        f"💳  Terminal {idx+1}/{len(terminals)} "
                        f"[{station_no}] çekiliyor... ({len(chunks)} dönem)"
                    )

                    for chunk in chunks:
                        try:
                            txs = vomsis_get_terminal_transactions(
                                self._api_base, token, tid,
                                chunk["begin"], chunk["end"]
                            )
                        except Exception as ce:
                            self.progress.emit(
                                f"⚠️  Terminal {station_no} chunk hatası: {ce}"
                            )
                            continue

                        for tx in txs:
                            # ── Alan eşlemeleri — PHP womsisPosIsle.php ile birebir ──
                            # PHP: $tx['workplace'] veya $workplaceNo (station'dan)
                            isyerino = str(
                                tx.get("workplace")   or tx.get("workplaceNo") or
                                tx.get("merchantNo")  or tx.get("merchantId")  or
                                workplace_no          or ""
                            )
                            # PHP: $bankTitle (station'dan)
                            carihesap = str(
                                tx.get("cariHesap")  or tx.get("accountNo") or
                                bank_title           or ""
                            )
                            # PHP: $tx['valor'] ?? $tx['transfer_to_account_date']
                            hesabagecistarihi = str(
                                tx.get("valor")                    or
                                tx.get("transfer_to_account_date") or
                                tx.get("settlementDate")           or
                                tx.get("valueDate")                or ""
                            )
                            # PHP: $tx['date']
                            islemtarihi = str(
                                tx.get("date")            or
                                tx.get("transactionDate") or
                                tx.get("islemTarihi")     or ""
                            )
                            # PHP: $tx['station'] ?? $stationNo
                            posno = str(
                                tx.get("station")    or tx.get("stationNo") or
                                tx.get("terminalId") or station_no or ""
                            )
                            # PHP: $tx['sub_card_type'] ?? $tx['card_type']
                            brand = str(
                                tx.get("sub_card_type") or tx.get("card_type") or
                                tx.get("cardBrand")     or tx.get("brand")     or
                                tx.get("scheme")        or ""
                            )
                            # PHP: $tx['card_number']
                            kartno = str(
                                tx.get("card_number")        or
                                tx.get("maskedCardNumber")   or
                                tx.get("maskedCardNo")       or
                                tx.get("cardNo")             or ""
                            )
                            # PHP: $tx['transaction_type']
                            islemtipi = str(
                                tx.get("transaction_type") or tx.get("transactionType") or
                                tx.get("islemTipi")        or tx.get("type") or ""
                            )
                            # PHP: $tx['description']
                            aciklama = str(tx.get("description") or tx.get("aciklama") or "")[:255]

                            # PHP: $tx['gross_amount'], $tx['commission'], $tx['net_amount']
                            try:
                                islemtutari = round(abs(float(
                                    tx.get("gross_amount") or tx.get("amount") or
                                    tx.get("islemTutari")  or 0
                                )), 2)
                                isyeriucretitutar = round(abs(float(
                                    tx.get("commission")        or tx.get("commissionAmount") or
                                    tx.get("isyeriUcretiTutar") or tx.get("fee") or 0
                                )), 2)
                                nettutar = round(abs(float(
                                    tx.get("net_amount") or tx.get("netAmount") or
                                    tx.get("netTutar")   or tx.get("net") or 0
                                )), 2)
                            except (ValueError, TypeError):
                                islemtutari = isyeriucretitutar = nettutar = 0.0

                            # ── Mükerrer kontrol (PHP: INSERT IGNORE mantığı) ──
                            existing = conn.execute(
                                "SELECT id FROM womsi_pos "
                                "WHERE userid=? AND isyerino=? AND islemtarihi=? "
                                "AND islemtutari=? AND kartno=? LIMIT 1",
                                (self._userid, isyerino, islemtarihi, islemtutari, kartno)
                            ).fetchone()
                            if existing:
                                total_skip += 1
                                continue

                            conn.execute(
                                """INSERT INTO womsi_pos
                                   (userid, musterino, isyerino, carihesap, hesabagecistarihi,
                                    islemtutari, islemtarihi, posno,
                                    isyeriucretitutar, nettutar, brand,
                                    kartno, islemtipi, aciklama, islemtarih)
                                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                                (
                                    self._userid, self._musterino,
                                    isyerino, carihesap, hesabagecistarihi,
                                    islemtutari, islemtarihi, posno,
                                    isyeriucretitutar, nettutar, brand,
                                    kartno, islemtipi, aciklama, now_str
                                )
                            )
                            total_ins += 1

                conn.commit()
            except Exception as exc:
                conn.rollback()
                self.finished.emit({
                    "success": False,
                    "message": str(exc),
                    "count": total_ins
                })
                return
            finally:
                conn.close()

            self.finished.emit({
                "success": True,
                "message": (
                    f"{total_ins} POS hareketi eklendi, {total_skip} mükerrer atlandı "
                    f"({len(terminals)} terminal, {len(chunks)} dönem)."
                ),
                "count": total_ins
            })
        except Exception as exc:
            self.finished.emit({"success": False, "message": f"Beklenmeyen hata: {exc}", "count": 0})




class PaytrTopluIsleWorker(QThread):
    """
    PHP: #btnPaytrPosIsle click → ajax/paytrPosIsle.php
    PayTR API'den 30 günlük batch'ler halinde işlem dökümü çekip paytr tablosuna yazar.
    """
    progress = pyqtSignal(str)
    batch_done = pyqtSignal(int, int)
    finished = pyqtSignal(dict)

    def __init__(self, start_dt: datetime.datetime, end_dt: datetime.datetime,
                 userid: int, musterino: str = ""):
        super().__init__()
        self._start_dt  = start_dt
        self._end_dt    = end_dt
        self._userid    = userid
        self._musterino = str(musterino)

    def run(self):
        from services.paytr_service import sync_chunk

        # 30 günlük batch'ler oluştur (PHP: batchEnd.setDate(batchEnd.getDate() + 29))
        batches: list[tuple[str, str]] = []
        cur = self._start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
        while cur <= self._end_dt:
            batch_end = min(cur + datetime.timedelta(days=29), self._end_dt)
            batches.append((cur.strftime("%Y-%m-%d"), batch_end.strftime("%Y-%m-%d")))
            cur = batch_end + datetime.timedelta(days=1)
            cur = cur.replace(hour=0, minute=0, second=0, microsecond=0)

        total_batches  = len(batches)
        total_inserted = 0
        total_skipped  = 0
        tum_uyarilar: list[str] = []

        for i, (bs, be) in enumerate(batches):
            self.progress.emit(
                f"💳  PayTR Parça {i+1}/{total_batches}: "
                f"{bs} → {be} işleniyor..."
            )
            result = sync_chunk(
                userid=self._userid,
                musterino=self._musterino,
                chunk_start=bs,
                chunk_end=be
            )
            if result.get("success"):
                total_inserted += result.get("inserted", 0)
                total_skipped  += result.get("skipped", 0)
                # Uyarıları topla (API hataları vb.)
                uyarilar = result.get("uyarilar", [])
                if uyarilar:
                    tum_uyarilar.extend(uyarilar)
            self.batch_done.emit(i + 1, total_batches)

        sonuc = {
            "success": True,
            "message": (
                f"{total_inserted} yeni PayTR kaydı aktarıldı, "
                f"{total_skipped} mevcut kayıt atlandı ({total_batches} parça)."
            ),
            "count": total_inserted
        }
        if tum_uyarilar:
            sonuc["uyarilar"] = tum_uyarilar
        self.finished.emit(sonuc)


# ─────────────────────────────────────────────────────────────────────────────
# ── Manuel Toplu İşlemler Dialog  (PHP: #modalMtopluisle + JS handler'ları)  ─
# ─────────────────────────────────────────────────────────────────────────────

class ManuelTopluIsleDialog(QDialog):
    """
    PHP ayarlar.php → #modalMtopluisle dialog'unun PyQt6 tam karşılığı.

    Butonlar:
      • Bankalar İşle       → TopluBankalarIsleWorker (ajax/topluWomIsle.php)
      • Womsis Pos İşle     → WomsisPosIsleWorker     (ajax/ayarlar/womsisPosIsle.php)
      • Sanal Pos İşle      → PaytrTopluIsleWorker    (ajax/ayarlar/paytrPosIsle.php)
      • Google Sheets       → Web sürümüne özel (bilgi mesajı gösterilir)
      • Toplu İşle          → Bankalar + WomsisPos + PayTR sırayla çalıştırır
    """

    def __init__(self, userid: int, api_base: str, app_key: str,
                 app_secret: str, musterino: int = 1, parent=None):
        super().__init__(parent)
        self._userid     = userid
        self._musterino  = musterino
        self._api_base   = api_base
        self._app_key    = app_key
        self._app_secret = app_secret
        self._worker: QThread | None = None
        self._queue: list[str] = []     # sıradaki işlemler

        self.setWindowTitle("Manuel Toplu İşlemler")
        self.setMinimumWidth(540)
        self.setModal(True)
        self.setStyleSheet(
            "QDialog{background:#ffffff;}"
            "QLabel{font-family:'Segoe UI',Arial,sans-serif;}"
        )
        self._build()

    # ── UI ────────────────────────────────────────────────────────────────

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 24)
        root.setSpacing(16)

        # Başlık
        title_row = QHBoxLayout()
        ic = QLabel("⚡")
        ic.setStyleSheet("font-size:22px;")
        title_row.addWidget(ic)
        t = QLabel("Manuel Toplu İşlemler")
        t.setStyleSheet("font-size:16px;font-weight:700;color:#1e293b;")
        title_row.addWidget(t)
        title_row.addStretch()
        root.addLayout(title_row)

        # ── Tarih aralığı seçici (PHP: #topluWomsisBas / #topluWomsisBit) ──
        date_frame = QFrame()
        date_frame.setStyleSheet(
            "QFrame{background:#f8f9fa;border:1px solid #e0e0e0;"
            "border-radius:8px;padding:4px;}"
        )
        date_layout = QVBoxLayout(date_frame)
        date_layout.setContentsMargins(14, 12, 14, 12)
        date_layout.setSpacing(8)

        range_lbl = QLabel("📅 Çekilecek Tarih Aralığı")
        range_lbl.setStyleSheet("font-size:13px;font-weight:600;color:#555;")
        date_layout.addWidget(range_lbl)

        DATE_STYLE = (
            "QDateEdit{background:white;border:1.5px solid #ccc;"
            "border-radius:6px;padding:0 8px;font-size:13px;color:#1f2937;}"
            "QDateEdit:focus{border-color:#6366f1;}"
        )
        date_row = QHBoxLayout()
        date_row.setSpacing(10)

        bas_col = QVBoxLayout()
        bas_lbl = QLabel("Başlangıç Tarihi")
        bas_lbl.setStyleSheet("font-size:11px;color:#888;")
        self._bas_date = QDateEdit()
        self._bas_date.setFixedHeight(34)
        self._bas_date.setDisplayFormat("dd.MM.yyyy")
        self._bas_date.setDate(QDate(QDate.currentDate().year(), 1, 1))
        self._bas_date.setCalendarPopup(True)
        self._bas_date.setStyleSheet(DATE_STYLE)
        bas_col.addWidget(bas_lbl)
        bas_col.addWidget(self._bas_date)
        date_row.addLayout(bas_col)

        bit_col = QVBoxLayout()
        bit_lbl = QLabel("Bitiş Tarihi")
        bit_lbl.setStyleSheet("font-size:11px;color:#888;")
        self._bit_date = QDateEdit()
        self._bit_date.setFixedHeight(34)
        self._bit_date.setDisplayFormat("dd.MM.yyyy")
        self._bit_date.setDate(QDate.currentDate())
        self._bit_date.setCalendarPopup(True)
        self._bit_date.setStyleSheet(DATE_STYLE)
        bit_col.addWidget(bit_lbl)
        bit_col.addWidget(self._bit_date)
        date_row.addLayout(bit_col)
        date_row.addStretch()

        date_layout.addLayout(date_row)
        root.addWidget(date_frame)

        # ── 4 ana buton (PHP: alertp içindeki 4 buton) ──
        BTN_BASE = (
            "QPushButton{border:none;border-radius:8px;font-size:13px;"
            "font-weight:600;padding:10px 14px;text-align:left;}"
            "QPushButton:hover{opacity:.85;}"
            "QPushButton:disabled{background:#e2e8f0;color:#94a3b8;}"
        )
        btn_grid = QVBoxLayout()
        btn_grid.setSpacing(8)

        # Bankalar İşle
        self._btn_bankalar = QPushButton("🏦  Bankalar İşle")
        self._btn_bankalar.setStyleSheet(
            BTN_BASE + "QPushButton{background:#3b82f6;color:white;}"
            "QPushButton:hover{background:#2563eb;}"
        )
        self._btn_bankalar.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_bankalar.clicked.connect(self._on_bankalar_isle)
        btn_grid.addWidget(self._btn_bankalar)

        # Womsis Pos İşle
        self._btn_wpos = QPushButton("📟  Womsis POS İşle")
        self._btn_wpos.setStyleSheet(
            BTN_BASE + "QPushButton{background:#1a3a5c;color:white;}"
            "QPushButton:hover{background:#152e4a;}"
        )
        self._btn_wpos.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_wpos.clicked.connect(self._on_wpos_isle)
        btn_grid.addWidget(self._btn_wpos)

        # Sanal Pos İşle
        self._btn_paytr = QPushButton("💳  Sanal Pos İşle  (PayTR)")
        self._btn_paytr.setStyleSheet(
            BTN_BASE + "QPushButton{background:#212121;color:white;}"
            "QPushButton:hover{background:#111111;}"
        )
        self._btn_paytr.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_paytr.clicked.connect(self._on_paytr_isle)
        btn_grid.addWidget(self._btn_paytr)

        # Google Sheets — web only
        self._btn_gsheets = QPushButton("📊  Tüm Hesap Tablolarını İşle  (Google Sheets)")
        self._btn_gsheets.setStyleSheet(
            BTN_BASE + "QPushButton{background:#34a853;color:white;}"
            "QPushButton:hover{background:#2d9248;}"
        )
        self._btn_gsheets.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_gsheets.clicked.connect(self._on_gsheets_isle)
        btn_grid.addWidget(self._btn_gsheets)

        root.addLayout(btn_grid)

        # ── İlerleme / durum etiketi ──
        self._progress_bar = QProgressBar()
        self._progress_bar.setFixedHeight(6)
        self._progress_bar.setTextVisible(False)
        self._progress_bar.setStyleSheet(
            "QProgressBar{background:#f1f5f9;border-radius:3px;border:none;}"
            "QProgressBar::chunk{background:#3b82f6;border-radius:3px;}"
        )
        self._progress_bar.hide()
        root.addWidget(self._progress_bar)

        self._status_lbl = QLabel("")
        self._status_lbl.setWordWrap(True)
        self._status_lbl.setStyleSheet(
            "font-size:12px;color:#374151;background:#f8fafc;"
            "border-radius:6px;padding:8px 10px;border:none;"
        )
        self._status_lbl.hide()
        root.addWidget(self._status_lbl)

        # ── Alt butonlar — Toplu İşle + İptal ──
        foot = QHBoxLayout()
        foot.setSpacing(10)

        self._btn_toplu = QPushButton("⚡  Toplu İşle  (Bankalar + POS + PayTR + Google Sheets)")
        self._btn_toplu.setFixedHeight(42)
        self._btn_toplu.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_toplu.setStyleSheet(
            "QPushButton{background:#4f46e5;color:white;border:none;"
            "border-radius:9px;font-size:13px;font-weight:700;padding:0 18px;}"
            "QPushButton:hover{background:#4338ca;}"
            "QPushButton:disabled{background:#cbd5e1;color:#94a3b8;}"
        )
        self._btn_toplu.clicked.connect(self._on_toplu_isle)
        foot.addWidget(self._btn_toplu)

        btn_iptal = QPushButton("İptal")
        btn_iptal.setFixedHeight(42)
        btn_iptal.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_iptal.setStyleSheet(
            "QPushButton{background:#f1f5f9;color:#374151;border:1.5px solid #e2e8f0;"
            "border-radius:9px;font-size:13px;font-weight:600;padding:0 18px;}"
            "QPushButton:hover{background:#e2e8f0;}"
        )
        btn_iptal.clicked.connect(self.reject)
        foot.addWidget(btn_iptal)
        root.addLayout(foot)

    # ── Yardımcılar ───────────────────────────────────────────────────────

    def _get_dates(self) -> tuple[datetime.datetime, datetime.datetime]:
        bas_q = self._bas_date.date()
        bit_q = self._bit_date.date()
        start = datetime.datetime(bas_q.year(), bas_q.month(), bas_q.day(), 0, 0, 0)
        end   = datetime.datetime(bit_q.year(), bit_q.month(), bit_q.day(), 23, 59, 59)
        return start, end

    def _validate_dates(self) -> bool:
        start, end = self._get_dates()
        if start > end:
            self._show_status(
                "⚠️  Başlangıç tarihi bitiş tarihinden büyük olamaz.", "#92400e", ok=False
            )
            return False
        return True

    def _validate_api(self) -> bool:
        if not self._api_base or not self._app_key or not self._app_secret:
            self._show_status(
                "⚠️  Lütfen önce Eklentiler sekmesinden VOMSİS API bilgilerini kaydedin.",
                "#92400e", ok=False
            )
            return False
        return True

    def _set_buttons_enabled(self, enabled: bool):
        for btn in (self._btn_bankalar, self._btn_wpos, self._btn_paytr,
                    self._btn_gsheets, self._btn_toplu):
            btn.setEnabled(enabled)

    def _show_status(self, text: str, color: str = "#374151", ok: bool = True):
        self._status_lbl.setText(text)
        bg = "#dcfce7" if ok else "#fef3c7"
        if "#dc2626" in color or "#92400e" in color:
            bg = "#fee2e2"
        self._status_lbl.setStyleSheet(
            f"font-size:12px;color:{color};background:{bg};"
            "border-radius:6px;padding:8px 10px;border:none;"
        )
        self._status_lbl.show()

    def _start_progress(self, total: int = 0):
        self._progress_bar.setMaximum(max(total, 1))
        self._progress_bar.setValue(0)
        self._progress_bar.show()

    # ── Bankalar İşle ─────────────────────────────────────────────────────

    def _on_bankalar_isle(self):
        if not self._validate_dates():
            return
        if not self._validate_api():
            return
        start, end = self._get_dates()
        self._set_buttons_enabled(False)
        self._show_status("⏳  Banka hareketleri çekiliyor...", "#6366f1")
        self._start_progress()

        self._worker = TopluBankalarIsleWorker(
            self._api_base, self._app_key, self._app_secret,
            start, end, self._userid
        )
        self._worker.progress.connect(lambda m: self._show_status(m, "#6366f1"))
        self._worker.batch_done.connect(
            lambda done, total: self._progress_bar.setValue(
                int(done / total * self._progress_bar.maximum())
                if total else done
            )
        )
        self._worker.finished.connect(self._on_done)
        self._worker.start()

    # ── Womsis Pos İşle ───────────────────────────────────────────────────

    def _on_wpos_isle(self):
        if not self._validate_dates():
            return
        if not self._validate_api():
            return
        start, end = self._get_dates()
        self._set_buttons_enabled(False)
        self._show_status("⏳  VOMSİS POS verileri çekiliyor...", "#1a3a5c")
        self._start_progress()

        self._worker = WomsisPosIsleWorker(
            self._api_base, self._app_key, self._app_secret,
            start, end, self._userid, musterino=self._musterino
        )
        self._worker.progress.connect(lambda m: self._show_status(m, "#1a3a5c"))
        self._worker.finished.connect(self._on_done)
        self._worker.start()

    # ── Sanal Pos (PayTR) İşle ────────────────────────────────────────────

    def _on_paytr_isle(self):
        if not self._validate_dates():
            return
        start, end = self._get_dates()
        self._set_buttons_enabled(False)
        self._show_status("⏳  PayTR sanal POS verileri çekiliyor...", "#374151")
        self._start_progress()

        self._worker = PaytrTopluIsleWorker(start, end, self._userid,
                                            musterino=self._musterino)
        self._worker.progress.connect(lambda m: self._show_status(m, "#374151"))
        self._worker.batch_done.connect(
            lambda done, total: self._progress_bar.setValue(
                int(done / total * self._progress_bar.maximum())
                if total else done
            )
        )
        self._worker.finished.connect(self._on_done)
        self._worker.start()

    # ── Google Sheets İşle ────────────────────────────────────────────────

    def _on_gsheets_isle(self):
        """Google Sheets → genel_hesap_hareketleri aktarımı."""
        if not self._validate_dates():
            return
        import datetime as _dt
        bas_q = self._bas_date.date()
        bit_q = self._bit_date.date()
        bas = _dt.date(bas_q.year(), bas_q.month(), bas_q.day())
        bit = _dt.date(bit_q.year(), bit_q.month(), bit_q.day())

        self._set_buttons_enabled(False)
        self._show_status("📥  Google Sheets verileri indiriliyor...", "#34a853")
        self._start_progress()

        self._worker = GoogleSheetsAktarWorker(
            userid=self._userid,
            musterino=self._musterino,
            bas=bas,
            bit=bit,
            kaynaklar=["kasa", "gider", "genelHesap"],
        )
        self._worker.progress.connect(lambda m: self._show_status(m, "#34a853"))
        self._worker.finished.connect(self._on_done)
        self._worker.start()

    # ── Toplu İşle (sırayla 3 işlem) ─────────────────────────────────────

    def _on_toplu_isle(self):
        """Bankalar + WomsisPos + PayTR + Google Sheets sırayla çalıştırır."""
        if not self._validate_dates():
            return
        if not self._validate_api():
            return
        # Sırayı belirle
        self._queue = ["bankalar", "wpos", "paytr", "gsheets"]
        self._set_buttons_enabled(False)
        self._show_status("⏳  Toplu işlem başlatıldı...", "#4f46e5")
        self._start_progress(4)
        self._run_next_in_queue()

    def _run_next_in_queue(self):
        if not self._queue:
            self._progress_bar.setValue(self._progress_bar.maximum())
            self._set_buttons_enabled(True)
            self._show_status(
                "✅  Toplu işlem tamamlandı. Tüm kaynaklar aktarıldı (Bankalar + POS + PayTR + Google Sheets).", "#059669", ok=True
            )
            return
        next_op = self._queue.pop(0)
        start, end = self._get_dates()
        if next_op == "bankalar":
            self._show_status("📡  (1/4) Bankalar İşleniyor...", "#6366f1")
            self._worker = TopluBankalarIsleWorker(
                self._api_base, self._app_key, self._app_secret,
                start, end, self._userid
            )
            self._worker.progress.connect(lambda m: self._show_status(m, "#6366f1"))
            self._worker.finished.connect(self._on_queue_step_done)
            self._worker.start()
        elif next_op == "wpos":
            self._show_status("📟  (2/4) Womsis POS İşleniyor...", "#1a3a5c")
            self._worker = WomsisPosIsleWorker(
                self._api_base, self._app_key, self._app_secret,
                start, end, self._userid, musterino=self._musterino
            )
            self._worker.progress.connect(lambda m: self._show_status(m, "#1a3a5c"))
            self._worker.finished.connect(self._on_queue_step_done)
            self._worker.start()
        elif next_op == "paytr":
            self._show_status("💳  (3/4) PayTR Sanal POS İşleniyor...", "#374151")
            self._worker = PaytrTopluIsleWorker(start, end, self._userid,
                                                musterino=self._musterino)
            self._worker.progress.connect(lambda m: self._show_status(m, "#374151"))
            self._worker.finished.connect(self._on_queue_step_done)
            self._worker.start()
        elif next_op == "gsheets":
            import datetime as _dt
            bas = _dt.date(start.year, start.month, start.day)
            bit = _dt.date(end.year, end.month, end.day)
            self._show_status("📊  (4/4) Google Sheets Aktarılıyor...", "#34a853")
            self._worker = GoogleSheetsAktarWorker(
                userid=self._userid,
                musterino=self._musterino,
                bas=bas,
                bit=bit,
                kaynaklar=["kasa", "gider", "genelHesap"],
            )
            self._worker.progress.connect(lambda m: self._show_status(m, "#34a853"))
            self._worker.finished.connect(self._on_queue_step_done)
            self._worker.start()

    def _on_queue_step_done(self, result: dict):
        done = 4 - len(self._queue)
        self._progress_bar.setValue(done)
        # Eski worker'ı Qt'ye bırak — GC'den önce thread tamamen bitsin
        if self._worker:
            self._worker.deleteLater()
            self._worker = None
        self._run_next_in_queue()

    # ── Tek işlem bitişi ──────────────────────────────────────────────────

    def _on_done(self, result: dict):
        self._progress_bar.hide()
        self._set_buttons_enabled(True)
        # Eski worker'ı Qt'ye bırak — GC thread'i silmesin
        if self._worker:
            self._worker.deleteLater()
            self._worker = None
        if result["success"]:
            msg = result['message']
            uyarilar = result.get("uyarilar", [])
            if uyarilar:
                n = len(uyarilar)
                msg += f"  |  ⚠️ {n} API uyarısı (bazı chunk'lar çekilememiş olabilir)"
            self._show_status(f"✅  {msg}", "#059669", ok=True)
        else:
            self._show_status(f"❌  {result['message']}", "#dc2626", ok=False)


# ── Google Sheets Aktarım Worker ─────────────────────────────────────────────

class GoogleSheetsAktarWorker(QThread):
    """PHP google_sheets_aktar.php → Python worker."""
    progress = pyqtSignal(str)
    finished = pyqtSignal(dict)

    def __init__(self, userid: int, musterino, bas: "datetime.date", bit: "datetime.date",
                 kaynaklar: list, parent=None):
        super().__init__(parent)
        self._userid    = userid
        self._musterino = musterino
        self._bas       = bas
        self._bit       = bit
        self._kaynaklar = kaynaklar

    def run(self):
        try:
            import datetime
            from services.google_sheets_service import google_sheets_aktar
            result = google_sheets_aktar(
                userid=self._userid,
                musterino=self._musterino,
                bas_tarih=self._bas,
                bit_tarih=self._bit,
                kaynaklar=self._kaynaklar,
                progress_cb=lambda m: self.progress.emit(m),
            )
            self.finished.emit(result)
        except Exception as exc:
            self.finished.emit({"success": False, "message": str(exc),
                                "eklenen": 0, "atlanan": 0, "hatalar": [str(exc)]})



# ── Google Sheets Ayarları Kartı ──────────────────────────────────────────────

class GoogleSheetsSettingsCard(QFrame):
    """
    Eklentiler - Manuel Toplu Isle nin ustunde yer alan kart.
    Nakit/Kasa, Gider ve Genel Hesap tablolari icin
    Google Sheets URL si veya ID girisi saglar.
    Ayarlar ~/NakitAkim/data/gsheets_config.json dosyasina kaydedilir.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self._setup_ui()
        self._load_saved()

    def _setup_ui(self):
        self.setObjectName("gsheetsSettingsCard")
        self.setStyleSheet("""
            QFrame#gsheetsSettingsCard {
                background: #ffffff;
                border: 1.5px solid #e0e7ff;
                border-radius: 16px;
            }
        """)
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(14)

        hdr = QHBoxLayout()
        icon = QLabel("\U0001f4ca")
        icon.setStyleSheet("font-size:22px;")
        title_lbl = QLabel("Google Sheets Bağlantıları")
        title_lbl.setStyleSheet(
            "font-size:15px;font-weight:700;color:#1e3a8a;letter-spacing:0.5px;"
        )
        sub_lbl = QLabel("Her sheet için URL veya ID girin — tam URL yapıştırabilirsiniz")
        sub_lbl.setStyleSheet("font-size:11px;color:#6b7280;")
        txt_v = QVBoxLayout()
        txt_v.setSpacing(2)
        txt_v.addWidget(title_lbl)
        txt_v.addWidget(sub_lbl)
        hdr.addWidget(icon)
        hdr.addSpacing(8)
        hdr.addLayout(txt_v)
        hdr.addStretch()
        root.addLayout(hdr)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color:#e0e7ff;")
        root.addWidget(sep)

        grid = QGridLayout()
        grid.setSpacing(10)
        grid.setColumnStretch(1, 1)
        self._inputs: dict[str, QLineEdit] = {}
        sheets_def = [
            ("\U0001f4b0  Nakit / Kasa",        "kasa_sheet_id",        False),
            ("    Tab Adı:",                      "kasa_tab_name",        True),
            ("\U0001f4b8  Gider",               "gider_sheet_id",       False),
            ("\U0001f4cb  Genel Hesap Tablosu", "genel_hesap_sheet_id", False),
        ]
        for r, (lbl_txt, key, small) in enumerate(sheets_def):
            lbl = QLabel(lbl_txt)
            lbl.setStyleSheet(
                "font-size:12px;font-weight:600;color:#374151;min-width:170px;"
            )
            grid.addWidget(lbl, r, 0, Qt.AlignmentFlag.AlignVCenter)
            inp = QLineEdit()
            if small:
                inp.setPlaceholderText("ör: Kasa")
                inp.setFixedHeight(34)
                inp.setStyleSheet(self._input_style(small=True))
            else:
                inp.setPlaceholderText(
                    "URL veya Sheet ID  (https://docs.google.com/spreadsheets/d/...)"
                )
                inp.setFixedHeight(38)
                inp.setStyleSheet(self._input_style())
            self._inputs[key] = inp
            grid.addWidget(inp, r, 1)
        root.addLayout(grid)

        self._status = QLabel("")
        self._status.setWordWrap(True)
        self._status.setStyleSheet("font-size:12px;color:#374151;")
        self._status.hide()
        root.addWidget(self._status)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)
        self._btn_save = QPushButton("\U0001f4be  Kaydet")
        self._btn_save.setFixedHeight(38)
        self._btn_save.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_save.setStyleSheet(
            "QPushButton{background:#1e3a8a;color:white;border:none;border-radius:10px;"
            "font-size:13px;font-weight:700;padding:0 20px;}"
            "QPushButton:hover{background:#1d4ed8;}"
        )
        self._btn_save.clicked.connect(self._on_save)
        btn_row.addWidget(self._btn_save)
        self._btn_test = QPushButton("\U0001f517  Bağlantıyı Doğrula")
        self._btn_test.setFixedHeight(38)
        self._btn_test.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_test.setStyleSheet(
            "QPushButton{background:#0f766e;color:white;border:none;border-radius:10px;"
            "font-size:13px;font-weight:600;padding:0 20px;}"
            "QPushButton:hover{background:#0d9488;}"
        )
        self._btn_test.clicked.connect(self._on_test)
        btn_row.addWidget(self._btn_test)
        btn_row.addStretch()
        self._btn_reset = QPushButton("\u21a9  Fabrika Değerlerine Dön")
        self._btn_reset.setFixedHeight(38)
        self._btn_reset.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_reset.setStyleSheet(
            "QPushButton{background:#f3f4f6;color:#374151;border:1px solid #d1d5db;"
            "border-radius:10px;font-size:12px;font-weight:600;padding:0 16px;}"
            "QPushButton:hover{background:#e5e7eb;}"
        )
        self._btn_reset.clicked.connect(self._on_reset)
        btn_row.addWidget(self._btn_reset)
        root.addLayout(btn_row)

    def _input_style(self, small: bool = False) -> str:
        fs = "12px" if small else "13px"
        return (
            f"QLineEdit{{background:#f8fafc;border:1.5px solid #e2e8f0;"
            f"border-radius:8px;padding:0 12px;font-size:{fs};color:#1e293b;}}"
            "QLineEdit:focus{border-color:#3b82f6;background:white;}"
        )

    def _show_status(self, msg: str, color: str = "#374151"):
        self._status.setText(msg)
        self._status.setStyleSheet(f"font-size:12px;color:{color};")
        self._status.show()

    def _load_saved(self):
        try:
            from services.gsheets_config_service import load_config
            cfg = load_config()
            for key, inp in self._inputs.items():
                val = cfg.get(key, "")
                if val:
                    inp.setText(val)
        except Exception as exc:
            self._show_status(f"Ayarlar yuklenemedi: {exc}", "#dc2626")

    def _on_save(self):
        try:
            from services.gsheets_config_service import save_config, extract_sheet_id
            cfg = {}
            for key, inp in self._inputs.items():
                val = inp.text().strip()
                if key.endswith("_sheet_id"):
                    val = extract_sheet_id(val)
                    inp.setText(val)
                cfg[key] = val
            save_config(cfg)
            self._show_status(
                "\u2705  Ayarlar kaydedildi. Bir sonraki Toplu Islede yeni baglantilar kullanilacak.",
                "#059669"
            )
        except Exception as exc:
            self._show_status(f"Kaydetme hatasi: {exc}", "#dc2626")

    def _on_test(self):
        self._show_status("\U0001f504  Test ediliyor, lutfen bekleyin...", "#4f46e5")
        self._btn_test.setEnabled(False)
        self._btn_save.setEnabled(False)
        from PyQt6.QtCore import QThread, pyqtSignal as _sig

        class _Tester(QThread):
            done = _sig(str, str)
            def __init__(self, cfg_vals):
                super().__init__()
                self._cfg_vals = cfg_vals
            def run(self):
                import urllib.request
                from services.google_sheets_service import _build_urls, _TIMEOUT
                from services.gsheets_config_service import extract_sheet_id, load_config
                base = load_config()
                for k, v in self._cfg_vals.items():
                    if k.endswith("_sheet_id"):
                        v = extract_sheet_id(v)
                    base[k] = v
                kasa_url, gider_tpl, genel_url = _build_urls(base)
                results = []
                for name, url in [
                    ("Kasa",        kasa_url),
                    ("Gider",       gider_tpl.format(year=2025)),
                    ("Genel Hesap", genel_url),
                ]:
                    try:
                        import ssl
                        req = urllib.request.Request(url, headers={"User-Agent": "NakitAkim/1.0"})
                        try:
                            import certifi
                            ssl_ctx = ssl.create_default_context(cafile=certifi.where())
                        except ImportError:
                            ssl_ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
                            ssl_ctx.check_hostname = False
                            ssl_ctx.verify_mode = ssl.CERT_NONE
                        except Exception:
                            ssl_ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
                            ssl_ctx.check_hostname = False
                            ssl_ctx.verify_mode = ssl.CERT_NONE
                        with urllib.request.urlopen(req, timeout=_TIMEOUT, context=ssl_ctx) as r:
                            results.append(f"\u2705  {name} — OK (HTTP {r.status})")
                    except Exception as e:
                        results.append(f"\u274c  {name} — {e}")
                ok = all(r.startswith("\u2705") for r in results)
                self.done.emit("\n".join(results), "#059669" if ok else "#b45309")

        vals = {k: inp.text().strip() for k, inp in self._inputs.items()}
        self._tester = _Tester(vals)

        def _done(m, c):
            self._show_status(m, c)
            self._btn_test.setEnabled(True)
            self._btn_save.setEnabled(True)

        self._tester.done.connect(_done)
        self._tester.start()

    def _on_reset(self):
        from services.gsheets_config_service import _DEFAULTS
        for key, inp in self._inputs.items():
            inp.setText(_DEFAULTS.get(key, ""))
        self._show_status(
            "\u21a9  Fabrika degerleri yuklendi. Kaydetmek icin Kaydet tusuna basin.",
            "#6b7280"
        )

    def refresh(self):
        self._load_saved()


# ── Manuel Toplu İşle Kartı ───────────────────────────────────────────────────

class ManuelTopluIsleCard(QFrame):
    """
    E-fatura grubunun üstünde ayrı bir kart olarak görünen
    'Manuel Toplu İşle' düğmesi. VOMSİS API kartından bağımsız çalışır.
    """

    def __init__(self, userid: int, musterino: int = 1, parent=None):
        super().__init__(parent)
        self._userid = userid
        self._musterino = musterino
        self.setStyleSheet(
            "QFrame{background:white;border:2px solid #bbf7d0;"
            "border-radius:16px;}"
        )
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 22)
        root.setSpacing(14)

        # ── Başlık satırı ──
        h = QHBoxLayout()
        ic = QLabel("🗂️")
        ic.setStyleSheet("font-size:22px;")
        h.addWidget(ic)
        t = QLabel("Manuel Toplu İşle")
        t.setStyleSheet(
            "font-size:15px;font-weight:800;color:#065f46;"
            "letter-spacing:.6px;"
        )
        h.addWidget(t)
        badge = QLabel("VOMSİS + PayTR")
        badge.setStyleSheet(
            "background:#d1fae5;color:#065f46;border-radius:8px;"
            "padding:3px 10px;font-size:11px;font-weight:700;border:none;"
        )
        h.addWidget(badge)
        h.addStretch()
        root.addLayout(h)

        # ── Bilgi/uyarı alanı (yeşil üzerine beyaz) ──
        info_frame = QFrame()
        info_frame.setStyleSheet(
            "QFrame{background:#059669;border-radius:12px;border:none;}"
        )
        info_lay = QVBoxLayout(info_frame)
        info_lay.setContentsMargins(16, 14, 16, 14)
        info_lay.setSpacing(6)

        baslik_lbl = QLabel("ℹ️  Bu mod ne yapar?")
        baslik_lbl.setStyleSheet(
            "color:white;font-size:13px;font-weight:700;background:transparent;"
        )
        info_lay.addWidget(baslik_lbl)

        maddeler = [
            "✅  Bankalar İşle  —  VOMSİS API'ünden banka hesabı hareketlerini çeker ve "
            "hareketler tablosuna aktarır (15 günlük toplu işlemler halinde).",
            "✅  VOMSİS POS İşle  —  Kayıtlı fiziksel POS terminallerinin işlem "
            "verilerini çeker, womsi_pos tablosuna yazar (eskiler silinir, yeniden yazılır).",
            "✅  PayTR Sanal POS İşle  —  PayTR API'ünden müşteri bazında satış işlemlerini "
            "senkronize eder; tekrarı olan kayıtlar güncellenir, yeni kayıtlar eklenir.",
            "📌  Tarih aralığı seçerek hangi dönem için veri çekeceğinizi belirleyebilirsiniz.",
            "⚠️  VOMSİS işlemlerinin çalışması için VOMSİS API bilgilerinizin ve IP whitelist "
            "tanımının doğru yapılandırılmış olması gerekir.",
        ]
        for m in maddeler:
            lbl = QLabel(m)
            lbl.setWordWrap(True)
            lbl.setStyleSheet(
                "color:white;font-size:12px;background:transparent;"
                "padding:1px 0;"
            )
            info_lay.addWidget(lbl)

        root.addWidget(info_frame)

        # ── Büyük, dikkat çekici buton ──
        btn_row = QHBoxLayout()
        self._toplu_btn = QPushButton("📦  Tüm Hesap Tablolarını İşle — Toplu İşle (Bankalar + POS + PayTR)")
        self._toplu_btn.setFixedHeight(48)
        self._toplu_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._toplu_btn.setStyleSheet(
            "QPushButton{"
            "  background:qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "    stop:0 #059669,stop:1 #047857);"
            "  color:white;border:none;border-radius:12px;"
            "  font-size:14px;font-weight:700;"
            "  padding:0 24px;letter-spacing:.5px;"
            "}"
            "QPushButton:hover{"
            "  background:qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "    stop:0 #10b981,stop:1 #059669);"
            "}"
            "QPushButton:pressed{background:#047857;}"
        )
        self._toplu_btn.clicked.connect(self._on_toplu_isle)
        btn_row.addWidget(self._toplu_btn)
        btn_row.addStretch()
        root.addLayout(btn_row)

    # ── Tıklama ──────────────────────────────────────────────────────────────

    def _on_toplu_isle(self):
        """
        VOMSİS API kartından API bilgilerini al,
        ManuelTopluIsleDialog'u aç.
        """
        url = appkey = seckey = ""
        # Üst widget hiyerarşisinde VomsisCard'ı bul
        p = self.parent()
        while p is not None:
            if hasattr(p, '_vomsis_card'):
                vc = p._vomsis_card
                url    = vc._url_inp.text().strip()
                appkey = vc._key_inp.text().strip()
                seckey = vc._sec_inp.text().strip()
                break
            p = p.parent() if callable(getattr(p, 'parent', None)) else None

        dlg = ManuelTopluIsleDialog(
            userid=self._userid,
            musterino=self._musterino,
            api_base=url,
            app_key=appkey,
            app_secret=seckey,
            parent=self
        )
        dlg.exec()

    def refresh(self):
        pass


# ── VOMSİS API Kartı ──────────────────────────────────────────────────────────

class VomsisCard(QFrame):

    """
    PHP ayarlar.php → Eklentiler → VOMSİS API bölümünün PyQt6 karşılığı.
    Alanlar: API URL, API KEY, API SECRET KEY
    Butonlar: Kontrol Et | VOMSİS İşle (tarih aralığı seçerek)
    """

    def __init__(self, userid: int, musterino: int = 1, parent=None):
        super().__init__(parent)
        self._userid    = userid
        self._musterino = musterino
        self._test_worker: VomsisTestWorker | None = None
        self._isle_worker: VomsisIsleWorker | None = None
        self.setStyleSheet(
            "QFrame{background:white;border:1.5px solid #e0e7ff;"
            "border-radius:14px;}"
        )
        self._build()
        self._load()

    # ── Yardımcılar ────────────────────────────────────────────────────────

    def _input_style(self) -> str:
        return (
            "QLineEdit{background:#f8fafc;border:1.5px solid #e2e8f0;"
            "border-radius:8px;padding:0 10px;font-size:13px;color:#1f2937;}"
            "QLineEdit:focus{border-color:#6366f1;}"
        )

    def _label_style(self) -> str:
        return "font-size:12px;font-weight:600;color:#374151;"

    def _btn_style(self, color: str) -> str:
        return (
            f"QPushButton{{background:{color};color:white;border:none;"
            f"border-radius:9px;font-size:13px;font-weight:600;"
            f"padding:0 16px;letter-spacing:.4px;}}"
            f"QPushButton:hover{{background:{color}cc;}}"
            f"QPushButton:disabled{{background:#cbd5e1;color:#94a3b8;}}"
        )

    # ── UI inşası ──────────────────────────────────────────────────────────

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(22, 20, 22, 20)
        root.setSpacing(14)

        # Başlık satırı
        h = QHBoxLayout()
        ic = QLabel("🔌")
        ic.setStyleSheet("font-size:20px;")
        h.addWidget(ic)
        t = QLabel("VOMSİS API")
        t.setStyleSheet("font-size:14px;font-weight:700;color:#000000;letter-spacing:.5px;")
        h.addWidget(t)
        h.addStretch()
        root.addLayout(h)

        # Bilgi chip
        info = QLabel(
            "💡  Vomsis bilgilerini girdikten sonra sistem otomatik olarak "
            "banka hareketlerinizi çekebilir. 'Kontrol Et' ile bağlantıyı doğrulayın."
        )
        info.setWordWrap(True)
        info.setStyleSheet(
            "background:#ede9fe;color:#4c1d95;border-radius:6px;"
            "padding:9px 12px;font-size:12px;border:none;"
        )
        root.addWidget(info)

        # ── Alan satırı 1 — URL + APP KEY ──
        row1 = QHBoxLayout()
        row1.setSpacing(12)

        url_col = QVBoxLayout()
        url_lbl = QLabel("API URL")
        url_lbl.setStyleSheet(self._label_style())
        self._url_inp = QLineEdit()
        self._url_inp.setFixedHeight(34)
        self._url_inp.setPlaceholderText("https://developers.vomsis.com/api/v2")
        self._url_inp.setStyleSheet(self._input_style())
        url_col.addWidget(url_lbl)
        url_col.addWidget(self._url_inp)
        row1.addLayout(url_col, 2)

        key_col = QVBoxLayout()
        key_lbl = QLabel("API KEY")
        key_lbl.setStyleSheet(self._label_style())
        self._key_inp = QLineEdit()
        self._key_inp.setFixedHeight(34)
        self._key_inp.setPlaceholderText("API KEY")
        self._key_inp.setStyleSheet(self._input_style())
        key_col.addWidget(key_lbl)
        key_col.addWidget(self._key_inp)
        row1.addLayout(key_col, 2)
        root.addLayout(row1)

        # ── Alan satırı 2 — SECRET KEY ──
        sec_lbl = QLabel("API SECRET KEY")
        sec_lbl.setStyleSheet(self._label_style())
        self._sec_inp = QLineEdit()
        self._sec_inp.setFixedHeight(34)
        self._sec_inp.setPlaceholderText("API SECRET KEY")
        self._sec_inp.setStyleSheet(self._input_style())
        root.addWidget(sec_lbl)
        root.addWidget(self._sec_inp)

        # ── Tarih aralığı (Manuel İşle için) ──
        date_row = QHBoxLayout()
        date_row.setSpacing(10)
        date_lbl = QLabel("Tarih Aralığı:")
        date_lbl.setStyleSheet(self._label_style())
        date_row.addWidget(date_lbl)

        DATE_STYLE = (
            "QDateEdit{background:#f8fafc;border:1.5px solid #e2e8f0;"
            "border-radius:8px;padding:0 8px;font-size:13px;color:#1f2937;}"
            "QDateEdit:focus{border-color:#6366f1;}"
        )
        self._start_date = QDateEdit()
        self._start_date.setFixedHeight(34)
        self._start_date.setFixedWidth(130)
        self._start_date.setDisplayFormat("dd.MM.yyyy")
        self._start_date.setDate(QDate(2026, 1, 1))  # Baştan sona çek
        self._start_date.setCalendarPopup(True)
        self._start_date.setStyleSheet(DATE_STYLE)
        date_row.addWidget(self._start_date)

        dash = QLabel("—")
        dash.setStyleSheet("color:#6b7280;font-size:14px;")
        date_row.addWidget(dash)

        self._end_date = QDateEdit()
        self._end_date.setFixedHeight(34)
        self._end_date.setFixedWidth(130)
        self._end_date.setDisplayFormat("dd.MM.yyyy")
        self._end_date.setDate(QDate.currentDate())
        self._end_date.setCalendarPopup(True)
        self._end_date.setStyleSheet(DATE_STYLE)
        date_row.addWidget(self._end_date)
        date_row.addStretch()
        root.addLayout(date_row)

        # ── Butonlar ──
        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)

        self._kontrol_btn = QPushButton("🔍  Kontrol Et")
        self._kontrol_btn.setFixedHeight(38)
        self._kontrol_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._kontrol_btn.setStyleSheet(self._btn_style("#6366f1"))
        self._kontrol_btn.clicked.connect(self._on_kontrol_et)
        btn_row.addWidget(self._kontrol_btn)

        self._kaydet_btn = QPushButton("💾  Kaydet")
        self._kaydet_btn.setFixedHeight(38)
        self._kaydet_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._kaydet_btn.setStyleSheet(self._btn_style("#2563eb"))
        self._kaydet_btn.clicked.connect(self._on_kaydet)
        btn_row.addWidget(self._kaydet_btn)

        self._isle_btn = QPushButton("⚡  VOMSİS İşle")
        self._isle_btn.setFixedHeight(38)
        self._isle_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._isle_btn.setStyleSheet(self._btn_style("#0891b2"))
        self._isle_btn.clicked.connect(self._on_isle)
        btn_row.addWidget(self._isle_btn)

        # ── webadmin REST API üzerinden Çek ──────────────────────────────────
        self._webadmin_btn = QPushButton("🌐  webadmin Çek")
        self._webadmin_btn.setFixedHeight(38)
        self._webadmin_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._webadmin_btn.setStyleSheet(self._btn_style("#7c3aed"))
        self._webadmin_btn.setToolTip(
            "webadmin-nakitAkim REST API üzerinden Womsis verilerini çeker.\n"
            "webadmin sunucusu http://localhost:5050 adresinde çalışıyor olmalı."
        )
        self._webadmin_btn.clicked.connect(self._on_webadmin_cek)
        btn_row.addWidget(self._webadmin_btn)

        # ── VOMSİS Excel Import ──────────────────────────────────────────────
        self._excel_import_btn = QPushButton("📂  Excel Import")
        self._excel_import_btn.setFixedHeight(38)
        self._excel_import_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._excel_import_btn.setStyleSheet(self._btn_style("#059669"))
        self._excel_import_btn.setToolTip(
            "VOMSİS banka hareketleri Excel dosyasını içe aktarır.\n"
            "Şablon için dialog içinde 'Şablon İndir' butonunu kullanın."
        )
        self._excel_import_btn.clicked.connect(self._on_excel_import)
        btn_row.addWidget(self._excel_import_btn)

        btn_row.addStretch()
        root.addLayout(btn_row)

        # ── Durum etiketi ──
        self._status_lbl = QLabel("")
        self._status_lbl.setWordWrap(True)
        self._status_lbl.setStyleSheet("font-size:12px;color:#374151;")
        self._status_lbl.hide()
        root.addWidget(self._status_lbl)

    # ── Veri yükleme ────────────────────────────────────────────────────────

    def _load(self):
        from services.vomsis_service import get_vomsis_bilgileri, DEFAULT_API_URL
        bilgi = get_vomsis_bilgileri(self._userid, self._musterino)
        self._url_inp.setText(bilgi.get("url") or DEFAULT_API_URL)
        self._key_inp.setText(bilgi.get("appkey") or "")
        self._sec_inp.setText(bilgi.get("seckey") or "")

    def refresh(self):
        self._load()

    # ── Kaydet ──────────────────────────────────────────────────────────────

    def _on_kaydet(self):
        from services.vomsis_service import save_vomsis_bilgileri
        url    = self._url_inp.text().strip()
        appkey = self._key_inp.text().strip()
        seckey = self._sec_inp.text().strip()

        if not url or not appkey or not seckey:
            self._show_status("⚠️  Tüm alanları doldurunuz.", "#92400e")
            return

        result = save_vomsis_bilgileri(self._userid, appkey, seckey, url,
                                        musterino=self._musterino)
        if result["success"]:
            self._show_status(f"✅  {result['message']}", "#059669")
        else:
            self._show_status(f"❌  {result['message']}", "#dc2626")

    # ── Kontrol Et ──────────────────────────────────────────────────────────

    def _on_kontrol_et(self):
        url    = self._url_inp.text().strip()
        appkey = self._key_inp.text().strip()
        seckey = self._sec_inp.text().strip()

        if not url or not appkey or not seckey:
            self._show_status("⚠️  Tüm alanları doldurunuz.", "#92400e")
            return

        self._set_busy(True, "🔑  VOMSİS sunucusuna bağlanılıyor...")

        self._test_worker = VomsisTestWorker(url, appkey, seckey)
        self._test_worker.result.connect(self._on_test_result)
        self._test_worker.start()

    def _on_test_result(self, r: dict):
        self._set_busy(False)
        if r["success"]:
            self._show_status(f"✅  {r['message']}", "#059669")
        else:
            self._show_status(f"❌  {r['message']}", "#dc2626")

    # ── VOMSİS İşle ─────────────────────────────────────────────────────────

    def _on_isle(self):
        url    = self._url_inp.text().strip()
        appkey = self._key_inp.text().strip()
        seckey = self._sec_inp.text().strip()

        if not url or not appkey or not seckey:
            self._show_status("⚠️  Önce API bilgilerini kaydedin.", "#92400e")
            return

        start_qd = self._start_date.date()
        end_qd   = self._end_date.date()
        start_dt = datetime.datetime(start_qd.year(), start_qd.month(), start_qd.day())
        end_dt   = datetime.datetime(end_qd.year(),   end_qd.month(),   end_qd.day(),
                                     23, 59, 59)

        if start_dt > end_dt:
            self._show_status("⚠️  Başlangıç tarihi bitiş tarihinden büyük olamaz.", "#92400e")
            return

        onay = SweetConfirmDialog.confirm(
            self,
            title="VOMSİS İşle",
            text=(
                f"{start_qd.toString('dd.MM.yyyy')} — {end_qd.toString('dd.MM.yyyy')} "
                f"aralığındaki banka hareketleri aktarılacak.\n"
                "Devam etmek istiyor musunuz?"
            ),
            confirm_text="⚡  İşle",
            cancel_text="Vazgeç",
        )
        if not onay:
            return

        self._set_busy(True, "⚡  VOMSİS banka hareketleri alınıyor...")

        self._isle_worker = VomsisIsleWorker(
            url, appkey, seckey, start_dt, end_dt,
            self._userid, musterino=self._musterino
        )
        self._isle_worker.progress.connect(
            lambda msg: self._show_status(msg, "#0891b2")
        )
        self._isle_worker.finished.connect(self._on_isle_done)
        self._isle_worker.start()

    def _on_isle_done(self, r: dict):
        self._set_busy(False)
        if r["success"]:
            self._show_status(f"✅  {r['message']}", "#059669")
        else:
            self._show_status(f"❌  {r['message']}", "#dc2626")

    # ── webadmin REST API üzerinden Çek ─────────────────────────────────────

    def _on_webadmin_cek(self):
        """
        webadmin-nakitAkim REST API'yi tetikler — progress dialog ile.
        """
        start_qd = self._start_date.date()
        end_qd   = self._end_date.date()

        if start_qd > end_qd:
            self._show_status("⚠️  Başlangıç tarihi bitiş tarihinden büyük olamaz.", "#92400e")
            return

        start_str = f"{start_qd.year():04d}-{start_qd.month():02d}-{start_qd.day():02d}"
        end_str   = f"{end_qd.year():04d}-{end_qd.month():02d}-{end_qd.day():02d}"

        dlg = WebAdminSyncDialog(
            userid=self._userid,
            start_str=start_str,
            end_str=end_str,
            parent=self,
        )
        dlg.exec()

        # Dialog kapandıktan sonra kart statusını güncelle
        self._show_status(
            "🌐  webadmin işlemi tamamlandı. Sonuçlar dialog'da gösterildi.",
            "#6366f1"
        )

    def _on_webadmin_done(self, r: dict):
        # Artık kullanılmıyor — WebAdminSyncDialog içinde hallediliyor
        pass

    # ── Excel Import — Seçim Dialogu ─────────────────────────────────────────

    def _on_excel_import(self):
        """Hangi verinin içe aktarılacağını soran küçük seçim dialogu."""
        dlg = QDialog(self)
        dlg.setWindowTitle("📂  Excel Import — Tür Seçin")
        dlg.setFixedWidth(320)
        dlg.setModal(True)
        dlg.setStyleSheet("""
            QDialog {
                background: #f8fafc;
                border-radius: 12px;
            }
            QLabel {
                background: transparent;
                font-family: 'Segoe UI', Arial, sans-serif;
            }
            QPushButton {
                border: none;
                border-radius: 10px;
                font-size: 13px;
                font-weight: 700;
                padding: 14px 20px;
                text-align: left;
            }
            QPushButton:hover { opacity: 0.88; }
        """)

        vlay = QVBoxLayout(dlg)
        vlay.setContentsMargins(22, 20, 22, 20)
        vlay.setSpacing(12)

        # Başlık
        title = QLabel("İçe aktarmak istediğiniz veri türünü seçin:")
        title.setStyleSheet("font-size:13px;color:#374151;font-weight:600;")
        title.setWordWrap(True)
        vlay.addWidget(title)

        vlay.addSpacing(4)

        # ── Banka Hareketleri ──
        banka_btn = QPushButton("🏦  Banka Hareketleri Çek")
        banka_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #1e40af, stop:1 #1d4ed8);
                color: white;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #1e3a8a, stop:1 #1e40af);
            }
        """)
        banka_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        banka_btn.setFixedHeight(52)
        banka_btn.clicked.connect(lambda: (dlg.accept(), self._on_banka_excel_import()))
        vlay.addWidget(banka_btn)

        # ── Pos Hareketleri ──
        pos_btn = QPushButton("💳  Pos Hareketleri Çek")
        pos_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #059669, stop:1 #047857);
                color: white;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #047857, stop:1 #065f46);
            }
        """)
        pos_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        pos_btn.setFixedHeight(52)
        pos_btn.clicked.connect(lambda: (dlg.accept(), self._on_pos_excel_import()))
        vlay.addWidget(pos_btn)

        vlay.addSpacing(4)

        # ── İptal ──
        iptal_btn = QPushButton("✕  İptal")
        iptal_btn.setStyleSheet("""
            QPushButton {
                background: #e2e8f0; color: #475569;
                font-size:12px; font-weight:600;
                padding: 8px; text-align:center;
            }
            QPushButton:hover { background: #cbd5e1; }
        """)
        iptal_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        iptal_btn.clicked.connect(dlg.reject)
        vlay.addWidget(iptal_btn)

        dlg.exec()

    def _on_banka_excel_import(self):
        """VOMSİS Banka Hareketleri Excel Import dialog'unu açar."""
        dlg = WomsisExcelImportDialog(userid=self._userid, parent=self)
        dlg.exec()

    def _on_pos_excel_import(self):
        """VOMSİS POS Hareketleri Excel Import dialog'unu açar."""
        dlg = WomsisPosExcelImportDialog(userid=self._userid, parent=self)
        dlg.exec()

    # ── Manuel Toplu İşle ───────────────────────────────────────────────────

    def _on_manuel_toplu_isle(self):
        """
        PHP: #btnManuelTopluIsle click → showMtopluisle()
        ManuelTopluIsleDialog'ı mevcut API bilgileriyle açar.
        """
        url    = self._url_inp.text().strip()
        appkey = self._key_inp.text().strip()
        seckey = self._sec_inp.text().strip()
        dlg = ManuelTopluIsleDialog(
            userid=self._userid,
            musterino=self._musterino,
            api_base=url,
            app_key=appkey,
            app_secret=seckey,
            parent=self
        )
        dlg.exec()

    # ── Yardımcı metotlar ────────────────────────────────────────────────────

    def _set_busy(self, busy: bool, msg: str = ""):
        self._kontrol_btn.setEnabled(not busy)
        self._kaydet_btn.setEnabled(not busy)
        self._isle_btn.setEnabled(not busy)
        self._webadmin_btn.setEnabled(not busy)
        if busy and msg:
            self._show_status(msg, "#6366f1")

    def _show_status(self, text: str, color: str = "#374151"):
        self._status_lbl.setText(text)
        self._status_lbl.setStyleSheet(
            f"font-size:12px;color:{color};padding:4px 0;"
        )
        self._status_lbl.show()


# ── Fatura Yönetim Kartı ──────────────────────────────────────────────────────

class FaturaYonetimCard(QFrame):
    """Yıl + mod seçerek fatura kayıtlarını silen yönetim paneli."""

    def __init__(self, userid: int, musterino: int = 1, parent=None):
        super().__init__(parent)
        self._userid = userid
        self.setStyleSheet(
            "QFrame{background:white;border:1.5px solid #fee2e2;"
            "border-radius:14px;}"
        )
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(22, 20, 22, 20)
        root.setSpacing(14)

        # Başlık
        h = QHBoxLayout()
        ic = QLabel("🗑️")
        ic.setStyleSheet("font-size:20px;")
        h.addWidget(ic)
        t = QLabel("Fatura Yönetimi")
        t.setStyleSheet("font-size:14px;font-weight:700;color:#000000;letter-spacing:.5px;")
        h.addWidget(t)
        h.addStretch()
        root.addLayout(h)

        # Uyarı chip
        warn = QLabel(
            "⚠️  Silinen fatura kayıtları geri alınamaz. "
            "Silmeden önce doğru yıl ve türü seçtiğinizden emin olun."
        )
        warn.setWordWrap(True)
        warn.setStyleSheet(
            "background:#fff3cd;color:#92400e;border-radius:6px;"
            "padding:9px 12px;font-size:12px;border:none;"
        )
        root.addWidget(warn)

        # Seçiciler
        sel_row = QHBoxLayout()
        sel_row.setSpacing(12)

        yil_lbl = QLabel("Yıl:")
        yil_lbl.setStyleSheet("font-size:13px;font-weight:600;color:#000000;")
        sel_row.addWidget(yil_lbl)

        self._yil_combo = QComboBox()
        self._yil_combo.setFixedHeight(34)
        self._yil_combo.setFixedWidth(110)
        self._yil_combo.addItem("Tüm Yıllar", None)
        for y in range(CURRENT_YEAR + 1, CURRENT_YEAR - 6, -1):
            self._yil_combo.addItem(str(y), y)
        idx = self._yil_combo.findData(CURRENT_YEAR)
        if idx >= 0:
            self._yil_combo.setCurrentIndex(idx)
        self._yil_combo.setStyleSheet(self._combo_style())
        self._yil_combo.currentIndexChanged.connect(self._on_yil_changed)
        sel_row.addWidget(self._yil_combo)

        ay_lbl = QLabel("Ay:")
        ay_lbl.setStyleSheet("font-size:13px;font-weight:600;color:#000000;")
        sel_row.addWidget(ay_lbl)

        self._ay_combo = QComboBox()
        self._ay_combo.setFixedHeight(34)
        self._ay_combo.setFixedWidth(130)
        AYLAR = [
            (None, "Tüm Aylar"),
            (1, "Ocak"), (2, "Şubat"), (3, "Mart"),
            (4, "Nisan"), (5, "Mayıs"), (6, "Haziran"),
            (7, "Temmuz"), (8, "Ağustos"), (9, "Eylül"),
            (10, "Ekim"), (11, "Kasım"), (12, "Aralık"),
        ]
        for val, txt in AYLAR:
            self._ay_combo.addItem(txt, val)
        self._ay_combo.setStyleSheet(self._combo_style())
        self._ay_combo.currentIndexChanged.connect(self._refresh_count)
        sel_row.addWidget(self._ay_combo)

        mod_lbl = QLabel("Fatura Türü:")
        mod_lbl.setStyleSheet("font-size:13px;font-weight:600;color:#000000;")
        sel_row.addWidget(mod_lbl)

        self._mod_combo = QComboBox()
        self._mod_combo.setFixedHeight(34)
        self._mod_combo.setFixedWidth(220)
        self._mod_combo.addItem("Tüm Türler", None)
        self._mod_combo.addItem("↑  Kesilen Faturalar (Gelir)", "gelir")
        self._mod_combo.addItem("↓  Gelen Faturalar (Gider)", "gider")
        self._mod_combo.setStyleSheet(self._combo_style())
        self._mod_combo.currentIndexChanged.connect(self._refresh_count)
        sel_row.addWidget(self._mod_combo)

        sel_row.addStretch()
        root.addLayout(sel_row)

        # Kayıt sayısı etiketi
        self._count_lbl = QLabel("")
        self._count_lbl.setTextFormat(Qt.TextFormat.RichText)
        self._count_lbl.setStyleSheet(
            "font-size:12px;color:#374151;background:#f3f4f6;"
            "border-radius:6px;padding:6px 12px;border:none;"
        )
        root.addWidget(self._count_lbl)

        # Sil butonu
        btn_row = QHBoxLayout()
        self._sil_btn = QPushButton("🗑️  Seçili Faturaları Sil")
        self._sil_btn.setFixedHeight(38)
        self._sil_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._sil_btn.setStyleSheet(
            "QPushButton{background:#dc2626;color:white;border:none;"
            "border-radius:9px;font-size:13px;font-weight:600;"
            "padding:0 18px;letter-spacing:.5px;}"
            "QPushButton:hover{background:#b91c1c;}"
            "QPushButton:disabled{background:#fca5a5;color:#fee2e2;}"
        )
        self._sil_btn.clicked.connect(self._on_sil)
        btn_row.addWidget(self._sil_btn)
        btn_row.addStretch()
        root.addLayout(btn_row)

        self._refresh_count()

    def _combo_style(self) -> str:
        return (
            f"QComboBox{{background:#f8fafc;border:1.5px solid {COLORS.get('border','#e2e8f0')};"
            f"border-radius:8px;padding:0 10px;font-size:13px;color:#000000;}}"
            f"QComboBox::drop-down{{border:none;}}"
            f"QComboBox QAbstractItemView{{background:white;color:#1f2937;"
            f"selection-background-color:#dbeafe;selection-color:#1e40af;}}"
        )

    def _on_yil_changed(self):
        yil = self._yil_combo.currentData()
        self._ay_combo.setEnabled(yil is not None)
        if yil is None:
            self._ay_combo.setCurrentIndex(0)
        self._refresh_count()

    def _get_count(self) -> int:
        yil = self._yil_combo.currentData()
        ay  = self._ay_combo.currentData()
        mod = self._mod_combo.currentData()
        conn = get_connection()
        try:
            q = "SELECT COUNT(*) FROM faturalar WHERE userid=?"
            params = [self._userid]
            if yil:
                q += " AND substr(tarih,1,4)=?"
                params.append(str(yil))
                if ay:
                    q += " AND substr(tarih,6,2)=?"
                    params.append(f"{ay:02d}")
            if mod:
                q += " AND gelirGiderMod=?"
                params.append(mod)
            return conn.execute(q, params).fetchone()[0]
        finally:
            conn.close()

    def _refresh_count(self):
        count  = self._get_count()
        kriter = (
            f"{self._yil_combo.currentText()} / "
            f"{self._ay_combo.currentText()} / "
            f"{self._mod_combo.currentText()}"
        )
        if count == 0:
            self._count_lbl.setText(
                f"📊  <i>Seçilen kriterde ({kriter}) hiç fatura kaydı bulunamadı.</i>"
            )
        else:
            self._count_lbl.setText(
                f"📊  Seçilen kriterde ({kriter}): "
                f"<b style='color:#dc2626'>{count:,} fatura kaydı</b> bulunuyor."
            )
        self._sil_btn.setEnabled(count > 0)

    def _on_sil(self):
        count    = self._get_count()
        yil_txt  = self._yil_combo.currentText()
        ay_txt   = self._ay_combo.currentText()
        mod_txt  = self._mod_combo.currentText()
        kriter   = f"{yil_txt} / {ay_txt} / {mod_txt}"

        if count == 0:
            QMessageBox.information(self, "Bilgi", "Seçilen kriterde silinecek fatura yok.")
            return

        # ─ 1. Onay ─
        dlg = QMessageBox(self)
        dlg.setWindowTitle("⚠️  Fatura Silme Onayı")
        dlg.setIcon(QMessageBox.Icon.Warning)
        dlg.setText(
            f"<b>{yil_txt} / {mod_txt}</b> kriterindeki<br>"
            f"<b style='color:#dc2626;font-size:15px'>{count:,} fatura kaydı</b> silinecek.<br><br>"
            "<b>Bu işlem geri alınamaz!</b>"
        )
        dlg.setInformativeText("Devam etmek istiyor musunuz?")
        dlg.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel)
        dlg.setDefaultButton(QMessageBox.StandardButton.Cancel)
        dlg.button(QMessageBox.StandardButton.Yes).setText("🗑️  Evet, Sil")
        dlg.button(QMessageBox.StandardButton.Cancel).setText("Vazgeç")

        if dlg.exec() != QMessageBox.StandardButton.Yes:
            return

        # ─ 2. İkinci onay ─
        second = QMessageBox(self)
        second.setWindowTitle("🔴  Son Onay — Geri Alınamaz")
        second.setIcon(QMessageBox.Icon.Critical)
        second.setText(
            f"<b style='color:#dc2626'>{count:,} fatura kaydı</b> kalıcı olarak silinecek.<br><br>"
            "Emin misiniz? Bu işlemi geri alamazsınız."
        )
        second.setStandardButtons(QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel)
        second.setDefaultButton(QMessageBox.StandardButton.Cancel)
        second.button(QMessageBox.StandardButton.Ok).setText("🔴  Kalıcı Olarak Sil")
        second.button(QMessageBox.StandardButton.Cancel).setText("Vazgeç")

        if second.exec() != QMessageBox.StandardButton.Ok:
            return

        # ─ 3. Sil ─
        yil_val = self._yil_combo.currentData()
        ay_val  = self._ay_combo.currentData()
        mod_val = self._mod_combo.currentData()
        conn = get_connection()
        try:
            q = "DELETE FROM faturalar WHERE userid=?"
            params = [self._userid]
            if yil_val:
                q += " AND substr(tarih,1,4)=?"
                params.append(str(yil_val))
                if ay_val:
                    q += " AND substr(tarih,6,2)=?"
                    params.append(f"{ay_val:02d}")
            if mod_val:
                q += " AND gelirGiderMod=?"
                params.append(mod_val)
            conn.execute(q, params)
            conn.commit()
        except Exception as exc:
            conn.rollback()
            QMessageBox.critical(self, "Hata", f"Silme işlemi başarısız:\n{exc}")
            return
        finally:
            conn.close()

        QMessageBox.information(
            self, "✅  Tamamlandı",
            f"{count:,} fatura kaydı başarıyla silindi.\n"
            "Artık doğru klasörden yeniden aktarabilirsiniz."
        )
        self._refresh_count()



# ── MOY Workers ──────────────────────────────────────────────────────────────

class MoyTestWorker(QThread):
    """'Test' butonu — Moy MySQL sunucusuna bağlanıp kullanıcıyı doğrular."""
    result = pyqtSignal(dict)   # {success, data:[{adi,soyadim,kayitNo}]}

    def __init__(self, host: str, user: str, password: str, musteri_no: int):
        super().__init__()
        self._host       = host
        self._user       = user
        self._password   = password
        self._musteri_no = musteri_no

    def run(self):
        from services.moy_service import moy_test_connection
        r = moy_test_connection(self._host, self._user, self._password, self._musteri_no)
        self.result.emit(r)


class MoyKaydetWorker(QThread):
    """'Verileri Çek' butonu — seçilen yıla ait hareketleri aktarır."""
    progress = pyqtSignal(str)
    finished = pyqtSignal(dict)   # {success, message, eklenen}

    def __init__(self, musteri_no: int, yil: int):
        super().__init__()
        self._musteri_no = musteri_no
        self._yil        = yil

    def run(self):
        import subprocess, sys, os, datetime

        # ── Otomatik Git Yedek (veri çekmeden ÖNCE) ──────────────────────
        try:
            repo_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            simdi    = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            self.progress.emit("💾  Git yedeği alınıyor...")

            # Sadece veritabanı ile ilgili dosyaları stage'e ekle
            subprocess.run(
                ["git", "add", "-A"],
                cwd=repo_dir, capture_output=True, timeout=15
            )
            commit_msg = (
                f"backup: Moy veri çekimi öncesi otomatik yedek\n\n"
                f"Müşteri No: {self._musteri_no} | Yıl: {self._yil} | {simdi}"
            )
            result = subprocess.run(
                ["git", "commit", "-m", commit_msg,
                 "--allow-empty"],   # Değişiklik yoksa bile commit oluştur
                cwd=repo_dir, capture_output=True, text=True, timeout=15
            )
            if result.returncode == 0:
                sha = result.stdout.strip().split()[1][:7] if result.stdout else "?"
                self.progress.emit(f"✅  Git yedeği alındı → {sha}")
            else:
                self.progress.emit("⚠️  Git yedeği alınamadı (devam ediliyor)")
        except Exception as _ge:
            # Yedek başarısız olsa da veri çekmeye devam et
            self.progress.emit(f"⚠️  Git yedek hatası: {_ge} (devam ediliyor)")
        # ─────────────────────────────────────────────────────────────────

        from services.moy_service import moy_kaydet_veriler
        r = moy_kaydet_veriler(
            self._musteri_no, self._yil,
            progress_cb=lambda msg: self.progress.emit(msg)
        )
        self.finished.emit(r)


# ── Moy Kartı ─────────────────────────────────────────────────────────────────

class MoyCard(QFrame):
    """
    PHP ayarlar.php → Eklentiler → Moy div'inin PyQt6 karşılığı.
    Alanlar  : URL (IP), Kullanıcı Adı, Şifre
    Butonlar : Test | Verileri Çek (Yıl seçerek)
    """
    # Veri başarıyla çekildiğinde dashboard’u haberdar eder
    data_changed = pyqtSignal()

    def __init__(self, userid: int, musterino: int = 1, parent=None):
        super().__init__(parent)
        self._userid         = userid
        self._musterino      = musterino
        self._test_worker: MoyTestWorker   | None = None
        self._kaydet_worker: MoyKaydetWorker | None = None
        self._test_basarili  = False
        self.setStyleSheet(
            "QFrame{background:white;border:1.5px solid #d1fae5;"
            "border-radius:14px;}"
        )
        self._build()
        self._load()

    # ─ Stiller ────────────────────────────────────────────────────────

    def _inp(self) -> str:
        return (
            "QLineEdit{background:#f8fafc;border:1.5px solid #e2e8f0;"
            "border-radius:8px;padding:0 10px;font-size:13px;color:#1f2937;}"
            "QLineEdit:focus{border-color:#10b981;}"
            "QLineEdit[error='true']{border-color:#ef4444;}"
        )

    def _lbl(self) -> str:
        return "font-size:12px;font-weight:600;color:#374151;"

    def _btn(self, color: str) -> str:
        return (
            f"QPushButton{{background:{color};color:white;border:none;"
            f"border-radius:9px;font-size:13px;font-weight:600;padding:0 16px;}}"
            f"QPushButton:hover{{background:{color}cc;}}"
            f"QPushButton:disabled{{background:#cbd5e1;color:#94a3b8;}}"
        )

    # ─ UI inşası ─────────────────────────────────────────────────────

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(22, 20, 22, 20)
        root.setSpacing(14)

        # Başlık
        h = QHBoxLayout()
        ic = QLabel("🏢")
        ic.setStyleSheet("font-size:20px;")
        h.addWidget(ic)
        t = QLabel("Moy")
        t.setStyleSheet("font-size:14px;font-weight:700;color:#000;letter-spacing:.5px;")
        h.addWidget(t)
        self._aktif_chip = QLabel("✔  AKTİF")
        self._aktif_chip.setStyleSheet(
            "background:#d1fae5;color:#065f46;border-radius:5px;"
            "padding:2px 8px;font-size:11px;font-weight:700;border:none;"
        )
        self._aktif_chip.hide()
        h.addWidget(self._aktif_chip)
        h.addStretch()
        root.addLayout(h)

        # Bilgi
        info = QLabel("💡  Moy muhasebe programından yıllık vergi hareketlerini çekebilirsiniz.")
        info.setWordWrap(True)
        info.setStyleSheet(
            "background:#ecfdf5;color:#065f46;border-radius:6px;"
            "padding:9px 12px;font-size:12px;border:none;"
        )
        root.addWidget(info)

        # ─ URL (IP) satırı ─
        url_row = QHBoxLayout()
        url_row.setSpacing(10)

        proto_lbl = QLabel("Protokol")
        proto_lbl.setStyleSheet(self._lbl())
        self._proto_combo = QComboBox()
        self._proto_combo.addItems(["https://", "http://"])
        self._proto_combo.setFixedHeight(34)
        self._proto_combo.setFixedWidth(90)
        self._proto_combo.setStyleSheet(
            "QComboBox{background:#f8fafc;border:1.5px solid #e2e8f0;"
            "border-radius:8px;padding:0 6px;font-size:12px;color:#1f2937;}"
        )

        ip_lbl = QLabel("URL / IP Adresi")
        ip_lbl.setStyleSheet(self._lbl())
        self._ip_inp = QLineEdit()
        self._ip_inp.setFixedHeight(34)
        self._ip_inp.setPlaceholderText("123.121.121.122")
        self._ip_inp.setStyleSheet(self._inp())
        self._ip_inp.textChanged.connect(self._on_input_changed)

        url_col = QVBoxLayout()
        url_col.addWidget(ip_lbl)
        url_col.addWidget(self._ip_inp)

        proto_col = QVBoxLayout()
        proto_col.addWidget(proto_lbl)
        proto_col.addWidget(self._proto_combo)

        url_row.addLayout(proto_col)
        url_row.addLayout(url_col, 3)
        root.addLayout(url_row)

        # ─ Kullanıcı Adı + Şifre satırı ─
        cred_row = QHBoxLayout()
        cred_row.setSpacing(12)

        u_col = QVBoxLayout()
        u_lbl = QLabel("Kullanıcı Adı")
        u_lbl.setStyleSheet(self._lbl())
        self._user_inp = QLineEdit()
        self._user_inp.setFixedHeight(34)
        self._user_inp.setPlaceholderText("Moy Kullanıcı Adı")
        self._user_inp.setStyleSheet(self._inp())
        self._user_inp.textChanged.connect(self._on_input_changed)
        u_col.addWidget(u_lbl)
        u_col.addWidget(self._user_inp)
        cred_row.addLayout(u_col)

        p_col = QVBoxLayout()
        p_lbl = QLabel("Şifre")
        p_lbl.setStyleSheet(self._lbl())
        self._pass_inp = QLineEdit()
        self._pass_inp.setFixedHeight(34)
        self._pass_inp.setPlaceholderText("Moy Şifreniz")
        self._pass_inp.setStyleSheet(self._inp())
        self._pass_inp.textChanged.connect(self._on_input_changed)
        p_col.addWidget(p_lbl)
        p_col.addWidget(self._pass_inp)
        cred_row.addLayout(p_col)
        root.addLayout(cred_row)

        # ─ Butonlar (Test) ─
        btn_row1 = QHBoxLayout()
        btn_row1.setSpacing(10)
        self._test_btn = QPushButton("🔍  Test")
        self._test_btn.setFixedHeight(38)
        self._test_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._test_btn.setStyleSheet(self._btn("#059669"))
        self._test_btn.clicked.connect(self._on_test)
        btn_row1.addWidget(self._test_btn)
        btn_row1.addStretch()
        root.addLayout(btn_row1)

        # ─ Sonuc satırı (PHP: .moysonucResponse) ─
        self._sonuc_lbl = QLabel("")
        self._sonuc_lbl.setWordWrap(True)
        self._sonuc_lbl.setStyleSheet("font-size:12px;color:#374151;")
        self._sonuc_lbl.hide()
        root.addWidget(self._sonuc_lbl)

        # ─ Ozel bilgi bolumu (PHP: .moyozelbilgi) ─ başlangıçta gizli ─
        self._ozel_frame = QFrame()
        self._ozel_frame.setStyleSheet(
            "QFrame{background:#f0fdf4;border:1px solid #bbf7d0;"
            "border-radius:10px;}"
        )
        ozel_layout = QVBoxLayout(self._ozel_frame)
        ozel_layout.setContentsMargins(14, 12, 14, 12)
        ozel_layout.setSpacing(10)

        self._ozel_info = QLabel(
            "Şimdi yıl seçip hangi aralıkta veri aktarılacağını seçin, "
            "daha sonra 'Verileri Çek' butonuna tıklayın."
        )
        self._ozel_info.setWordWrap(True)
        self._ozel_info.setStyleSheet("font-size:12px;color:#065f46;")
        ozel_layout.addWidget(self._ozel_info)

        self._ozel_veri = QLabel("")
        self._ozel_veri.setStyleSheet("font-size:13px;font-weight:600;color:#065f46;")
        self._ozel_veri.hide()
        ozel_layout.addWidget(self._ozel_veri)

        # Yıl seçimi (PHP: #yearSelect — son 5 yıl)
        yil_row = QHBoxLayout()
        yil_lbl = QLabel("📅  Yıl Seçin:")
        yil_lbl.setStyleSheet("font-size:12px;font-weight:600;color:#374151;")
        yil_row.addWidget(yil_lbl)
        self._yil_sec = QComboBox()
        self._yil_sec.setFixedHeight(32)
        self._yil_sec.setFixedWidth(100)
        self._yil_sec.setStyleSheet(
            "QComboBox{background:white;border:1.5px solid #d1fae5;"
            "border-radius:8px;padding:0 8px;font-size:13px;color:#1f2937;}"
        )
        import datetime as _dt
        cy = _dt.datetime.now().year
        for i in range(5):
            yr = cy - i
            self._yil_sec.addItem(str(yr), yr)
        yil_row.addWidget(self._yil_sec)
        yil_row.addStretch()
        ozel_layout.addLayout(yil_row)

        # Verileri Çek butonu — siyah arka plan
        self._cek_btn = QPushButton("⚡  Verileri Çek")
        self._cek_btn.setFixedHeight(38)
        self._cek_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._cek_btn.setStyleSheet(
            "QPushButton{background:#1f2937;color:white;border:none;"
            "border-radius:9px;font-size:13px;font-weight:600;padding:0 16px;}"
            "QPushButton:hover{background:#111827;}"
            "QPushButton:disabled{background:#cbd5e1;color:#94a3b8;}"
        )
        self._cek_btn.setEnabled(False)   # Başlangıçta pasif (PHP .pasif sınıfı)
        self._cek_btn.clicked.connect(self._on_cek)
        ozel_layout.addWidget(self._cek_btn)

        # Detay alanı — çekilen verileri listeler
        self._detay_lbl = QLabel("")
        self._detay_lbl.setWordWrap(True)
        self._detay_lbl.setTextFormat(Qt.TextFormat.RichText)
        self._detay_lbl.setStyleSheet(
            "font-size:11px;color:#1f2937;"
            "background:#f0fdf4;border-radius:6px;padding:6px;"
        )
        self._detay_lbl.hide()
        ozel_layout.addWidget(self._detay_lbl)

        self._ozel_frame.hide()   # PHP: .moyozelbilgi — test başarılı olunca görünür
        root.addWidget(self._ozel_frame)

    # ─ Veri yükleme ────────────────────────────────────────────────

    def _load(self):
        from services.moy_service import get_moy_bilgileri
        bilgi = get_moy_bilgileri(self._userid)
        if bilgi.get("success"):
            self._ip_inp.setText(bilgi.get("url") or "")
            self._user_inp.setText(bilgi.get("username") or "")
            self._pass_inp.setText(bilgi.get("sifre") or "")
            self._aktif_chip.show()

    def refresh(self):
        self._load()

    # ─ Input değişimi (PHP: .moyinput.on('input') — Kaydet butonunu pasif yap) ─

    def _on_input_changed(self):
        if self._test_basarili:
            self._test_basarili = False
            self._cek_btn.setEnabled(False)
            self._ozel_frame.hide()

    # ─ Test butonu (PHP: btnMoyTest click) ────────────────────────

    def _on_test(self):
        host  = self._ip_inp.text().strip()
        user  = self._user_inp.text().strip()
        passwd = self._pass_inp.text().strip()

        # Validasyon (PHP: validateMoy())
        import re as _re
        ip_rx = _re.compile(
            r"^(25[0-5]|2[0-4]\d|1\d\d|[1-9]?\d)"
            r"(\.(25[0-5]|2[0-4]\d|1\d\d|[1-9]?\d)){3}$"
        )
        if not host:
            self._show_sonuc("❌  IP boş olamaz", "#ef4444")
            return
        if not ip_rx.match(host):
            self._show_sonuc("❌  Bu IP hatalı", "#ef4444")
            return
        if not user:
            self._show_sonuc("❌  Kullanıcı adı boş olamaz", "#ef4444")
            return
        if not passwd:
            self._show_sonuc("❌  Şifre boş olamaz", "#ef4444")
            return

        self._set_busy(True, "🔍  Kontrol ediliyor...")

        self._test_worker = MoyTestWorker(host, user, passwd, self._userid)
        self._test_worker.result.connect(self._on_test_result)
        self._test_worker.start()

    def _on_test_result(self, r: dict):
        self._set_busy(False)
        if r.get("success"):
            data  = r.get("data", [])
            first = data[0] if data else {}
            ad    = first.get("adi", "") + " " + first.get("soyadim", "")
            self._show_sonuc("✅  Başarılı!", "#059669")
            self._ozel_veri.setText(f"👤  Moy bilgileriniz onaylandı: {ad.strip()}")
            self._ozel_veri.show()
            self._ozel_frame.show()
            self._cek_btn.setEnabled(True)   # PHP: addClass('aktif')
            self._aktif_chip.show()
            self._test_basarili = True
        else:
            msg = r.get("message", "Hata oluştu.")
            self._show_sonuc(f"❌  {msg}", "#ef4444")
            self._cek_btn.setEnabled(False)

    # ─ Verileri Çek (PHP: btnMoyKaydet click) ────────────────────

    def _on_cek(self):
        if not self._test_basarili:
            return
        yil = self._yil_sec.currentData()
        self._set_busy(True, f"⚡  {yil} yılı hareketleri alınıyor...")
        self._ozel_veri.hide()

        self._kaydet_worker = MoyKaydetWorker(self._musterino, yil)
        self._kaydet_worker.progress.connect(
            lambda msg: self._show_sonuc(msg, "#0891b2")
        )
        self._kaydet_worker.finished.connect(self._on_cek_done)
        self._kaydet_worker.start()

    def _on_cek_done(self, r: dict):
        self._set_busy(False)
        if r.get("success"):
            eklenen  = r.get("eklenen", 0)
            detaylar = r.get("detaylar", [])

            # Özet metin
            msg = f"Başarıyla {eklenen} tane kayıt eklendi"
            self._ozel_veri.setText(f"✅  {msg}")
            self._ozel_veri.show()
            self._show_sonuc("", "")

            # Detay tablosu
            if detaylar:
                satirlar = "".join(
                    f"<tr>"
                    f"<td style='padding:2px 8px;color:#065f46;font-weight:600;'>{d['kod']}</td>"
                    f"<td style='padding:2px 8px;'>{d['tarih']}</td>"
                    f"<td style='padding:2px 8px;text-align:right;'>{d['tutar']:.2f}</td>"
                    f"</tr>"
                    for d in detaylar
                )
                html = (
                    f"<b>Çekilen Veriler ({len(detaylar)} satır):</b><br>"
                    f"<table width='100%' cellspacing='0'>"
                    f"<tr style='background:#d1fae5;'>"
                    f"<th style='text-align:left;padding:2px 8px;'>Hesap Kodu</th>"
                    f"<th style='text-align:left;padding:2px 8px;'>Tarih</th>"
                    f"<th style='text-align:right;padding:2px 8px;'>Tutar</th>"
                    f"</tr>"
                    f"{satirlar}"
                    f"</table>"
                )
                self._detay_lbl.setText(html)
                self._detay_lbl.show()
            else:
                self._detay_lbl.hide()

            # Dashboard’u otomatik yenile
            self.data_changed.emit()
        else:
            self._show_sonuc(f"❌  {r.get('message','Hata')}", "#ef4444")
            self._detay_lbl.hide()

    # ─ Yardımcılar ──────────────────────────────────────────────

    def _set_busy(self, busy: bool, msg: str = ""):
        self._test_btn.setEnabled(not busy)
        self._cek_btn.setEnabled(not busy and self._test_basarili)
        if busy and msg:
            self._show_sonuc(msg, "#6366f1")

    def _show_sonuc(self, text: str, color: str):
        if not text:
            self._sonuc_lbl.hide()
            return
        self._sonuc_lbl.setText(text)
        self._sonuc_lbl.setStyleSheet(f"font-size:12px;color:{color};padding:4px 0;")
        self._sonuc_lbl.show()


# ── Vergi Muhtasar Workers ───────────────────────────────────────────────────

class VergiMuhtasarYukleWorker(QThread):
    """CSV toplu yükleme arka plan iş parçacığı (PHP vergiMuhtasarTopluYukle.php)."""
    finished = pyqtSignal(dict)  # {'success', 'message', 'added', 'updated', 'skipped'}

    def __init__(self, userid: int, dosya_yolu: str):
        super().__init__()
        self._userid     = userid
        self._dosya_yolu = dosya_yolu

    def run(self):
        from services.vergi_muhtasar_service import toplu_yukle_csv
        r = toplu_yukle_csv(self._userid, self._dosya_yolu)
        self.finished.emit(r)


# ── Vergi Muhtasar Kartı ─────────────────────────────────────────────────────

class VergiMuhtasarCard(QFrame):
    """
    PHP ayarlar.php → Eklentiler → Vergi Muhtasar bölümünün PyQt6 karşılığı.

    Özellikler (PHP ile birebir):
    - CSV toplu yükle (vergiMuhtasarTopluYukle.php)
    - Dönem & açıklama filtresi
    - Kayıt listesi (DataTable karşılığı QTableWidget)
    - Inline düzenleme: gaytutar, vergkestutar (çift tık)
    - Silme (sil butonu)
    - Özet toplamlar barı: Gayri Safi Tutar | Vergi Kesinti | Fark
    - Şema CSV indir
    """

    # Sütun indeksleri (PHP DataTable sütun sırası)
    COL_HESAP  = 0
    COL_ACK    = 1
    COL_DONEM  = 2
    COL_GAY    = 3   # gaytutar — inline düzenlenebilir
    COL_VERG   = 4   # vergkestutar — inline düzenlenebilir
    COL_SIL    = 5

    def __init__(self, userid: int, musterino: int = 1, parent=None):
        super().__init__(parent)
        self._userid       = userid
        self._yukle_worker: VergiMuhtasarYukleWorker | None = None
        self._all_data: list[dict] = []   # Cache: tüm kayıtlar
        self._row_ids:  list[int]  = []   # Tablo satır → DB id eşleşmesi

        self.setStyleSheet(
            "QFrame{background:white;border:1.5px solid #e0e7ff;"
            "border-radius:14px;}"
        )
        self._build()
        self.refresh()

    # ── Stil yardımcıları ─────────────────────────────────────────────────

    def _lbl_style(self) -> str:
        return "font-size:12px;font-weight:600;color:#374151;"

    def _btn_style(self, color: str) -> str:
        return (
            f"QPushButton{{background:{color};color:white;border:none;"
            f"border-radius:9px;font-size:13px;font-weight:600;"
            f"padding:0 14px;letter-spacing:.4px;}}"
            f"QPushButton:hover{{background:{color}cc;}}"
            f"QPushButton:disabled{{background:#cbd5e1;color:#94a3b8;}}"
        )

    def _combo_style(self) -> str:
        return (
            "QComboBox{background:#f8fafc;border:1.5px solid #e2e8f0;"
            "border-radius:8px;padding:0 10px;font-size:13px;color:#1f2937;}"
            "QComboBox::drop-down{border:none;}"
            "QComboBox QAbstractItemView{background:white;color:#1f2937;"
            "selection-background-color:#dbeafe;selection-color:#1e40af;}"
        )

    @staticmethod
    def _fmt_tutar(val) -> str:
        """float/None → '1.234,56 ₺' ya da '—' (PHP vmFormatTutar karşılığı)."""
        if val is None or val == "":
            return "—"
        try:
            f = float(str(val).replace(",", "."))
            return f"{f:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") + " ₺"
        except (ValueError, TypeError):
            return "—"

    @staticmethod
    def _parse_float(val) -> float:
        """Herhangi formatı float'a çevirir (PHP vmParseFloat karşılığı)."""
        if val is None or val == "":
            return 0.0
        try:
            return float(str(val).replace(",", "."))
        except (ValueError, TypeError):
            return 0.0

    # ── UI İnşası ─────────────────────────────────────────────────────────

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(22, 20, 22, 20)
        root.setSpacing(14)

        # Başlık satırı
        h = QHBoxLayout()
        ic = QLabel("🧾")
        ic.setStyleSheet("font-size:20px;")
        h.addWidget(ic)
        t = QLabel("Vergi Muhtasar")
        t.setStyleSheet("font-size:14px;font-weight:700;color:#000000;letter-spacing:.5px;")
        h.addWidget(t)
        h.addStretch()
        root.addLayout(h)

        # Bilgi chip (PHP açıklama paragrafı)
        info = QLabel(
            "💡  Muhtasar ve prim hizmetleri beyannamesi kapsamındaki stopaj vergisi "
            "verilerini CSV dosyası ile toplu olarak içeri aktarabilirsiniz. "
            "Hesap kodu, dönem ve vergi kesinti tutarları takip edilir."
        )
        info.setWordWrap(True)
        info.setStyleSheet(
            "background:#ede9fe;color:#4c1d95;border-radius:6px;"
            "padding:9px 12px;font-size:12px;border:none;"
        )
        root.addWidget(info)

        # ── Butonlar (üst sıra) ──
        top_btn_row = QHBoxLayout()
        top_btn_row.setSpacing(8)

        # Verileri Yükle (PHP: vergiMuhtasarTopluBtn)
        self._yukle_btn = QPushButton("📤  Verileri Yükle")
        self._yukle_btn.setFixedHeight(36)
        self._yukle_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._yukle_btn.setStyleSheet(self._btn_style("#7c3aed"))
        self._yukle_btn.clicked.connect(self._on_yukle)
        top_btn_row.addWidget(self._yukle_btn)

        # Şema İndir (PHP: vmSemaIndir)
        self._sema_btn = QPushButton("⬇  Şema İndir")
        self._sema_btn.setFixedHeight(36)
        self._sema_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._sema_btn.setStyleSheet(self._btn_style("#64748b"))
        self._sema_btn.setToolTip(
            "💡 Şema dosyası — hızlı bir şekilde uygun formda veri girmeniz için"
        )
        self._sema_btn.clicked.connect(self._on_sema_indir)
        top_btn_row.addWidget(self._sema_btn)

        top_btn_row.addStretch()
        root.addLayout(top_btn_row)

        # ── Özet Toplamlar Barı (PHP: #vmTotalsBar) ──
        totals_frame = QFrame()
        totals_frame.setStyleSheet(
            "QFrame{background:#f8f9fc;border:1px solid #e8eaf0;"
            "border-radius:8px;}"
        )
        totals_layout = QHBoxLayout(totals_frame)
        totals_layout.setContentsMargins(14, 10, 14, 10)
        totals_layout.setSpacing(20)

        def _make_total_item(label_txt: str, value_id: str, color: str) -> QLabel:
            col = QVBoxLayout()
            lbl = QLabel(label_txt)
            lbl.setStyleSheet(
                "font-size:11px;color:#888;font-weight:600;letter-spacing:0.5px;"
                "border:none;background:transparent;"
            )
            val = QLabel("0,00 ₺")
            val.setStyleSheet(
                f"font-size:18px;font-weight:700;color:{color};"
                "border:none;background:transparent;"
            )
            val.setObjectName(value_id)
            col.addWidget(lbl)
            col.addWidget(val)
            wrapper = QWidget()
            wrapper.setLayout(col)
            return wrapper, val

        gay_w,  self._lbl_gay  = _make_total_item("GAYRİ SAFİ TUTAR",   "vmToplamGay",  "#2d5be3")
        verg_w, self._lbl_verg = _make_total_item("VERGİ KESİNTİ TUTARI","vmToplamVerg", "#e35d2d")
        fark_w, self._lbl_fark = _make_total_item("FARK",                  "vmToplamFark", "#22a06b")
        totals_layout.addWidget(gay_w,  1)
        totals_layout.addWidget(verg_w, 1)
        totals_layout.addWidget(fark_w, 1)
        root.addWidget(totals_frame)

        # ── Filtre Satırı (PHP: #vmDonemFilter + #vmSearchInput) ──
        filter_row = QHBoxLayout()
        filter_row.setSpacing(10)

        self._donem_combo = QComboBox()
        self._donem_combo.setFixedHeight(32)
        self._donem_combo.setMinimumWidth(130)
        self._donem_combo.addItem("Tüm Dönemler", "")
        self._donem_combo.setStyleSheet(self._combo_style())
        self._donem_combo.currentIndexChanged.connect(self._on_donem_changed)
        filter_row.addWidget(self._donem_combo)

        self._ack_combo = QComboBox()
        self._ack_combo.setFixedHeight(32)
        self._ack_combo.setMinimumWidth(130)
        self._ack_combo.addItem("Tüm Açıklamalar", "")
        self._ack_combo.setStyleSheet(self._combo_style())
        self._ack_combo.currentIndexChanged.connect(self._on_ack_changed)
        filter_row.addWidget(self._ack_combo)

        filter_row.addStretch()
        root.addLayout(filter_row)

        # ── DataTable (QTableWidget) ──
        self._table = QTableWidget()
        self._table.setColumnCount(6)
        self._table.setHorizontalHeaderLabels([
            "Hesap Kodu", "Açıklama", "Dönem",
            "Gayri Safi Tutar", "Vergi Kesinti Tutarı", "İşlem"
        ])
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        self._table.setColumnWidth(5, 60)
        self._table.verticalHeader().setVisible(False)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setAlternatingRowColors(True)
        self._table.setMinimumHeight(200)
        self._table.setStyleSheet(
            "QTableWidget{background:white;gridline-color:#f1f5f9;border:1px solid #e2e8f0;"
            "border-radius:8px;font-size:13px;color:#1f2937;}"
            "QTableWidget::item{padding:6px 10px;}"
            "QTableWidget::item:selected{background:#dbeafe;color:#1e40af;}"
            "QTableWidget::item.editing{background:#fef9c3;border:2px solid #f59e0b;}"
            "QHeaderView::section{background:#f8fafc;border:none;border-bottom:1px solid #e2e8f0;"
            "font-size:12px;font-weight:600;color:#374151;padding:8px 10px;}"
            "QTableWidget QScrollBar:horizontal{height:8px;}"
            "QTableWidget QScrollBar:vertical{width:8px;}"
        )
        # Çift tık ile inline düzenleme (PHP: dblclick 'td')
        self._table.cellDoubleClicked.connect(self._on_cell_double_click)
        root.addWidget(self._table)

        # ── Durum etiketi ──
        self._status_lbl = QLabel("")
        self._status_lbl.setWordWrap(True)
        self._status_lbl.setStyleSheet("font-size:12px;color:#374151;")
        self._status_lbl.hide()
        root.addWidget(self._status_lbl)

    # ── Veri Yükleme & Tablo Güncelleme ──────────────────────────────────

    def refresh(self):
        """Tablo + filtreler + toplamları yeniler."""
        from services.vergi_muhtasar_service import get_vergi_muhtasar
        donem  = self._donem_combo.currentData() or ""
        result = get_vergi_muhtasar(self._userid, donem)
        if not result["success"]:
            self._show_status(f"❌  {result.get('message', 'Hata')}", "#dc2626")
            return

        self._all_data = result["data"]
        self._fill_donem_combo(result["donemler"])
        self._fill_ack_combo(self._all_data)
        self._apply_filters()

    def _fill_donem_combo(self, donemler: list[str]):
        """PHP vmDonemFilter doldurma karşılığı."""
        current = self._donem_combo.currentData()
        self._donem_combo.blockSignals(True)
        self._donem_combo.clear()
        self._donem_combo.addItem("Tüm Dönemler", "")
        for d in donemler:
            self._donem_combo.addItem(d, d)
        # Eski seçimi koru
        idx = self._donem_combo.findData(current)
        if idx >= 0:
            self._donem_combo.setCurrentIndex(idx)
        self._donem_combo.blockSignals(False)

    def _fill_ack_combo(self, data: list[dict]):
        """PHP vmSearchInput doldurma karşılığı."""
        current = self._ack_combo.currentData()
        self._ack_combo.blockSignals(True)
        self._ack_combo.clear()
        self._ack_combo.addItem("Tüm Açıklamalar", "")
        seen = set()
        for row in data:
            ack = row.get("ack") or ""
            if ack and ack not in seen:
                seen.add(ack)
                self._ack_combo.addItem(ack, ack)
        idx = self._ack_combo.findData(current)
        if idx >= 0:
            self._ack_combo.setCurrentIndex(idx)
        self._ack_combo.blockSignals(False)

    def _apply_filters(self):
        """Seçili dönem + açıklamaya göre tabloyu filtreler ve doldurur."""
        donem_filter = self._donem_combo.currentData() or ""
        ack_filter   = self._ack_combo.currentData() or ""

        filtered = [
            r for r in self._all_data
            if (not donem_filter or r.get("donem", "") == donem_filter)
            and (not ack_filter  or (r.get("ack") or "") == ack_filter)
        ]
        self._fill_table(filtered)
        self._update_totals(filtered)

    def _fill_table(self, data: list[dict]):
        """QTableWidget'i veriyle doldurur."""
        self._row_ids = []
        self._table.setRowCount(0)
        self._table.setRowCount(len(data))

        for row_idx, row in enumerate(data):
            self._row_ids.append(row.get("id", -1))

            def _item(text: str, align=Qt.AlignmentFlag.AlignLeft) -> QTableWidgetItem:
                it = QTableWidgetItem(str(text) if text is not None else "")
                it.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                it.setTextAlignment(align | Qt.AlignmentFlag.AlignVCenter)
                return it

            self._table.setItem(row_idx, self.COL_HESAP, _item(row.get("hesapkodu", "")))
            self._table.setItem(row_idx, self.COL_ACK,   _item(row.get("ack", "")))
            self._table.setItem(row_idx, self.COL_DONEM, _item(row.get("donem", "")))
            self._table.setItem(
                row_idx, self.COL_GAY,
                _item(self._fmt_tutar(row.get("gaytutar")),
                      Qt.AlignmentFlag.AlignRight)
            )
            self._table.setItem(
                row_idx, self.COL_VERG,
                _item(self._fmt_tutar(row.get("vergkestutar")),
                      Qt.AlignmentFlag.AlignRight)
            )

            # Sil butonu
            sil_btn = QPushButton("🗑")
            sil_btn.setFixedSize(36, 28)
            sil_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            sil_btn.setStyleSheet(
                "QPushButton{background:#fee2e2;color:#dc2626;border:none;"
                "border-radius:6px;font-size:14px;}"
                "QPushButton:hover{background:#fecaca;}"
            )
            sil_btn.clicked.connect(lambda _, r=row_idx: self._on_sil(r))
            self._table.setCellWidget(row_idx, self.COL_SIL, sil_btn)

    def _update_totals(self, data: list[dict]):
        """Özet toplamları hesaplar ve günceller (PHP vmUpdateTotals karşılığı)."""
        gay  = sum(self._parse_float(r.get("gaytutar"))  for r in data)
        verg = sum(self._parse_float(r.get("vergkestutar")) for r in data)
        fark = gay - verg

        self._lbl_gay.setText(self._fmt_tutar(gay))
        self._lbl_verg.setText(self._fmt_tutar(verg))
        self._lbl_fark.setText(self._fmt_tutar(fark))

    # ── Dönem & Açıklama Filtresi ─────────────────────────────────────────

    def _on_donem_changed(self):
        """PHP: vmDonemFilter change → vmTable.ajax.url(…).load()"""
        self.refresh()

    def _on_ack_changed(self):
        """PHP: vmSearchInput change → vmTable.column(1).search(…).draw()"""
        self._apply_filters()

    # ── CSV Toplu Yükleme (PHP: vergiMuhtasarTopluBtn click) ─────────────

    def _on_yukle(self):
        dosya, _ = QFileDialog.getOpenFileName(
            self, "CSV Dosyası Seç", "", "CSV Dosyaları (*.csv)"
        )
        if not dosya:
            return

        self._yukle_btn.setEnabled(False)
        self._show_status("📤  CSV yükleniyor...", "#7c3aed")

        self._yukle_worker = VergiMuhtasarYukleWorker(self._userid, dosya)
        self._yukle_worker.finished.connect(self._on_yukle_done)
        self._yukle_worker.start()

    def _on_yukle_done(self, r: dict):
        self._yukle_btn.setEnabled(True)
        if r["success"]:
            msg = (
                f"✅  {r['added']} yeni kayıt eklendi, "
                f"{r['updated']} kayıt güncellendi, "
                f"{r['skipped']} satır atlandı."
            )
            self._show_status(msg, "#059669")
            self.refresh()
        else:
            self._show_status(f"❌  {r['message']}", "#dc2626")

    # ── Inline Düzenleme (PHP: dblclick 'td' — sadece col 3 & 4) ─────────

    def _on_cell_double_click(self, row: int, col: int):
        """Gayri Safi Tutar (col 3) ve Vergi Kesinti (col 4) düzenlenebilir."""
        if col not in (self.COL_GAY, self.COL_VERG):
            return
        if row >= len(self._row_ids):
            return

        kayit_id = self._row_ids[row]
        kolon    = "gaytutar" if col == self.COL_GAY else "vergkestutar"
        item     = self._table.item(row, col)
        if not item:
            return

        eski_display = item.text()
        # Türkçe tutar stringini ham float'a çevir
        eski_ham = eski_display.replace(" ₺", "").replace(".", "").replace(",", ".")
        try:
            eski_float = float(eski_ham)
        except ValueError:
            eski_float = 0.0

        # QLineEdit ile inline düzenleme (QDialog kullan)
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Düzenle — {kolon}")
        dlg.setFixedWidth(280)
        dlg_layout = QVBoxLayout(dlg)
        dlg_layout.setSpacing(10)
        dlg_layout.setContentsMargins(16, 16, 16, 16)

        dlg_layout.addWidget(QLabel(f"<b>{kolon}</b> değerini düzenleyin:"))
        edit = QLineEdit()
        edit.setText(str(eski_float).replace(".", ","))
        edit.setFixedHeight(34)
        edit.setStyleSheet(
            "QLineEdit{background:#f8fafc;border:1.5px solid #e2e8f0;"
            "border-radius:8px;padding:0 10px;font-size:13px;}"
            "QLineEdit:focus{border-color:#7c3aed;}"
        )
        edit.selectAll()
        dlg_layout.addWidget(edit)

        bilgi = QLabel("⚠️  Eski: " + eski_display)
        bilgi.setStyleSheet("font-size:11px;color:#6b7280;")
        dlg_layout.addWidget(bilgi)

        bb = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        bb.button(QDialogButtonBox.StandardButton.Ok).setText("Kaydet")
        bb.button(QDialogButtonBox.StandardButton.Cancel).setText("Vazgeç")
        bb.accepted.connect(dlg.accept)
        bb.rejected.connect(dlg.reject)
        dlg_layout.addWidget(bb)

        edit.setFocus()
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        raw = edit.text().strip()
        normalized = raw.replace(".", "").replace(",", ".")
        try:
            yeni_float = float(normalized)
        except ValueError:
            QMessageBox.warning(self, "Geçersiz Değer", "Lütfen geçerli bir sayı girin.")
            return

        if yeni_float == eski_float:
            return  # Değişim yok

        from services.vergi_muhtasar_service import update_vergi_muhtasar_alan
        r = update_vergi_muhtasar_alan(self._userid, kayit_id, kolon, yeni_float)
        if r["success"]:
            item.setText(self._fmt_tutar(yeni_float))
            # Önbellekteki veriyi güncelle
            for d in self._all_data:
                if d.get("id") == kayit_id:
                    d[kolon] = yeni_float
                    break
            # Toplamları yenile
            visible = self._visible_data()
            self._update_totals(visible)
            self._show_status("✅  Güncellendi.", "#059669")
        else:
            self._show_status(f"❌  {r['message']}", "#dc2626")

    def _visible_data(self) -> list[dict]:
        """Şu an tabloda görünen satırlara karşılık gelen veri listesi."""
        donem_filter = self._donem_combo.currentData() or ""
        ack_filter   = self._ack_combo.currentData() or ""
        return [
            r for r in self._all_data
            if (not donem_filter or r.get("donem", "") == donem_filter)
            and (not ack_filter  or (r.get("ack") or "") == ack_filter)
        ]

    # ── Silme (PHP: .sil click → showAlert2 → pendingSilinecek) ─────────

    def _on_sil(self, row: int):
        if row >= len(self._row_ids):
            return
        kayit_id = self._row_ids[row]
        hesap    = (self._table.item(row, self.COL_HESAP) or QTableWidgetItem("")).text()
        donem    = (self._table.item(row, self.COL_DONEM) or QTableWidgetItem("")).text()

        dlg = QMessageBox(self)
        dlg.setWindowTitle("Kaydı Sil")
        dlg.setIcon(QMessageBox.Icon.Warning)
        dlg.setText(
            f"<b>{hesap}</b> — {donem} kaydını silmek istiyor musunuz?<br><br>"
            "Bu işlem geri alınamaz."
        )
        dlg.setStandardButtons(
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel
        )
        dlg.button(QMessageBox.StandardButton.Yes).setText("🗑  Sil")
        dlg.button(QMessageBox.StandardButton.Cancel).setText("Vazgeç")
        if dlg.exec() != QMessageBox.StandardButton.Yes:
            return

        from services.vergi_muhtasar_service import delete_vergi_muhtasar
        r = delete_vergi_muhtasar(self._userid, kayit_id)
        if r["success"]:
            self._show_status("✅  Kayıt silindi.", "#059669")
            self.refresh()
        else:
            self._show_status(f"❌  {r['message']}", "#dc2626")

    # ── Şema İndir (PHP: #vmSemaIndir click) ─────────────────────────────

    def _on_sema_indir(self):
        from services.vergi_muhtasar_service import SEMA_CSV_ICERIK, SEMA_CSV_DOSYA_ADI
        kayit_yolu, _ = QFileDialog.getSaveFileName(
            self, "Şema CSV Kaydet", SEMA_CSV_DOSYA_ADI, "CSV Dosyaları (*.csv)"
        )
        if not kayit_yolu:
            return
        try:
            with open(kayit_yolu, "w", encoding="utf-8-sig") as f:
                f.write(SEMA_CSV_ICERIK)
            self._show_status(f"✅  Şema kaydedildi: {kayit_yolu}", "#059669")
        except Exception as exc:
            self._show_status(f"❌  Kayıt hatası: {exc}", "#dc2626")

    # ── Durum Etiketi ─────────────────────────────────────────────────────

    def _show_status(self, text: str, color: str = "#374151"):
        self._status_lbl.setText(text)
        self._status_lbl.setStyleSheet(f"font-size:12px;color:{color};padding:4px 0;")
        self._status_lbl.show()


# ── Şirket Profili Kartı ──────────────────────────────────────────────────────

class SirketProfilCard(QFrame):
    """Hesap & Güvenlik sekmesindeki şirket profili formu."""

    def __init__(self, userid: int, musterino: int = 1, parent=None):
        super().__init__(parent)
        self._userid = userid
        self._musterino = musterino
        self.setStyleSheet(
            "QFrame{background:white;border:1.5px solid #e2e8f0;"
            "border-radius:14px;}"
        )
        self._fields = {}
        self._build()
        self._load()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(22, 20, 22, 20)
        root.setSpacing(14)

        from PyQt6.QtWidgets import QLineEdit, QGridLayout

        # Baslik
        h = QHBoxLayout()
        ic = QLabel("🏢")
        ic.setStyleSheet("font-size:20px;")
        h.addWidget(ic)
        t = QLabel("Şirket Profili")
        t.setStyleSheet("font-size:14px;font-weight:700;color:#000000;letter-spacing:.5px;")
        h.addWidget(t)
        h.addStretch()
        root.addLayout(h)

        info = QLabel(
            "💡  Bu bilgiler fatura XML'indeki VKN ile karşılaştırılarak "
            "Gelen / Kesilen ayrımı otomatik yapılır."
        )
        info.setWordWrap(True)
        info.setStyleSheet(
            "background:#e6edfa;color:#000000;border-radius:6px;"
            "padding:9px 12px;font-size:12px;border:none;"
        )
        root.addWidget(info)

        grid = QGridLayout()
        grid.setSpacing(10)

        FIELDS = [
            ("unvan",        "Şirket / Ünvan *",     0, 0, 1, 2),
            ("vergino",      "Vergi No (VKN) *",      1, 0, 1, 1),
            ("tckn",         "TCKN (Şahıs ise)",      1, 1, 1, 1),
            ("vergidairesi", "Vergi Dairesi",          2, 0, 1, 1),
            ("adres",        "Adres",                  3, 0, 1, 2),
            ("il",           "İl",                     4, 0, 1, 1),
            ("ilce",         "İlçe",                   4, 1, 1, 1),
        ]

        INPUT_STYLE = (
            "QLineEdit{background:#f8fafc;border:1.5px solid #e2e8f0;"
            "border-radius:8px;padding:0 10px;font-size:13px;color:#1f2937;}"
            "QLineEdit:focus{border-color:#2563eb;}"
        )
        LBL_STYLE = "font-size:12px;font-weight:600;color:#374151;"

        for key, lbl_txt, row, col, rspan, cspan in FIELDS:
            lbl = QLabel(lbl_txt)
            lbl.setStyleSheet(LBL_STYLE)
            inp = QLineEdit()
            inp.setFixedHeight(34)
            inp.setStyleSheet(INPUT_STYLE)
            inp.setPlaceholderText(lbl_txt)
            self._fields[key] = inp
            grid.addWidget(lbl, row * 2,     col, 1, cspan)
            grid.addWidget(inp, row * 2 + 1, col, 1, cspan)

        root.addLayout(grid)

        # Kaydet butonu
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self._save_btn = QPushButton("💾  Profili Kaydet")
        self._save_btn.setFixedHeight(38)
        self._save_btn.setMinimumWidth(160)
        self._save_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._save_btn.setStyleSheet(
            "QPushButton{background:#2563eb;color:white;border:none;"
            "border-radius:9px;font-size:13px;font-weight:600;padding:0 18px;}"
            "QPushButton:hover{background:#1d4ed8;}"
        )
        self._save_btn.clicked.connect(self._on_kaydet)
        btn_row.addWidget(self._save_btn)
        root.addLayout(btn_row)

        self._result_lbl = QLabel("")
        self._result_lbl.setStyleSheet("font-size:12px;color:#059669;")
        root.addWidget(self._result_lbl)

    def _load(self):
        from services.sirket_service import get_sirket_profili
        p = get_sirket_profili(self._userid, self._musterino)
        for key, inp in self._fields.items():
            inp.setText(p.get(key, ""))

    def _on_kaydet(self):
        from services.sirket_service import save_sirket_profili
        unvan = self._fields["unvan"].text().strip()
        vergino = self._fields["vergino"].text().strip()
        if not unvan:
            self._result_lbl.setText("⚠️  Ünvan alanı zorunludur.")
            self._result_lbl.setStyleSheet("font-size:12px;color:#dc2626;")
            return
        if not vergino and not self._fields["tckn"].text().strip():
            self._result_lbl.setText("⚠️  VKN veya TCKN alanlarından biri zorunludur.")
            self._result_lbl.setStyleSheet("font-size:12px;color:#dc2626;")
            return

        ok = save_sirket_profili(
            userid=self._userid,
            unvan=unvan,
            vergino=vergino,
            tckn=self._fields["tckn"].text().strip(),
            vergidairesi=self._fields["vergidairesi"].text().strip(),
            adres=self._fields["adres"].text().strip(),
            il=self._fields["il"].text().strip(),
            ilce=self._fields["ilce"].text().strip(),
            musterino=self._musterino,
        )
        if ok:
            self._result_lbl.setText("✅  Şirket profili başarıyla kaydedildi.")
            self._result_lbl.setStyleSheet("font-size:12px;color:#059669;")
        else:
            self._result_lbl.setText("❌  Kayıt sırasında hata oluştu.")
            self._result_lbl.setStyleSheet("font-size:12px;color:#dc2626;")


# ── Kredi Kartı Worker ────────────────────────────────────────────────────────

class KrediKartYukleWorker(QThread):
    """
    Dosya yükleme arka plan iş parçacığı.
    PHP: bankaiceribtn → hiddenFileInput → AJAX (krediKartVeriAktar*.php)
    """
    progress = pyqtSignal(str)
    finished = pyqtSignal(dict)

    def __init__(self, userid, musterino, dosya_yolları, hesap_kodu,
                 banka_adi, dosya_turu, banka_adi_listesi):
        super().__init__()
        self._userid           = userid
        self._musterino        = musterino
        self._dosya_yolları    = dosya_yolları
        self._hesap_kodu       = hesap_kodu
        self._banka_adi        = banka_adi
        self._dosya_turu       = dosya_turu
        self._banka_listesi    = banka_adi_listesi

    def run(self):
        from services.kredi_kart_service import yukle_dosyalar
        r = yukle_dosyalar(
            userid          = self._userid,
            musterino       = self._musterino,
            dosya_yolları   = self._dosya_yolları,
            hesap_kodu      = self._hesap_kodu,
            banka_adi       = self._banka_adi,
            dosya_turu      = self._dosya_turu,
            banka_adi_liste = self._banka_listesi,
        )
        self.finished.emit(r)


# ── Kredi Kartı Kartı ─────────────────────────────────────────────────────────

class KrediKartCard(QFrame):
    """
    PHP ayarlar.php → Eklentiler → Kredi Kartı div'inin PyQt6 karşılığı.

    Özellikler (PHP JS ayarlar.js + PHP backend ile birebir):
    - Kart seçimi (key_kartlari tablosundan — PHP: guncelleKartlar() + nocache.php)
    - Dosya türü seçimi: CSV | PDF | XLSX
    - Çoklu dosya seçimi (birden fazla PDF)
    - Banka uyum kontrolü (YapıKredi / İş Bankası)
    - İçeri Aktar butonu → KrediKartYukleWorker
    - Yeni kart tanımlama formu (key_kartlari INSERT)
    - Sonuç gösterimi (durum etiketi)
    """

    def __init__(self, userid: int, musterino: int, parent=None):
        super().__init__(parent)
        self._userid       = userid
        self._musterino    = musterino
        self._worker: KrediKartYukleWorker | None = None
        self._kart_listesi: list[dict] = []

        self.setStyleSheet(
            "QFrame{background:white;border:1.5px solid #dbeafe;"
            "border-radius:14px;}"
        )
        self._build()
        self.refresh()

    def _btn_style(self, color: str) -> str:
        return (
            f"QPushButton{{background:{color};color:white;border:none;"
            f"border-radius:9px;font-size:13px;font-weight:600;"
            f"padding:0 14px;letter-spacing:.4px;}}"
            f"QPushButton:hover{{background:{color}cc;}}"
            f"QPushButton:disabled{{background:#cbd5e1;color:#94a3b8;}}"
        )

    def _combo_style(self) -> str:
        return (
            "QComboBox{background:#f8fafc;border:1.5px solid #e2e8f0;"
            "border-radius:8px;padding:0 10px;font-size:13px;color:#1f2937;height:32px;}"
            "QComboBox::drop-down{border:none;}"
            "QComboBox QAbstractItemView{background:white;color:#1f2937;"
            "selection-background-color:#dbeafe;selection-color:#1e40af;}"
        )

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(22, 20, 22, 20)
        root.setSpacing(14)

        # Başlık
        h = QHBoxLayout()
        ic = QLabel("💳")
        ic.setStyleSheet("font-size:20px;")
        h.addWidget(ic)
        t = QLabel("Kredi Kartı")
        t.setStyleSheet("font-size:14px;font-weight:700;color:#000000;letter-spacing:.5px;")
        h.addWidget(t)
        h.addStretch()
        root.addLayout(h)

        # Bilgi chip
        info = QLabel(
            "💡  Kayıtlı kredi kartınızı seçip CSV (YapıKredi), "
            "PDF (İş Bankası / YapıKredi) veya XLSX formatındaki "
            "banka ekstrenizi içeri aktarabilirsiniz."
        )
        info.setWordWrap(True)
        info.setStyleSheet(
            "background:#eff6ff;color:#1e3a8a;border-radius:6px;"
            "padding:9px 12px;font-size:12px;border:none;"
        )
        root.addWidget(info)

        # ── Kart Seçimi (PHP: #kartSelect) ──
        kart_row = QHBoxLayout()
        kart_row.setSpacing(10)
        kart_lbl = QLabel("Kredi Kartı Seçin:")
        kart_lbl.setStyleSheet("font-size:12px;font-weight:600;color:#374151;")
        kart_row.addWidget(kart_lbl)
        self._kart_combo = QComboBox()
        self._kart_combo.setMinimumWidth(220)
        self._kart_combo.addItem("Bir kart tanımı seçin", "")
        self._kart_combo.setStyleSheet(self._combo_style())
        self._kart_combo.currentIndexChanged.connect(self._on_kart_changed)
        kart_row.addWidget(self._kart_combo, 1)
        kart_row.addStretch()
        root.addLayout(kart_row)

        # ── Yeni Kart Ekleme Formu ──
        self._ekle_frame = QFrame()
        self._ekle_frame.setStyleSheet(
            "QFrame{background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px;}"
        )
        ekle_layout = QVBoxLayout(self._ekle_frame)
        ekle_layout.setContentsMargins(14, 12, 14, 12)
        ekle_layout.setSpacing(10)
        ekle_title = QLabel("➕  Yeni Kart Ekle")
        ekle_title.setStyleSheet("font-size:13px;font-weight:700;color:#1e40af;")
        ekle_layout.addWidget(ekle_title)

        ekle_grid = QHBoxLayout()
        ekle_grid.setSpacing(10)

        def _inp_col(label_txt, placeholder=""):
            col = QVBoxLayout()
            lbl = QLabel(label_txt)
            lbl.setStyleSheet("font-size:11px;font-weight:600;color:#374151;")
            inp = QLineEdit()
            inp.setFixedHeight(32)
            inp.setPlaceholderText(placeholder)
            inp.setStyleSheet(
                "QLineEdit{background:white;border:1.5px solid #bfdbfe;"
                "border-radius:8px;padding:0 10px;font-size:13px;color:#1f2937;}"
                "QLineEdit:focus{border-color:#2563eb;}"
            )
            col.addWidget(lbl)
            col.addWidget(inp)
            return col, inp

        banka_col,  self._yeni_banka_inp  = _inp_col("Banka Adı *", "Yapı Kredi")
        etiket_col, self._yeni_etiket_inp = _inp_col("Kart Etiketi", "YapıKredi-5432")
        hesap_col,  self._yeni_hesap_inp  = _inp_col("Hesap Kodu", "100.01")
        ekle_grid.addLayout(banka_col, 2)
        ekle_grid.addLayout(etiket_col, 2)
        ekle_grid.addLayout(hesap_col, 1)
        ekle_layout.addLayout(ekle_grid)

        ekle_btn_row = QHBoxLayout()
        self._ekle_btn = QPushButton("➕  Kartı Kaydet")
        self._ekle_btn.setFixedHeight(34)
        self._ekle_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._ekle_btn.setStyleSheet(self._btn_style("#2563eb"))
        self._ekle_btn.clicked.connect(self._on_kart_ekle)
        ekle_btn_row.addWidget(self._ekle_btn)
        ekle_btn_row.addStretch()
        ekle_layout.addLayout(ekle_btn_row)
        root.addWidget(self._ekle_frame)

        # ── İçeri Aktar (PHP: #bankkapan) ──
        self._aktar_frame = QFrame()
        self._aktar_frame.setStyleSheet("QFrame{background:transparent;border:none;}")
        aktar_layout = QVBoxLayout(self._aktar_frame)
        aktar_layout.setContentsMargins(0, 0, 0, 0)
        aktar_layout.setSpacing(10)

        dosya_row = QHBoxLayout()
        dosya_row.setSpacing(10)
        dosya_lbl2 = QLabel("Dosya Türü:")
        dosya_lbl2.setStyleSheet("font-size:12px;font-weight:600;color:#374151;")
        dosya_row.addWidget(dosya_lbl2)

        self._dosya_combo = QComboBox()
        self._dosya_combo.setFixedWidth(120)
        self._dosya_combo.addItems(["PDF", "XLSX", "CSV"])
        self._dosya_combo.setStyleSheet(self._combo_style())
        self._dosya_combo.currentIndexChanged.connect(self._on_dosya_turu_changed)
        dosya_row.addWidget(self._dosya_combo)

        self._aktar_btn = QPushButton("↤  İçeri Aktar")
        self._aktar_btn.setFixedHeight(36)
        self._aktar_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._aktar_btn.setStyleSheet(self._btn_style("#2563eb"))
        self._aktar_btn.clicked.connect(self._on_aktar)
        dosya_row.addWidget(self._aktar_btn)
        dosya_row.addStretch()
        aktar_layout.addLayout(dosya_row)

        self._dosya_lbl = QLabel("")
        self._dosya_lbl.setStyleSheet("font-size:11px;color:#6b7280;")
        self._dosya_lbl.setWordWrap(True)
        self._dosya_lbl.hide()
        aktar_layout.addWidget(self._dosya_lbl)

        self._durum_lbl = QLabel("")
        self._durum_lbl.setWordWrap(True)
        self._durum_lbl.setTextFormat(Qt.TextFormat.RichText)
        self._durum_lbl.setStyleSheet("font-size:12px;color:#374151;")
        self._durum_lbl.hide()
        aktar_layout.addWidget(self._durum_lbl)

        root.addWidget(self._aktar_frame)
        self._aktar_frame.setEnabled(False)

    # ── Kart Listesi ─────────────────────────────────────────────────────

    def refresh(self):
        from services.kredi_kart_service import get_kart_listesi
        r = get_kart_listesi(self._userid, self._musterino)
        self._kart_listesi = r.get("data", [])
        self._fill_kart_combo()

    def _fill_kart_combo(self):
        self._kart_combo.blockSignals(True)
        self._kart_combo.clear()
        self._kart_combo.addItem("Bir kart tanımı seçin", None)
        for item in self._kart_listesi:
            banka = (item.get("banka") or "").strip()
            no    = (item.get("no") or "").strip()
            display_text = f"{banka} {no}".strip()
            self._kart_combo.addItem(display_text, item)
        self._kart_combo.blockSignals(False)

    def _get_secili_banka_adi(self) -> str:
        idx = self._kart_combo.currentIndex()
        if idx <= 0:
            return ""
        item = self._kart_combo.itemData(idx)
        if not item or not isinstance(item, dict):
            return ""
        return (item.get("bankaadi") or item.get("banka") or "").strip()

    def _on_kart_changed(self):
        kart = self._kart_combo.currentData()
        self._aktar_frame.setEnabled(bool(kart))
        self._durum_lbl.hide()
        self._dosya_lbl.hide()

    def _on_dosya_turu_changed(self):
        self._durum_lbl.hide()
        self._dosya_lbl.hide()

    # ── Yeni Kart Ekleme ──────────────────────────────────────────────

    def _on_kart_ekle(self):
        banka_adi  = self._yeni_banka_inp.text().strip()
        etiket     = self._yeni_etiket_inp.text().strip()
        hesap_kodu = self._yeni_hesap_inp.text().strip()
        if not banka_adi:
            self._show_durum("⚠️  Banka adı zorunludur.", "#dc2626")
            return
        banka_col_val = etiket or banka_adi
        from db.database import get_connection as _gc
        conn = _gc()
        try:
            conn.execute(
                "INSERT INTO key_kartlari (banka, no, userid, hesapkodu, bankaadi, musterino) "
                "VALUES (%s, '', %s, %s, %s, %s)",
                (banka_col_val, self._userid, hesap_kodu, banka_adi, self._musterino)
            )
            conn.commit()
            self._show_durum(f"✅  '{banka_adi}' kartı kaydedildi.", "#059669")
            self._yeni_banka_inp.clear()
            self._yeni_etiket_inp.clear()
            self._yeni_hesap_inp.clear()
            self.refresh()
        except Exception as exc:
            self._show_durum(f"❌  Kayıt hatası: {exc}", "#dc2626")
        finally:
            conn.close()

    # ── İçeri Aktar ───────────────────────────────────────────────────

    def _on_aktar(self):
        kart = self._get_secili_kart()
        if not kart:
            self._show_durum("⚠️  Lütfen bir kart seçin.", "#f59e0b")
            return
        hesap_kodu = kart.get("hesapKodu", "")
        banka_adi  = kart.get("bankaAdi", "")
        dosya_turu = self._dosya_combo.currentText().lower()
        filtre_map = {"pdf": "PDF Dosyaları (*.pdf)", "xlsx": "Excel Dosyaları (*.xlsx)",
                      "csv": "CSV Dosyaları (*.csv)"}
        filtre = filtre_map.get(dosya_turu, "Tüm Dosyalar (*)")
        dosyalar, _ = QFileDialog.getOpenFileNames(self, "Dosya Seç", "", filtre)
        if not dosyalar:
            return
        self._dosya_lbl.setText(
            f"📄  {dosyalar[0].split('/')[-1]}" if len(dosyalar) == 1
            else f"📄  {len(dosyalar)} dosya seçildi"
        )
        self._dosya_lbl.show()
        if dosya_turu == "pdf" and not self._pdf_uyum_kontrol(dosyalar[0], banka_adi):
            self._show_durum("⚠️  Seçilen dosya ile kart uyumsuz olabilir.", "#f59e0b")
        self._set_busy(True)
        self._show_durum("⌛  Dosya yükleniyor...", "#2563eb")
        self._worker = KrediKartYukleWorker(
            userid=self._userid, musterino=self._musterino, dosya_yolları=dosyalar,
            hesap_kodu=hesap_kodu, banka_adi=banka_adi,
            dosya_turu=dosya_turu, banka_adi_listesi=[banka_adi] * len(dosyalar),
        )
        self._worker.finished.connect(self._on_aktar_done)
        self._worker.start()

    def _pdf_uyum_kontrol(self, dosya_yolu: str, banka_adi: str) -> bool:
        try:
            import pdfplumber
            norm = banka_adi.replace(" ", "").lower()
            is_yk = "yapıkredi" in norm or "yapikredi" in norm
            with pdfplumber.open(dosya_yolu) as pdf:
                text = (pdf.pages[0].extract_text() or "").lower().replace(" ", "") if pdf.pages else ""
            if len(text) < 50:
                return True
            return ("yapıkredi" in text or "yapikredi" in text) if is_yk else any(
                k in text for k in ["işbank", "isbank", "maximum", "hesapözet", "kredikart"]
            )
        except Exception:
            return True

    def _on_aktar_done(self, r: dict):
        self._set_busy(False)
        if r.get("success"):
            self._show_durum(
                f"✅  <b>İşlem tamamlandı!</b><br>"
                f"{r.get('message','')}<br>"
                f"<small style='color:#6b7280;'>{r.get('added',0)} kayıt, "
                f"{r.get('skipped',0)} mükerrer atlandı</small>",
                "#059669"
            )
        else:
            self._show_durum(
                f"❌  <b>Hata:</b> {r.get('errors', r.get('message','Bilinmeyen hata.'))}",
                "#dc2626"
            )

    def _set_busy(self, busy: bool):
        for w in [self._aktar_btn, self._ekle_btn, self._kart_combo, self._dosya_combo]:
            try:
                w.setEnabled(not busy)
            except Exception:
                pass

    def _show_durum(self, text: str, color: str = "#374151"):
        self._durum_lbl.setText(text)
        self._durum_lbl.setStyleSheet(
            f"font-size:12px;color:{color};padding:6px 10px;background:#f8fafc;border-radius:6px;"
        )
        self._durum_lbl.show()


    def refresh(self):
        """Kart listesini DB'den yükler (PHP: guncelleKartlar() + nocache.php)."""
        from services.kredi_kart_service import get_kart_listesi
        r = get_kart_listesi(self._userid, self._musterino)
        self._kart_listesi = r.get("data", [])
        self._fill_kart_combo()

    def _fill_kart_combo(self):
        self._kart_combo.blockSignals(True)
        self._kart_combo.clear()
        self._kart_combo.addItem("Bir kart tanımı seçin", None)
        for item in self._kart_listesi:
            banka = (item.get("banka") or "").strip()
            no    = (item.get("no") or "").strip()
            display_text = f"{banka} {no}".strip()
            self._kart_combo.addItem(display_text, item)  # item dict'in tamamını sakla
        self._kart_combo.blockSignals(False)

    def _get_secili_kart(self) -> dict:
        """Seçili kartı {id, hesapKodu, bankaAdi} olarak döndürür."""
        idx = self._kart_combo.currentIndex()
        if idx <= 0:
            return {}
        item = self._kart_combo.itemData(idx)
        if not item or not isinstance(item, dict):
            return {}
        return {
            "id":        item.get("id", ""),
            "hesapKodu": (item.get("hesapkodu") or item.get("hesapKodu") or ""),
            "bankaAdi":  (item.get("bankaadi") or item.get("bankaAdi") or item.get("banka") or ""),
        }

    def _get_secili_banka_adi(self) -> str:
        return self._get_secili_kart().get("bankaAdi", "")

    # ── Kart Seçimi Değişikliği (PHP: kartSelect + bankaDosyaTuru change) ───

    def _on_kart_changed(self):
        """Kart seçildiğinde İçeri Aktar bölümünü aktif et."""
        kart = self._get_secili_kart()
        aktif = bool(kart)
        self._aktar_frame.setEnabled(aktif)
        self._durum_lbl.hide()
        self._dosya_lbl.hide()

    def _on_dosya_turu_changed(self):
        """Dosya türü değişince durum temizle."""
        self._durum_lbl.hide()
        self._dosya_lbl.hide()

    # ── Yeni Kart Ekleme (key_kartlari INSERT) ──────────────────────────

    def _on_kart_ekle(self):
        """Yeni kart tanımını key_kartlari tablosuna kaydeder."""
        banka_adi  = self._yeni_banka_inp.text().strip()
        etiket     = self._yeni_etiket_inp.text().strip()
        hesap_kodu = self._yeni_hesap_inp.text().strip()

        if not banka_adi:
            self._show_durum("⚠️  Banka adı zorunludur.", "#dc2626")
            return

        banka_col_val = etiket or banka_adi  # 'banka' kolonu = etiket

        from db.database import get_connection
        conn = get_connection()
        try:
            conn.execute(
                "INSERT INTO key_kartlari (banka, no, userid, hesapKodu, bankaAdi) "
                "VALUES (?, '', ?, ?, ?)",
                (banka_col_val, self._userid, hesap_kodu, banka_adi)
            )
            conn.commit()
            self._show_durum(
                f"✅  '{banka_adi}' kartı kaydedildi.", "#059669"
            )
            # Alanları temizle
            self._yeni_banka_inp.clear()
            self._yeni_etiket_inp.clear()
            self._yeni_hesap_inp.clear()
            # Listeyi yenile
            self.refresh()
        except Exception as exc:
            self._show_durum(f"❌  Kayıt hatası: {exc}", "#dc2626")
        finally:
            conn.close()

# ── Ana Ayarlar Ekranı ────────────────────────────────────────────────────────

class AyarlarScreen(QWidget):
    # Herhangi bir import kartı veri çekince bu sinyal tetiklenir
    # main_window bunu dash.refresh()'e bağlar → dashboard otomatik güncellenir
    data_changed = pyqtSignal()

    def __init__(self, user: dict, parent=None):
        super().__init__(parent)
        self._user = user
        self._userid    = user.get("GercekUserId", user.get("Kayitno", 1))
        self._musterino = user.get("musterino", user.get("GercekUserId", 1))
        self._setup_ui()

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(20)

        # Başlık
        title = QLabel("Ayarlar")
        title.setStyleSheet(
            "color:#000000;"
            "font-size:24px;font-weight:700;letter-spacing:1px;"
        )
        root.addWidget(title)

        # Tab butonları
        tab_row = QHBoxLayout()
        tab_row.setSpacing(8)
        self._tab_btns: dict[str, QPushButton] = {}
        tabs = [
            ("hesap",     "👤  Hesap & Güvenlik"),
            ("eklentiler","🔌  Eklentiler"),
            ("veritabani","🗄️  Veritabanı"),
        ]

        for key, label in tabs:
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.setFixedHeight(34)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setStyleSheet(self._tab_style())
            btn.clicked.connect(lambda _, k=key: self._switch_tab(k))
            self._tab_btns[key] = btn
            tab_row.addWidget(btn)
        tab_row.addStretch()
        root.addLayout(tab_row)

        # Scroll alan
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setStyleSheet("QScrollArea{background:transparent;}")

        self._content = QWidget()
        self._content.setStyleSheet("background:transparent;")
        self._content_layout = QVBoxLayout(self._content)
        self._content_layout.setContentsMargins(0, 0, 0, 0)
        self._content_layout.setSpacing(16)

        scroll.setWidget(self._content)
        root.addWidget(scroll)

        # Hesap & Güvenlik sekmesi — Şirket Profili
        self._sirket_card = SirketProfilCard(self._userid, musterino=self._musterino)
        self._content_layout.addWidget(self._sirket_card)

        # Manuel Toplu İşle kartı — EFatura'nın üstünde
        self._gsheets_settings_card = GoogleSheetsSettingsCard()
        self._content_layout.addWidget(self._gsheets_settings_card)

        self._manuel_toplu_card = ManuelTopluIsleCard(self._userid, musterino=self._musterino)
        self._content_layout.addWidget(self._manuel_toplu_card)

        # Eklentiler kartı
        self._efatura_card = EFaturaCard(self._userid, self._musterino)
        self._content_layout.addWidget(self._efatura_card)

        # Fatura Yönetim kartı
        self._yonetim_card = FaturaYonetimCard(self._userid)
        self._content_layout.addWidget(self._yonetim_card)

        # VOMSİS API kartı — Fatura Yönetim'in altında
        self._vomsis_card = VomsisCard(self._userid, musterino=self._musterino)
        self._content_layout.addWidget(self._vomsis_card)

        # Moy kartı — VOMSİS'in altında
        self._moy_card = MoyCard(self._userid, musterino=self._musterino)
        self._content_layout.addWidget(self._moy_card)

        # Vergi Muhtasar kartı — Moy'un altında
        self._vergi_muhtasar_card = VergiMuhtasarCard(self._userid)
        self._content_layout.addWidget(self._vergi_muhtasar_card)

        # Kredi Kartı kartı — Vergi Muhtasar'ın altında
        self._kredi_kart_card = KrediKartCard(self._userid, self._musterino)
        self._content_layout.addWidget(self._kredi_kart_card)

        # Veritabanı sekmesi
        self._veritabani_card = VeriTabaniCard()
        self._content_layout.addWidget(self._veritabani_card)

        self._content_layout.addStretch()

        # Kart sinyallerini AyarlarScreen.data_changed'e bağla
        # → Herhangi bir import bitince dashboard otomatik güncellenir
        for card in [
            self._vomsis_card, self._moy_card,
            self._vergi_muhtasar_card, self._kredi_kart_card,
            self._efatura_card, self._yonetim_card,
        ]:
            if hasattr(card, 'data_changed'):
                card.data_changed.connect(self.data_changed)

        # Aktif tab
        self._switch_tab("eklentiler")

    def _tab_style(self) -> str:
        return (
            "QPushButton{background:#f1f5f9;border:1.5px solid #e2e8f0;"
            "border-radius:8px;font-size:13px;font-weight:600;color:#000000;"
            "padding:0 16px;}"
            "QPushButton:checked{background:#1e293b;color:white;border-color:#1e293b;}"
            "QPushButton:hover:!checked{background:#e2e8f0;}"
        )

    def _switch_tab(self, key: str):
        for k, btn in self._tab_btns.items():
            btn.setChecked(k == key)

        self._sirket_card.setVisible(key == "hesap")
        self._efatura_card.setVisible(key == "eklentiler")
        self._yonetim_card.setVisible(key == "eklentiler")
        self._vomsis_card.setVisible(key == "eklentiler")
        self._moy_card.setVisible(key == "eklentiler")
        self._vergi_muhtasar_card.setVisible(key == "eklentiler")
        self._kredi_kart_card.setVisible(key == "eklentiler")
        self._veritabani_card.setVisible(key == "veritabani")

        if key == "veritabani":
            self._veritabani_card.refresh()

        if key == "eklentiler":
            if hasattr(self._efatura_card, "refresh"):
                self._efatura_card.refresh()
            if hasattr(self._vomsis_card, "refresh"):
                self._vomsis_card.refresh()
            if hasattr(self._moy_card, "refresh"):
                self._moy_card.refresh()
            if hasattr(self._vergi_muhtasar_card, "refresh"):
                self._vergi_muhtasar_card.refresh()
            if hasattr(self._kredi_kart_card, "refresh"):
                self._kredi_kart_card.refresh()


    # ── Yardımcılar ──────────────────────────────────────────────────

    def _set_busy(self, busy: bool):
        self._aktar_btn.setEnabled(not busy)
        self._ekle_btn.setEnabled(not busy)
        self._kart_combo.setEnabled(not busy)
        self._dosya_combo.setEnabled(not busy)

    def _show_durum(self, text: str, color: str = "#374151"):
        self._durum_lbl.setText(text)
        self._durum_lbl.setStyleSheet(
            f"font-size:12px;color:{color};padding:6px 10px;"
            "background:#f8fafc;border-radius:6px;"
        )
        self._durum_lbl.show()

# ── Kullanıcı Yönetimi Kartı ─────────────────────────────────────────────────

class KullaniciYonetimCard(QFrame):
    """
    Hesap & Güvenlik sekmesinde alt kullanıcı yönetimi.
    alt_kullanici tablosu üzerinden CRUD işlemleri yapar.

    Yetki seviyeleri:
      1 → Admin       (tüm ekranlar + veri girişi)
      2 → Kullanıcı   (sadece görüntüleme)
    """

    YETKI_SECIMLERI = [("1", "Admin  —  Tüm yetkiler"), ("2", "Kullanıcı  —  Sadece görüntüleme")]
    YETKI_BADGE = {
        "1": ("#dbeafe", "#1d4ed8", "👑 Admin"),
        "2": ("#dcfce7", "#15803d", "👁 Kullanıcı"),
        "3": ("#fef9c3", "#a16207", "📊 Analist"),
    }

    def __init__(self, userid: int, musterino: int = 1, parent=None):
        super().__init__(parent)
        self._userid = userid
        self.setStyleSheet(
            "QFrame{background:white;border:1.5px solid #bfdbfe;"
            "border-radius:14px;}"
        )
        self._build()

    # ── UI ───────────────────────────────────────────────────────────────────

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(22, 20, 22, 20)
        root.setSpacing(14)

        # ── Başlık ──
        h = QHBoxLayout()
        ic = QLabel("👥")
        ic.setStyleSheet("font-size:20px;")
        h.addWidget(ic)
        t = QLabel("Giriş Yetkileri & Kullanıcılar")
        t.setStyleSheet("font-size:14px;font-weight:700;color:#000000;letter-spacing:.5px;")
        h.addWidget(t)
        h.addStretch()
        root.addLayout(h)

        # ── Banner (kota) ──
        self._kota_banner = QLabel()
        self._kota_banner.setWordWrap(True)
        self._kota_banner.setStyleSheet(
            "background:#dbeafe;color:#1e40af;border-radius:8px;"
            "padding:10px 14px;font-size:12px;font-weight:600;border:none;"
        )
        root.addWidget(self._kota_banner)

        # ── Bilgi bandı ──
        bilgi = QLabel(
            "💡  <b>Admin</b> rolündeki kullanıcılar ayarlar ve veri giriş ekranlarına "
            "erişebilir.  <b>Kullanıcı</b> rolündekiler yalnızca raporları ve dashboard'ı görür."
        )
        bilgi.setWordWrap(True)
        bilgi.setTextFormat(Qt.TextFormat.RichText)
        bilgi.setStyleSheet(
            "background:#f0fdf4;color:#166534;border-radius:6px;"
            "padding:9px 12px;font-size:12px;border:none;"
        )
        root.addWidget(bilgi)

        # ── Kullanıcı Listesi Tablosu ──
        self._tbl = QTableWidget()
        self._tbl.setColumnCount(5)  # Kullanıcı Adı | E-Posta | Yetki | Kayıt Tarihi | Sil
        self._tbl.setHorizontalHeaderLabels(["Kullanıcı Adı", "E-Posta", "Yetki", "Kayıt Tarihi", ""])
        self._tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._tbl.setAlternatingRowColors(True)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.setSortingEnabled(False)
        hdr = self._tbl.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self._tbl.setColumnWidth(4, 70)
        self._tbl.setMinimumHeight(200)
        self._tbl.setStyleSheet("""
            QTableWidget {
                background: white;
                gridline-color: #e2e8f0;
                font-size: 12px;
                color: #000000;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
            }
            QTableWidget::item { color: #000000; padding: 4px 6px; background: white; }
            QTableWidget::item:alternate { background: #f8fafc; color: #000000; }
            QTableWidget::item:selected { background: #dbeafe; color: #1e40af; }
            QHeaderView::section {
                background: #1e293b; color: white;
                font-weight: 700; font-size: 11px;
                padding: 5px 6px; border: none;
                border-right: 1px solid #334155;
            }
        """)
        root.addWidget(self._tbl, 1)

        # ── Yeni Kullanıcı Formu ──
        form_frame = QFrame()
        form_frame.setStyleSheet(
            "QFrame{background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;}"
        )
        form_lay = QVBoxLayout(form_frame)
        form_lay.setContentsMargins(14, 12, 14, 12)
        form_lay.setSpacing(10)

        form_title = QLabel("➕  Yeni Kullanıcı Ekle")
        form_title.setStyleSheet("font-size:12px;font-weight:700;color:#374151;")
        form_lay.addWidget(form_title)

        row1 = QHBoxLayout()
        row1.setSpacing(10)

        self._inp_adi = QLineEdit()
        self._inp_adi.setPlaceholderText("Kullanıcı adı")
        self._inp_adi.setFixedHeight(34)
        self._inp_adi.setStyleSheet(self._inp_style())
        row1.addWidget(self._inp_adi, 2)

        self._inp_eposta = QLineEdit()
        self._inp_eposta.setPlaceholderText("E-Posta (isteğe bağlı)")
        self._inp_eposta.setFixedHeight(34)
        self._inp_eposta.setStyleSheet(self._inp_style())
        row1.addWidget(self._inp_eposta, 2)

        self._inp_sifre = QLineEdit()
        self._inp_sifre.setPlaceholderText("Şifre")
        self._inp_sifre.setEchoMode(QLineEdit.EchoMode.Password)
        self._inp_sifre.setFixedHeight(34)
        self._inp_sifre.setStyleSheet(self._inp_style())
        row1.addWidget(self._inp_sifre, 2)

        self._inp_yetki = QComboBox()
        self._inp_yetki.setFixedHeight(34)
        self._inp_yetki.setFixedWidth(200)
        for val, lbl in self.YETKI_SECIMLERI:
            self._inp_yetki.addItem(lbl, val)
        self._inp_yetki.setStyleSheet(self._combo_style())
        row1.addWidget(self._inp_yetki)

        self._ekle_btn = QPushButton("✔  Ekle")
        self._ekle_btn.setFixedHeight(34)
        self._ekle_btn.setFixedWidth(80)
        self._ekle_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._ekle_btn.setStyleSheet(
            "QPushButton{background:#1d4ed8;color:white;border:none;"
            "border-radius:7px;font-size:13px;font-weight:700;}"
            "QPushButton:hover{background:#1e40af;}"
        )
        self._ekle_btn.clicked.connect(self._on_ekle)
        row1.addWidget(self._ekle_btn)

        form_lay.addLayout(row1)
        root.addWidget(form_frame)

        # ── Durum mesajı ──
        self._durum_lbl = QLabel("")
        self._durum_lbl.setWordWrap(True)
        self._durum_lbl.setStyleSheet("font-size:11px;color:#374151;padding:4px 0;")
        root.addWidget(self._durum_lbl)

    # ── Stiller ──────────────────────────────────────────────────────────────

    def _inp_style(self) -> str:
        return (
            "QLineEdit{background:white;border:1.5px solid #cbd5e1;"
            "border-radius:7px;font-size:12px;color:#000000;padding:0 10px;}"
            "QLineEdit:focus{border-color:#3b82f6;}"
        )

    def _combo_style(self) -> str:
        return (
            "QComboBox{background:white;border:1.5px solid #cbd5e1;"
            "border-radius:7px;font-size:12px;color:#000000;padding:0 10px;}"
            "QComboBox::drop-down{border:none;}"
        )

    # ── Veri yükleme ─────────────────────────────────────────────────────────

    def refresh(self):
        from services.alt_hesap_service import get_alt_kullanicilar
        from PyQt6.QtGui import QColor, QBrush
        from PyQt6.QtCore import Qt as _Qt

        BLACK = QBrush(QColor("#000000"))
        sonuc = get_alt_kullanicilar(self._userid)
        data  = sonuc.get("data", [])
        kalan = sonuc.get("kalan_hak", 0)
        toplam_hak = 10
        kullanilanlar = toplam_hak - kalan

        # Kota banner
        bar = "█" * kullanilanlar + "░" * kalan
        self._kota_banner.setText(
            f"👤  Kullanıcı Kotası:  {bar}  {kullanilanlar} / {toplam_hak} kullanılıyor   "
            f"—   {kalan} kullanıcı hakkı kaldı"
        )
        bg = "#dbeafe" if kalan > 2 else ("#fef9c3" if kalan > 0 else "#fee2e2")
        fg = "#1e40af" if kalan > 2 else ("#92400e" if kalan > 0 else "#b91c1c")
        self._kota_banner.setStyleSheet(
            f"background:{bg};color:{fg};border-radius:8px;"
            f"padding:10px 14px;font-size:12px;font-weight:600;border:none;"
        )

        # Tabloyu doldur
        self._tbl.setSortingEnabled(False)
        self._tbl.setRowCount(0)
        for row in data:
            ri = self._tbl.rowCount()
            self._tbl.insertRow(ri)
            self._tbl.setRowHeight(ri, 30)

            yetki = str(row.get("yetki", "1"))
            badge_bg, badge_fg, badge_txt = self.YETKI_BADGE.get(
                yetki, ("#e2e8f0", "#374151", "Bilinmiyor")
            )

            it_adi   = QTableWidgetItem(row.get("kullanici_adi", ""))
            it_email = QTableWidgetItem(row.get("eposta", ""))
            it_yetki = QTableWidgetItem(badge_txt)
            it_tarih = QTableWidgetItem(row.get("uyelik_tarihi", ""))

            for it in (it_adi, it_email, it_tarih):
                it.setForeground(BLACK)

            # Yetki hücresi renkli
            it_yetki.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            it_yetki.setForeground(QBrush(QColor(badge_fg)))
            it_yetki.setBackground(QBrush(QColor(badge_bg)))

            self._tbl.setItem(ri, 0, it_adi)
            self._tbl.setItem(ri, 1, it_email)
            self._tbl.setItem(ri, 2, it_yetki)
            self._tbl.setItem(ri, 3, it_tarih)

            sil_btn = QPushButton("🗑  Sil")
            sil_btn.setFixedHeight(26)
            sil_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            sil_btn.setStyleSheet(
                "QPushButton{background:#fee2e2;color:#dc2626;border:none;"
                "border-radius:5px;font-size:11px;font-weight:600;}"
                "QPushButton:hover{background:#fca5a5;}"
            )
            sil_btn.clicked.connect(lambda _, rid=row.get("id"): self._on_sil(rid))
            self._tbl.setCellWidget(ri, 4, sil_btn)

        self._tbl.setSortingEnabled(True)
        sayi = len(data)
        self._durum_lbl.setText(
            f"✅  {sayi} kullanıcı tanımlı." if sayi else
            "ℹ️  Henüz alt kullanıcı tanımlanmamış."
        )

    # ── Ekle ─────────────────────────────────────────────────────────────────

    def _on_ekle(self):
        from services.alt_hesap_service import kaydet_alt_kullanici
        adi    = self._inp_adi.text().strip()
        eposta = self._inp_eposta.text().strip() or f"{adi}@local"
        sifre  = self._inp_sifre.text().strip()
        yetki  = self._inp_yetki.currentData()

        if not adi or not sifre:
            self._durum_lbl.setText("⚠️  Kullanıcı adı ve şifre zorunludur.")
            self._durum_lbl.setStyleSheet("font-size:11px;color:#dc2626;padding:4px 0;")
            return

        sonuc = kaydet_alt_kullanici(self._userid, adi, eposta, sifre, yetki)
        if sonuc.get("success"):
            self._inp_adi.clear()
            self._inp_eposta.clear()
            self._inp_sifre.clear()
            self._durum_lbl.setText(f"✅  '{adi}' kullanıcısı eklendi.")
            self._durum_lbl.setStyleSheet("font-size:11px;color:#15803d;padding:4px 0;")
            self.refresh()
        else:
            self._durum_lbl.setText(f"❌  {sonuc.get('message', 'Hata oluştu.')}")
            self._durum_lbl.setStyleSheet("font-size:11px;color:#dc2626;padding:4px 0;")

    # ── Sil ──────────────────────────────────────────────────────────────────

    def _on_sil(self, kullanici_id: int):
        from services.alt_hesap_service import sil_alt_kullanici
        from PyQt6.QtWidgets import QMessageBox
        cevap = QMessageBox.question(
            self, "Kullanıcı Sil",
            "Bu kullanıcıyı silmek istediğinizden emin misiniz?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if cevap != QMessageBox.StandardButton.Yes:
            return
        sonuc = sil_alt_kullanici(self._userid, kullanici_id)
        if sonuc.get("success"):
            self._durum_lbl.setText("✅  Kullanıcı silindi.")
            self._durum_lbl.setStyleSheet("font-size:11px;color:#15803d;padding:4px 0;")
            self.refresh()
        else:
            self._durum_lbl.setText(f"❌  {sonuc.get('message', 'Silinemedi.')}")
            self._durum_lbl.setStyleSheet("font-size:11px;color:#dc2626;padding:4px 0;")


# ── Hesap Tanımları Kartı ────────────────────────────────────────────────────

class HesapTanimCard(QFrame):
    """
    althesapkodu tablosunu yönetir.
      - Mevcut kayıtları tablo olarak listeler
      - Yeni hesap kodu ekle (Kod | Açıklama | Gelir/Gider)
      - Seçili satırı sil
      - CSV ile toplu yükle
      - Şema CSV'sini indir
    """

    def __init__(self, userid: int, musterino: int = 1, parent=None):
        super().__init__(parent)
        self._userid = userid
        self.setStyleSheet(
            "QFrame{background:white;border:1.5px solid #d1fae5;"
            "border-radius:14px;}"
        )
        self._build()
        self.refresh()

    # ── UI ───────────────────────────────────────────────────────────────

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(22, 20, 22, 20)
        root.setSpacing(14)

        # Başlık
        h = QHBoxLayout()
        ic = QLabel("📒")
        ic.setStyleSheet("font-size:20px;")
        h.addWidget(ic)
        t = QLabel("Hesap Tanımları")
        t.setStyleSheet("font-size:14px;font-weight:700;color:#000000;letter-spacing:.5px;")
        h.addWidget(t)
        h.addStretch()
        root.addLayout(h)

        # Bilgi bandı
        info = QLabel(
            "💡  Hesap kodlarını burada tanımlayın. "
            "Gelir/Gider hareketlerinde seçim listesinde görünürler."
        )
        info.setWordWrap(True)
        info.setStyleSheet(
            "background:#d1fae5;color:#065f46;border-radius:6px;"
            "padding:9px 12px;font-size:12px;border:none;"
        )
        root.addWidget(info)

        # ── Yeni Kayıt Formu ──
        form_frame = QFrame()
        form_frame.setStyleSheet(
            "QFrame{background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;}"
        )
        form_lay = QVBoxLayout(form_frame)
        form_lay.setContentsMargins(14, 12, 14, 12)
        form_lay.setSpacing(10)

        form_title = QLabel("Yeni Hesap Kodu Ekle")
        form_title.setStyleSheet("font-size:12px;font-weight:700;color:#374151;")
        form_lay.addWidget(form_title)

        form_row = QHBoxLayout()
        form_row.setSpacing(10)

        self._inp_kod = QLineEdit()
        self._inp_kod.setPlaceholderText("Hesap Kodu (örn: 100, 770.01)")
        self._inp_kod.setFixedHeight(34)
        self._inp_kod.setStyleSheet(self._inp_style())
        form_row.addWidget(self._inp_kod, 2)

        self._inp_ack = QLineEdit()
        self._inp_ack.setPlaceholderText("Açıklama")
        self._inp_ack.setFixedHeight(34)
        self._inp_ack.setStyleSheet(self._inp_style())
        form_row.addWidget(self._inp_ack, 3)

        self._inp_gg = QComboBox()
        self._inp_gg.setFixedHeight(34)
        self._inp_gg.setFixedWidth(130)
        self._inp_gg.addItem("Gelir", "gelir")
        self._inp_gg.addItem("Gider", "gider")
        self._inp_gg.setStyleSheet(self._combo_style())
        form_row.addWidget(self._inp_gg)

        self._ekle_btn = QPushButton("➕  Ekle")
        self._ekle_btn.setFixedHeight(34)
        self._ekle_btn.setFixedWidth(90)
        self._ekle_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._ekle_btn.setStyleSheet(self._btn_style("#059669"))
        self._ekle_btn.clicked.connect(self._on_ekle)
        form_row.addWidget(self._ekle_btn)

        form_lay.addLayout(form_row)
        root.addWidget(form_frame)

        # ── Arama / Filtre satırı ──
        filter_row = QHBoxLayout()
        filter_row.setSpacing(8)

        self._filtre_inp = QLineEdit()
        self._filtre_inp.setPlaceholderText("🔍  Kod veya açıklamada ara...")
        self._filtre_inp.setFixedHeight(34)
        self._filtre_inp.setStyleSheet(self._inp_style())
        self._filtre_inp.textChanged.connect(self._filter_tablo)
        filter_row.addWidget(self._filtre_inp, 3)

        self._filtre_gg = QComboBox()
        self._filtre_gg.setFixedHeight(34)
        self._filtre_gg.setFixedWidth(120)
        self._filtre_gg.addItem("Tümü", None)
        self._filtre_gg.addItem("↑ Gelir", "gelir")
        self._filtre_gg.addItem("↓ Gider", "gider")
        self._filtre_gg.setStyleSheet(self._combo_style())
        self._filtre_gg.currentIndexChanged.connect(self._filter_tablo)
        filter_row.addWidget(self._filtre_gg)

        self._filtre_sonuc_lbl = QLabel("")
        self._filtre_sonuc_lbl.setStyleSheet("font-size:11px;color:#6b7280;")
        filter_row.addWidget(self._filtre_sonuc_lbl)

        filter_row.addStretch()
        root.addLayout(filter_row)

        # ── Tablo ──
        self._tbl = QTableWidget()
        self._tbl.setColumnCount(4)  # Kod | Açıklama | Tür | Sil
        self._tbl.setHorizontalHeaderLabels(["Hesap Kodu", "Açıklama", "Tür", ""])
        self._tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._tbl.setAlternatingRowColors(True)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.setSortingEnabled(True)
        hdr = self._tbl.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        self._tbl.setColumnWidth(3, 70)
        self._tbl.setMinimumHeight(260)
        self._tbl.setStyleSheet("""
            QTableWidget {
                background: white;
                gridline-color: #e2e8f0;
                font-size: 12px;
                color: #000000;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
            }
            QTableWidget::item {
                color: #000000;
                padding: 3px 6px;
                background: white;
            }
            QTableWidget::item:alternate {
                background: #f1f5f9;
                color: #000000;
            }
            QTableWidget::item:selected {
                background: #dbeafe;
                color: #1e40af;
            }
            QHeaderView::section {
                background: #1e293b;
                color: white;
                font-weight: 700;
                font-size: 11px;
                padding: 5px 6px;
                border: none;
                border-right: 1px solid #334155;
            }
        """)
        root.addWidget(self._tbl, 1)

        # ── Alt buton satırı ──
        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)

        self._csv_yukle_btn = QPushButton("📂  CSV Yükle")
        self._csv_yukle_btn.setFixedHeight(34)
        self._csv_yukle_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._csv_yukle_btn.setStyleSheet(self._btn_style("#2563eb"))
        self._csv_yukle_btn.clicked.connect(self._on_csv_yukle)
        btn_row.addWidget(self._csv_yukle_btn)

        self._sema_btn = QPushButton("⬇️  Şema İndir")
        self._sema_btn.setFixedHeight(34)
        self._sema_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._sema_btn.setStyleSheet(self._btn_style("#6366f1"))
        self._sema_btn.clicked.connect(self._on_sema_indir)
        btn_row.addWidget(self._sema_btn)

        btn_row.addStretch()

        self._durum_lbl = QLabel("")
        self._durum_lbl.setStyleSheet("font-size:12px;color:#374151;")
        btn_row.addWidget(self._durum_lbl)

        root.addLayout(btn_row)

    # ── Yardımcı stiller ──

    def _inp_style(self) -> str:
        return (
            "QLineEdit{background:#f8fafc;border:1.5px solid #e2e8f0;"
            "border-radius:8px;padding:0 10px;font-size:13px;color:#1f2937;}"
            "QLineEdit:focus{border-color:#059669;}"
        )

    def _combo_style(self) -> str:
        return (
            "QComboBox{background:#f8fafc;border:1.5px solid #e2e8f0;"
            "border-radius:8px;padding:0 8px;font-size:13px;color:#1f2937;}"
            "QComboBox::drop-down{border:none;}"
            "QComboBox QAbstractItemView{background:white;color:#1f2937;"
            "selection-background-color:#d1fae5;selection-color:#065f46;}"
        )

    def _btn_style(self, color: str) -> str:
        return (
            f"QPushButton{{background:{color};color:white;border:none;"
            f"border-radius:9px;font-size:13px;font-weight:600;"
            f"padding:0 14px;letter-spacing:.4px;}}"
            f"QPushButton:hover{{background:{color}cc;}}"
            f"QPushButton:disabled{{background:#cbd5e1;color:#94a3b8;}}"
        )

    # ── Veri yükleme ──

    def refresh(self):
        from services.alt_hesap_kodu_service import get_alt_hesap_kodlari
        from PyQt6.QtGui import QColor, QBrush
        BLACK = QBrush(QColor("#000000"))
        sonuc = get_alt_hesap_kodlari(self._userid)
        data = sonuc.get("data", [])
        self._tbl.setSortingEnabled(False)
        self._tbl.setRowCount(0)
        for row in data:
            ri = self._tbl.rowCount()
            self._tbl.insertRow(ri)
            self._tbl.setRowHeight(ri, 28)

            # Hesap Kodu
            it_kod = QTableWidgetItem(row.get("kod", ""))
            it_kod.setData(Qt.ItemDataRole.UserRole, row.get("id"))
            it_kod.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
            it_kod.setForeground(BLACK)
            self._tbl.setItem(ri, 0, it_kod)

            # Açıklama
            it_ack = QTableWidgetItem(row.get("aciklama", ""))
            it_ack.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
            it_ack.setForeground(BLACK)
            self._tbl.setItem(ri, 1, it_ack)

            # Tür (Gelir/Gider) — badge tarzı metin
            gg = row.get("gelirGider", "gelir")
            gg_txt = "↑ Gelir" if gg == "gelir" else "↓ Gider"
            it_gg = QTableWidgetItem(gg_txt)
            it_gg.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            it_gg.setForeground(
                QBrush(QColor("#15803d")) if gg == "gelir" else QBrush(QColor("#dc2626"))
            )
            self._tbl.setItem(ri, 2, it_gg)

            # Sil butonu
            sil_btn = QPushButton("🗑  Sil")
            sil_btn.setFixedHeight(24)
            sil_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            sil_btn.setStyleSheet(
                "QPushButton{background:#fee2e2;color:#dc2626;border:none;"
                "border-radius:5px;font-size:11px;font-weight:600;}"
                "QPushButton:hover{background:#fca5a5;}"
            )
            sil_btn.clicked.connect(lambda _, rid=row.get("id"): self._on_sil(rid))
            self._tbl.setCellWidget(ri, 3, sil_btn)

        self._tbl.setSortingEnabled(True)
        self._durum_lbl.setText(
            f"✅  {len(data)} hesap tanımı yüklendi." if data else "⚠️  Henüz hesap tanımı bulunmuyor."
        )

    # ── Filtre ──

    def _filter_tablo(self):
        metin = self._filtre_inp.text().strip().lower()
        gg_filtre = self._filtre_gg.currentData()   # None | 'gelir' | 'gider'
        gorünen = 0
        toplam  = self._tbl.rowCount()
        for row in range(toplam):
            kod_item = self._tbl.item(row, 0)
            ack_item = self._tbl.item(row, 1)
            gg_item  = self._tbl.item(row, 2)
            kod_txt  = (kod_item.text() if kod_item else "").lower()
            ack_txt  = (ack_item.text() if ack_item else "").lower()
            gg_txt   = (gg_item.text()  if gg_item  else "").lower()   # "↑ gelir" / "↓ gider"

            metin_ok = (not metin) or (metin in kod_txt) or (metin in ack_txt)
            gg_ok    = (gg_filtre is None) or (gg_filtre in gg_txt)

            gizle = not (metin_ok and gg_ok)
            self._tbl.setRowHidden(row, gizle)
            if not gizle:
                gorünen += 1

        if metin or gg_filtre:
            self._filtre_sonuc_lbl.setText(f"{gorünen} / {toplam} kayıt")
        else:
            self._filtre_sonuc_lbl.setText("")

    # ── Ekle ──


    def _on_ekle(self):
        from services.alt_hesap_kodu_service import ekle_alt_hesap_kodu
        kod = self._inp_kod.text().strip()
        ack = self._inp_ack.text().strip()
        gg  = self._inp_gg.currentData()
        if not kod or not ack:
            self._show_durum("⚠️  Hesap kodu ve açıklama alanları zorunludur.", "#92400e")
            return
        sonuc = ekle_alt_hesap_kodu(self._userid, kod, ack, gg)
        if sonuc.get("success"):
            self._inp_kod.clear()
            self._inp_ack.clear()
            self._show_durum(f"✅  '{kod}' hesap kodu başarıyla eklendi.", "#059669")
            self.refresh()
        else:
            self._show_durum(f"❌  {sonuc.get('message', 'Bilinmeyen hata.')}", "#dc2626")

    # ── Sil ──

    def _on_sil(self, kayit_id: int):
        from services.alt_hesap_kodu_service import sil_alt_hesap_kodu
        dlg = QMessageBox(self)
        dlg.setWindowTitle("Hesap Tanımını Sil")
        dlg.setIcon(QMessageBox.Icon.Warning)
        dlg.setText(
            "Bu hesap tanımını silmek istediğinizden emin misiniz?\n"
            "Bu işlem geri alınamaz."
        )
        dlg.setStandardButtons(
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel
        )
        dlg.button(QMessageBox.StandardButton.Yes).setText("Evet, Sil")
        dlg.button(QMessageBox.StandardButton.Cancel).setText("Vazgeç")
        if dlg.exec() != QMessageBox.StandardButton.Yes:
            return
        sonuc = sil_alt_hesap_kodu(self._userid, kayit_id)
        if sonuc.get("success"):
            self._show_durum("✅  Hesap tanımı silindi.", "#059669")
            self.refresh()
        else:
            self._show_durum(f"❌  {sonuc.get('message', 'Silme hatası.')}", "#dc2626")

    # ── CSV Toplu Yükle ──

    def _on_csv_yukle(self):
        from services.alt_hesap_kodu_service import toplu_yukle_csv
        dosya, _ = QFileDialog.getOpenFileName(
            self, "CSV Dosyası Seç", "", "CSV Dosyaları (*.csv)"
        )
        if not dosya:
            return
        sonuc = toplu_yukle_csv(self._userid, dosya)
        if sonuc.get("success"):
            self._show_durum(
                f"✅  {sonuc.get('added', 0)} kayıt eklendi, "
                f"{sonuc.get('skipped', 0)} atlandı.",
                "#059669"
            )
            self.refresh()
        else:
            self._show_durum(f"❌  {sonuc.get('message', 'CSV yükleme hatası.')}", "#dc2626")

    # ── Şema CSV İndir ──

    def _on_sema_indir(self):
        dosya, _ = QFileDialog.getSaveFileName(
            self, "Şema CSV'yi Kaydet", "alt_hesap_kodlari_sema.csv", "CSV Dosyaları (*.csv)"
        )
        if not dosya:
            return
        try:
            import csv
            with open(dosya, "w", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f)
                w.writerow(["kod", "aciklama", "gelirGider"])
                w.writerow(["100", "Nakit Para", "gelir"])
                w.writerow(["770.01", "Vergi Giderleri", "gider"])
            self._show_durum("✅  Şema CSV dosyası kaydedildi.", "#059669")
        except Exception as exc:
            self._show_durum(f"❌  {exc}", "#dc2626")

    # ── Durum etiketi ──

    def _show_durum(self, msg: str, color: str = "#374151"):
        self._durum_lbl.setText(msg)
        self._durum_lbl.setStyleSheet(f"font-size:12px;color:{color};")


# ── Ana Ayarlar Ekranı ────────────────────────────────────────────────────────

class AyarlarScreen(QWidget):
    def __init__(self, user: dict, parent=None):
        super().__init__(parent)
        self._user = user
        self._userid    = user.get("GercekUserId", user.get("Kayitno", 1))
        self._musterino = user.get("musterino", user.get("GercekUserId", 1))
        self._setup_ui()

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(20)

        # Başlık
        title = QLabel("Ayarlar")
        title.setStyleSheet(
            "color:#000000;"
            "font-size:24px;font-weight:700;letter-spacing:1px;"
        )
        root.addWidget(title)

        # Tab butonları
        tab_row = QHBoxLayout()
        tab_row.setSpacing(8)
        self._tab_btns: dict[str, QPushButton] = {}
        tabs = [
            ("hesap",      "👤  Hesap & Güvenlik"),
            ("eklentiler", "🔌  Eklentiler"),
            ("veritabani", "🗄️  Veritabanı"),
        ]

        for key, label in tabs:
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.setFixedHeight(34)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setStyleSheet(self._tab_style())
            btn.clicked.connect(lambda _, k=key: self._switch_tab(k))
            self._tab_btns[key] = btn
            tab_row.addWidget(btn)
        tab_row.addStretch()
        root.addLayout(tab_row)

        # Scroll alan
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setStyleSheet("QScrollArea{background:transparent;}")

        self._content = QWidget()
        self._content.setStyleSheet("background:transparent;")
        self._content_layout = QVBoxLayout(self._content)
        self._content_layout.setContentsMargins(0, 0, 0, 0)
        self._content_layout.setSpacing(16)

        scroll.setWidget(self._content)
        root.addWidget(scroll)

        # Hesap & Güvenlik sekmesi — Şirket Profili
        self._sirket_card = SirketProfilCard(self._userid, musterino=self._musterino)
        self._content_layout.addWidget(self._sirket_card)

        # Kullanıcı Yetkileri kartı — Şirket Profili'nin altında
        self._kullanici_yonetim_card = KullaniciYonetimCard(self._userid)
        self._content_layout.addWidget(self._kullanici_yonetim_card)

        # Manuel Toplu İşle kartı — EFatura'nın üstünde
        self._gsheets_settings_card = GoogleSheetsSettingsCard()
        self._content_layout.addWidget(self._gsheets_settings_card)

        self._manuel_toplu_card = ManuelTopluIsleCard(self._userid, musterino=self._musterino)
        self._content_layout.addWidget(self._manuel_toplu_card)

        # Eklentiler kartı
        self._efatura_card = EFaturaCard(self._userid, self._musterino)
        self._content_layout.addWidget(self._efatura_card)

        # Fatura Yönetim kartı
        self._yonetim_card = FaturaYonetimCard(self._userid)
        self._content_layout.addWidget(self._yonetim_card)

        # VOMSİS API kartı — Fatura Yönetim'in altında
        self._vomsis_card = VomsisCard(self._userid, musterino=self._musterino)
        self._content_layout.addWidget(self._vomsis_card)

        # Moy kartı — VOMSİS'in altında
        self._moy_card = MoyCard(self._userid, musterino=self._musterino)
        self._content_layout.addWidget(self._moy_card)

        # Vergi Muhtasar kartı — Moy'un altında
        self._vergi_muhtasar_card = VergiMuhtasarCard(self._userid)
        self._content_layout.addWidget(self._vergi_muhtasar_card)

        # Kredi Kartı kartı — Vergi Muhtasar'ın altında
        self._kredi_kart_card = KrediKartCard(self._userid, self._musterino)
        self._content_layout.addWidget(self._kredi_kart_card)

        # Hesap Tanımları kartı — Eklentiler sekmesinde
        self._hesap_tanim_card = HesapTanimCard(self._userid)
        self._content_layout.addWidget(self._hesap_tanim_card)

        # Veritabanı sekmesi
        self._veritabani_card = VeriTabaniCard()
        self._content_layout.addWidget(self._veritabani_card)

        self._content_layout.addStretch()

        # Aktif tab
        self._switch_tab("eklentiler")

    def _tab_style(self) -> str:
        return (
            "QPushButton{background:#f1f5f9;border:1.5px solid #e2e8f0;"
            "border-radius:8px;font-size:13px;font-weight:600;color:#000000;"
            "padding:0 16px;}"
            "QPushButton:checked{background:#1e293b;color:white;border-color:#1e293b;}"
            "QPushButton:hover:!checked{background:#e2e8f0;}"
        )

    def _switch_tab(self, key: str):
        for k, btn in self._tab_btns.items():
            btn.setChecked(k == key)

        self._sirket_card.setVisible(key == "hesap")
        self._kullanici_yonetim_card.setVisible(key == "hesap")
        self._efatura_card.setVisible(key == "eklentiler")
        self._yonetim_card.setVisible(key == "eklentiler")
        self._vomsis_card.setVisible(key == "eklentiler")
        self._moy_card.setVisible(key == "eklentiler")
        self._vergi_muhtasar_card.setVisible(key == "eklentiler")
        self._kredi_kart_card.setVisible(key == "eklentiler")
        self._hesap_tanim_card.setVisible(key == "eklentiler")
        self._veritabani_card.setVisible(key == "veritabani")

        if key == "hesap":
            self._kullanici_yonetim_card.refresh()

        if key == "veritabani":
            self._veritabani_card.refresh()

        if key == "eklentiler":
            if hasattr(self._efatura_card, "refresh"):
                self._efatura_card.refresh()
            if hasattr(self._vomsis_card, "refresh"):
                self._vomsis_card.refresh()
            if hasattr(self._moy_card, "refresh"):
                self._moy_card.refresh()
            if hasattr(self._vergi_muhtasar_card, "refresh"):
                self._vergi_muhtasar_card.refresh()
            self._hesap_tanim_card.refresh()




# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  VERİTABANI AYARLARI — SweetAlert tasarımı, batch migration, local/PG mod  ║
# ╚══════════════════════════════════════════════════════════════════════════════╝


def _mk_lbl(text, style=""):
    l = QLabel(text)
    if style:
        l.setStyleSheet(style)
    return l


# ─────────────────────────────────────────────────────────────────────────────
# ── webadmin Bağlantı Ayar Kartı ─────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

class WebAdminConfigCard(QFrame):
    """
    webadmin sunucusuna bağlantı, SSH tünel ve HTTPS ayarlarını yönetir.
    Ayarlar → Eklentiler → VOMSİS API kartının altında görünür.
    """

    def __init__(self, userid: int, musterino: int = 1, parent=None):
        super().__init__(parent)
        self._userid = userid
        self.setStyleSheet(
            "QFrame{background:white;border:1.5px solid #ddd6fe;"
            "border-radius:14px;}"
        )
        self._build()
        self._load()

    # ── Stil yardımcıları ─────────────────────────────────────────────────────

    def _inp(self) -> str:
        return (
            "QLineEdit{background:#f8fafc;border:1.5px solid #e2e8f0;"
            "border-radius:8px;padding:0 10px;font-size:13px;color:#1f2937;}"
            "QLineEdit:focus{border-color:#7c3aed;}"
        )

    def _lbl(self) -> str:
        return "font-size:12px;font-weight:600;color:#374151;"

    def _btn(self, color: str) -> str:
        c2 = {"#7c3aed": "#6d28d9", "#059669": "#047857",
               "#dc2626": "#b91c1c", "#0891b2": "#0e7490"}.get(color, color)
        return (
            f"QPushButton{{background:{color};color:white;border:none;"
            f"border-radius:8px;font-size:12px;font-weight:700;"
            f"padding:0 14px;}}"
            f"QPushButton:hover{{background:{c2};}}"
            f"QPushButton:disabled{{background:#e2e8f0;color:#94a3b8;}}"
        )

    # ── UI İnşa ───────────────────────────────────────────────────────────────

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(22, 18, 22, 18)
        root.setSpacing(14)

        # ── Başlık ──
        hdr = QHBoxLayout()
        ic = QLabel("🔌")
        ic.setStyleSheet("font-size:20px;")
        hdr.addWidget(ic)
        ttl = QLabel("webadmin Bağlantı Ayarları")
        ttl.setStyleSheet("font-size:14px;font-weight:700;color:#1e293b;")
        hdr.addWidget(ttl)
        hdr.addStretch()
        # Durum badge
        self._status_badge = QLabel("⚪ Bağlı değil")
        self._status_badge.setStyleSheet(
            "font-size:11px;font-weight:600;color:#64748b;"
            "background:#f1f5f9;border-radius:10px;padding:2px 8px;"
        )
        hdr.addWidget(self._status_badge)
        root.addLayout(hdr)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color:#ede9fe;")
        root.addWidget(sep)

        # ── Temel Bağlantı ──
        grp1_lbl = QLabel("🌐  Sunucu Bağlantısı")
        grp1_lbl.setStyleSheet("font-size:12px;font-weight:700;color:#7c3aed;")
        root.addWidget(grp1_lbl)

        row1 = QHBoxLayout()
        row1.setSpacing(10)

        col_url = QVBoxLayout()
        col_url.addWidget(self._mk_lbl("Sunucu URL"))
        self._url_inp = QLineEdit()
        self._url_inp.setFixedHeight(34)
        self._url_inp.setPlaceholderText("http://localhost:5050")
        self._url_inp.setStyleSheet(self._inp())
        col_url.addWidget(self._url_inp)
        row1.addLayout(col_url, 3)

        col_key = QVBoxLayout()
        col_key.addWidget(self._mk_lbl("API Key"))
        self._key_inp = QLineEdit()
        self._key_inp.setFixedHeight(34)
        self._key_inp.setEchoMode(QLineEdit.EchoMode.Password)
        self._key_inp.setPlaceholderText("••••••••••••")
        self._key_inp.setStyleSheet(self._inp())
        col_key.addWidget(self._key_inp)
        row1.addLayout(col_key, 2)

        root.addLayout(row1)

        # ── Otomatik Sync On/Off ──────────────────────────────────────────────
        sync_row = QHBoxLayout()
        sync_row.setSpacing(10)

        sync_ic = QLabel("⏱️")
        sync_ic.setStyleSheet("font-size:18px;")
        sync_row.addWidget(sync_ic)

        sync_lbl = QLabel("Otomatik Sync (Login sonrası)")
        sync_lbl.setStyleSheet("font-size:13px;font-weight:600;color:#1e293b;")
        sync_row.addWidget(sync_lbl)
        sync_row.addStretch()

        self._sync_on_btn  = QPushButton("▶  AÇIK")
        self._sync_off_btn = QPushButton("⏸  KAPALI")

        for btn in (self._sync_on_btn, self._sync_off_btn):
            btn.setFixedHeight(30)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setCheckable(True)

        self._sync_on_btn.setStyleSheet(
            "QPushButton{background:#dcfce7;color:#166534;border:2px solid #16a34a;"
            "border-radius:7px;font-size:12px;font-weight:700;padding:0 12px;}"
            "QPushButton:checked{background:#16a34a;color:white;}"
            "QPushButton:hover{background:#bbf7d0;}"
        )
        self._sync_off_btn.setStyleSheet(
            "QPushButton{background:#f1f5f9;color:#64748b;border:2px solid #e2e8f0;"
            "border-radius:7px;font-size:12px;font-weight:700;padding:0 12px;}"
            "QPushButton:checked{background:#64748b;color:white;}"
            "QPushButton:hover{background:#e2e8f0;}"
        )

        self._sync_on_btn.clicked.connect(lambda: self._set_sync_enabled(True))
        self._sync_off_btn.clicked.connect(lambda: self._set_sync_enabled(False))

        sync_row.addWidget(self._sync_on_btn)
        sync_row.addWidget(self._sync_off_btn)
        root.addLayout(sync_row)

        sync_note = QLabel(
            "💡  KAPALI yapıldığında login sonrası otomatik banka ve POS verisi çekilmez."
        )
        sync_note.setWordWrap(True)
        sync_note.setStyleSheet(
            "font-size:11px;color:#64748b;background:#f8fafc;"
            "border-radius:6px;padding:6px 10px;"
        )
        root.addWidget(sync_note)

        # ── SSH Tünel Bölümü (toggle) ──
        self._ssh_toggle = QPushButton("🔐  SSH Tünel  ▸")
        self._ssh_toggle.setCheckable(True)
        self._ssh_toggle.setStyleSheet(
            "QPushButton{background:#f5f3ff;color:#7c3aed;border:1px solid #ddd6fe;"
            "border-radius:8px;font-size:12px;font-weight:700;padding:6px 12px;"
            "text-align:left;}"
            "QPushButton:checked{background:#ede9fe;}"
            "QPushButton:hover{background:#ede9fe;}"
        )
        self._ssh_toggle.clicked.connect(self._toggle_ssh)
        root.addWidget(self._ssh_toggle)


        self._ssh_frame = QFrame()
        self._ssh_frame.setStyleSheet(
            "QFrame{background:#faf5ff;border:1px solid #ddd6fe;"
            "border-radius:10px;}"
        )
        ssh_lay = QVBoxLayout(self._ssh_frame)
        ssh_lay.setContentsMargins(14, 12, 14, 12)
        ssh_lay.setSpacing(10)

        ssh_info = QLabel(
            "💡  SSH Tünel: webadmin sunucusu uzak makinedeyse, "
            "SSH üzerinden güvenli bağlantı kurulur."
        )
        ssh_info.setWordWrap(True)
        ssh_info.setStyleSheet("font-size:11px;color:#6d28d9;")
        ssh_lay.addWidget(ssh_info)

        ssh_row1 = QHBoxLayout()
        ssh_row1.setSpacing(8)

        c_host = QVBoxLayout()
        c_host.addWidget(self._mk_lbl("SSH Host / IP"))
        self._ssh_host = QLineEdit()
        self._ssh_host.setFixedHeight(32)
        self._ssh_host.setPlaceholderText("212.xxx.xxx.xxx")
        self._ssh_host.setStyleSheet(self._inp())
        c_host.addWidget(self._ssh_host)
        ssh_row1.addLayout(c_host, 3)

        c_port = QVBoxLayout()
        c_port.addWidget(self._mk_lbl("SSH Port"))
        self._ssh_port = QLineEdit("22")
        self._ssh_port.setFixedHeight(32)
        self._ssh_port.setFixedWidth(70)
        self._ssh_port.setStyleSheet(self._inp())
        c_port.addWidget(self._ssh_port)
        ssh_row1.addLayout(c_port)

        c_user = QVBoxLayout()
        c_user.addWidget(self._mk_lbl("Kullanıcı Adı"))
        self._ssh_user = QLineEdit()
        self._ssh_user.setFixedHeight(32)
        self._ssh_user.setPlaceholderText("ubuntu")
        self._ssh_user.setStyleSheet(self._inp())
        c_user.addWidget(self._ssh_user)
        ssh_row1.addLayout(c_user, 2)

        ssh_lay.addLayout(ssh_row1)

        c_key = QVBoxLayout()
        c_key.addWidget(self._mk_lbl("SSH Private Key Dosyası (.pem / id_rsa)"))
        key_row = QHBoxLayout()
        key_row.setSpacing(6)
        self._ssh_key_path = QLineEdit()
        self._ssh_key_path.setFixedHeight(32)
        self._ssh_key_path.setPlaceholderText("~/.ssh/id_rsa")
        self._ssh_key_path.setStyleSheet(self._inp())
        key_row.addWidget(self._ssh_key_path)
        btn_browse = QPushButton("📂")
        btn_browse.setFixedSize(32, 32)
        btn_browse.setStyleSheet(
            "QPushButton{background:#ede9fe;border:1px solid #ddd6fe;"
            "border-radius:6px;font-size:14px;}"
            "QPushButton:hover{background:#ddd6fe;}"
        )
        btn_browse.clicked.connect(self._browse_key)
        key_row.addWidget(btn_browse)
        c_key.addLayout(key_row)
        ssh_lay.addLayout(c_key)

        self._ssh_frame.hide()
        root.addWidget(self._ssh_frame)

        # ── HTTPS Sertifika ──
        sep2 = QFrame()
        sep2.setFrameShape(QFrame.Shape.HLine)
        sep2.setStyleSheet("color:#ede9fe;")
        root.addWidget(sep2)

        https_lbl = QLabel("🔒  HTTPS / SSL Sertifikası")
        https_lbl.setStyleSheet("font-size:12px;font-weight:700;color:#7c3aed;")
        root.addWidget(https_lbl)

        https_info = QLabel(
            "webadmin sunucusu için self-signed SSL sertifikası oluşturur. "
            "Sertifika webadmin klasörüne (cert.pem + key.pem) kaydedilir ve "
            "sunucu HTTPS üzerinden çalışmaya başlar."
        )
        https_info.setWordWrap(True)
        https_info.setStyleSheet("font-size:11px;color:#64748b;")
        root.addWidget(https_info)

        self._cert_status = QLabel("📋  Sertifika: Yok")
        self._cert_status.setStyleSheet(
            "font-size:11px;color:#64748b;background:#f8fafc;"
            "border-radius:6px;padding:4px 8px;"
        )
        root.addWidget(self._cert_status)

        https_row = QHBoxLayout()
        https_row.setSpacing(8)
        self._btn_gen_cert = QPushButton("🔒  HTTPS Sertifikası Oluştur")
        self._btn_gen_cert.setFixedHeight(36)
        self._btn_gen_cert.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_gen_cert.setStyleSheet(self._btn("#059669"))
        self._btn_gen_cert.clicked.connect(self._gen_cert)
        https_row.addWidget(self._btn_gen_cert)

        self._btn_check_cert = QPushButton("🔍  Sertifika Kontrol")
        self._btn_check_cert.setFixedHeight(36)
        self._btn_check_cert.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_check_cert.setStyleSheet(self._btn("#0891b2"))
        self._btn_check_cert.clicked.connect(self._check_cert)
        https_row.addWidget(self._btn_check_cert)
        https_row.addStretch()
        root.addLayout(https_row)

        # ── Alt Butonlar ──
        sep3 = QFrame()
        sep3.setFrameShape(QFrame.Shape.HLine)
        sep3.setStyleSheet("color:#ede9fe;")
        root.addWidget(sep3)

        self._status_lbl = QLabel("")
        self._status_lbl.setWordWrap(True)
        self._status_lbl.setStyleSheet(
            "font-size:12px;color:#374151;background:#f8fafc;"
            "border-radius:6px;padding:6px 10px;border:none;"
        )
        self._status_lbl.hide()
        root.addWidget(self._status_lbl)

        foot = QHBoxLayout()
        foot.setSpacing(8)

        self._btn_test = QPushButton("📡  Bağlantı Test Et")
        self._btn_test.setFixedHeight(36)
        self._btn_test.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_test.setStyleSheet(self._btn("#0891b2"))
        self._btn_test.clicked.connect(self._test_connection)
        foot.addWidget(self._btn_test)

        self._btn_save = QPushButton("💾  Kaydet")
        self._btn_save.setFixedHeight(36)
        self._btn_save.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_save.setStyleSheet(self._btn("#7c3aed"))
        self._btn_save.clicked.connect(self._save)
        foot.addWidget(self._btn_save)

        foot.addStretch()
        root.addLayout(foot)

    # ── Yardımcılar ───────────────────────────────────────────────────────────

    def _mk_lbl(self, text: str) -> QLabel:
        l = QLabel(text)
        l.setStyleSheet(self._lbl())
        return l

    def _toggle_ssh(self, checked: bool):
        self._ssh_frame.setVisible(checked)
        arrow = "▾" if checked else "▸"
        self._ssh_toggle.setText(f"🔐  SSH Tünel  {arrow}")

    def _browse_key(self):
        from PyQt6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getOpenFileName(
            self, "SSH Private Key Seç", str(Path.home() / ".ssh"),
            "Key Dosyaları (*.pem *.key id_rsa);;Tüm Dosyalar (*)"
        )
        if path:
            self._ssh_key_path.setText(path)

    def _show_status(self, msg: str, color: str = "#374151"):
        self._status_lbl.setStyleSheet(
            f"font-size:12px;color:{color};background:#f8fafc;"
            "border-radius:6px;padding:6px 10px;border:none;"
        )
        self._status_lbl.setText(msg)
        self._status_lbl.show()

    # ── Veri Yükleme ─────────────────────────────────────────────────────────

    def _load(self):
        import json
        cfg_path = Path.home() / "NakitAkim" / "data" / "webadmin_config.json"
        if not cfg_path.exists():
            return
        try:
            cfg = json.loads(cfg_path.read_text())
            self._url_inp.setText(cfg.get("base_url", ""))
            self._key_inp.setText(cfg.get("api_key", ""))
            # SSH
            ssh = cfg.get("ssh", {})
            if ssh:
                self._ssh_host.setText(ssh.get("host", ""))
                self._ssh_port.setText(str(ssh.get("port", 22)))
                self._ssh_user.setText(ssh.get("username", ""))
                self._ssh_key_path.setText(ssh.get("key_path", ""))
                if ssh.get("enabled"):
                    self._ssh_toggle.setChecked(True)
                    self._toggle_ssh(True)
        except Exception:
            pass
        self._check_cert()
        # ── Otomatik sync toggle durumunu DB'den oku ──
        try:
            from services.webadmin_client import get_webadmin_config
            wcfg = get_webadmin_config(self._userid)
            is_on = wcfg.get("auto_sync_enabled", True)
            self._sync_on_btn.setChecked(is_on)
            self._sync_off_btn.setChecked(not is_on)
        except Exception:
            self._sync_on_btn.setChecked(True)
            self._sync_off_btn.setChecked(False)

    def _set_sync_enabled(self, enabled: bool):
        """Otomatik sync toggle — DB'ye yazar, butonları günceller."""
        self._sync_on_btn.setChecked(enabled)
        self._sync_off_btn.setChecked(not enabled)
        try:
            from services.webadmin_client import set_auto_sync_enabled
            ok = set_auto_sync_enabled(self._userid, enabled)
            if ok:
                durum = "AÇIK ▶" if enabled else "KAPALI ⏸"
                self._show_status(
                    f"✅  Otomatik Sync {durum} — ayar kaydedildi.",
                    "#059669" if enabled else "#64748b"
                )
            else:
                self._show_status("⚠️  Ayar kaydedilemedi (DB hatası).", "#92400e")
        except Exception as e:
            self._show_status(f"Hata: {e}", "#dc2626")

    # ── Kaydetme ─────────────────────────────────────────────────────────────

    def _save(self):
        import json
        url = self._url_inp.text().strip().rstrip("/")
        key = self._key_inp.text().strip()

        if not url:
            self._show_status("⚠️  Sunucu URL boş olamaz.", "#92400e")
            return

        cfg = {
            "base_url": url,
            "api_key":  key,
            "timeout":  60,
            "enabled":  True,
        }

        if self._ssh_toggle.isChecked():
            cfg["ssh"] = {
                "enabled":  True,
                "host":     self._ssh_host.text().strip(),
                "port":     int(self._ssh_port.text().strip() or "22"),
                "username": self._ssh_user.text().strip(),
                "key_path": self._ssh_key_path.text().strip(),
                "remote_bind_port": 5050,
                "local_bind_port":  5051,
            }
        else:
            cfg["ssh"] = {"enabled": False}

        cfg_path = Path.home() / "NakitAkim" / "data" / "webadmin_config.json"
        cfg_path.parent.mkdir(parents=True, exist_ok=True)
        cfg_path.write_text(json.dumps(cfg, ensure_ascii=False, indent=2))
        self._show_status("✅  Ayarlar kaydedildi.", "#059669")

    # ── Bağlantı Testi ───────────────────────────────────────────────────────

    def _test_connection(self):
        self._btn_test.setEnabled(False)
        self._show_status("📡  Test ediliyor...", "#6366f1")

        url = self._url_inp.text().strip().rstrip("/") or "http://localhost:5050"
        key = self._key_inp.text().strip()

        import threading
        def _do_test():
            try:
                import requests
                resp = requests.get(f"{url}/", timeout=5)
                ok = resp.status_code in (200, 302, 404)
            except Exception as e:
                ok = False

            # UI thread'e gönder
            from PyQt6.QtCore import QMetaObject, Qt, Q_ARG
            QMetaObject.invokeMethod(
                self, "_on_test_result",
                Qt.ConnectionType.QueuedConnection,
                Q_ARG(bool, ok),
                Q_ARG(str, url),
            )

        threading.Thread(target=_do_test, daemon=True).start()

    from PyQt6.QtCore import pyqtSlot

    @pyqtSlot(bool, str)
    def _on_test_result(self, ok: bool, url: str):
        self._btn_test.setEnabled(True)
        if ok:
            self._show_status(f"✅  {url} adresine bağlantı başarılı!", "#059669")
            self._status_badge.setText("🟢 Bağlı")
            self._status_badge.setStyleSheet(
                "font-size:11px;font-weight:600;color:#059669;"
                "background:#f0fdf4;border-radius:10px;padding:2px 8px;"
            )
        else:
            self._show_status(
                f"❌  {url} adresine ulaşılamadı. "
                "webadmin sunucusunun çalıştığından emin olun.", "#dc2626"
            )
            self._status_badge.setText("🔴 Bağlı değil")
            self._status_badge.setStyleSheet(
                "font-size:11px;font-weight:600;color:#dc2626;"
                "background:#fef2f2;border-radius:10px;padding:2px 8px;"
            )

    # ── HTTPS Sertifika Oluştur ───────────────────────────────────────────────

    def _gen_cert(self):
        confirm = SweetConfirmDialog.confirm(
            self,
            title="HTTPS Sertifikası Oluştur",
            text=(
                "webadmin klasörüne self-signed SSL sertifikası (cert.pem + key.pem) "
                "oluşturulacak.\n\n"
                "Bu işlem openssl komutunu kullanır. Devam edilsin mi?"
            ),
            confirm_text="🔒  Oluştur",
            cancel_text="İptal",
        )
        if not confirm:
            return

        self._btn_gen_cert.setEnabled(False)
        self._show_status("🔒  Sertifika oluşturuluyor...", "#059669")

        import subprocess, threading

        webadmin_dir = Path.home() / "webadmin-nakitAkim"
        cert_path = webadmin_dir / "cert.pem"
        key_path  = webadmin_dir / "key.pem"

        def _run():
            try:
                result = subprocess.run([
                    "openssl", "req", "-x509",
                    "-newkey", "rsa:4096",
                    "-keyout", str(key_path),
                    "-out",    str(cert_path),
                    "-days",   "365",
                    "-nodes",
                    "-subj",   "/C=TR/ST=Istanbul/L=Istanbul/O=webadmin-nakitAkim/CN=localhost"
                ], capture_output=True, text=True, timeout=60)

                success = cert_path.exists() and key_path.exists()
                msg = "✅  Sertifika oluşturuldu." if success else f"❌  Hata: {result.stderr[:200]}"
            except FileNotFoundError:
                success = False
                msg = "❌  openssl komutu bulunamadı. Lütfen openssl yükleyin: brew install openssl"
            except Exception as e:
                success = False
                msg = f"❌  Hata: {e}"

            from PyQt6.QtCore import QMetaObject, Qt, Q_ARG
            QMetaObject.invokeMethod(
                self, "_on_cert_done",
                Qt.ConnectionType.QueuedConnection,
                Q_ARG(bool, success),
                Q_ARG(str, msg),
            )

        threading.Thread(target=_run, daemon=True).start()

    @pyqtSlot(bool, str)
    def _on_cert_done(self, success: bool, msg: str):
        self._btn_gen_cert.setEnabled(True)
        self._show_status(msg, "#059669" if success else "#dc2626")
        if success:
            self._check_cert()
            # Config'e HTTPS URL öner
            cur_url = self._url_inp.text().strip()
            if cur_url.startswith("http://"):
                https_url = cur_url.replace("http://", "https://")
                self._url_inp.setText(https_url)
                self._show_status(
                    f"✅  Sertifika oluşturuldu. URL otomatik HTTPS'e güncellendi: {https_url}\n"
                    "💡  webadmin'i yeniden başlatın: python3 app.py",
                    "#059669"
                )

    def _check_cert(self):
        webadmin_dir = Path.home() / "webadmin-nakitAkim"
        cert_path = webadmin_dir / "cert.pem"
        key_path  = webadmin_dir / "key.pem"

        if cert_path.exists() and key_path.exists():
            try:
                import subprocess
                result = subprocess.run(
                    ["openssl", "x509", "-in", str(cert_path),
                     "-noout", "-enddate"],
                    capture_output=True, text=True, timeout=5
                )
                expire_line = result.stdout.strip()
                expire = expire_line.replace("notAfter=", "") if expire_line else "?"
                self._cert_status.setText(f"🔐  Sertifika: Mevcut  |  Son geçerlilik: {expire}")
                self._cert_status.setStyleSheet(
                    "font-size:11px;color:#059669;background:#f0fdf4;"
                    "border-radius:6px;padding:4px 8px;"
                )
            except Exception:
                self._cert_status.setText("🔐  Sertifika: Mevcut (detay alınamadı)")
        else:
            self._cert_status.setText("📋  Sertifika: Yok — 'Oluştur' butonuna basın")
            self._cert_status.setStyleSheet(
                "font-size:11px;color:#64748b;background:#f8fafc;"
                "border-radius:6px;padding:4px 8px;"
            )



# ─────────────────────────────────────────────────────────────────────────────
# ── VOMSİS Excel Import — Worker (modül seviyesinde olmalı, PyQt6 gereği) ────
# ─────────────────────────────────────────────────────────────────────────────

class WomsisExcelWorker(QThread):
    """
    Excel dosyasını arka planda işler.
    NOT: PyQt6'da pyqtSignal'ların doğru çalışması için
    sınıf MUTLAKA modül seviyesinde tanımlanmalıdır.
    """
    satir_done   = pyqtSignal(str, int, int, str)   # sayfa, eklenen, atlanan, durum
    header_error = pyqtSignal(str, object, object)  # sayfa, bulunan_list, eksik_list
    progress     = pyqtSignal(str)                  # canlı durum mesajı
    finished     = pyqtSignal(dict)

    ZORUNLU_KOLONLAR = ["HAREKET ID", "TARİH", "AÇIKLAMA", "TUTAR"]

    def __init__(self, filepath: str, userid: int):
        super().__init__()
        self._path   = filepath
        self._userid = userid

    def run(self):
        try:
            import openpyxl
        except ImportError:
            self.finished.emit({
                "success": False,
                "message": "openpyxl yüklü değil. pip install openpyxl",
                "toplam": 0, "eklenen": 0, "atlanan": 0, "hata": 1,
            })
            return

        from db.database import get_connection

        try:
            wb = openpyxl.load_workbook(self._path, data_only=True)
        except Exception as e:
            self.finished.emit({
                "success": False,
                "message": f"Excel dosyası açılamadı: {e}",
                "toplam": 0, "eklenen": 0, "atlanan": 0, "hata": 1,
            })
            return

        conn = get_connection()
        g_toplam = g_eklenen = g_atlanan = g_hata = 0

        try:
            for sheet_name in wb.sheetnames:
                ws  = wb[sheet_name]
                ins = skip = err = 0

                # ── Başlık Satırı Bul & Doğrula ──────────────────────────────
                header_row = None
                col_map    = {}
                for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
                    if row and any(str(v).strip() == "HAREKET ID" for v in row if v):
                        header_row = r_idx
                        col_map = {
                            str(v).strip(): i
                            for i, v in enumerate(row) if v and str(v).strip()
                        }
                        break

                if header_row is None:
                    self.header_error.emit(sheet_name, [], self.ZORUNLU_KOLONLAR)
                    self.satir_done.emit(sheet_name, 0, 0, "❌ Başlık satırı bulunamadı")
                    g_hata += 1
                    continue

                bulunan_kolonlar = list(col_map.keys())
                eksik_kolonlar   = [k for k in self.ZORUNLU_KOLONLAR if k not in col_map]
                if eksik_kolonlar:
                    self.header_error.emit(sheet_name, bulunan_kolonlar, eksik_kolonlar)
                    self.satir_done.emit(
                        sheet_name, 0, 0,
                        f"❌ Eksik kolon: {', '.join(eksik_kolonlar)}"
                    )
                    g_hata += 1
                    continue

                self.progress.emit(f"📂  {sheet_name} işleniyor...")

                # ── Veri Satırları (Ters Çevrilerek) ──────────────────────────
                all_rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
                all_rows.reverse()  # Eskiden yeniye doğru işle
                
                bakiye_onceki = None

                for row in all_rows:
                    if not row or all(v is None for v in row):
                        continue

                    def _get(key, default=None, _row=row, _map=col_map):
                        idx = _map.get(key)
                        return _row[idx] if idx is not None and idx < len(_row) else default

                    hareket_id  = _get("HAREKET ID")
                    tarih_raw   = _get("TARİH", "")
                    aciklama    = str(_get("AÇIKLAMA", "") or "")
                    tutar_val   = _get("TUTAR", 0)
                    # ── BANKA/ŞUBE: Aynı şube adına sahip birden fazla hesap
                    # (örn. İŞ BANKASI-1667134 ve İŞ BANKASI-1667311 her ikisi de
                    # 'Türkiye İş Bankası / Sultanhamam/istanbul' döndürür).
                    # Çakışmayı önlemek için HESAP NO (sheet_name'in son kısmı)
                    # eklenerek banka_sube eşsiz hale getirilir.
                    _raw_sube   = str(_get("BANKA/ŞUBE", "") or "").strip()
                    _hesap_no   = str(_get("HESAP NO", "") or "").strip()
                    if not _raw_sube:
                        _raw_sube = sheet_name
                    if _hesap_no and _hesap_no not in _raw_sube:
                        banka_sube = f"{_raw_sube} [{_hesap_no}]"
                    else:
                        banka_sube = _raw_sube
                    karsi_unvan = str(_get("KARŞI TARAF ÜNVAN", "") or "")
                    dekont_no   = str(_get("DEKONT NO", "") or "")
                    iban_val    = str(_get("IBAN", "") or "")
                    hesap_turu  = str(_get("HESAP TÜRÜ", "") or "")
                    bakiye_val  = _get("BAKİYE")

                    g_toplam += 1

                    # Parse values
                    try:
                        tutar_excel = round(float(str(tutar_val).replace(",", ".").replace(" ", "")), 2)
                    except Exception:
                        err += 1; g_hata += 1; continue

                    bakiye_float = None
                    if bakiye_val is not None and str(bakiye_val).strip() != "":
                        try:
                            bakiye_float = round(float(str(bakiye_val).replace(",", ".").replace(" ", "")), 2)
                        except Exception:
                            pass

                    # Gelir/Gider ve Tutar Hesaplama
                    # Bakiye varsa ve önceki bakiye biliniyorsa farktan hesapla, yoksa Tutar kolonuna güven.
                    if bakiye_float is not None and bakiye_onceki is not None:
                        fark = round(bakiye_float - bakiye_onceki, 2)
                        gelir_gider = "gelir" if fark >= 0 else "gider"
                        tutar_abs = abs(fark)
                    else:
                        gelir_gider = "gelir" if tutar_excel >= 0 else "gider"
                        tutar_abs = abs(tutar_excel)

                    bakiye_onceki = bakiye_float

                    # Tarih dönüşümü DD/MM/YYYY → YYYY-MM-DD
                    tarih_str = ""
                    if tarih_raw:
                        t = str(tarih_raw).strip()
                        for fmt in ("%d/%m/%Y", "%d.%m.%Y", "%Y-%m-%d", "%Y/%m/%d"):
                            try:
                                tarih_str = datetime.datetime.strptime(t[:10], fmt).strftime("%Y-%m-%d")
                                break
                            except Exception:
                                continue
                        if not tarih_str:
                            tarih_str = t

                    # womsisKey — HESAP NO prefix ile farklı hesaplar ayrışır
                    # (Örn: 1667134_13245 ≠ 1667311_13245)
                    _key_prefix = _hesap_no if _hesap_no else sheet_name
                    if hareket_id:
                        womsis_key = f"{_key_prefix}_{hareket_id}"
                    else:
                        womsis_key = f"{sheet_name}_{tarih_str}_{tutar_abs}"

                    # ── Mükerrer Kontrol (3 kademe) ───────────────────────────
                    exists = None
                    try:
                        exists = conn.execute(
                            "SELECT id FROM womsis_banka WHERE womsiskey=? AND userid=? LIMIT 1",
                            (womsis_key, self._userid)
                        ).fetchone()
                    except Exception:
                        exists = None

                    if not exists and dekont_no and dekont_no not in ("-", ""):
                        try:
                            exists = conn.execute(
                                """SELECT id FROM womsis_banka
                                   WHERE userid=? AND tarih=?
                                     AND (womsiskey=? OR aciklama LIKE ?)
                                     AND ABS(tutar - ?) < 0.01
                                   LIMIT 1""",
                                (self._userid, tarih_str,
                                 dekont_no, f"%{dekont_no}%",
                                 tutar_abs)
                            ).fetchone()
                        except Exception:
                            exists = None

                    if exists:
                        skip += 1; g_atlanan += 1
                        if (ins + skip + err) % 50 == 0:
                            self.progress.emit(
                                f"⏳ {sheet_name}  —  "
                                f"✅ {ins} eklendi   "
                                f"⏭️ {skip} atlandı   "
                                f"❌ {err} hata"
                            )
                        continue

                    # ── Kayıt Ekle ────────────────────────────────────────────
                    try:
                        conn.execute(
                            """INSERT INTO womsis_banka
                               (tarih, aciklama, gelirgider, tutar,
                                kaynak, womsiskey, userid, sube, faturaunvan,
                                bakiye, iban, hesap_turu, dekont_no)
                               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                            (tarih_str, aciklama, gelir_gider, tutar_abs,
                             "vomsis_excel", womsis_key, self._userid,
                             banka_sube, karsi_unvan,
                             bakiye_float, iban_val, hesap_turu, dekont_no)
                        )
                        ins += 1; g_eklenen += 1
                    except Exception as e:
                        print("Womsis DB Hata:", e)
                        err += 1; g_hata += 1

                    if (ins + skip + err) % 50 == 0:
                        self.progress.emit(
                            f"⏳ {sheet_name}  —  "
                            f"✅ {ins} eklendi   "
                            f"⏭️ {skip} atlandı   "
                            f"❌ {err} hata"
                        )

                conn.commit()
                durum = "✅ Tamam" if err == 0 else f"⚠️ {err} hata"
                self.satir_done.emit(sheet_name, ins, skip, durum)

        except Exception as e:
            conn.rollback()
            self.finished.emit({
                "success": False, "message": str(e),
                "toplam": g_toplam, "eklenen": g_eklenen,
                "atlanan": g_atlanan, "hata": g_hata,
            })
            return
        finally:
            conn.close()

        self.finished.emit({
            "success": True,
            "message": f"{g_eklenen:,} hareket eklendi, {g_atlanan:,} mükerrer atlandı.",
            "toplam": g_toplam, "eklenen": g_eklenen,
            "atlanan": g_atlanan, "hata": g_hata,
        })


# ─────────────────────────────────────────────────────────────────────────────
# ── VOMSİS Excel Import Dialog ───────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

class WomsisExcelImportDialog(QDialog):
    """
    VOMSİS banka hareketleri Excel dosyasını içe aktarır.
    • Beklenen format önizlemesi (şablon tablosu) gösterir
    • Şablon Excel dosyası indirir
    • FileDialog ile seçilen dosyayı parse edip hareketler tablosuna yazar
    """

    # Beklenen Excel kolon başlıkları (4. satırda olması beklenen)
    BEKLENEN_KOLONLAR = [
        "HAREKET ID", "HRKT KODU", "BANKA/ŞUBE", "HESAP ADI", "MHSB KODU",
        "HESAP NO", "IBAN", "HESAP TÜRÜ", "DEKONT NO", "TARİH",
        "İŞLEM SAATİ", "KARŞI TARAF ÜNVAN", "AÇIKLAMA", "TUTAR", "BAKİYE",
        "KARŞI TARAF VKN", "KARŞI TARAF IBAN", "NOT", "NOT OLUŞTURAN", "ETİKET",
    ]

    ORNEK_SATIRLAR = [
        ["13289", "INT", "Enpara / Enpara", "", "-", "94240595",
         "TR630015700000000094240595", "TL", "196000227989...", "30/06/2026",
         "23:17:11", "KARŞI FİRMA A.Ş.", "EFT Açıklaması", "1500.00", "100000",
         "-", "TR...", "-", "-", "-"],
        ["13290", "FST", "İş Bankası / Sultanhamam", "", "-", "1667134",
         "TR460006400000110481667134", "TL", "110481667134...", "29/06/2026",
         "14:30:00", "ÜMİT ÇOLAK", "FAST transferi", "-430.00", "19090.93",
         "-", "TR...", "-", "-", "-"],
    ]

    def __init__(self, userid: int, musterino: int = 1, parent=None):
        super().__init__(parent)
        self._userid = userid
        self._import_worker = None

        self.setWindowTitle("📂  VOMSİS Excel Import")
        self.setMinimumSize(1050, 680)
        self.setModal(True)
        self.setStyleSheet("""
            QDialog { background: #f8fafc; }
            QLabel  { background: transparent; font-family: 'Segoe UI', Arial, sans-serif; }
        """)
        self._build()

    # ── UI İnşa ──────────────────────────────────────────────────────────────

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(14)

        # ── Başlık ──────────────────────────────────────────────────────────
        hdr = QHBoxLayout()
        ic = QLabel("📂")
        ic.setStyleSheet("font-size:24px;")
        hdr.addWidget(ic)
        ttl = QLabel("VOMSİS Banka Hareketleri — Excel Import")
        ttl.setStyleSheet("font-size:16px;font-weight:700;color:#1e293b;")
        hdr.addWidget(ttl)
        hdr.addStretch()
        close_btn = QPushButton("✕")
        close_btn.setFixedSize(32, 32)
        close_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        close_btn.setStyleSheet(
            "QPushButton{background:#fee2e2;color:#dc2626;border:none;"
            "border-radius:8px;font-size:15px;font-weight:700;}"
            "QPushButton:hover{background:#fca5a5;}"
        )
        close_btn.clicked.connect(self.accept)
        hdr.addWidget(close_btn)
        root.addLayout(hdr)

        # ── Açıklama chip ────────────────────────────────────────────────────
        desc = QLabel(
            "ℹ️  VOMSİS sisteminden indirilen Excel dosyasını seçin. "
            "Her sayfa ayrı bir banka hesabı olarak işlenir. "
            "HAREKET ID ile mükerrer kayıt kontrolü yapılır."
        )
        desc.setWordWrap(True)
        desc.setStyleSheet(
            "font-size:12px;color:#475569;background:#eff6ff;"
            "border:1px solid #bfdbfe;border-radius:8px;padding:10px 14px;"
        )
        root.addWidget(desc)

        # ── Sekme başlıkları ─────────────────────────────────────────────────
        tab_row = QHBoxLayout()
        tab_row.setSpacing(0)

        self._tab_sablon = QPushButton("📋  Beklenen Format")
        self._tab_sablon.setCheckable(True)
        self._tab_sablon.setChecked(True)
        self._tab_sablon.setCursor(Qt.CursorShape.PointingHandCursor)
        self._tab_sablon.setFixedHeight(34)

        self._tab_sonuc = QPushButton("📊  Import Sonuçları")
        self._tab_sonuc.setCheckable(True)
        self._tab_sonuc.setChecked(False)
        self._tab_sonuc.setCursor(Qt.CursorShape.PointingHandCursor)
        self._tab_sonuc.setFixedHeight(34)

        for btn in (self._tab_sablon, self._tab_sonuc):
            btn.setStyleSheet("""
                QPushButton {
                    background: #e2e8f0; color: #475569; border: none;
                    border-radius: 0; font-size: 12px; font-weight: 600;
                    padding: 0 18px;
                }
                QPushButton:checked {
                    background: #1e40af; color: white;
                }
                QPushButton:hover:!checked { background: #cbd5e1; }
            """)
        self._tab_sablon.setStyleSheet(
            self._tab_sablon.styleSheet() + "QPushButton{border-radius: 8px 0 0 8px;}"
        )
        self._tab_sonuc.setStyleSheet(
            self._tab_sonuc.styleSheet() + "QPushButton{border-radius: 0 8px 8px 0;}"
        )
        self._tab_sablon.clicked.connect(lambda: self._switch_tab(0))
        self._tab_sonuc.clicked.connect(lambda: self._switch_tab(1))
        tab_row.addWidget(self._tab_sablon)
        tab_row.addWidget(self._tab_sonuc)
        tab_row.addStretch()
        root.addLayout(tab_row)

        # ── Stack ────────────────────────────────────────────────────────────
        self._stack = QStackedWidget()
        root.addWidget(self._stack, 1)

        # Sayfa 0 — Şablon önizleme
        self._stack.addWidget(self._build_sablon_page())

        # Sayfa 1 — Import sonuçları
        self._stack.addWidget(self._build_sonuc_page())

        # ── Alt buton çubuğu ─────────────────────────────────────────────────
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color:#e2e8f0;")
        root.addWidget(sep)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)

        self._sablon_indir_btn = QPushButton("📥  Şablon Excel İndir")
        self._sablon_indir_btn.setFixedHeight(40)
        self._sablon_indir_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._sablon_indir_btn.setStyleSheet(self._btn_style("#0891b2"))
        self._sablon_indir_btn.clicked.connect(self._on_sablon_indir)
        btn_row.addWidget(self._sablon_indir_btn)

        btn_row.addStretch()

        self._dosya_sec_btn = QPushButton("📂  Dosya Seç & Import Et")
        self._dosya_sec_btn.setFixedHeight(40)
        self._dosya_sec_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._dosya_sec_btn.setStyleSheet(self._btn_style("#059669"))
        self._dosya_sec_btn.clicked.connect(self._on_dosya_sec)
        btn_row.addWidget(self._dosya_sec_btn)

        root.addLayout(btn_row)

        # ── İlerleme barı (gizli) ────────────────────────────────────────────
        self._progress = QProgressBar()
        self._progress.setRange(0, 0)   # marquee mod
        self._progress.setFixedHeight(6)
        self._progress.setTextVisible(False)
        self._progress.setStyleSheet(
            "QProgressBar{background:#e2e8f0;border:none;border-radius:3px;}"
            "QProgressBar::chunk{background:#059669;border-radius:3px;}"
        )
        self._progress.hide()
        root.addWidget(self._progress)

    def _build_sablon_page(self) -> QWidget:
        page = QWidget()
        page.setStyleSheet("background:transparent;")
        vlay = QVBoxLayout(page)
        vlay.setContentsMargins(0, 0, 0, 0)
        vlay.setSpacing(10)

        lbl = QLabel("📋  Beklenen Excel Yapısı (Her sayfa = bir banka hesabı):")
        lbl.setStyleSheet("font-size:12px;font-weight:700;color:#1e293b;")
        vlay.addWidget(lbl)

        # Satır 1-3: Meta bilgi
        meta_lbl = QLabel(
            "Satır 1: OLUŞTURMA TARİHİ  •  Satır 2: FİRMA  •  "
            "Satır 3: FİLTRENEN TARİH ARALIĞI  •  Satır 4: KOLON BAŞLIKLARI  •  "
            "Satır 5+: Hareket verileri"
        )
        meta_lbl.setWordWrap(True)
        meta_lbl.setStyleSheet(
            "font-size:11px;color:#64748b;background:#fef9c3;"
            "border:1px solid #fde68a;border-radius:6px;padding:8px 12px;"
        )
        vlay.addWidget(meta_lbl)

        # Şablon tablosu
        tbl = QTableWidget(len(self.ORNEK_SATIRLAR) + 1, len(self.BEKLENEN_KOLONLAR))
        tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        tbl.setHorizontalHeaderLabels(self.BEKLENEN_KOLONLAR)
        tbl.verticalHeader().setVisible(False)
        tbl.setAlternatingRowColors(True)
        tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        tbl.horizontalHeader().setStretchLastSection(True)
        tbl.setStyleSheet("""
            QTableWidget {
                background: white; alternate-background-color: #f0f9ff;
                gridline-color: #e2e8f0; font-size: 11px;
                border: 1.5px solid #bfdbfe; border-radius: 8px;
            }
            QTableWidget::item { padding: 4px 8px; color: #1e293b; }
            QTableWidget::item:selected { background: #dbeafe; color: #1e40af; }
            QHeaderView::section {
                background: #1e40af; color: white;
                font-size: 10px; font-weight: 700;
                padding: 6px 4px; border: none;
                border-bottom: 1px solid #1d4ed8;
                border-right: 1px solid #1d4ed8;
            }
        """)

        # Zorunlu kolon açıklamaları (ilk satır)
        zorunlu = {
            "HAREKET ID": "✅ Zorunlu (mükerrer kontrol)",
            "TARİH": "✅ Zorunlu (DD/MM/YYYY)",
            "AÇIKLAMA": "✅ Zorunlu",
            "TUTAR": "✅ Zorunlu (+gelir / -gider)",
            "BANKA/ŞUBE": "💡 Önerilen",
            "HESAP NO": "💡 Önerilen",
            "IBAN": "💡 Önerilen",
            "KARŞI TARAF ÜNVAN": "💡 Önerilen",
        }
        for c_idx, col in enumerate(self.BEKLENEN_KOLONLAR):
            aciklama = zorunlu.get(col, "— İsteğe bağlı")
            item = QTableWidgetItem(aciklama)
            if "Zorunlu" in aciklama:
                item.setForeground(QBrush(QColor("#059669")))
                f = item.font(); f.setBold(True); item.setFont(f)
            elif "Önerilen" in aciklama:
                item.setForeground(QBrush(QColor("#d97706")))
            else:
                item.setForeground(QBrush(QColor("#94a3b8")))
            tbl.setItem(0, c_idx, item)

        # Örnek veri satırları
        for r_idx, row_data in enumerate(self.ORNEK_SATIRLAR):
            for c_idx, val in enumerate(row_data):
                item = QTableWidgetItem(str(val))
                # TUTAR kolonunu renklendir
                if c_idx == 13:
                    try:
                        t = float(val)
                        item.setForeground(QBrush(QColor("#059669" if t >= 0 else "#dc2626")))
                        f = item.font(); f.setBold(True); item.setFont(f)
                    except Exception:
                        pass
                tbl.setItem(r_idx + 1, c_idx, item)

        tbl.resizeRowsToContents()
        vlay.addWidget(tbl, 1)

        # Alan eşleştirme özeti
        eslestirme = QLabel(
            "🔗  Alan Eşleştirme:  "
            "HAREKET ID → womsisKey (mükerrer kontrol)  │  "
            "TARİH → tarih (YYYY-MM-DD'ye çevrilir)  │  "
            "AÇIKLAMA → aciklama  │  "
            "|TUTAR| → alinan_tutar1  │  "
            "TUTAR>0 → gelir, <0 → gider  │  "
            "BANKA/ŞUBE → sube alanı  │  "
            "KARŞI TARAF ÜNVAN → faturaUnvan"
        )
        eslestirme.setWordWrap(True)
        eslestirme.setStyleSheet(
            "font-size:11px;color:#1e40af;background:#eff6ff;"
            "border:1px solid #bfdbfe;border-radius:6px;padding:8px 12px;"
        )
        vlay.addWidget(eslestirme)

        return page

    def _build_sonuc_page(self) -> QWidget:
        page = QWidget()
        page.setStyleSheet("background:transparent;")
        vlay = QVBoxLayout(page)
        vlay.setContentsMargins(0, 0, 0, 0)
        vlay.setSpacing(10)

        # Durum etiketi
        self._sonuc_lbl = QLabel("Henüz import yapılmadı.")
        self._sonuc_lbl.setWordWrap(True)
        self._sonuc_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._sonuc_lbl.setStyleSheet(
            "font-size:13px;color:#64748b;background:#f1f5f9;"
            "border-radius:10px;padding:20px;"
        )
        vlay.addWidget(self._sonuc_lbl)

        # Log tablosu
        self._log_tbl = QTableWidget(0, 4)
        self._log_tbl.setHorizontalHeaderLabels(["Sayfa/Hesap", "Eklenen", "Atlanan", "Durum"])
        self._log_tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._log_tbl.verticalHeader().setVisible(False)
        self._log_tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self._log_tbl.setAlternatingRowColors(True)
        self._log_tbl.setStyleSheet("""
            QTableWidget {
                background: white; alternate-background-color: #f0fdf4;
                gridline-color: #e2e8f0; font-size: 12px;
                border: 1.5px solid #bbf7d0; border-radius: 8px;
            }
            QTableWidget::item { padding: 6px 10px; }
            QHeaderView::section {
                background: #059669; color: white;
                font-size: 11px; font-weight: 700; padding: 7px;
                border: none; border-right: 1px solid #047857;
            }
        """)
        vlay.addWidget(self._log_tbl, 1)

        # Özet çipler — değer label'ları direkt saklanıyor
        self._ozet_row = QHBoxLayout()
        self._ozet_row.setSpacing(10)

        chip_toplam,  self._lbl_toplam  = self._make_chip("📦 Toplam",   "0", "#1e40af", "#eff6ff")
        chip_eklenen, self._lbl_eklenen = self._make_chip("✅ Eklenen",  "0", "#059669", "#f0fdf4")
        chip_atlanan, self._lbl_atlanan = self._make_chip("⏭️ Atlanan",  "0", "#d97706", "#fffbeb")
        chip_hata,    self._lbl_hata    = self._make_chip("❌ Hata",      "0", "#dc2626", "#fef2f2")

        for chip in (chip_toplam, chip_eklenen, chip_atlanan, chip_hata):
            self._ozet_row.addWidget(chip)
        self._ozet_row.addStretch()
        vlay.addLayout(self._ozet_row)

        return page

    def _make_chip(self, label: str, value: str, color: str, bg: str):
        """(QFrame, val_QLabel) döndürür — label direkt saklanarak güvenilir güncelleme."""
        frame = QFrame()
        frame.setStyleSheet(
            f"QFrame{{background:{bg};border:1.5px solid {color}40;"
            f"border-radius:10px;padding:2px;}}"
        )
        lay = QVBoxLayout(frame)
        lay.setContentsMargins(14, 8, 14, 8)
        lay.setSpacing(2)
        lbl = QLabel(label)
        lbl.setStyleSheet(f"font-size:10px;font-weight:600;color:{color};")
        val = QLabel(value)
        val.setStyleSheet(f"font-size:22px;font-weight:800;color:{color};")
        val.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(lbl)
        lay.addWidget(val)
        frame.setFixedWidth(140)
        return frame, val  # frame ve değer label'ı birlikte döndür

    def _set_chips(self, toplam: str, eklenen: str, atlanan: str, hata: str):
        """4 chip değerini tek çağrıyla günceller."""
        self._lbl_toplam.setText(toplam)
        self._lbl_eklenen.setText(eklenen)
        self._lbl_atlanan.setText(atlanan)
        self._lbl_hata.setText(hata)

    def _btn_style(self, color: str) -> str:
        return (
            f"QPushButton{{background:{color};color:white;border:none;"
            f"border-radius:9px;font-size:13px;font-weight:600;"
            f"padding:0 18px;letter-spacing:.4px;}}"
            f"QPushButton:hover{{background:{color}cc;}}"
            f"QPushButton:disabled{{background:#cbd5e1;color:#94a3b8;}}"
        )

    def _switch_tab(self, idx: int):
        self._tab_sablon.setChecked(idx == 0)
        self._tab_sonuc.setChecked(idx == 1)
        self._stack.setCurrentIndex(idx)

    # ── Şablon Excel İndir ────────────────────────────────────────────────────

    def _on_sablon_indir(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.warning(self, "Hata", "openpyxl kütüphanesi yüklü değil.\npip install openpyxl")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Şablon Excel Kaydet",
            "vomsis_banka_hareketi_sablon.xlsx",
            "Excel Dosyaları (*.xlsx)"
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BANKA-HESAP_NO"

            # Meta satırlar
            ws.append(["OLUŞTURMA TARİHİ", "07/07/2026 12:00:00"])
            ws.append(["FİRMA", "FİRMA ADI"])
            ws.append(["FİLTRENEN TARİH ARALIĞI", "01/01/2026-30/06/2026"])

            # Kolon başlıkları
            ws.append(self.BEKLENEN_KOLONLAR)

            # Başlık satırı stili
            hdr_fill  = PatternFill("solid", fgColor="1E40AF")
            hdr_font  = Font(bold=True, color="FFFFFF", size=10)
            hdr_align = Alignment(horizontal="center", vertical="center")
            thin = Side(style="thin", color="1D4ED8")
            hdr_border = Border(left=thin, right=thin, bottom=thin)
            for cell in ws[4]:
                cell.fill   = hdr_fill
                cell.font   = hdr_font
                cell.alignment = hdr_align
                cell.border = hdr_border

            # Örnek veri satırları
            for row_data in self.ORNEK_SATIRLAR:
                ws.append(row_data)

            # Sütun genişlikleri
            col_widths = [12, 8, 30, 15, 10, 14, 32, 8, 28, 12,
                          10, 30, 40, 14, 14, 14, 32, 14, 14, 14]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            # Satır yüksekliği (başlık)
            ws.row_dimensions[4].height = 20

            # Not sayfası
            ws2 = wb.create_sheet("AÇIKLAMALAR")
            notlar = [
                ("Alan", "Açıklama", "Zorunlu?"),
                ("HAREKET ID", "Her hareket için benzersiz ID (mükerrer önleme)", "EVET"),
                ("TARİH", "DD/MM/YYYY formatında tarih", "EVET"),
                ("AÇIKLAMA", "İşlem açıklaması", "EVET"),
                ("TUTAR", "Pozitif = Gelir, Negatif = Gider", "EVET"),
                ("BANKA/ŞUBE", "Banka adı ve şube bilgisi", "ÖNERİLEN"),
                ("HESAP NO", "Banka hesap numarası", "ÖNERİLEN"),
                ("IBAN", "IBAN numarası", "ÖNERİLEN"),
                ("KARŞI TARAF ÜNVAN", "Karşı taraf şirket/kişi adı", "ÖNERİLEN"),
                ("KARŞI TARAF VKN", "Vergi/TC numarası", "İSTEĞE BAĞLI"),
                ("BAKİYE", "İşlem sonrası hesap bakiyesi", "İSTEĞE BAĞLI"),
                ("ETİKET", "İşlem etiketi", "İSTEĞE BAĞLI"),
            ]
            ws2.append([])
            for row in notlar:
                ws2.append(list(row))
            # Not başlıkları
            for cell in ws2[2]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="059669")

            wb.save(path)
            QMessageBox.information(self, "✅ Şablon İndirildi",
                                    f"Şablon başarıyla kaydedildi:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Şablon oluşturulamadı:\n{e}")

    # ── Dosya Seç & Import ────────────────────────────────────────────────────

    def _on_dosya_sec(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "VOMSİS Excel Dosyası Seç",
            "", "Excel Dosyaları (*.xlsx *.xls)"
        )
        if not path:
            return
        self._start_import(path)

    def _start_import(self, path: str):
        self._dosya_sec_btn.setEnabled(False)
        self._sablon_indir_btn.setEnabled(False)
        self._progress.show()
        self._switch_tab(1)
        self._sonuc_lbl.setText(f"⏳  Import ediliyor...\n{path}")
        self._sonuc_lbl.setStyleSheet(
            "font-size:12px;color:#1e40af;"
            "background:#eff6ff;border:1px solid #bfdbfe;"
            "border-radius:8px;padding:12px;"
        )
        self._log_tbl.setRowCount(0)
        self._set_chips("0", "0", "0", "0")

        self._worker = WomsisExcelWorker(path, self._userid)
        self._worker.satir_done.connect(self._on_satir_done)
        self._worker.header_error.connect(self._on_header_error)
        self._worker.progress.connect(self._on_progress)
        self._worker.finished.connect(self._on_import_done)
        self._worker.start()


    def _on_progress(self, msg: str):
        """Worker'dan gelen canlı durum mesajını gösterir."""
        self._sonuc_lbl.setText(msg)
        self._sonuc_lbl.setStyleSheet(
            "font-size:12px;color:#1e40af;"
            "background:#eff6ff;border:1px solid #bfdbfe;"
            "border-radius:8px;padding:12px;"
        )

    def _on_header_error(self, sayfa: str, bulunan: object, eksik: object):
        """
        Worker başlık hatası bildirdiğinde çağrılır.
        Hata detayını gösterir ve boş şablon Excel'i indirmeyi teklif eder.
        """
        if not eksik:
            return

        # Hata mesajı oluştur
        if not bulunan:
            baslik = f"'{sayfa}' sayfasında başlık satırı bulunamadı!"
            detay  = (
                "Excel dosyasında 'HAREKET ID' başlıklı kolon satırı tespit edilemedi.\n"
                "Dosya VOMSİS formatında olmayabilir."
            )
        else:
            baslik = f"'{sayfa}' sayfasında zorunlu kolonlar eksik!"
            detay  = (
                f"Eksik kolonlar: {', '.join(eksik)}\n\n"
                f"Dosyada bulunan kolonlar:\n"
                + "  ".join(f"• {k}" for k in bulunan[:10])
                + (f"  ... (+{len(bulunan)-10} daha)" if len(bulunan) > 10 else "")
            )

        # Mesaj kutusu — şablon indir seçeneğiyle
        msg = QMessageBox(self)
        msg.setWindowTitle("⚠️  Başlık Hatası")
        msg.setIcon(QMessageBox.Icon.Warning)
        msg.setText(f"<b>{baslik}</b>")
        msg.setInformativeText(
            detay + "\n\nDoğru formatı görmek için şablon Excel'i indirebilirsiniz."
        )
        msg.setStyleSheet(
            "QMessageBox { background: #fff; font-family: 'Segoe UI'; }"
            "QLabel { color: #1e293b; font-size: 13px; }"
            "QPushButton { background: #1e40af; color: white; border-radius: 6px;"
            "              padding: 6px 16px; font-weight: 600; }"
            "QPushButton:hover { background: #1d4ed8; }"
        )
        sablon_btn = msg.addButton("📥  Şablon Excel İndir", QMessageBox.ButtonRole.ActionRole)
        msg.addButton("Kapat", QMessageBox.ButtonRole.RejectRole)
        msg.exec()

        if msg.clickedButton() == sablon_btn:
            self._on_sablon_indir()

    def _on_satir_done(self, sayfa: str, eklenen: int, atlanan: int, durum: str):
        row = self._log_tbl.rowCount()
        self._log_tbl.insertRow(row)
        items = [sayfa, str(eklenen), str(atlanan), durum]
        colors = [None, "#059669", "#d97706", None]
        for c, (txt, clr) in enumerate(zip(items, colors)):
            item = QTableWidgetItem(txt)
            if clr:
                item.setForeground(QBrush(QColor(clr)))
                f = item.font(); f.setBold(True); item.setFont(f)
            if "Tamam" in txt:
                item.setForeground(QBrush(QColor("#059669")))
            elif "hata" in txt.lower() or "⚠️" in txt:
                item.setForeground(QBrush(QColor("#dc2626")))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self._log_tbl.setItem(row, c, item)
        self._log_tbl.scrollToBottom()

    def _on_import_done(self, result: dict):
        self._progress.hide()
        self._dosya_sec_btn.setEnabled(True)
        self._sablon_indir_btn.setEnabled(True)

        self._set_chips(
            f"{result.get('toplam',  0):,}",
            f"{result.get('eklenen', 0):,}",
            f"{result.get('atlanan', 0):,}",
            f"{result.get('hata',    0):,}",
        )

        if result["success"]:
            self._sonuc_lbl.setText(f"✅  {result['message']}")
            self._sonuc_lbl.setStyleSheet(
                "font-size:13px;font-weight:700;color:#059669;"
                "background:#f0fdf4;border:1.5px solid #bbf7d0;"
                "border-radius:10px;padding:16px;"
            )
        else:
            self._sonuc_lbl.setText(f"❌  {result['message']}")
            self._sonuc_lbl.setStyleSheet(
                "font-size:13px;font-weight:700;color:#dc2626;"
                "background:#fef2f2;border:1.5px solid #fca5a5;"
                "border-radius:10px;padding:16px;"
            )


# ─────────────────────────────────────────────────────────────────────────────
# ── webadmin Womsis Çekme İlerleme Dialog'u ────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

class WebAdminSyncDialog(QDialog):
    """
    'webadmin Çek' butonu için tam ilerleme paneli.
    Aşamaları:
      1. Onay göster (tarih aralığı + başlat butonu)
      2. Worker başlat — canlı log satırları + marquee progress bar
      3. Sonuç özeti göster (toplam çekilen / eklenen / atlanan)
    """

    def __init__(self, userid: int, start_str: str, end_str: str, parent=None):
        super().__init__(parent)
        self._userid    = userid
        self._start_str = start_str
        self._end_str   = end_str
        self._worker    = None
        self._log_lines: list[str] = []

        self.setWindowTitle("webadmin Üzerinden Çek")
        self.setMinimumWidth(480)
        self.setMinimumHeight(300)
        self.setModal(True)
        self.setStyleSheet(
            "QDialog{background:#ffffff;border-radius:12px;}"
            "QLabel{font-family:'Segoe UI',Arial,sans-serif;background:transparent;}"
        )
        self._build()

        # Dialog ekrana gelir gelmez otomatik olarak veri çekmeye başla
        from PyQt6.QtCore import QTimer
        QTimer.singleShot(150, self._start)

    # ── UI İnşa ────────────────────────────────────────────────────

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 22)
        root.setSpacing(14)

        # ── Başlık satırı ──
        hdr = QHBoxLayout()
        ic = QLabel("🌐")
        ic.setStyleSheet("font-size:22px;")
        hdr.addWidget(ic)
        ttl = QLabel("webadmin üzerinden Womsis Çek")
        ttl.setStyleSheet("font-size:15px;font-weight:700;color:#1e293b;")
        hdr.addWidget(ttl)
        hdr.addStretch()
        root.addLayout(hdr)

        # ── Tarih aralığı bilgi çipi ──
        period_lbl = QLabel(
            f"📅  {self._start_str}  —  {self._end_str} aralığı"
        )
        period_lbl.setStyleSheet(
            "font-size:12px;color:#475569;background:#f1f5f9;"
            "border-radius:6px;padding:7px 12px;"
        )
        root.addWidget(period_lbl)

        # ── Log alanı ──
        self._log_widget = QLabel("🌐  Bağlanılıyor...")
        self._log_widget.setWordWrap(True)
        self._log_widget.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
        self._log_widget.setStyleSheet(
            "font-size:12px;color:#374151;"
            "background:#f8fafc;border:1px solid #e2e8f0;"
            "border-radius:8px;padding:10px 12px;"
            "line-height:1.6;"
        )
        self._log_widget.setMinimumHeight(110)
        root.addWidget(self._log_widget)

        # ── Progress bar (marquee) ──
        self._pbar = QProgressBar()
        self._pbar.setFixedHeight(6)
        self._pbar.setTextVisible(False)
        self._pbar.setRange(0, 0)   # marquee / belirsiz mod
        self._pbar.setStyleSheet(
            "QProgressBar{background:#e2e8f0;border-radius:3px;border:none;}"
            "QProgressBar::chunk{background:#6366f1;border-radius:3px;}"
        )
        self._pbar.hide()
        root.addWidget(self._pbar)

        # ── Sonuç kartı (başlangıçta gizli) ──
        self._result_frame = QFrame()
        self._result_frame.setStyleSheet(
            "QFrame{background:#f0fdf4;border:1px solid #86efac;"
            "border-radius:8px;padding:2px;}"
        )
        result_lay = QHBoxLayout(self._result_frame)
        result_lay.setContentsMargins(14, 10, 14, 10)
        result_lay.setSpacing(18)

        def _stat(label: str, val: str, color: str) -> QVBoxLayout:
            col = QVBoxLayout()
            col.setSpacing(2)
            v = QLabel(val)
            v.setObjectName(f"stat_{label}")
            v.setStyleSheet(
                f"font-size:22px;font-weight:800;color:{color};"
                "border:none;background:transparent;"
            )
            v.setAlignment(Qt.AlignmentFlag.AlignCenter)
            l = QLabel(label)
            l.setStyleSheet("font-size:10px;color:#64748b;border:none;background:transparent;")
            l.setAlignment(Qt.AlignmentFlag.AlignCenter)
            col.addWidget(v)
            col.addWidget(l)
            return col

        self._stat_cekilen  = QLabel("0")
        self._stat_eklenen  = QLabel("0")
        self._stat_atlanan  = QLabel("0")

        def _stat_block(val_lbl: QLabel, label: str, color: str):
            col = QVBoxLayout()
            col.setSpacing(2)
            val_lbl.setStyleSheet(
                f"font-size:22px;font-weight:800;color:{color};"
                "border:none;background:transparent;"
            )
            val_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl = QLabel(label)
            lbl.setStyleSheet("font-size:10px;color:#64748b;border:none;background:transparent;")
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            col.addWidget(val_lbl)
            col.addWidget(lbl)
            return col

        result_lay.addLayout(_stat_block(self._stat_cekilen, "Çekilen", "#2563eb"))
        result_lay.addLayout(_stat_block(self._stat_eklenen, "Eklenen", "#059669"))
        result_lay.addLayout(_stat_block(self._stat_atlanan, "Atlanan", "#d97706"))
        self._result_frame.hide()
        root.addWidget(self._result_frame)

        # ── Alt butonlar ──
        foot = QHBoxLayout()
        foot.setSpacing(10)

        self._btn_start = QPushButton("🔄  Tekrar Dene")
        self._btn_start.setFixedHeight(40)
        self._btn_start.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_start.setStyleSheet(
            "QPushButton{background:#6366f1;color:white;border:none;"
            "border-radius:9px;font-size:13px;font-weight:700;padding:0 20px;}"
            "QPushButton:hover{background:#4f46e5;}"
            "QPushButton:disabled{background:#e2e8f0;color:#94a3b8;}"
        )
        self._btn_start.clicked.connect(self._start)
        self._btn_start.hide()   # başlangıçta gizli — sadece sonra gösterilir
        foot.addWidget(self._btn_start)

        self._btn_close = QPushButton("Kapat")
        self._btn_close.setFixedHeight(40)
        self._btn_close.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_close.setStyleSheet(
            "QPushButton{background:#f1f5f9;color:#475569;border:1.5px solid #e2e8f0;"
            "border-radius:9px;font-size:13px;font-weight:600;padding:0 20px;}"
            "QPushButton:hover{background:#e2e8f0;}"
        )
        self._btn_close.clicked.connect(self.accept)
        foot.addWidget(self._btn_close)
        root.addLayout(foot)

    # ── İş Mantığı ────────────────────────────────────────────────

    def _append_log(self, msg: str):
        self._log_lines.append(msg)
        # Son 6 satırı göster
        self._log_widget.setText("\n".join(self._log_lines[-6:]))

    def _start(self):
        self._btn_start.hide()     # işlem başladığında butonu gizle
        self._btn_close.setEnabled(False)
        self._pbar.show()
        self._result_frame.hide()
        self._log_lines = []
        self._append_log("🌐  webadmin sunucusuna bağlanılıyor...")

        from services.webadmin_client import WebAdminSyncWorker
        self._worker = WebAdminSyncWorker(
            userid=self._userid,
            start_date=self._start_str,
            end_date=self._end_str,
        )
        self._worker.progress.connect(self._append_log)
        self._worker.finished.connect(self._on_done)
        self._worker.start()

    def _on_done(self, r: dict):
        self._pbar.hide()
        self._btn_close.setEnabled(True)

        if r.get("success"):
            # ── Başarılı ──
            inserted = r.get("inserted", 0)
            skipped  = r.get("skipped",  0)
            fetched  = r.get("count",    0)

            self._append_log(f"✅  Tamamlandı! {inserted} yeni kayıt eklendi, {skipped} atlandı.")

            # İstatistik kartları güncelle
            self._stat_cekilen.setText(str(fetched))
            self._stat_eklenen.setText(str(inserted))
            self._stat_atlanan.setText(str(skipped))
            self._result_frame.setStyleSheet(
                "QFrame{background:#f0fdf4;border:1px solid #86efac;"
                "border-radius:8px;padding:2px;}"
            )
            self._result_frame.show()

            # Başarı butonu
            self._btn_start.setText("✅  Tamamlandı")
            self._btn_start.setStyleSheet(
                "QPushButton{background:#059669;color:white;border:none;"
                "border-radius:9px;font-size:13px;font-weight:700;padding:0 20px;}"
            )
            self._btn_start.setEnabled(False)   # sadece görselliği için
            self._btn_start.show()

        else:
            # ── Hata ──
            err = r.get("error") or r.get("message") or "Bilinmeyen hata"
            self._append_log(f"❌  Hata: {err}")
            self._result_frame.setStyleSheet(
                "QFrame{background:#fef2f2;border:1px solid #fca5a5;"
                "border-radius:8px;padding:2px;}"
            )
            self._stat_cekilen.setText("❌")
            self._stat_eklenen.setText("0")
            self._stat_atlanan.setText("0")
            self._result_frame.show()

            self._btn_start.setText("🔄  Tekrar Dene")
            self._btn_start.setStyleSheet(
                "QPushButton{background:#dc2626;color:white;border:none;"
                "border-radius:9px;font-size:13px;font-weight:700;padding:0 20px;}"
                "QPushButton:hover{background:#b91c1c;}"
            )


class SweetConfirmDialog(QDialog):
    """Minimal onay diyaloğu — SweetAlert tarzı, frameless."""

    def __init__(self, parent=None, title="", text="",
                 confirm_text="Evet", cancel_text="İptal", is_danger=False):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.Dialog | Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self._build(title, text, confirm_text, cancel_text, is_danger)

    def _build(self, title, text, confirm_text, cancel_text, is_danger):
        from PyQt6.QtWidgets import QGraphicsDropShadowEffect
        from PyQt6.QtGui import QColor
        master = QVBoxLayout(self)
        master.setContentsMargins(12, 12, 12, 12)
        box = QFrame()
        box.setObjectName("scd_box")
        box.setStyleSheet(
            "QFrame#scd_box{background:#ffffff;border:1.5px solid #e2e8f0;border-radius:16px;}"
        )
        sh = QGraphicsDropShadowEffect(self)
        sh.setBlurRadius(20)
        sh.setColor(QColor(0, 0, 0, 40))
        sh.setOffset(0, 4)
        box.setGraphicsEffect(sh)
        master.addWidget(box)
        lay = QVBoxLayout(box)
        lay.setContentsMargins(24, 24, 24, 20)
        lay.setSpacing(12)
        t = QLabel(title)
        t.setAlignment(Qt.AlignmentFlag.AlignCenter)
        t.setStyleSheet(
            "font-size:15px;font-weight:700;color:#1e293b;"
            "background:transparent;border:none;"
        )
        lay.addWidget(t)
        d = QLabel(text)
        d.setWordWrap(True)
        d.setAlignment(Qt.AlignmentFlag.AlignCenter)
        d.setStyleSheet(
            "font-size:12px;color:#475569;background:transparent;border:none;"
        )
        lay.addWidget(d)
        lay.addSpacing(4)
        brow = QHBoxLayout()
        brow.setSpacing(8)
        if cancel_text:
            cb = QPushButton(cancel_text)
            cb.setFixedHeight(36)
            cb.setCursor(Qt.CursorShape.PointingHandCursor)
            cb.setStyleSheet(
                "QPushButton{background:#f1f5f9;color:#475569;border:none;"
                "border-radius:8px;font-size:12px;font-weight:600;padding:0 16px;}"
                "QPushButton:hover{background:#e2e8f0;}"
            )
            cb.clicked.connect(self.reject)
            brow.addWidget(cb)
        ok_bg = "#ef4444" if is_danger else "#6366f1"
        ok_hv = "#dc2626" if is_danger else "#4f46e5"
        ob = QPushButton(confirm_text)
        ob.setFixedHeight(36)
        ob.setCursor(Qt.CursorShape.PointingHandCursor)
        ob.setStyleSheet(
            f"QPushButton{{background:{ok_bg};color:white;border:none;"
            f"border-radius:8px;font-size:12px;font-weight:700;padding:0 16px;}}"
            f"QPushButton:hover{{background:{ok_hv};}}"
        )
        ob.clicked.connect(self.accept)
        brow.addWidget(ob)
        lay.addLayout(brow)

    @classmethod
    def confirm(cls, parent, title, text, confirm_text="Evet",
                cancel_text="İptal", is_danger=False):
        dlg = cls(parent, title, text, confirm_text, cancel_text, is_danger)
        return dlg.exec() == QDialog.DialogCode.Accepted


# ── PostgreSQL bağlantı test worker ──────────────────────────────────────────

class VeriTabaniTestWorker(QThread):
    islendi = pyqtSignal(dict)

    def __init__(self, host, port, dbname, user, password, sslmode):
        super().__init__()
        self._host = host
        self._port = port
        self._db = dbname
        self._user = user
        self._pw = password
        self._ssl = sslmode

    def run(self):
        try:
            import psycopg2
        except ImportError:
            self.islendi.emit({
                "success": False,
                "message": "psycopg2 yüklü değil.\npip install psycopg2-binary",
                "ver": ""
            })
            return
        params = dict(
            host=self._host, port=self._port, dbname=self._db,
            user=self._user, sslmode=self._ssl, connect_timeout=8
        )
        if self._pw:
            params["password"] = self._pw

        # Neon SNI Hatası Çözümü
        if "neon.tech" in self._host:
            endpoint_id = self._host.split(".")[0]
            params["options"] = f"project={endpoint_id}"
        try:
            conn = psycopg2.connect(**params)
            v = conn.server_version
            conn.close()
            self.islendi.emit({
                "success": True,
                "message": "Bağlantı başarılı!",
                "ver": f"PostgreSQL {v//10000}.{(v % 10000)//100}"
            })
        except Exception as exc:
            self.islendi.emit({"success": False, "message": str(exc), "ver": ""})


# ── SQLite → PostgreSQL migrasyon worker ─────────────────────────────────────

class MigrasyonWorker(QThread):
    ilerleme = pyqtSignal(str, int, int, bool)  # tablo, aktarilan, toplam, ara
    islendi  = pyqtSignal(bool, str)

    def __init__(self, batch_size: int = 250):
        super().__init__()
        self._batch_size = batch_size

    def run(self):
        from db.sqlite_to_pg import migrate_all
        toplam_aktarilan = 0
        hatalar: list[str] = []
        try:
            for s in migrate_all(batch_size=self._batch_size):
                if s.bitti:
                    break
                if s.hata:
                    hatalar.append(f"{s.tablo}: {s.hata}")
                if not s.ara:
                    toplam_aktarilan += s.aktarilan
                self.ilerleme.emit(s.tablo, s.aktarilan, s.toplam, s.ara)
            if hatalar:
                ozet = (
                    f"Tamamlandı — {toplam_aktarilan:,} kayıt aktarıldı.\n"
                    "Hatalar:\n" + "\n".join(hatalar[:5])
                )
                self.islendi.emit(False, ozet)
            else:
                self.islendi.emit(
                    True,
                    f"✅  Tüm veriler aktarıldı!\n{toplam_aktarilan:,} kayıt PostgreSQL'e taşındı."
                )
        except Exception as exc:
            self.islendi.emit(False, f"Hata: {exc}")


# ── Migrasyon ilerleme diyaloğu ───────────────────────────────────────────────

class MigrasyonDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.Dialog | Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setMinimumWidth(640)
        self.setModal(True)
        self._worker: MigrasyonWorker | None = None
        self._tablo_sayisi = 0
        self._satir_map: dict[str, bool] = {}
        self._satir_lbl_map: dict[str, tuple] = {}
        self._toplam_aktarilan = 0
        self._build()

    def _build(self):
        from PyQt6.QtWidgets import QGraphicsDropShadowEffect, QScrollArea
        from PyQt6.QtGui import QColor

        master = QVBoxLayout(self)
        master.setContentsMargins(16, 16, 16, 16)

        box = QFrame()
        box.setObjectName("mgr_box")
        box.setStyleSheet(
            "QFrame#mgr_box{background:#ffffff;border:1.5px solid #e2e8f0;border-radius:18px;}"
        )
        sh = QGraphicsDropShadowEffect(self)
        sh.setBlurRadius(28)
        sh.setColor(QColor(0, 0, 0, 45))
        sh.setOffset(0, 6)
        box.setGraphicsEffect(sh)
        master.addWidget(box)

        root = QVBoxLayout(box)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Mor gradient başlık
        hdr = QFrame()
        hdr.setFixedHeight(64)
        hdr.setStyleSheet(
            "QFrame{"
            "  background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #6366f1,stop:1 #8b5cf6);"
            "  border-top-left-radius:17px;border-top-right-radius:17px;"
            "}"
        )
        hl = QHBoxLayout(hdr)
        hl.setContentsMargins(22, 0, 22, 0)
        hl.addWidget(_mk_lbl("🚚", "font-size:24px;background:transparent;border:none;"))
        hl.addWidget(_mk_lbl(
            "Veri Taşıma  —  SQLite → PostgreSQL",
            "font-size:15px;font-weight:700;color:#ffffff;"
            "background:transparent;border:none;letter-spacing:.4px;"
        ))
        hl.addStretch()
        root.addWidget(hdr)

        # Gövde
        body = QFrame()
        body.setStyleSheet("QFrame{background:transparent;border:none;}")
        bl = QVBoxLayout(body)
        bl.setContentsMargins(22, 18, 22, 18)
        bl.setSpacing(14)

        bl.addWidget(_mk_lbl(
            "SQLite veritabanındaki tüm kayıtlar PostgreSQL'e aktarılıyor.\n"
            "İnternet bağlantı hızınıza göre birkaç dakika sürebilir.",
            "font-size:12px;color:#000000;background:transparent;border:none;"
        ))

        sr = QHBoxLayout()
        self._durum_lbl = _mk_lbl(
            "⏳  Başlatılıyor...",
            "font-size:12px;font-weight:700;color:#6366f1;"
            "background:transparent;border:none;"
        )
        sr.addWidget(self._durum_lbl)
        sr.addStretch()
        self._sayac_lbl = _mk_lbl(
            "0 kayıt aktarıldı",
            "font-size:12px;font-weight:600;color:#000000;"
            "background:transparent;border:none;"
        )
        sr.addWidget(self._sayac_lbl)
        bl.addLayout(sr)

        self._prog = QProgressBar()
        self._prog.setRange(0, 1000)
        self._prog.setValue(0)
        self._prog.setFixedHeight(10)
        self._prog.setTextVisible(False)
        self._prog.setStyleSheet(
            "QProgressBar{background:#e2e8f0;border-radius:5px;border:none;}"
            "QProgressBar::chunk{"
            "  background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #6366f1,stop:1 #8b5cf6);"
            "  border-radius:5px;"
            "}"
        )
        bl.addWidget(self._prog)

        # Tablo başlık şeridi
        hf = QFrame()
        hf.setFixedHeight(30)
        hf.setStyleSheet("QFrame{background:#f1f5f9;border-radius:8px 8px 0 0;border:none;}")
        hfl = QHBoxLayout(hf)
        hfl.setContentsMargins(10, 0, 10, 0)
        for col_txt, col_stretch in [("Tablo", 3), ("Aktarılan", 2), ("Toplam", 2), ("Durum", 1)]:
            hfl.addWidget(_mk_lbl(
                col_txt,
                "font-size:11px;font-weight:700;color:#000000;background:transparent;border:none;"
            ), col_stretch)
        bl.addWidget(hf)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setStyleSheet(
            "QScrollArea{background:white;border:1px solid #e2e8f0;"
            "border-top:none;border-radius:0 0 8px 8px;}"
        )
        scroll.setMinimumHeight(240)
        scroll.setMaximumHeight(320)
        self._tablo_widget = QWidget()
        self._tablo_widget.setStyleSheet("background:white;")
        self._tablo_lay = QVBoxLayout(self._tablo_widget)
        self._tablo_lay.setContentsMargins(0, 0, 0, 0)
        self._tablo_lay.setSpacing(0)
        self._tablo_lay.addStretch()
        scroll.setWidget(self._tablo_widget)
        bl.addWidget(scroll)
        self._scroll = scroll

        self._sonuc_frame = QFrame()
        self._sonuc_frame.setStyleSheet(
            "QFrame{background:#f0fdf4;border:1.5px solid #bbf7d0;border-radius:10px;}"
        )
        sfl = QHBoxLayout(self._sonuc_frame)
        sfl.setContentsMargins(14, 10, 14, 10)
        self._sonuc_ic = _mk_lbl("✅", "font-size:20px;background:transparent;border:none;")
        sfl.addWidget(self._sonuc_ic)
        self._sonuc_lbl = QLabel("")
        self._sonuc_lbl.setWordWrap(True)
        self._sonuc_lbl.setStyleSheet(
            "font-size:12px;color:#000000;background:transparent;border:none;"
        )
        sfl.addWidget(self._sonuc_lbl, 1)
        self._sonuc_frame.hide()
        bl.addWidget(self._sonuc_frame)

        br = QHBoxLayout()
        br.addStretch()
        self._kapat_btn = QPushButton("  Kapat  ")
        self._kapat_btn.setEnabled(False)
        self._kapat_btn.setFixedHeight(38)
        self._kapat_btn.setMinimumWidth(120)
        self._kapat_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._kapat_btn.setStyleSheet(
            "QPushButton{background:#6366f1;color:#ffffff;border:none;"
            "border-radius:9px;font-size:13px;font-weight:700;}"
            "QPushButton:hover{background:#4f46e5;}"
            "QPushButton:disabled{background:#e2e8f0;color:#94a3b8;}"
        )
        self._kapat_btn.clicked.connect(self.accept)
        br.addWidget(self._kapat_btn)
        bl.addLayout(br)
        root.addWidget(body)

    def _add_row(self, tablo):
        idx = len(self._satir_map)
        satir = QFrame()
        bg = "#f8fafc" if idx % 2 == 0 else "#ffffff"
        satir.setStyleSheet(
            f"QFrame{{background:{bg};border:none;border-bottom:1px solid #f1f5f9;}}"
        )
        satir.setFixedHeight(36)
        lay = QHBoxLayout(satir)
        lay.setContentsMargins(10, 0, 10, 0)
        lay.setSpacing(0)
        nm = _mk_lbl(tablo, "font-size:12px;color:#000000;background:transparent;border:none;")
        ak = _mk_lbl("0", "font-size:12px;font-weight:600;color:#000000;background:transparent;border:none;")
        tp = _mk_lbl("0", "font-size:12px;color:#000000;background:transparent;border:none;")
        du = _mk_lbl("⏳", "font-size:12px;color:#6366f1;background:transparent;border:none;")
        lay.addWidget(nm, 3)
        lay.addWidget(ak, 2)
        lay.addWidget(tp, 2)
        lay.addWidget(du, 1)
        self._tablo_lay.insertWidget(self._tablo_lay.count() - 1, satir)
        return ak, tp, du

    def start(self, batch_size: int = 250):
        self._worker = MigrasyonWorker(batch_size)
        self._worker.ilerleme.connect(self._on_ilerleme)
        self._worker.islendi.connect(self._on_islendi)
        self._worker.start()

    def _on_ilerleme(self, tablo, aktarilan, toplam, ara):
        self._durum_lbl.setText(f"⏳  {tablo} aktarılıyor...")
        if tablo not in self._satir_map:
            lbls = self._add_row(tablo)
            self._satir_map[tablo] = True
            self._satir_lbl_map[tablo] = lbls
        a_lbl, t_lbl, d_lbl = self._satir_lbl_map[tablo]
        a_lbl.setText(f"{aktarilan:,}")
        t_lbl.setText(f"{toplam:,}")
        if ara:
            d_lbl.setText("⏳")
            d_lbl.setStyleSheet(
                "font-size:12px;color:#6366f1;background:transparent;border:none;"
            )
        else:
            self._tablo_sayisi += 1
            if aktarilan >= toplam:
                d_lbl.setText("✅")
                d_lbl.setStyleSheet(
                    "font-size:12px;color:#059669;background:transparent;border:none;"
                )
                self._toplam_aktarilan += aktarilan
            else:
                d_lbl.setText("⚠️")
                d_lbl.setStyleSheet(
                    "font-size:12px;color:#d97706;background:transparent;border:none;"
                )
        self._sayac_lbl.setText(f"{self._toplam_aktarilan:,} kayıt aktarıldı")
        vb = self._scroll.verticalScrollBar()
        vb.setValue(vb.maximum())
        tp_val = int(self._tablo_sayisi * (960 / 24))
        ic_val = int(aktarilan / toplam * (960 / 24)) if toplam > 0 else 0
        self._prog.setValue(min(990, tp_val + ic_val))

    def _on_islendi(self, ok, msg):
        self._prog.setValue(1000)
        self._durum_lbl.setText("✅  Tamamlandı!" if ok else "❌  Hatalar oluştu")
        self._durum_lbl.setStyleSheet(
            f"font-size:12px;font-weight:700;"
            f"color:{'#059669' if ok else '#dc2626'};"
            "background:transparent;border:none;"
        )
        self._sonuc_lbl.setText(msg)
        if not ok:
            self._sonuc_ic.setText("❌")
            self._sonuc_frame.setStyleSheet(
                "QFrame{background:#fef2f2;border:1.5px solid #fecaca;border-radius:10px;}"
            )
        self._sonuc_frame.show()
        self._kapat_btn.setEnabled(True)


# ── Veritabanı Ayarları Kartı ─────────────────────────────────────────────────

class VeriTabaniCard(QFrame):
    """
    Ayarlar → 🗄️ Veritabanı sekmesi.
    Mod: SQLite (lokal) veya PostgreSQL (sunucu / Supabase).
    Migration: SQLite → PostgreSQL, paket boyutu seçilebilir.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self._test_worker: VeriTabaniTestWorker | None = None
        self._current_mod = "sqlite"
        self.setStyleSheet(
            "QFrame{background:white;border:1.5px solid #c7d2fe;border-radius:14px;}"
        )
        self._build()
        self._load()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 22)
        root.setSpacing(18)

        # ── Başlık ──────────────────────────────────────────────────────────
        h = QHBoxLayout()
        h.addWidget(_mk_lbl("🗄️", "font-size:22px;"))
        h.addWidget(_mk_lbl(
            "Veritabanı Bağlantısı",
            "font-size:14px;font-weight:700;color:#000000;letter-spacing:.5px;"
        ))
        h.addStretch()
        root.addLayout(h)

        root.addWidget(_mk_lbl(
            "💡  Lokal mod tek kullanıcı içindir. Birden fazla kişi aynı verileri "
            "kullanacaksa PostgreSQL sunucu modunu seçin. "
            "Mod değiştirdikten sonra uygulamayı yeniden başlatın.",
            "background:#eef2ff;color:#3730a3;border-radius:8px;"
            "padding:10px 14px;font-size:12px;border:none;"
        ))

        # ── Mod Seçimi ──────────────────────────────────────────────────────
        mf = QFrame()
        mf.setStyleSheet(
            "QFrame{background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;}"
        )
        ml = QVBoxLayout(mf)
        ml.setContentsMargins(16, 14, 16, 14)
        ml.setSpacing(10)
        ml.addWidget(_mk_lbl(
            "Bağlantı Modu",
            "font-size:12px;font-weight:600;color:#64748b;"
        ))
        br2 = QHBoxLayout()
        br2.setSpacing(10)
        self._sqlite_btn = QPushButton("💻  Lokal (SQLite)")
        self._sqlite_btn.setCheckable(True)
        self._sqlite_btn.setFixedHeight(44)
        self._sqlite_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._sqlite_btn.clicked.connect(lambda: self._on_mod("sqlite"))
        br2.addWidget(self._sqlite_btn)
        self._pg_btn = QPushButton("🌐  Sunucu (PostgreSQL)")
        self._pg_btn.setCheckable(True)
        self._pg_btn.setFixedHeight(44)
        self._pg_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._pg_btn.clicked.connect(lambda: self._on_mod("postgres"))
        br2.addWidget(self._pg_btn)
        ml.addLayout(br2)
        self._mod_hint = QLabel("Mevcut mod: Lokal (SQLite)")
        self._mod_hint.setStyleSheet("font-size:11px;color:#94a3b8;")
        ml.addWidget(self._mod_hint)
        root.addWidget(mf)

        # ── PostgreSQL Alanları ─────────────────────────────────────────────
        self._pg_frame = QFrame()
        self._pg_frame.setStyleSheet(
            "QFrame{background:#f0f9ff;border:1px solid #bae6fd;border-radius:10px;}"
        )
        pgl = QVBoxLayout(self._pg_frame)
        pgl.setContentsMargins(18, 16, 18, 16)
        pgl.setSpacing(12)

        pgl.addWidget(_mk_lbl(
            "🌐  PostgreSQL Sunucu Bilgileri",
            "font-size:13px;font-weight:700;color:#0c4a6e;"
        ))
        pgl.addWidget(_mk_lbl(
            "Neon, Supabase veya başka bir PostgreSQL servisinin bağlantı cümleciğini\n"
            "aşağıya yapıştırın — tüm alanlar otomatik dolar.",
            "font-size:11px;color:#0369a1;"
        ))

        _INP = (
            "QLineEdit{background:white;border:1.5px solid #bae6fd;border-radius:8px;"
            "padding:0 10px;font-size:13px;color:#000000;}"
            "QLineEdit:focus{border-color:#0ea5e9;}"
        )
        _LBL = "font-size:11px;font-weight:600;color:#000000;"

        # ── Bağlantı Cümleciği (hızlı doldur) ──────────────────────────────
        dsnc = QVBoxLayout()
        dsnc.setSpacing(4)
        dsnc.addWidget(_mk_lbl("Bağlantı Cümleciği (Connection String)", _LBL))
        dsn_row = QHBoxLayout()
        dsn_row.setSpacing(8)
        self._dsn_inp = QLineEdit()
        self._dsn_inp.setFixedHeight(36)
        self._dsn_inp.setPlaceholderText(
            "postgresql://kullanici:sifre@host:5432/dbadi?sslmode=require"
        )
        self._dsn_inp.setStyleSheet(_INP)
        dsn_row.addWidget(self._dsn_inp, 1)
        self._dsn_btn = QPushButton("⚡  Doldur")
        self._dsn_btn.setFixedHeight(36)
        self._dsn_btn.setFixedWidth(90)
        self._dsn_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._dsn_btn.setStyleSheet(
            "QPushButton{background:#0ea5e9;color:white;border:none;"
            "border-radius:8px;font-size:12px;font-weight:700;}"
            "QPushButton:hover{background:#0284c7;}"
        )
        self._dsn_btn.clicked.connect(self._on_dsn_doldur)
        dsn_row.addWidget(self._dsn_btn)
        dsnc.addLayout(dsn_row)
        self._dsn_hint = QLabel("")
        self._dsn_hint.setStyleSheet("font-size:11px;color:#dc2626;")
        dsnc.addWidget(self._dsn_hint)
        pgl.addLayout(dsnc)

        # Ayırıcı çizgi
        _sep = QFrame()
        _sep.setFrameShape(QFrame.Shape.HLine)
        _sep.setStyleSheet("color:#bae6fd;")
        pgl.addWidget(_sep)

        # Host + Port
        r1 = QHBoxLayout()
        r1.setSpacing(10)
        hc = QVBoxLayout()
        hc.addWidget(_mk_lbl("Sunucu (Host)", _LBL))
        self._host_inp = QLineEdit()
        self._host_inp.setFixedHeight(36)
        self._host_inp.setPlaceholderText("aws-0-eu-central-1.pooler.supabase.com")
        self._host_inp.setStyleSheet(_INP)
        hc.addWidget(self._host_inp)
        r1.addLayout(hc, 3)
        pc = QVBoxLayout()
        pc.addWidget(_mk_lbl("Port", _LBL))
        self._port_inp = QLineEdit()
        self._port_inp.setFixedHeight(36)
        self._port_inp.setPlaceholderText("5432")
        self._port_inp.setStyleSheet(_INP)
        pc.addWidget(self._port_inp)
        r1.addLayout(pc, 1)
        pgl.addLayout(r1)

        # DB + User
        r2 = QHBoxLayout()
        r2.setSpacing(10)
        dc = QVBoxLayout()
        dc.addWidget(_mk_lbl("Veritabanı Adı", _LBL))
        self._db_inp = QLineEdit()
        self._db_inp.setFixedHeight(36)
        self._db_inp.setPlaceholderText("postgres")
        self._db_inp.setStyleSheet(_INP)
        dc.addWidget(self._db_inp)
        r2.addLayout(dc)
        uc = QVBoxLayout()
        uc.addWidget(_mk_lbl("Kullanıcı Adı", _LBL))
        self._user_inp = QLineEdit()
        self._user_inp.setFixedHeight(36)
        self._user_inp.setPlaceholderText("postgres.proje-id")
        self._user_inp.setStyleSheet(_INP)
        uc.addWidget(self._user_inp)
        r2.addLayout(uc)
        pgl.addLayout(r2)

        # Şifre + SSL
        r3 = QHBoxLayout()
        r3.setSpacing(10)
        pasc = QVBoxLayout()
        pasc.addWidget(_mk_lbl("Şifre", _LBL))
        self._pass_inp = QLineEdit()
        self._pass_inp.setFixedHeight(36)
        self._pass_inp.setEchoMode(QLineEdit.EchoMode.Password)
        self._pass_inp.setPlaceholderText("••••••••••••")
        self._pass_inp.setStyleSheet(_INP)
        pasc.addWidget(self._pass_inp)
        r3.addLayout(pasc, 3)
        sslc = QVBoxLayout()
        sslc.addWidget(_mk_lbl("SSL Modu", _LBL))
        self._ssl_combo = QComboBox()
        self._ssl_combo.addItems(["require", "prefer", "disable"])
        self._ssl_combo.setFixedHeight(36)
        self._ssl_combo.setStyleSheet(
            "QComboBox{background:white;border:1.5px solid #bae6fd;"
            "border-radius:8px;padding:0 8px;font-size:13px;color:#000000;}"
            "QComboBox:focus{border-color:#0ea5e9;}"
        )
        sslc.addWidget(self._ssl_combo)
        r3.addLayout(sslc, 1)
        pgl.addLayout(r3)

        # Test butonu + sonuç
        tr = QHBoxLayout()
        tr.setSpacing(10)
        self._test_btn = QPushButton("🔌  Bağlantıyı Test Et")
        self._test_btn.setFixedHeight(36)
        self._test_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._test_btn.setStyleSheet(
            "QPushButton{background:#0ea5e9;color:white;border:none;"
            "border-radius:8px;font-size:13px;font-weight:600;padding:0 16px;}"
            "QPushButton:hover{background:#0284c7;}"
            "QPushButton:disabled{background:#cbd5e1;color:#94a3b8;}"
        )
        self._test_btn.clicked.connect(self._on_test)
        tr.addWidget(self._test_btn)
        self._test_result = QLabel("")
        self._test_result.setWordWrap(True)
        self._test_result.setStyleSheet("font-size:12px;color:#000000;")
        tr.addWidget(self._test_result, 1)
        pgl.addLayout(tr)

        root.addWidget(self._pg_frame)

        # ── Kaydet ──────────────────────────────────────────────────────────
        kr = QHBoxLayout()
        kr.setSpacing(10)
        self._kaydet_btn = QPushButton("💾  Kaydet")
        self._kaydet_btn.setFixedHeight(40)
        self._kaydet_btn.setMinimumWidth(140)
        self._kaydet_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._kaydet_btn.setStyleSheet(
            "QPushButton{background:#6366f1;color:white;border:none;"
            "border-radius:10px;font-size:13px;font-weight:700;letter-spacing:.5px;}"
            "QPushButton:hover{background:#4f46e5;}"
        )
        self._kaydet_btn.clicked.connect(self._on_kaydet)
        kr.addWidget(self._kaydet_btn)
        self._kaydet_durum = QLabel("")
        self._kaydet_durum.setStyleSheet("font-size:12px;color:#000000;")
        kr.addWidget(self._kaydet_durum, 1)
        root.addLayout(kr)

        # Ayırıcı çizgi
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color:#e2e8f0;")
        root.addWidget(sep)

        # ── Migrasyon Bölümü ────────────────────────────────────────────────
        tf = QFrame()
        tf.setStyleSheet(
            "QFrame{background:#fefce8;border:1px solid #fde68a;border-radius:10px;}"
        )
        tl = QVBoxLayout(tf)
        tl.setContentsMargins(18, 14, 18, 14)
        tl.setSpacing(10)
        tl.addWidget(_mk_lbl(
            "🚚  Mevcut Veriyi PostgreSQL'e Taşı",
            "font-size:13px;font-weight:700;color:#92400e;"
        ))
        tl.addWidget(_mk_lbl(
            "SQLite'daki tüm kayıtları PostgreSQL'e kopyalar.  Önce bağlantıyı "
            "kaydedin ve test edin.  Mevcut veriler kaybolmaz — sadece kopyalanır.  "
            "Yeniden taşıma yapılsa bile duplicate oluşmaz (ON CONFLICT DO NOTHING).",
            "font-size:11px;color:#78350f;"
        ))

        pr = QHBoxLayout()
        pr.setSpacing(8)
        pr.addWidget(_mk_lbl(
            "📦  Paket boyutu:",
            "font-size:12px;font-weight:600;color:#92400e;"
        ))
        self._batch_combo = QComboBox()
        self._batch_combo.addItems([
            "50  (Çok yavaş bağlantı / Supabase free)",
            "100  (Yavaş bağlantı)",
            "250  (Normal — varsayılan)",
            "500  (Hızlı bağlantı)",
            "1000  (Çok hızlı / LAN)"
        ])
        self._batch_combo.setCurrentIndex(2)
        self._batch_combo.setFixedHeight(32)
        self._batch_combo.setStyleSheet(
            "QComboBox{background:white;border:1.5px solid #fde68a;"
            "border-radius:7px;padding:0 8px;font-size:12px;color:#000000;}"
            "QComboBox:focus{border-color:#d97706;}"
        )
        pr.addWidget(self._batch_combo)
        pr.addStretch()
        tl.addLayout(pr)

        self._tasima_btn = QPushButton("🚀  Veriyi PostgreSQL'e Taşı")
        self._tasima_btn.setFixedHeight(40)
        self._tasima_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._tasima_btn.setStyleSheet(
            "QPushButton{background:#d97706;color:white;border:none;"
            "border-radius:9px;font-size:13px;font-weight:700;}"
            "QPushButton:hover{background:#b45309;}"
        )
        self._tasima_btn.clicked.connect(self._on_tasima)
        tl.addWidget(self._tasima_btn)
        root.addWidget(tf)

        self._update_mod_ui("sqlite")

    # ── Yardımcı metodlar ────────────────────────────────────────────────────

    def _update_mod_ui(self, mod):
        self._current_mod = mod
        ACT = (
            "QPushButton{background:#6366f1;color:white;border:none;"
            "border-radius:9px;font-size:13px;font-weight:700;}"
        )
        IDL = (
            "QPushButton{background:#f1f5f9;color:#475569;"
            "border:1.5px solid #e2e8f0;border-radius:9px;"
            "font-size:13px;font-weight:600;}"
            "QPushButton:hover{background:#e2e8f0;}"
        )
        if mod == "sqlite":
            self._sqlite_btn.setStyleSheet(ACT)
            self._sqlite_btn.setChecked(True)
            self._pg_btn.setStyleSheet(IDL)
            self._pg_btn.setChecked(False)
            self._mod_hint.setText("Mevcut mod: Lokal (SQLite) — yalnızca bu bilgisayar")
        else:
            self._pg_btn.setStyleSheet(ACT)
            self._pg_btn.setChecked(True)
            self._sqlite_btn.setStyleSheet(IDL)
            self._sqlite_btn.setChecked(False)
            self._mod_hint.setText("Mevcut mod: Sunucu (PostgreSQL) — çok kullanıcılı")
        if hasattr(self, "_pg_frame"):
            self._pg_frame.setVisible(mod == "postgres")

    def _load(self):
        from db.db_config import load_config
        cfg = load_config()
        self._update_mod_ui(cfg.get("mode", "sqlite"))
        host    = cfg.get("pg_host", "")
        port    = str(cfg.get("pg_port", 5432))
        db      = cfg.get("pg_db", "postgres")
        user    = cfg.get("pg_user", "postgres")
        pw      = cfg.get("pg_pass", "")
        sslmode = cfg.get("pg_sslmode", "require")
        self._host_inp.setText(host)
        self._port_inp.setText(port)
        self._db_inp.setText(db)
        self._user_inp.setText(user)
        self._pass_inp.setText(pw)
        idx = self._ssl_combo.findText(sslmode)
        if idx >= 0:
            self._ssl_combo.setCurrentIndex(idx)
        # DSN alanını da hazır doldur
        if host and user:
            from urllib.parse import quote
            pw_enc = quote(pw, safe="") if pw else ""
            dsn = f"postgresql://{user}:{pw_enc}@{host}:{port}/{db}?sslmode={sslmode}"
            self._dsn_inp.setText(dsn)

    def _on_dsn_doldur(self):
        """Bağlantı cümleciğini parse edip alanları doldurur ve otomatik kaydeder."""
        from urllib.parse import urlparse, parse_qs
        raw = self._dsn_inp.text().strip()
        if not raw:
            self._dsn_hint.setText("⚠  Lütfen bir bağlantı cümleciği girin.")
            return
        try:
            # postgresql:// veya postgres:// her ikisini de destekle
            if raw.startswith("postgres://"):
                raw = "postgresql" + raw[8:]
            parsed = urlparse(raw)
            host = parsed.hostname or ""
            port = str(parsed.port) if parsed.port else "5432"
            db   = (parsed.path or "/").lstrip("/")
            user = parsed.username or ""
            pw   = parsed.password or ""
            qs   = parse_qs(parsed.query)
            ssl  = qs.get("sslmode", ["require"])[0]
            if not host or not user:
                raise ValueError("Host veya kullanıcı adı boş")
            self._host_inp.setText(host)
            self._port_inp.setText(port)
            self._db_inp.setText(db)
            self._user_inp.setText(user)
            self._pass_inp.setText(pw)
            idx = self._ssl_combo.findText(ssl)
            self._ssl_combo.setCurrentIndex(idx if idx >= 0 else 0)
            # Modu postgres yap ve otomatik kaydet
            self._update_mod_ui("postgres")
            self._on_kaydet()
            self._dsn_hint.setStyleSheet("font-size:11px;color:#059669;font-weight:600;")
            self._dsn_hint.setText("✅  Alanlar dolduruldu ve kaydedildi.")
        except Exception as e:
            self._dsn_hint.setStyleSheet("font-size:11px;color:#dc2626;")
            self._dsn_hint.setText(f"❌  Geçersiz format: {e}")

    def _on_mod(self, mod):
        self._update_mod_ui(mod)

    def _on_test(self):
        h = self._host_inp.text().strip()
        db = self._db_inp.text().strip()
        u = self._user_inp.text().strip()
        if not h or not db or not u:
            self._test_result.setText("❌  Host, DB Adı ve Kullanıcı zorunludur.")
            self._test_result.setStyleSheet("font-size:12px;color:#dc2626;")
            return
        try:
            port = int(self._port_inp.text().strip() or "5432")
        except ValueError:
            self._test_result.setText("❌  Port sayı olmalıdır.")
            self._test_result.setStyleSheet("font-size:12px;color:#dc2626;")
            return
        self._test_btn.setEnabled(False)
        self._test_result.setText("⏳  Bağlanılıyor...")
        self._test_result.setStyleSheet("font-size:12px;color:#6366f1;")
        self._test_worker = VeriTabaniTestWorker(
            h, port, db, u, self._pass_inp.text(), self._ssl_combo.currentText()
        )
        self._test_worker.islendi.connect(self._on_test_done)
        self._test_worker.start()

    def _on_test_done(self, res):
        self._test_btn.setEnabled(True)
        if res.get("success"):
            msg = f"✅  {res['message']}  ({res['ver']})"
            self._test_result.setText(msg)
            self._test_result.setStyleSheet(
                "font-size:12px;color:#059669;font-weight:600;"
            )
        else:
            self._test_result.setText(f"❌  {res['message']}")
            self._test_result.setStyleSheet("font-size:12px;color:#dc2626;")

    def _on_kaydet(self):
        from db.db_config import load_config, save_config
        mod = self._current_mod
        cfg = load_config()
        old_mod = cfg.get("mode", "sqlite")
        try:
            port = int(self._port_inp.text().strip() or "5432")
        except ValueError:
            port = 5432
        cfg.update({
            "mode":       mod,
            "pg_host":    self._host_inp.text().strip(),
            "pg_port":    port,
            "pg_db":      self._db_inp.text().strip(),
            "pg_user":    self._user_inp.text().strip(),
            "pg_pass":    self._pass_inp.text(),
            "pg_sslmode": self._ssl_combo.currentText(),
        })
        save_config(cfg)
        if old_mod != mod:
            restart = SweetConfirmDialog.confirm(
                self,
                "Mod Değiştirildi",
                f"Veritabanı modu değiştirildi:\n  {old_mod.upper()} → {mod.upper()}\n\n"
                "Değişikliğin geçerli olması için uygulama yeniden başlatılmalı.\n"
                "Şimdi kapatılsın mı?",
                confirm_text="Evet, Kapat",
                cancel_text="Sonra",
                is_danger=False
            )
            if restart:
                from PyQt6.QtWidgets import QApplication
                QApplication.quit()
            else:
                self._kaydet_durum.setText("✅  Kaydedildi — yeniden başlatınca aktif olur.")
                self._kaydet_durum.setStyleSheet("font-size:12px;color:#d97706;font-weight:600;")
        else:
            self._kaydet_durum.setText("✅  Kaydedildi.")
            self._kaydet_durum.setStyleSheet("font-size:12px;color:#059669;font-weight:600;")

    def _on_tasima(self):
        from db.db_config import load_config
        cfg = load_config()
        if cfg.get("mode") != "postgres" or not cfg.get("pg_host"):
            SweetConfirmDialog.confirm(
                self,
                "Önce Bağlantı Gerekli",
                "PostgreSQL bağlantı bilgilerini girin,\n"
                "bağlantıyı test edin ve kaydedin.",
                confirm_text="Tamam",
                cancel_text="",
                is_danger=False
            )
            return
        _MAP = {0: 50, 1: 100, 2: 250, 3: 500, 4: 1000}
        bs = _MAP.get(self._batch_combo.currentIndex(), 250)
        if not SweetConfirmDialog.confirm(
            self,
            "Veri Taşıma",
            f"SQLite'daki tüm kayıtlar PostgreSQL'e kopyalanacak.\n\n"
            f"📦  Paket boyutu: {bs} satır\n"
            "Bu işlem birkaç dakika sürebilir.",
            confirm_text="Evet, Taşı",
            cancel_text="Vazgeç",
            is_danger=False
        ):
            return
        dlg = MigrasyonDialog(self)
        dlg.show()
        dlg.start(batch_size=bs)

    def refresh(self):
        self._load()


# ─────────────────────────────────────────────────────────────────────────────
# ── VOMSİS POS Excel Import — Worker + Dialog ────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

class WomsisPosExcelWorker(QThread):
    """
    POS Excel dosyasını arka planda işler → womsi_pos tablosuna yazar.
    PyQt6 gereği modül seviyesinde tanımlanmıştır.

    Hem kendi şablonumuzdaki büyük harfli başlıkları (İŞLEM TARİHİ, İŞLEM TUTARI…)
    hem de gerçek VOMSİS çıktısındaki başlıkları (İşlem Tarihi, Brüt Tutar, Terminal No…)
    case-insensitive normalize ederek eşler.
    """
    satir_done   = pyqtSignal(str, int, int, str)   # sayfa, eklenen, atlanan, durum
    header_error = pyqtSignal(str, object, object)  # sayfa, bulunan_list, eksik_list
    skipped_row  = pyqtSignal(dict)                 # atlanan satır detayı
    progress     = pyqtSignal(str)
    finished     = pyqtSignal(dict)

    # Zorunlu olan canonical (normalize) kolon adları
    ZORUNLU_KANONLAR = ["islemtarihi", "islemtutari", "nettutar"]

    # Normalize edilmiş başlık → DB alan adı eşlemesi
    # (normalize = küçük harf + boşluk/noktalama silinmiş)
    KOLON_MAP = {
        # Şablon formatı (büyük harf)
        "islemtarihi":         "islemtarihi",
        "islemtutari":         "islemtutari",
        "nettutar":            "nettutar",
        "isyeriucrettutari":   "isyeriucretitutar",
        "isyeriucretitutar":   "isyeriucretitutar",
        "posno":               "posno",
        "kartno":              "kartno",
        "brand":               "brand",
        "marka":               "brand",
        "islemtipi":           "islemtipi",
        "aciklama":            "aciklama",
        "isyerino":            "isyerino",
        "carihesap":           "carihesap",
        "hesabagecistarihi":   "hesabagecistarihi",
        # Gerçek VOMSİS POS Excel dosyası başlıkları
        "islemtarihi":         "islemtarihi",       # İşlem Tarihi
        "isyerinumarasi":      "isyerino",          # İşyeri Numarası
        "terminalno":          "posno",             # Terminal No
        "valor":               "hesabagecistarihi", # Valör
        "islem":               "aciklama",          # İşlem
        "bruttutar":           "islemtutari",       # Brüt Tutar
        "kartnumarasi":        "kartno",            # Kart Numarası
        "altkarttip":          "brand",             # Alt Kart Tipi
        "altkarttipi":         "brand",             # Alt Kart Tipi (normalize edilmiş)
        "altkarttıpı":         "brand",
        "komisyon":            "isyeriucretitutar", # Komisyon
        "komisyonoran":        "",                  # Komisyon Oranı — atla
        "komisyonorani":       "",                  # Komisyon Oranı (normalize edilmiş)
        "kur":                 "",                  # Kur — atla
        "onayno":              "carihesap",         # Onay No → cariHesap
    }

    def __init__(self, filepath: str, userid: int):
        super().__init__()
        self._path   = filepath
        self._userid = userid

    @staticmethod
    def _normalize(s: str) -> str:
        """Başlığı küçük harfe çevirir, boşluk/noktalama/özel karakter siler."""
        import unicodedata
        s = str(s).strip().lower()
        # Türkçe karakterleri ASCII karşılıklarına çevir
        tr_map = str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosucgiosu")
        s = s.translate(tr_map)
        # Boşluk ve noktalama sil
        s = "".join(c for c in s if c.isalnum())
        return s

    def run(self):
        try:
            import openpyxl
        except ImportError:
            self.finished.emit({
                "success": False,
                "message": "openpyxl yüklü değil. pip install openpyxl",
                "toplam": 0, "eklenen": 0, "atlanan": 0, "hata": 1,
            })
            return

        from db.database import get_connection
        from services.fiziksel_pos_service import ensure_tables
        ensure_tables()

        try:
            wb = openpyxl.load_workbook(self._path, data_only=True)
        except Exception as e:
            self.finished.emit({
                "success": False,
                "message": f"Excel dosyası açılamadı: {e}",
                "toplam": 0, "eklenen": 0, "atlanan": 0, "hata": 1,
            })
            return

        conn = get_connection()
        g_toplam = g_eklenen = g_atlanan = g_hata = 0

        try:
            for sheet_name in wb.sheetnames:
                ws  = wb[sheet_name]
                ins = skip = err = 0

                # ── Başlık Satırı Bul ─────────────────────────────────────────
                # İlk 10 satırda normalize başlık kontrolü yap
                header_row = None
                col_map    = {}   # db_alan_adi → kolon_indeksi (0-tabanlı)

                for r_idx, row in enumerate(
                    ws.iter_rows(min_row=1, max_row=10, values_only=True), 1
                ):
                    if not row:
                        continue
                    # Herhangi bir hücre normalize edilince 'islemtarihi' oluyorsa başlık satırı
                    norm_vals = [
                        self._normalize(str(v)) for v in row if v is not None
                    ]
                    if "islemtarihi" in norm_vals:
                        header_row = r_idx
                        for i, v in enumerate(row):
                            if v is None:
                                continue
                            norm = self._normalize(str(v))
                            db_alan = self.KOLON_MAP.get(norm, "")
                            if db_alan:  # boş string → atla
                                # İlk eşleşme kazanır
                                if db_alan not in col_map:
                                    col_map[db_alan] = i
                        break

                if header_row is None:
                    bulunan = [
                        str(v) for row in ws.iter_rows(
                            min_row=1, max_row=3, values_only=True
                        ) for v in (row or []) if v
                    ]
                    self.header_error.emit(
                        sheet_name, bulunan[:12],
                        ["islemtarihi", "islemtutari", "nettutar"]
                    )
                    self.satir_done.emit(
                        sheet_name, 0, 0, "❌ Başlık satırı bulunamadı"
                    )
                    g_hata += 1
                    continue

                eksik_kanonlar = [
                    k for k in self.ZORUNLU_KANONLAR if k not in col_map
                ]
                if eksik_kanonlar:
                    bulunan_kolonlar = list(col_map.keys())
                    self.header_error.emit(
                        sheet_name, bulunan_kolonlar, eksik_kanonlar
                    )
                    self.satir_done.emit(
                        sheet_name, 0, 0,
                        f"❌ Eksik kolon: {', '.join(eksik_kanonlar)}"
                    )
                    g_hata += 1
                    continue

                self.progress.emit(f"📂  {sheet_name} işleniyor...")

                # ── Veri Satırları ────────────────────────────────────────────
                for row in ws.iter_rows(
                    min_row=header_row + 1, values_only=True
                ):
                    if not row or all(v is None for v in row):
                        continue

                    def _get(db_alan, default=None, _row=row, _map=col_map):
                        idx = _map.get(db_alan)
                        if idx is None or idx >= len(_row):
                            return default
                        return _row[idx]

                    islem_tarihi_raw = _get("islemtarihi",       "")
                    islem_tutari_raw = _get("islemtutari",       0)
                    net_tutar_raw    = _get("nettutar",          0)
                    isyeri_ucret_raw = _get("isyeriucretitutar", 0)
                    posno            = str(_get("posno",    "") or "")
                    kartno           = str(_get("kartno",   "") or "")
                    brand            = str(_get("brand",    "") or "")
                    islem_tipi       = str(_get("islemtipi", "") or "")
                    aciklama         = str(_get("aciklama",  "") or "")
                    isyerino         = str(_get("isyerino",  "") or "")
                    carihesap        = str(_get("carihesap", "") or "")
                    gecis_tarihi_raw = _get("hesabagecistarihi", "")

                    g_toplam += 1

                    # ── Tutar parse ───────────────────────────────────────────
                    def _parse_tutar(v):
                        if v is None:
                            return 0.0
                        try:
                            return float(
                                str(v)
                                .replace("\xa0", "")
                                .replace(" ", "")
                                .replace(",", ".")
                            )
                        except Exception:
                            return 0.0

                    islem_tutari = _parse_tutar(islem_tutari_raw)
                    net_tutar    = _parse_tutar(net_tutar_raw)
                    isyeri_ucret = _parse_tutar(isyeri_ucret_raw)

                    # ── Tarih parse ───────────────────────────────────────────
                    def _parse_tarih(raw):
                        if not raw:
                            return ""
                        t = str(raw).strip()
                        for fmt in (
                            "%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d",
                            "%d.%m.%Y %H:%M", "%d/%m/%Y %H:%M",
                            "%Y-%m-%d %H:%M:%S",
                        ):
                            try:
                                return datetime.datetime.strptime(
                                    t[:len(fmt)], fmt
                                ).strftime("%Y-%m-%d")
                            except Exception:
                                continue
                        return t[:10]

                    islem_tarihi = _parse_tarih(islem_tarihi_raw)
                    gecis_tarihi = _parse_tarih(gecis_tarihi_raw)

                    if not islem_tarihi:
                        err += 1; g_hata += 1
                        continue

                    # ── Mükerrer Kontrol ──────────────────────────────────────
                    exists = None
                    try:
                        exists = conn.execute(
                            """SELECT id FROM womsi_pos
                               WHERE userid=? AND islemTarihi=?
                                 AND ABS(islemTutari - ?) < 0.01
                                 AND posNo=?
                               LIMIT 1""",
                            (self._userid, islem_tarihi,
                             abs(islem_tutari), posno)
                        ).fetchone()
                    except Exception:
                        exists = None

                    if exists:
                        skip += 1; g_atlanan += 1
                        # Atlanan satır detayını bildir
                        self.skipped_row.emit({
                            "sayfa":      sheet_name,
                            "tarih":      islem_tarihi,
                            "tutar":      islem_tutari,
                            "net_tutar":  net_tutar,
                            "posno":      posno,
                            "kartno":     kartno,
                            "brand":      brand,
                            "islem_tipi": islem_tipi,
                            "aciklama":   aciklama,
                            "neden":      "Mükerrer — DB'de mevcut",
                        })
                        if (ins + skip + err) % 50 == 0:
                            self.progress.emit(
                                f"⏳ {sheet_name}  —  "
                                f"✅ {ins} eklendi   "
                                f"⏭️ {skip} atlandı   "
                                f"❌ {err} hata"
                            )
                        continue

                    # ── Kayıt Ekle ────────────────────────────────────────────
                    try:
                        conn.execute(
                            """INSERT INTO womsi_pos
                               (userid, isyeriNo, cariHesap, hesabaGecisTarihi,
                                islemTutari, islemTarihi, posNo, isyeriUcretiTutar,
                                netTutar, brand, kartNo, islemTipi, aciklama)
                               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                            (
                                self._userid,
                                isyerino,
                                carihesap,
                                gecis_tarihi,
                                islem_tutari,
                                islem_tarihi,
                                posno,
                                isyeri_ucret,
                                net_tutar,
                                brand,
                                kartno,
                                islem_tipi,
                                aciklama,
                            )
                        )
                        ins += 1; g_eklenen += 1
                    except Exception as ex:
                        err += 1; g_hata += 1

                    if (ins + skip + err) % 50 == 0:
                        self.progress.emit(
                            f"⏳ {sheet_name}  —  "
                            f"✅ {ins} eklendi   "
                            f"⏭️ {skip} atlandı   "
                            f"❌ {err} hata"
                        )

                conn.commit()
                durum = "✅ Tamam" if err == 0 else f"⚠️ {err} hata"
                self.satir_done.emit(sheet_name, ins, skip, durum)
                g_atlanan += skip
                g_hata    += err

        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            self.finished.emit({
                "success": False,
                "message": f"İşlem hatası: {e}",
                "toplam": g_toplam, "eklenen": g_eklenen,
                "atlanan": g_atlanan, "hata": g_hata,
            })
            return
        finally:
            try:
                conn.close()
            except Exception:
                pass

        basarili = g_hata == 0
        self.finished.emit({
            "success": basarili,
            "message": (
                f"{g_eklenen} POS hareketi eklendi"
                + (f", {g_atlanan} atlandı" if g_atlanan else "")
                + (f", {g_hata} hata" if g_hata else "")
                + "."
            ),
            "toplam": g_toplam, "eklenen": g_eklenen,
            "atlanan": g_atlanan, "hata": g_hata,
        })


# ─────────────────────────────────────────────────────────────────────────────
# ── VOMSİS POS Excel Import Dialog ───────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

class WomsisPosExcelImportDialog(QDialog):
    """
    WomsisExcelImportDialog'ın POS versiyonu.
    Aynı UI/UX, womsi_pos tablosuna yazar.
    Zorunlu kolonlar: İŞLEM TARİHİ, İŞLEM TUTARI, NET TUTAR
    """

    def __init__(self, userid: int, musterino: int = 1, parent=None):
        super().__init__(parent)
        self._userid  = userid
        self._worker: WomsisPosExcelWorker | None = None
        self.setWindowTitle("💳  VOMSİS POS Hareketleri — Excel Import")
        self.setMinimumSize(800, 580)
        self.resize(950, 640)
        self.setModal(True)
        self.setStyleSheet("""
            QDialog { background: #f8fafc; }
            QLabel  { background: transparent; font-family: 'Segoe UI', Arial, sans-serif; }
        """)
        self._build()

    # ── UI ───────────────────────────────────────────────────────────────────

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Üst banner ──
        banner = QFrame()
        banner.setFixedHeight(56)
        banner.setStyleSheet(
            "background: qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "stop:0 #059669, stop:1 #047857);"
        )
        b_lay = QHBoxLayout(banner)
        b_lay.setContentsMargins(20, 0, 20, 0)
        b_lbl = QLabel("💳  VOMSİS POS Hareketleri — Excel Import")
        b_lbl.setStyleSheet("color:white;font-size:15px;font-weight:700;")
        b_lay.addWidget(b_lbl)
        b_lay.addStretch()
        root.addWidget(banner)

        # ── Tab çubuğu ──
        tab_bar = QFrame()
        tab_bar.setFixedHeight(40)
        tab_bar.setStyleSheet("background:#f1f5f9;border-bottom:2px solid #e2e8f0;")
        t_lay = QHBoxLayout(tab_bar)
        t_lay.setContentsMargins(16, 0, 16, 0)
        t_lay.setSpacing(0)

        def _tab_btn(text, checked=False):
            b = QPushButton(text)
            b.setCheckable(True)
            b.setChecked(checked)
            b.setStyleSheet("""
                QPushButton {
                    background:transparent;border:none;
                    font-size:12px;font-weight:600;
                    color:#64748b;padding:0 16px;
                }
                QPushButton:checked {
                    color:#059669;
                    border-bottom:2px solid #059669;
                }
                QPushButton:hover:!checked { color:#374151; }
            """)
            return b

        self._tab_sablon = _tab_btn("📋  Şablon & Yükle", checked=True)
        self._tab_sonuc  = _tab_btn("📊  Sonuçlar")
        t_lay.addWidget(self._tab_sablon)
        t_lay.addWidget(self._tab_sonuc)
        t_lay.addStretch()
        root.addWidget(tab_bar)

        # ── Sayfa yığını ──
        self._stack = QStackedWidget()
        root.addWidget(self._stack, 1)

        self._stack.addWidget(self._build_sablon_page())
        self._stack.addWidget(self._build_sonuc_page())

        self._tab_sablon.clicked.connect(lambda: self._switch_tab(0))
        self._tab_sonuc.clicked.connect(lambda: self._switch_tab(1))

    def _build_sablon_page(self) -> QWidget:
        page = QWidget()
        vlay = QVBoxLayout(page)
        vlay.setContentsMargins(24, 20, 24, 20)
        vlay.setSpacing(16)

        # Bilgi kutusu
        info = QFrame()
        info.setStyleSheet(
            "background:#f0fdf4;border:1.5px solid #bbf7d0;border-radius:10px;"
        )
        i_lay = QVBoxLayout(info)
        i_lay.setContentsMargins(16, 12, 16, 12)
        i_lbl = QLabel(
            "📋  <b>POS Hareketleri Excel Şablonu</b><br>"
            "Zorunlu kolonlar: <b>İŞLEM TARİHİ · İŞLEM TUTARI · NET TUTAR</b><br>"
            "İsteğe bağlı: POS NO · KART NO · BRAND · İŞLEM TİPİ · "
            "İŞYERİ ÜCRET TUTARI · AÇIKLAMA · HESABA GEÇİŞ TARİHİ"
        )
        i_lbl.setStyleSheet("font-size:12px;color:#065f46;")
        i_lbl.setWordWrap(True)
        i_lbl.setTextFormat(Qt.TextFormat.RichText)
        i_lay.addWidget(i_lbl)
        vlay.addWidget(info)

        # Şablon indirme butonu
        self._sablon_indir_btn = QPushButton("⬇️  Şablon Excel İndir")
        self._sablon_indir_btn.setFixedHeight(36)
        self._sablon_indir_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._sablon_indir_btn.setStyleSheet(self._btn_style("#0891b2"))
        self._sablon_indir_btn.clicked.connect(self._on_sablon_indir)
        vlay.addWidget(self._sablon_indir_btn)

        # Dosya seçimi
        dosya_row = QHBoxLayout()
        self._dosya_lbl = QLabel("Henüz dosya seçilmedi.")
        self._dosya_lbl.setStyleSheet(
            "font-size:12px;color:#475569;background:#f1f5f9;"
            "border-radius:8px;padding:8px 14px;"
        )
        self._dosya_lbl.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        dosya_row.addWidget(self._dosya_lbl)

        self._dosya_sec_btn = QPushButton("📂  Dosya Seç")
        self._dosya_sec_btn.setFixedHeight(36)
        self._dosya_sec_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._dosya_sec_btn.setStyleSheet(self._btn_style("#059669"))
        self._dosya_sec_btn.clicked.connect(self._on_dosya_sec)
        dosya_row.addWidget(self._dosya_sec_btn)
        vlay.addLayout(dosya_row)

        # İlerleme
        self._progress = QLabel("")
        self._progress.setStyleSheet(
            "font-size:12px;color:#0891b2;background:#f0f9ff;"
            "border-radius:8px;padding:8px 14px;"
        )
        self._progress.hide()
        vlay.addWidget(self._progress)

        vlay.addStretch()

        # Kapat
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        kapat = QPushButton("✕  Kapat")
        kapat.setStyleSheet(self._btn_style("#64748b"))
        kapat.clicked.connect(self.reject)
        btn_row.addWidget(kapat)
        vlay.addLayout(btn_row)
        return page

    def _build_sonuc_page(self) -> QWidget:
        page = QWidget()
        vlay = QVBoxLayout(page)
        vlay.setContentsMargins(24, 20, 24, 20)
        vlay.setSpacing(12)

        # ── Sonuç Mesajı ──
        self._sonuc_lbl = QLabel("")
        self._sonuc_lbl.setStyleSheet(
            "font-size:13px;font-weight:600;color:#065f46;"
            "background:#f0fdf4;border-radius:8px;padding:10px;"
        )
        self._sonuc_lbl.setWordWrap(True)
        vlay.addWidget(self._sonuc_lbl)

        # ── İç Alt-Tab Çubuğu ──
        inner_tab_bar = QFrame()
        inner_tab_bar.setFixedHeight(34)
        inner_tab_bar.setStyleSheet(
            "background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;"
        )
        it_lay = QHBoxLayout(inner_tab_bar)
        it_lay.setContentsMargins(6, 3, 6, 3)
        it_lay.setSpacing(4)

        def _itab(text, checked=False):
            b = QPushButton(text)
            b.setCheckable(True)
            b.setChecked(checked)
            b.setFixedHeight(26)
            b.setStyleSheet("""
                QPushButton {
                    background: transparent; border: none;
                    font-size: 11px; font-weight: 600;
                    color: #64748b; padding: 0 12px; border-radius: 6px;
                }
                QPushButton:checked {
                    background: #059669; color: white;
                }
                QPushButton:hover:!checked { background: #f1f5f9; color: #374151; }
            """)
            return b

        self._itab_ozet    = _itab("📋  Sayfa Özeti", checked=True)
        self._itab_atlanan = _itab("⏭️  Atlanan Kayıtlar")
        it_lay.addWidget(self._itab_ozet)
        it_lay.addWidget(self._itab_atlanan)
        it_lay.addStretch()
        vlay.addWidget(inner_tab_bar)

        # ── İç Stack ──
        self._inner_stack = QStackedWidget()
        vlay.addWidget(self._inner_stack, 1)

        self._itab_ozet.clicked.connect(lambda: self._switch_inner_tab(0))
        self._itab_atlanan.clicked.connect(lambda: self._switch_inner_tab(1))

        # ── İç Sayfa 0: Sayfa Özet Tablosu ──
        p0 = QWidget()
        p0_lay = QVBoxLayout(p0)
        p0_lay.setContentsMargins(0, 0, 0, 0)
        self._log_tbl = QTableWidget()
        self._log_tbl.setColumnCount(4)
        self._log_tbl.setHorizontalHeaderLabels(["Sayfa", "Eklenen", "Atlanan", "Durum"])
        self._log_tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self._log_tbl.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self._log_tbl.verticalHeader().setVisible(False)
        self._log_tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._log_tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._log_tbl.setAlternatingRowColors(True)
        self._log_tbl.setStyleSheet("""
            QTableWidget {
                background: white; border: 1px solid #d1fae5;
                border-radius: 8px;
            }
            QTableWidget::item { padding: 6px 10px; }
            QHeaderView::section {
                background: #059669; color: white;
                font-size: 11px; font-weight: 700; padding: 7px;
                border: none; border-right: 1px solid #047857;
            }
        """)
        p0_lay.addWidget(self._log_tbl)
        self._inner_stack.addWidget(p0)

        # ── İç Sayfa 1: Atlanan Kayıtlar Tablosu ──
        p1 = QWidget()
        p1_lay = QVBoxLayout(p1)
        p1_lay.setContentsMargins(0, 0, 0, 0)
        p1_lay.setSpacing(6)

        info_lbl = QLabel(
            "⚠️  Atlanan kayıtlar — DB'de aynı İşlem Tarihi + Tutar + POS No kombinasyonu "
            "zaten mevcut olduğu için tekrar eklenmedi."
        )
        info_lbl.setWordWrap(True)
        info_lbl.setStyleSheet(
            "background:#fffbeb;color:#92400e;border-radius:6px;"
            "padding:8px 12px;font-size:11px;border:1.5px solid #fde68a;"
        )
        p1_lay.addWidget(info_lbl)

        self._atlanan_tbl = QTableWidget()
        self._atlanan_tbl.setColumnCount(9)
        self._atlanan_tbl.setHorizontalHeaderLabels([
            "Sayfa", "İşlem Tarihi", "Brüt Tutar", "Net Tutar",
            "POS No", "Kart No", "Brand", "İşlem Tipi", "Neden"
        ])
        hdr = self._atlanan_tbl.horizontalHeader()
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)
        self._atlanan_tbl.verticalHeader().setVisible(False)
        self._atlanan_tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._atlanan_tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._atlanan_tbl.setAlternatingRowColors(True)
        self._atlanan_tbl.setSortingEnabled(True)
        self._atlanan_tbl.setStyleSheet("""
            QTableWidget {
                background: white; border: 1px solid #fde68a;
                border-radius: 8px;
            }
            QTableWidget::item { padding: 5px 8px; font-size: 11px; }
            QTableWidget::item:alternate { background: #fffbeb; }
            QHeaderView::section {
                background: #d97706; color: white;
                font-size: 11px; font-weight: 700; padding: 7px;
                border: none; border-right: 1px solid #b45309;
            }
        """)
        p1_lay.addWidget(self._atlanan_tbl, 1)
        self._inner_stack.addWidget(p1)

        # ── Chip'ler ──
        self._chip_row = QHBoxLayout()
        self._chip_row.setSpacing(10)
        chip_toplam,  self._lbl_toplam  = self._make_chip("📦 Toplam",   "0", "#1e40af", "#eff6ff")
        chip_eklenen, self._lbl_eklenen = self._make_chip("✅ Eklenen",  "0", "#059669", "#f0fdf4")
        chip_atlanan, self._lbl_atlanan = self._make_chip("⏭️ Atlanan",  "0", "#d97706", "#fffbeb")
        chip_hata,    self._lbl_hata    = self._make_chip("❌ Hata",      "0", "#dc2626", "#fef2f2")
        for chip in (chip_toplam, chip_eklenen, chip_atlanan, chip_hata):
            self._chip_row.addWidget(chip)
        self._chip_row.addStretch()
        vlay.addLayout(self._chip_row)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        kapat = QPushButton("✕  Kapat")
        kapat.setStyleSheet(self._btn_style("#64748b"))
        kapat.clicked.connect(self.reject)
        btn_row.addWidget(kapat)
        vlay.addLayout(btn_row)
        return page

    def _make_chip(self, label, value, color, bg):
        frame = QFrame()
        frame.setStyleSheet(
            f"QFrame{{background:{bg};border:1.5px solid {color}40;"
            f"border-radius:10px;padding:2px;}}"
        )
        lay = QVBoxLayout(frame)
        lay.setContentsMargins(14, 8, 14, 8)
        lay.setSpacing(2)
        lbl = QLabel(label)
        lbl.setStyleSheet(f"font-size:10px;font-weight:600;color:{color};")
        val = QLabel(value)
        val.setStyleSheet(f"font-size:22px;font-weight:800;color:{color};")
        val.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(lbl)
        lay.addWidget(val)
        frame.setFixedWidth(140)
        return frame, val

    def _set_chips(self, toplam, eklenen, atlanan, hata):
        self._lbl_toplam.setText(toplam)
        self._lbl_eklenen.setText(eklenen)
        self._lbl_atlanan.setText(atlanan)
        self._lbl_hata.setText(hata)

    def _switch_inner_tab(self, idx: int):
        self._itab_ozet.setChecked(idx == 0)
        self._itab_atlanan.setChecked(idx == 1)
        self._inner_stack.setCurrentIndex(idx)

    def _btn_style(self, color):
        return (
            f"QPushButton{{background:{color};color:white;border:none;"
            f"border-radius:9px;font-size:13px;font-weight:600;"
            f"padding:0 18px;letter-spacing:.4px;}}"
            f"QPushButton:hover{{background:{color}cc;}}"
            f"QPushButton:disabled{{background:#cbd5e1;color:#94a3b8;}}"
        )

    def _switch_tab(self, idx):
        self._tab_sablon.setChecked(idx == 0)
        self._tab_sonuc.setChecked(idx == 1)
        self._stack.setCurrentIndex(idx)

    # ── Şablon İndir ─────────────────────────────────────────────────────────

    def _on_sablon_indir(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Hata", "openpyxl yüklü değil.\npip install openpyxl")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "POS Şablon Excel Kaydet",
            "vomsis_pos_hareketi_sablon.xlsx",
            "Excel Dosyaları (*.xlsx)"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "POS Hareketleri"

        headers = [
            "İŞLEM TARİHİ", "İŞLEM TUTARI", "NET TUTAR",
            "İŞYERİ ÜCRET TUTARI", "POS NO", "KART NO",
            "BRAND", "İŞLEM TİPİ", "AÇIKLAMA",
            "İŞYERİ NO", "CARİ HESAP", "HESABA GEÇİŞ TARİHİ",
        ]
        fill   = PatternFill("solid", fgColor="059669")
        font   = Font(bold=True, color="FFFFFF", size=10)
        align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill  = fill
            cell.font  = font
            cell.alignment = align
            ws.column_dimensions[cell.column_letter].width = max(16, len(h) + 4)

        ws.row_dimensions[1].height = 24

        # Örnek veri
        ws.append([
            "2024-01-15", 1500.00, 1455.00, 45.00,
            "POS001", "4111111111111111", "VISA",
            "SATIŞ", "Satış işlemi",
            "12345678", "", "2024-01-16",
        ])

        wb.save(path)
        from PyQt6.QtWidgets import QMessageBox
        QMessageBox.information(self, "Şablon İndirildi", f"✅  Şablon kaydedildi:\n{path}")

    # ── Dosya Seç & Import ────────────────────────────────────────────────────

    def _on_dosya_sec(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "POS Excel Dosyası Seç", "",
            "Excel Dosyaları (*.xlsx *.xls)"
        )
        if not path:
            return
        self._dosya_lbl.setText(path)
        self._start_import(path)

    def _start_import(self, path: str):
        self._dosya_sec_btn.setEnabled(False)
        self._sablon_indir_btn.setEnabled(False)
        self._progress.setText("⏳  İçe aktarma başlatılıyor...")
        self._progress.show()
        self._set_chips("0", "0", "0", "0")
        self._log_tbl.setRowCount(0)
        self._atlanan_tbl.setRowCount(0)   # atlanan listeyi sıfırla
        self._itab_atlanan.setText("⏭️  Atlanan Kayıtlar")  # badge sıfırla
        self._switch_inner_tab(0)          # özet tablosundan başla
        self._switch_tab(1)

        self._worker = WomsisPosExcelWorker(path, self._userid)
        self._worker.satir_done.connect(self._on_satir_done)
        self._worker.header_error.connect(self._on_header_error)
        self._worker.skipped_row.connect(self._on_skipped_row)
        self._worker.progress.connect(self._on_progress)
        self._worker.finished.connect(self._on_import_done)
        self._worker.start()

    def _on_progress(self, msg: str):
        self._progress.setText(msg)
        self._progress.show()

    def _on_satir_done(self, sayfa: str, eklenen: int, atlanan: int, durum: str):
        ri = self._log_tbl.rowCount()
        self._log_tbl.insertRow(ri)
        self._log_tbl.setRowHeight(ri, 28)
        for ci, val in enumerate([sayfa, str(eklenen), str(atlanan), durum]):
            item = QTableWidgetItem(val)
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            if ci == 3 and "❌" in durum:
                item.setForeground(QBrush(QColor("#dc2626")))
            elif ci == 3:
                item.setForeground(QBrush(QColor("#059669")))
            self._log_tbl.setItem(ri, ci, item)

    def _on_header_error(self, sayfa: str, bulunan: object, eksik: object):
        from PyQt6.QtWidgets import QMessageBox
        eksik_str   = ", ".join(eksik) if eksik else "?"
        bulunan_str = ", ".join(bulunan[:8]) if bulunan else "Yok"
        QMessageBox.warning(
            self, "Başlık Hatası",
            f"'{sayfa}' sayfasında zorunlu kolonlar eksik:\n\n"
            f"❌ Eksik: {eksik_str}\n"
            f"📋 Bulunan: {bulunan_str}\n\n"
            "Şablon Excel'i indirip kullanın."
        )

    def _on_skipped_row(self, info: dict):
        """Atlanan her satırı _atlanan_tbl'a ekler ve inner tab badge'i günceller."""
        ri = self._atlanan_tbl.rowCount()
        self._atlanan_tbl.insertRow(ri)
        self._atlanan_tbl.setRowHeight(ri, 26)

        def _fmt_tutar(v):
            try:
                return (
                    f"{float(v):,.2f}"
                    .replace(",", "X").replace(".", ",").replace("X", ".")
                )
            except (TypeError, ValueError):
                return str(v or "")

        vals = [
            info.get("sayfa",      ""),
            info.get("tarih",      ""),
            _fmt_tutar(info.get("tutar",      0)),
            _fmt_tutar(info.get("net_tutar",  0)),
            info.get("posno",      ""),
            info.get("kartno",     ""),
            info.get("brand",      ""),
            info.get("islem_tipi", ""),
            info.get("neden",      "Mükerrer"),
        ]
        CLR_MAP = {
            0: "#64748b",   # sayfa
            2: "#dc3545",   # brüt tutar — kırmızı
            3: "#28a745",   # net tutar — yeşil
            8: "#d97706",   # neden — turuncu
        }
        for ci, txt in enumerate(vals):
            it = QTableWidgetItem(str(txt))
            it.setTextAlignment(
                (Qt.AlignmentFlag.AlignRight if ci in (2, 3) else Qt.AlignmentFlag.AlignCenter)
                | Qt.AlignmentFlag.AlignVCenter
            )
            if ci in CLR_MAP:
                it.setForeground(QColor(CLR_MAP[ci]))
            self._atlanan_tbl.setItem(ri, ci, it)

        # Inner tab badge güncelle — kaç tane atlandı
        count = self._atlanan_tbl.rowCount()
        self._itab_atlanan.setText(f"⏭️  Atlanan Kayıtlar ({count:,})")

    def _on_import_done(self, result: dict):
        self._progress.hide()
        self._dosya_sec_btn.setEnabled(True)
        self._sablon_indir_btn.setEnabled(True)

        self._set_chips(
            f"{result.get('toplam',  0):,}",
            f"{result.get('eklenen', 0):,}",
            f"{result.get('atlanan', 0):,}",
            f"{result.get('hata',    0):,}",
        )

        if result["success"]:
            self._sonuc_lbl.setText(f"✅  {result['message']}")
            self._sonuc_lbl.setStyleSheet(
                "font-size:13px;font-weight:600;color:#065f46;"
                "background:#f0fdf4;border-radius:8px;padding:10px;"
            )
        else:
            self._sonuc_lbl.setText(f"❌  {result['message']}")
            self._sonuc_lbl.setStyleSheet(
                "font-size:13px;font-weight:600;color:#7f1d1d;"
                "background:#fef2f2;border-radius:8px;padding:10px;"
            )
